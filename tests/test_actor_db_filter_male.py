import asyncio
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from mdcx.config.resources import COL_JP, COL_TMDBID, DB_HEADERS
from mdcx.core import tmdb_actor
from mdcx.tools import actor_db_tool

_XML = """<?xml version="1.0" encoding="UTF-8"?>
<actor-mapping>
  <actor>
    <a zh_cn="阿部純子" zh_tw="阿部純子" jp="阿部純子" keyword="阿部純子" tmdb_id="1417328" verified="1" />
    <a zh_cn="阿部涼音" zh_tw="阿部涼音" jp="阿部涼音" keyword="阿部涼音" tmdb_id="1417329" verified="1" />
  </actor>
</actor-mapping>
"""


@pytest.fixture
def _tmp_actor_db(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setattr(tmdb_actor.manager, "data_folder", tmp_path)
    monkeypatch.setattr(tmdb_actor.resources, "actor_db", {})
    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    return userdata / "actor_database.xlsx"


@pytest.fixture(autouse=True)
def _reset_actor_db_row_index():
    with tmdb_actor._ACTOR_DB_ROW_INDEX_LOCK:
        tmdb_actor._ACTOR_DB_ROW_INDEX.clear()


@pytest.fixture
def _avdb_xml(tmp_path: Path):
    path = tmp_path / "mapping.xml"
    path.write_text(_XML, encoding="utf-8")
    return path


def _write_db(path: Path, rows, headers=DB_HEADERS):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _read_rows(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def _mock_tmdb(
    monkeypatch,
    genders: dict[int, int | None],
    identity_names: dict[int, str] | None = None,
    identity_spy: list[int] | None = None,
):
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("http://base", "test-key"))
    calls: list[int] = []

    async def fake_fetch(pid, base_url, api_key, client):
        calls.append(pid)
        return genders.get(pid)

    async def fake_fetch_identity(pid, base_url, api_key, client):
        if identity_spy is not None:
            identity_spy.append(pid)
        if identity_names is None:
            return None  # 与旧行为一致：身份不可知 -> 放行
        name = identity_names.get(pid)
        if name is None:
            return None
        return {"gender": None, "name": name, "original_name": name, "also_known_as": []}

    monkeypatch.setattr(actor_db_tool, "fetch_person_gender", fake_fetch)
    monkeypatch.setattr(actor_db_tool, "fetch_person_identity", fake_fetch_identity)
    return calls


def test_sync_skips_male_when_filter_male(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_tmdb(monkeypatch, {1417328: 2, 1417329: 1})
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 1
    assert result.created == 1
    rows = _read_rows(_tmp_actor_db)
    assert len(rows) == 1
    assert rows[0][COL_JP] == "阿部涼音"


def test_sync_keeps_female_and_unknown(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_tmdb(monkeypatch, {1417328: 1, 1417329: 0})
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 0
    assert result.created == 2


def test_sync_keeps_when_gender_unknown(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_tmdb(monkeypatch, {1417328: None, 1417329: None})
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 0
    assert result.created == 2


def test_sync_no_filter_when_no_tmdb_key(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("", ""))
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 0
    assert result.created == 2


def test_sync_does_not_requery_existing_tmdbid(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["阿部純子", "阿部純子", "", "", "", 1417328, "", "", ""]])
    calls = _mock_tmdb(monkeypatch, {1417328: 2, 1417329: 1})
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert 1417328 not in calls  # 本地已有该 tmdbid，不重复请求性别
    assert result.skipped_male == 0
    assert result.created == 1


def test_clean_removes_male_and_backs_up(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [["男优A", "男优A", "", "", "", 1001, "", "", ""], ["女优B", "女优B", "", "", "", 1002, "", "", ""]],
    )
    _mock_tmdb(monkeypatch, {1001: 2, 1002: 1})
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 1
    rows = _read_rows(_tmp_actor_db)
    assert [r[COL_JP] for r in rows] == ["女优B"]
    wb = load_workbook(_tmp_actor_db)
    assert "男优备份" in wb.sheetnames
    backup = list(wb["男优备份"].iter_rows(min_row=1, values_only=True))
    wb.close()
    assert len(backup) == 1
    assert backup[0][0] == "男优A"


def test_clean_keeps_female_unknown_and_missing(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [
            ["女优A", "女优A", "", "", "", 2001, "", "", ""],
            ["未知B", "未知B", "", "", "", 2002, "", "", ""],
            ["无idC", "无idC", "", "", "", "", "", "", ""],
        ],
    )
    _mock_tmdb(monkeypatch, {2001: 1, 2002: None})
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 0
    rows = _read_rows(_tmp_actor_db)
    assert len(rows) == 3


def test_clean_no_tmdb_key_noop(_tmp_actor_db: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["男优A", "男优A", "", "", "", 3001, "", "", ""]])
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("", ""))
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 0
    assert len(_read_rows(_tmp_actor_db)) == 1


def test_clean_male_list_works_without_tmdb_key(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [
            ["吉村卓", "吉村卓", "", "", "", "", "", "", ""],
            ["女优A", "女优A", "", "", "", "", "", "", ""],
        ],
    )
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("", ""))
    _mock_male_list(monkeypatch, ["吉村卓"])
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 1
    rows = _read_rows(_tmp_actor_db)
    assert [r[COL_JP] for r in rows] == ["女优A"]


def test_clean_limit_applies(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [["男优A", "男优A", "", "", "", 4001, "", "", ""], ["男优B", "男优B", "", "", "", 4002, "", "", ""]],
    )
    _mock_tmdb(monkeypatch, {4001: 2, 4002: 2})
    result = asyncio.run(actor_db_tool.clean_male_actors(limit=1))
    assert result.removed_male == 1
    rows = _read_rows(_tmp_actor_db)
    assert len(rows) == 1


def test_clean_idempotent(_tmp_actor_db: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["女优A", "女优A", "", "", "", 5001, "", "", ""]])
    _mock_tmdb(monkeypatch, {5001: 1})
    asyncio.run(actor_db_tool.clean_male_actors())
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 0
    assert len(_read_rows(_tmp_actor_db)) == 1


def _mock_male_list(monkeypatch, names: list[str]):
    monkeypatch.setattr(actor_db_tool, "_load_male_actor_set", lambda: {n.casefold() for n in names})


def test_clean_removes_male_by_list_without_tmdbid(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [
            ["吉村卓", "吉村卓", "", "", "", "", "", "", ""],
            ["女优A", "女优A", "", "", "", "", "", "", ""],
        ],
    )
    _mock_tmdb(monkeypatch, {})
    _mock_male_list(monkeypatch, ["吉村卓"])
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 1
    rows = _read_rows(_tmp_actor_db)
    assert [r[COL_JP] for r in rows] == ["女优A"]
    wb = load_workbook(_tmp_actor_db)
    backup = list(wb["男优备份"].iter_rows(min_row=1, values_only=True))
    wb.close()
    assert [b[0] for b in backup] == ["吉村卓"]


def test_clean_male_list_beats_tmdb_gender0(_tmp_actor_db: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["加藤鷹", "加藤鷹", "", "", "", 6001, "", "", ""]])
    _mock_tmdb(monkeypatch, {6001: 0})  # TMDB 未标性别，但名单命中
    _mock_male_list(monkeypatch, ["加藤鷹"])
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 1
    assert len(_read_rows(_tmp_actor_db)) == 0


def test_clean_male_list_only_name_checked(_tmp_actor_db: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["吉村卓", "吉村卓", "", "", "", 7001, "", "", ""]])
    calls = _mock_tmdb(monkeypatch, {7001: 2})
    _mock_male_list(monkeypatch, ["吉村卓"])
    asyncio.run(actor_db_tool.clean_male_actors())
    assert 7001 not in calls  # 名单命中后不应再请求 TMDB


def test_sync_skips_male_by_list_without_tmdbid(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_male_list(monkeypatch, ["阿部純子"])
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 1
    assert result.created == 1
    rows = _read_rows(_tmp_actor_db)
    assert [r[COL_JP] for r in rows] == ["阿部涼音"]


def test_sync_male_list_avoids_tmdb_request(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_male_list(monkeypatch, ["阿部純子"])
    calls = _mock_tmdb(monkeypatch, {1417328: 2, 1417329: 1})
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 1
    assert 1417328 not in calls  # 名单命中，不应发起 TMDB gender 请求


class _FakeResp:
    def __init__(self, status: int):
        self.status = status

    async def __aenter__(self):
        return self

    async def __aexit__(self, *args):
        return None

    def release(self):
        return None


def _mock_verify_network(monkeypatch, status_map: dict[int, int]):
    """mock verify_tmdb_ids 的 TMDB person/{id} 请求。status_map: tmdbid -> http status。"""
    import aiohttp

    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("https://api.tmdb.org", "test-key"))

    class _FakeGet:
        def __call__(self, url, timeout=None):
            import re

            m = re.search(r"/person/(\d+)", url)
            pid = int(m.group(1)) if m else 0
            return _FakeResp(status_map.get(pid, 200))

    class _FakeClient:
        def __init__(self):
            self.get = _FakeGet()

        async def __aenter__(self):
            return self

        async def __aexit__(self, *args):
            return None

    monkeypatch.setattr(aiohttp, "ClientSession", _FakeClient)


def test_verify_tmdbid_clears_invalid_ids(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [
            ["桃乃木香奈", "", "", "", "", 2616715, "https://www.themoviedb.org/person/2616715", "", ""],
            ["三佳詩", "", "", "", "", 6231965, "https://www.themoviedb.org/person/6231965", "", ""],  # 404 失效
            ["涼森れむ", "", "", "", "", 2640963, "https://www.themoviedb.org/person/2640963", "", ""],
        ],
    )
    _mock_verify_network(monkeypatch, {2616715: 200, 6231965: 404, 2640963: 200})
    result = asyncio.run(actor_db_tool.verify_tmdb_ids())
    assert result.checked == 3
    assert result.invalid == 1
    rows = _read_rows(_tmp_actor_db)
    by_jp = {r[0]: r[5] for r in rows}
    assert by_jp["桃乃木香奈"] == 2616715  # 有效保留
    assert by_jp["三佳詩"] is None  # 失效清除
    assert by_jp["涼森れむ"] == 2640963  # 有效保留


def test_verify_tmdbid_keeps_all_valid(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [
            ["桃乃木香奈", "", "", "", "", 2616715, "https://www.themoviedb.org/person/2616715", "", ""],
            ["涼森れむ", "", "", "", "", 2640963, "https://www.themoviedb.org/person/2640963", "", ""],
        ],
    )
    _mock_verify_network(monkeypatch, {2616715: 200, 2640963: 200})
    result = asyncio.run(actor_db_tool.verify_tmdb_ids())
    assert result.invalid == 0
    rows = _read_rows(_tmp_actor_db)
    assert {r[5] for r in rows} == {2616715, 2640963}


def test_verify_tmdbid_network_error_keeps_id(_tmp_actor_db: Path, monkeypatch):
    """网络失败(非404)保守保留 id，不误清。"""
    import aiohttp

    _write_db(
        _tmp_actor_db,
        [["某演员", "", "", "", "", 1001, "https://www.themoviedb.org/person/1001", "", ""]],
    )
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("https://api.tmdb.org", "test-key"))

    class _BoomGet:
        def __call__(self, url, timeout=None):
            raise aiohttp.ClientError("network down")

    class _FakeClient:
        def __init__(self):
            self.get = _BoomGet()

        async def __aenter__(self):
            return self

        async def __aexit__(self, *args):
            return None

    monkeypatch.setattr(aiohttp, "ClientSession", _FakeClient)
    result = asyncio.run(actor_db_tool.verify_tmdb_ids())
    assert result.invalid == 0
    rows = _read_rows(_tmp_actor_db)
    assert rows[0][5] == 1001  # 保留


def test_verify_tmdbid_recovers_new_id(_tmp_actor_db: Path, monkeypatch):
    """失效 id 清除后按名重搜补回新 id（TMDB 重建档案场景）。"""
    import aiohttp

    _write_db(
        _tmp_actor_db,
        [["三佳詩", "", "", "", "", 6231965, "https://www.themoviedb.org/person/6231965", "", ""]],
    )
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("https://api.tmdb.org", "test-key"))

    # 请求 404（旧 id 已删）
    class _FakeGet:
        def __call__(self, url, timeout=None):
            return _FakeResp(404)

    class _FakeClient:
        def __init__(self):
            self.get = _FakeGet()

        async def __aenter__(self):
            return self

        async def __aexit__(self, *args):
            return None

    monkeypatch.setattr(aiohttp, "ClientSession", _FakeClient)

    # mock 按名重搜返回新 id
    from mdcx.core import tmdb_actor

    async def fake_query(name, base_url, api_key, client):
        return {"pid": 5882313, "name": "三佳诗", "adult": True}

    monkeypatch.setattr(tmdb_actor, "query_single_actor_cached", fake_query)

    result = asyncio.run(actor_db_tool.verify_tmdb_ids())
    assert result.invalid == 1
    assert result.recovered == 1
    rows = _read_rows(_tmp_actor_db)
    assert rows[0][5] == 5882313  # 补回新 id


def test_update_nfo_tmdbid_text_replaces_and_adds():
    """文本级替换：nfo 旧 id 被新 id 覆盖，无 id 的补上。"""
    nfo = (
        "<movie><actor><name>桃園怜奈</name><type>Actor</type><tmdbid>12345</tmdbid></actor>"
        "<actor><name>三佳詩</name><type>Actor</type></actor>"
        "<actor><name>已同步</name><type>Actor</type><tmdbid>5882313</tmdbid></actor>"
        "<actor><name>未收录</name><type>Actor</type><tmdbid>999</tmdbid></actor></movie>"
    )
    id_map = {"桃園怜奈": 5122968, "三佳詩": 5882313, "已同步": 5882313}
    new, cnt = actor_db_tool._update_nfo_tmdbids_text(nfo, id_map)
    assert cnt == 2  # 桃園怜奈替换 + 三佳詩补入；已同步跳过
    assert "<tmdbid>5122968</tmdbid>" in new
    assert "<tmdbid>5882313</tmdbid>" in new
    assert "<tmdbid>999</tmdbid>" in new  # 未收录保留
    # 补入的 tmdbid 在 type 之后（顺序正确）
    assert new.index("<type>Actor</type>") < new.index("<tmdbid>5882313</tmdbid>")


def test_update_nfo_tmdbid_updates_files(tmp_path, monkeypatch):
    """端到端：遍历 nfo 目录，更新 actor tmdbid。"""
    nfo_dir = tmp_path / "nfo"
    nfo_dir.mkdir(parents=True)
    (nfo_dir / "a.nfo").write_text(
        "<movie><actor><name>桃園怜奈</name><type>Actor</type><tmdbid>123</tmdbid></actor></movie>",
        encoding="utf-8",
    )
    (nfo_dir / "b.nfo").write_text(
        "<movie><actor><name>三佳詩</name><type>Actor</type></actor></movie>", encoding="utf-8"
    )

    # mock 库反查
    def fake_reverse(name):
        if name == "桃園怜奈":
            return {"jp": "桃園怜奈", "tmdbid": 5122968}
        if name == "三佳詩":
            return {"jp": "三佳詩", "tmdbid": 5882313}
        return None

    from mdcx.core import tmdb_actor

    monkeypatch.setattr(tmdb_actor, "search_actor_db_reverse", fake_reverse)

    result = asyncio.run(actor_db_tool.update_nfo_tmdb_ids(nfo_dir))
    assert result.checked == 2
    assert result.updated_files == 2
    assert result.updated_actors == 2
    content_a = (nfo_dir / "a.nfo").read_text(encoding="utf-8")
    content_b = (nfo_dir / "b.nfo").read_text(encoding="utf-8")
    assert "<tmdbid>5122968</tmdbid>" in content_a
    assert "<tmdbid>5882313</tmdbid>" in content_b


def test_sync_keeps_tmdbid_when_identity_matches(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    calls = _mock_tmdb(
        monkeypatch,
        {1417328: 1, 1417329: 1},
        identity_names={1417328: "阿部純子", 1417329: "阿部涼音"},
    )
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_tmdbid == 0
    rows = _read_rows(_tmp_actor_db)
    by_jp = {r[COL_JP]: r[COL_TMDBID] for r in rows}
    assert by_jp["阿部純子"] == 1417328
    assert by_jp["阿部涼音"] == 1417329
    assert 1417328 in calls and 1417329 in calls


def test_sync_drops_tmdbid_when_identity_mismatch(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_tmdb(
        monkeypatch,
        {1417328: 1, 1417329: 1},
        identity_names={1417328: "Christa Allen", 1417329: "阿部涼音"},  # 错误映射样本：平山加奈->Christa Allen
    )
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_tmdbid == 1
    rows = _read_rows(_tmp_actor_db)
    by_jp = {r[COL_JP]: r[COL_TMDBID] for r in rows}
    assert by_jp["阿部純子"] in (None, "")  # 身份不匹配，丢弃该 id
    assert by_jp["阿部涼音"] == 1417329  # 匹配的保留


def test_sync_verify_disabled_keeps_tmdbid(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_tmdb(
        monkeypatch,
        {1417328: 1, 1417329: 1},
        identity_names={1417328: "Christa Allen", 1417329: "阿部涼音"},
    )
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml), verify_tmdbid=False))
    assert result.skipped_tmdbid == 0
    rows = _read_rows(_tmp_actor_db)
    by_jp = {r[COL_JP]: r[COL_TMDBID] for r in rows}
    assert by_jp["阿部純子"] == 1417328  # 关闭校验时原样写入


def test_sync_identity_match_with_katakana_roman(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    """片假名音译 ↔ 英文原名视为同人，不丢弃 id。"""
    xml = (
        _XML.replace('tmdb_id="1417328"', 'tmdb_id="1417328"')
        .replace('jp="阿部純子"', 'jp="キャシー・ヘブン"')
        .replace('zh_cn="阿部純子"', 'zh_cn="キャシー・ヘブン"')
    )
    xml_path = _avdb_xml.parent / "mapping2.xml"
    xml_path.write_text(xml, encoding="utf-8")
    _mock_tmdb(
        monkeypatch,
        {1417328: 1, 1417329: 1},
        identity_names={1417328: "Cathy Heaven", 1417329: "阿部涼音"},
    )
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(xml_path)))
    assert result.skipped_tmdbid == 0
    rows = _read_rows(_tmp_actor_db)
    by_jp = {r[COL_JP]: r[COL_TMDBID] for r in rows}
    assert by_jp["キャシー・ヘブン"] == 1417328  # 片假名↔英文放行


def test_sync_existing_tmdbid_not_requeried_for_identity(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    """本地已存在的 tmdbid 不重复做身份反查。"""
    _write_db(_tmp_actor_db, [["阿部純子", "阿部純子", "", "", "", 1417328, "", "", ""]])
    spy: list[int] = []
    _mock_tmdb(monkeypatch, {1417328: 2, 1417329: 1}, identity_spy=spy)
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_tmdbid == 0
    assert 1417328 not in spy  # 已在库中的 id 不重复反查
    assert 1417329 in spy  # 新 id 需要校验


def test_verify_tmdbid_checkpoint_skips_verified(_tmp_actor_db: Path, monkeypatch):
    """断点续传：首次校验后写入 verified 文件，重跑跳过已校验 id。"""
    import aiohttp

    _write_db(
        _tmp_actor_db,
        [
            ["演员甲", "", "", "", "", 1001, "https://www.themoviedb.org/person/1001", "", ""],
            ["演员乙", "", "", "", "", 1002, "https://www.themoviedb.org/person/1002", "", ""],
        ],
    )
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("https://api.tmdb.org", "test-key"))

    class _FakeGet:
        def __call__(self, url, timeout=None):
            import re

            m = re.search(r"/person/(\d+)", url)
            return _FakeResp(200 if m else 500)

    class _FakeClient:
        def __init__(self):
            self.get = _FakeGet()

        async def __aenter__(self):
            return self

        async def __aexit__(self, *args):
            return None

    monkeypatch.setattr(aiohttp, "ClientSession", _FakeClient)

    # 第一轮 limit=1：只校验前 1 个
    result1 = asyncio.run(actor_db_tool.verify_tmdb_ids(limit=1))
    assert result1.checked == 1

    # verified 文件已生成，含 1 个 id
    verified_file = _tmp_actor_db.parent / ".tmdbid_verified.json"
    import json

    assert verified_file.exists()
    assert len(json.loads(verified_file.read_text(encoding="utf-8"))) == 1

    # 第二轮 limit=1：跳过已校验的，只校验剩余 1 个
    result2 = asyncio.run(actor_db_tool.verify_tmdb_ids(limit=1))
    assert result2.checked == 1

    # 第三轮：全部校验完，无待处理
    result3 = asyncio.run(actor_db_tool.verify_tmdb_ids())
    assert result3.checked == 0


def test_clean_male_checkpoint_skips_verified(_tmp_actor_db: Path, monkeypatch):
    """剔除男演员断点续传：已校验 gender 的 id 重跑自动跳过。"""
    import aiohttp

    _write_db(
        _tmp_actor_db,
        [
            ["演员甲", "", "", "", "", 1001, "", "", ""],
            ["演员乙", "", "", "", "", 1002, "", "", ""],
        ],
    )
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("https://api.tmdb.org", "test-key"))

    # 名单不命中这两个名字，走 gender 校验
    monkeypatch.setattr(actor_db_tool, "is_male_actor", lambda n: False)

    class _FakeGet:
        def __call__(self, url, timeout=None):
            import re

            m = re.search(r"/person/(\d+)", url)
            return _FakeResp(200 if m else 500)

    class _FakeClient:
        def __init__(self):
            self.get = _FakeGet()

        async def __aenter__(self):
            return self

        async def __aexit__(self, *args):
            return None

    monkeypatch.setattr(aiohttp, "ClientSession", _FakeClient)
    # gender=1（非男）需 mock fetch_person_gender，保证 add checked_ids
    monkeypatch.setattr(tmdb_actor, "fetch_person_gender", lambda pid, b, k, c: 1)

    import json

    # 第一轮 limit=1：只校验前 1 个 id
    result1 = asyncio.run(actor_db_tool.clean_male_actors(limit=1))
    assert result1.checked == 1
    checked_file = _tmp_actor_db.parent / ".gender_checked.json"
    assert checked_file.exists()
    assert len(json.loads(checked_file.read_text(encoding="utf-8"))) == 1

    # 第二轮 limit=1：跳过已校验的，处理剩余 1 个
    result2 = asyncio.run(actor_db_tool.clean_male_actors(limit=1))
    assert result2.checked == 1

    # 第三轮：全部校验完
    result3 = asyncio.run(actor_db_tool.clean_male_actors())
    assert result3.checked == 0
