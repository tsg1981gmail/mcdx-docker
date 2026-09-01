import asyncio
from pathlib import Path

import pytest
from openpyxl import load_workbook

from mdcx.core import tmdb_actor
from mdcx.models.flags import Flags
from mdcx.signals import signal
from mdcx.tools import actor_db_tool


@pytest.fixture
def _reset_stop_flags():
    Flags.stop_requested = False
    signal.stop = False
    yield
    Flags.stop_requested = False
    signal.stop = False


@pytest.fixture
def _tmp_actor_db(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setattr(tmdb_actor.manager, "data_folder", tmp_path)
    monkeypatch.setattr(tmdb_actor.resources, "actor_db", {})
    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    return userdata / "actor_database.xlsx"


def test_collect_actors_from_nfo_dir_gathers_and_dedups(tmp_path: Path):
    sub = tmp_path / "sub"
    sub.mkdir(parents=True, exist_ok=True)
    (tmp_path / "a.nfo").write_text(
        '<?xml version="1.0"?><movie><actor><name>三上悠亚</name></actor><actor><name>明日花绮罗</name></actor></movie>',
        encoding="utf-8",
    )
    (sub / "b.nfo").write_text(
        '<?xml version="1.0"?><movie><actor><name>三上悠亚</name></actor><actor><name>桥本有菜</name></actor></movie>',
        encoding="utf-8",
    )
    (tmp_path / "not.nfo").write_text("hello", encoding="utf-8")

    import asyncio

    actors = asyncio.run(actor_db_tool.collect_actors_from_nfo_dir(tmp_path))

    assert set(actors) == {"三上悠亚", "明日花绮罗", "桥本有菜"}


def test_collect_actors_from_nfo_dir_empty_dir(tmp_path: Path):
    import asyncio

    assert asyncio.run(actor_db_tool.collect_actors_from_nfo_dir(tmp_path)) == []


def test_collect_actors_from_nfo_dir_non_existent(tmp_path: Path):
    import asyncio

    assert asyncio.run(actor_db_tool.collect_actors_from_nfo_dir(tmp_path / "missing")) == []


@pytest.mark.asyncio
async def test_run_with_empty_names(_tmp_actor_db: Path):
    result = await actor_db_tool.run([])
    assert result.total == 0
    assert result.translated == 0
    assert result.linked == 0
    assert result.skipped == 0


@pytest.mark.asyncio
async def test_run_skips_actor_without_tmdbid(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")
    monkeypatch.setattr(tmdb_actor.resources, "actor_db", {})
    monkeypatch.setattr(tmdb_actor.resources, "actor_db_reverse_index", None)

    result = await actor_db_tool.run(["未知演员"], translate=True, link=True)

    assert result.total == 1
    assert result.skipped == 1
    assert result.translated == 0
    assert result.linked == 0


@pytest.mark.asyncio
async def test_run_translate_backfills_zh_names(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="三上悠亜", zh_cn="", zh_tw="", tmdbid=12345)
    tmdb_actor.resources.actor_db = {
        "三上悠亜": {"zh_cn": "", "zh_tw": "", "keyword": "", "href": "", "tmdbid": 12345, "tmdb_url": ""}
    }
    tmdb_actor.resources.actor_db_reverse_index = None

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    async def _fake_translations(pid, base_url, api_key, client):
        return {"zh_cn": "三上悠亚", "zh_tw": "三上悠亞"}

    monkeypatch.setattr(actor_db_tool, "_fetch_person_translations", _fake_translations)
    monkeypatch.setattr(actor_db_tool, "fetch_libredmm_link", _no_link)

    result = await actor_db_tool.run(["三上悠亜"], translate=True, link=False)

    assert result.total == 1
    assert result.translated == 1

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip() == "三上悠亜":
            assert row[1] == "三上悠亚"
            assert row[2] == "三上悠亞"
            break
    wb.close()


@pytest.mark.asyncio
async def test_run_link_backfills_href(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="三上悠亜", zh_cn="三上悠亚", zh_tw="三上悠亞", tmdbid=12345)
    tmdb_actor.resources.actor_db = {
        "三上悠亜": {
            "zh_cn": "三上悠亚",
            "zh_tw": "三上悠亞",
            "keyword": "",
            "href": "",
            "tmdbid": 12345,
            "tmdb_url": "",
        }
    }
    tmdb_actor.resources.actor_db_reverse_index = None

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    async def _fake_link(actor_name: str) -> str:
        return "https://www.libredmm.com/actresses/mikami-yua"

    monkeypatch.setattr(actor_db_tool, "fetch_libredmm_link", _fake_link)

    result = await actor_db_tool.run(["三上悠亜"], translate=False, link=True)

    assert result.total == 1
    assert result.linked == 1

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip() == "三上悠亜":
            assert row[4] == "https://www.libredmm.com/actresses/mikami-yua"
            break
    wb.close()


@pytest.mark.asyncio
async def test_run_skips_when_translate_disabled(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="三上悠亜", zh_cn="", zh_tw="", tmdbid=12345)
    tmdb_actor.resources.actor_db = {
        "三上悠亜": {"zh_cn": "", "zh_tw": "", "keyword": "", "href": "", "tmdbid": 12345, "tmdb_url": ""}
    }
    tmdb_actor.resources.actor_db_reverse_index = None

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    call_count = {"translations": 0}

    async def _counting_translations(pid, base_url, api_key, client):
        call_count["translations"] += 1
        return {"zh_cn": "三上悠亚", "zh_tw": "三上悠亞"}

    monkeypatch.setattr(actor_db_tool, "_fetch_person_translations", _counting_translations)
    monkeypatch.setattr(actor_db_tool, "fetch_libredmm_link", _no_link)

    result = await actor_db_tool.run(["三上悠亜"], translate=False, link=False)

    assert result.total == 1
    assert result.skipped == 1
    assert call_count["translations"] == 0


@pytest.mark.asyncio
async def test_run_translate_and_link_together(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="三上悠亜", zh_cn="", zh_tw="", href="", tmdbid=12345)
    tmdb_actor.resources.actor_db = {
        "三上悠亜": {"zh_cn": "", "zh_tw": "", "keyword": "", "href": "", "tmdbid": 12345, "tmdb_url": ""}
    }
    tmdb_actor.resources.actor_db_reverse_index = None

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    async def _fake_translations(pid, base_url, api_key, client):
        return {"zh_cn": "三上悠亚", "zh_tw": "三上悠亞"}

    async def _fake_link(actor_name: str) -> str:
        return "https://www.libredmm.com/actresses/mikami-yua"

    monkeypatch.setattr(actor_db_tool, "_fetch_person_translations", _fake_translations)
    monkeypatch.setattr(actor_db_tool, "fetch_libredmm_link", _fake_link)

    result = await actor_db_tool.run(["三上悠亜"], translate=True, link=True)

    assert result.total == 1
    assert result.translated == 1
    assert result.linked == 1

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip() == "三上悠亜":
            assert row[1] == "三上悠亚", f"zh_cn actual: {row[1]}"
            assert row[2] == "三上悠亞", f"zh_tw actual: {row[2]}"
            assert row[4] == "https://www.libredmm.com/actresses/mikami-yua", f"href actual: {row[4]}"
            break
    wb.close()


async def _no_link(actor_name: str) -> str:
    return ""


@pytest.mark.asyncio
async def test_update_actor_db_row_skips_placeholder_name(_tmp_actor_db: Path):
    """占位符名字被语义清洗为空时应跳过写入。"""
    status = await tmdb_actor.update_actor_db_row(jp="素人奥様", zh_cn="", zh_tw="", tmdbid=None)
    assert status == "skipped_placeholder"
    assert not _tmp_actor_db.exists()  # 未创建数据库文件，即未写入任何行


@pytest.mark.asyncio
async def test_update_actor_db_row_cleans_series_tag(_tmp_actor_db: Path):
    """写入时剥离名字中的系列标签。"""
    status = await tmdb_actor.update_actor_db_row(jp="本田仁美(パコパコママ)", zh_cn="", zh_tw="", tmdbid=None)
    assert status in ("inserted_new_row", "updated_zh_cn")
    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    assert rows and rows[0][0] == "本田仁美"


@pytest.mark.asyncio
async def test_run_actor_db_xlsx_stop_cancels_pending_and_saves(
    _tmp_actor_db: Path, monkeypatch: pytest.MonkeyPatch, _reset_stop_flags
):
    """run_actor_db_xlsx 在手动停止后取消 pending 并保存已处理部分。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"])
    for jp, pid in [("演员甲", 11), ("演员乙", 12), ("演员丙", 13), ("演员丁", 14), ("演员戊", 15)]:
        ws.append([jp, "", "", "", "", str(pid), "", "", ""])
    wb.save(_tmp_actor_db)
    wb.close()

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    calls: list[str] = []
    write_lock = asyncio.Lock()

    async def fake_translations(pid, base_url, api_key, client):
        async with write_lock:
            calls.append(str(pid))
            if len(calls) >= 3:
                Flags.stop_requested = True
        await asyncio.sleep(0.01)
        return {"zh_cn": f"中文{pid}", "zh_tw": f"中文{pid}"}

    async def fake_person_url(*args, **kwargs):
        return "https://www.themoviedb.org/person/1"

    monkeypatch.setattr(actor_db_tool, "_fetch_person_translations", fake_translations)
    monkeypatch.setattr(actor_db_tool, "_tmdb_person_url", fake_person_url)

    await actor_db_tool.run_actor_db_xlsx("translate")

    assert Flags.stop_requested is True
    assert len(calls) <= 5  # 停止后不再提交新的

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    saved = [r for r in rows if str(r[1] or "").strip()]
    assert len(saved) >= 2  # 已处理部分被保存
    for r in saved:
        assert str(r[1]).startswith("中文")


@pytest.mark.asyncio
async def test_run_actor_db_xlsx_limit_slices_and_reruns_idempotently(
    _tmp_actor_db: Path, monkeypatch: pytest.MonkeyPatch, _reset_stop_flags
):
    """limit 限量分片：仅处理前 limit 条，重跑时不重复已处理条目。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"])
    for jp, pid in [("演员甲", 11), ("演员乙", 12), ("演员丙", 13), ("演员丁", 14)]:
        ws.append([jp, "", "", "", "", str(pid), "", "", ""])
    wb.save(_tmp_actor_db)
    wb.close()

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    processed: list[str] = []

    async def fake_translations(pid, base_url, api_key, client):
        processed.append(str(pid))
        await asyncio.sleep(0)
        return {"zh_cn": f"中文{pid}", "zh_tw": f"中文{pid}"}

    async def fake_person_url(*args, **kwargs):
        return "https://www.themoviedb.org/person/1"

    monkeypatch.setattr(actor_db_tool, "_fetch_person_translations", fake_translations)
    monkeypatch.setattr(actor_db_tool, "_tmdb_person_url", fake_person_url)

    # 第一轮：限量 2 条
    await actor_db_tool.run_actor_db_xlsx("translate", limit=2)
    assert len(processed) == 2

    # 第二轮：重跑，已处理的条目不再进入，仅处理剩余
    processed.clear()
    await actor_db_tool.run_actor_db_xlsx("translate", limit=2)
    assert len(processed) == 2

    # 第三轮：全部处理完，无剩余
    processed.clear()
    await actor_db_tool.run_actor_db_xlsx("translate", limit=2)
    assert len(processed) == 0

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    assert all(str(r[1] or "").startswith("中文") for r in rows)


@pytest.mark.asyncio
async def test_run_actor_db_xlsx_sync_aliases_only_handles_empty(
    _tmp_actor_db: Path, monkeypatch: pytest.MonkeyPatch, _reset_stop_flags
):
    """sync_aliases 仅处理别名列为空的条目，已有别名的行不重复请求。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"])
    ws.append(["已有别名演员", "", "", "别名甲", "", "11", "", "", ""])
    ws.append(["空别名演员", "", "", "", "", "12", "", "", ""])
    ws.append(["无别名但无id", "", "", "", "", "", "", "", ""])
    wb.save(_tmp_actor_db)
    wb.close()

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    queried: list[str] = []

    async def fake_query_single_actor_cached(actor_name, base_url, api_key, client):
        queried.append(actor_name)
        return {"name": actor_name, "original_name": actor_name, "also_known_as": ["新别名"]}

    monkeypatch.setattr(tmdb_actor, "query_single_actor_cached", fake_query_single_actor_cached)

    await actor_db_tool.run_actor_db_xlsx("sync_aliases")

    # 只查询了空别名的有 id 行，已有别名的行被跳过
    assert queried == ["空别名演员"]

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    empty_row = next(r for r in rows if r[0] == "空别名演员")
    assert "新别名" in str(empty_row[3])
    alias_row = next(r for r in rows if r[0] == "已有别名演员")
    assert "新别名" not in str(alias_row[3])


@pytest.mark.asyncio
async def test_run_actor_db_xlsx_sync_aliases_offset_skips_processed_rows(
    _tmp_actor_db: Path, monkeypatch: pytest.MonkeyPatch, _reset_stop_flags
):
    """sync_aliases 的 offset 参数：跳过前 N 个数据行，用于手动中断后的分片续跑。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"])
    # 4 行数据，row_idx=2..5
    for jp, pid in [("演员A", 21), ("演员B", 22), ("演员C", 23), ("演员D", 24)]:
        ws.append([jp, "", "", "", "", str(pid), "", "", ""])
    wb.save(_tmp_actor_db)
    wb.close()

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    queried: list[str] = []

    async def fake_query_single_actor_cached(actor_name, base_url, api_key, client):
        queried.append(actor_name)
        return {"name": actor_name, "original_name": actor_name, "also_known_as": [f"{actor_name}别名"]}

    monkeypatch.setattr(tmdb_actor, "query_single_actor_cached", fake_query_single_actor_cached)

    # overwrite=True 全量；offset=2 表示跳过前 2 个数据行（演员A/B），从演员C 开始处理
    await actor_db_tool.run_actor_db_xlsx("sync_aliases", overwrite=True, offset=2)

    # 只处理后两行（演员C/D），演员A/B 不应被请求
    assert set(queried) == {"演员C", "演员D"}
    assert len(queried) == 2

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    assert str(rows[0][3] or "") == ""  # 演员A 未被处理
    assert str(rows[1][3] or "") == ""  # 演员B 未被处理
    assert "演员C别名" in str(rows[2][3])  # 演员C 已处理
    assert "演员D别名" in str(rows[3][3])  # 演员D 已处理


@pytest.mark.asyncio
async def test_run_actor_db_xlsx_sync_aliases_offset_non_overwrite_double_skip(
    _tmp_actor_db: Path, monkeypatch: pytest.MonkeyPatch, _reset_stop_flags
):
    """非 overwrite（补缺别名）+ offset：offset 跳过 + keyword 非空跳过，双重过滤。

    场景：5 行中 row3 已有别名（自动跳过）、row4 空、row5 空；
    offset=2 跳过前 2 个数据行（演员甲/已有别名行），
    剩余 row4/row5 应被处理；offset 有效的作用与 overwrite 开关无关。
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"])
    ws.append(["演员甲", "", "", "", "", "41", "", "", ""])
    ws.append(["演员乙", "", "", "已有别名X", "", "42", "", "", ""])
    ws.append(["演员丙", "", "", "", "", "43", "", "", ""])
    ws.append(["演员丁", "", "", "", "", "44", "", "", ""])
    ws.append(["演员戊", "", "", "", "", "45", "", "", ""])
    wb.save(_tmp_actor_db)
    wb.close()

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    queried: list[str] = []

    async def fake_query_single_actor_cached(actor_name, base_url, api_key, client):
        queried.append(actor_name)
        return {"name": actor_name, "original_name": actor_name, "also_known_as": [f"{actor_name}新别名"]}

    monkeypatch.setattr(tmdb_actor, "query_single_actor_cached", fake_query_single_actor_cached)

    # 非 overwrite 补缺 + offset=2：跳过演员甲、演员乙；演员丙/丁/戊空别名参与处理
    await actor_db_tool.run_actor_db_xlsx("sync_aliases", overwrite=False, offset=2)

    # 演员甲（虽空别名）被 offset 拦住；演员乙被"已有别名"拦住；丙丁戊被处理
    assert set(queried) == {"演员丙", "演员丁", "演员戊"}
    assert len(queried) == 3  # 无重复请求
    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    assert str(rows[0][3] or "") == ""  # 甲 未被处理（offset 跳过)
    assert str(rows[1][3]) == "已有别名X"  # 乙 别名未被覆盖
    for i in (2, 3, 4):  # 丙丁戊 已并入新别名
        assert "新别名" in str(rows[i][3])


@pytest.mark.asyncio
async def test_run_actor_db_xlsx_sync_aliases_offset_limit_slice(
    _tmp_actor_db: Path, monkeypatch: pytest.MonkeyPatch, _reset_stop_flags
):
    """sync_aliases 配合 offset+limit 分片：offset 跳前 N，limit 限本批 N 条。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"])
    for jp, pid in [("演员1", 31), ("演员2", 32), ("演员3", 33), ("演员4", 34), ("演员5", 35)]:
        ws.append([jp, "", "", "", "", str(pid), "", "", ""])
    wb.save(_tmp_actor_db)
    wb.close()

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    queried: list[str] = []

    async def fake_query_single_actor_cached(actor_name, base_url, api_key, client):
        queried.append(actor_name)
        return {"name": actor_name, "original_name": actor_name, "also_known_as": [f"{actor_name}别名"]}

    monkeypatch.setattr(tmdb_actor, "query_single_actor_cached", fake_query_single_actor_cached)

    # 全量 + offset=1 + limit=2 → 跳过演员1，仅处理演员2、演员3；演员4/5 不被本批触及
    await actor_db_tool.run_actor_db_xlsx("sync_aliases", overwrite=True, offset=1, limit=2)

    assert set(queried) == {"演员2", "演员3"}
    assert len(queried) == 2


def test_is_structured_bio_detects_formatted_and_free_text():
    """_is_structured_bio 应识别统一格式简介，放行自由文本/空文本。"""
    assert actor_db_tool._is_structured_bio("身高: 155cm | 罩杯: F | 三围: 83/58/84")
    assert actor_db_tool._is_structured_bio("身高: 160cm")
    assert not actor_db_tool._is_structured_bio("")
    assert not actor_db_tool._is_structured_bio("身高155cm，三围：B83/W58/H84")
    assert not actor_db_tool._is_structured_bio("浅野心爱（Asano Kokoa）")
    assert not actor_db_tool._is_structured_bio(
        "姓名: 前嶋美紀（まえじまみき），身高: 160cm，三围: B83/W59/H84（C罩杯）"
    )


def test_fill_minnano_overwrite_semantics(_tmp_actor_db: Path, monkeypatch: pytest.MonkeyPatch):
    """fill_minnano 的 overwrite 语义：False 只补空，True 覆盖已有生日/简介。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"])
    ws.append(["已有数据演员", "", "", "", "", "11", "", "1990-01-01", "旧简介"])
    ws.append(["空数据演员", "", "", "", "", "12", "", "", ""])
    wb.save(_tmp_actor_db)
    wb.close()

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    parsed = {
        "name": "已有数据演员",
        "aliases": ["新别名"],
        "birthday": "2000-02-02",
        "height": "160",
        "bust": "83",
        "waist": "58",
        "hip": "84",
        "cup": "F",
        "place": "東京",
        "career": "",
    }

    from mdcx.tools import minnano_crawler

    async def fake_search(name):
        return ("123", "<html>mock</html>")

    def fake_parse(html, mid):
        return parsed

    monkeypatch.setattr(minnano_crawler, "_search_minnano_by_name", fake_search)
    monkeypatch.setattr(minnano_crawler, "parse_minnano_page", fake_parse)

    # overwrite=False（默认）：只补空数据演员，已有数据的不动
    import asyncio

    asyncio.run(actor_db_tool.run_actor_db_xlsx("fill_minnano"))

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    existing = next(r for r in rows if r[0] == "已有数据演员")
    assert existing[7] == "1990-01-01" and existing[8] == "旧简介"  # 未被覆盖
    empty = next(r for r in rows if r[0] == "空数据演员")
    assert empty[7] == "2000-02-02"
    assert "身高: 160cm" in str(empty[8])

    # overwrite=True：覆盖已有生日/简介
    asyncio.run(actor_db_tool.run_actor_db_xlsx("fill_minnano", overwrite=True))

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    existing = next(r for r in rows if r[0] == "已有数据演员")
    assert existing[7] == "2000-02-02"  # 生日被覆盖
    assert "身高: 160cm" in str(existing[8])  # 简介被覆盖为结构化

    # overwrite=True 断点续传：再次运行，已结构化的行不再选中（不会再触发搜索）
    searched: list[str] = []

    async def counting_search(name):
        searched.append(name)
        return ("123", "<html>mock</html>")

    monkeypatch.setattr(minnano_crawler, "_search_minnano_by_name", counting_search)
    asyncio.run(actor_db_tool.run_actor_db_xlsx("fill_minnano", overwrite=True))
    assert searched == []  # 两行均已结构化，重跑不再请求


def test_resolve_bio_from_parsed_priority():
    """_resolve_bio_from_parsed：minnano 优先；无 bio 字段 fallback 本地 reformat；再退清洗。"""
    from mdcx.tools.actor_db_tool import _resolve_bio_from_parsed

    # minnano 有 bio 字段 → 用 minnano
    parsed = {"height": "160", "bust": "83", "waist": "58", "hip": "84", "cup": "F"}
    bio, source = _resolve_bio_from_parsed(parsed, "旧简介自由文本，身高160cm三围B83/W58/H84", "测试")
    assert source == "minnano"
    assert "身高: 160cm" in bio

    # minnano 查到但 bio 字段全空 → fallback 本地 reformat 原简介
    empty_parsed = {"height": "", "bust": "", "waist": "", "hip": "", "cup": "", "tags": []}
    bio, source = _resolve_bio_from_parsed(empty_parsed, "身高161cm，三围B85/W60/H88", "测试")
    assert source == "local"
    assert "身高: 161cm" in bio

    # minnano 完全没查到（None）→ 本地 reformat
    bio, source = _resolve_bio_from_parsed(None, "身高162cm", "测试")
    assert source == "local"
    assert "身高: 162cm" in bio

    # 本地也提不出字段 → 清洗原简介残留
    bio, source = _resolve_bio_from_parsed(None, "テスト（てすと / Tesuto），鞋码：S", "テスト")
    assert source == "clean"
    assert "鞋码" not in bio

    # 全空 → 空
    bio, source = _resolve_bio_from_parsed(None, "", "测试")
    assert source == "" and bio == ""
