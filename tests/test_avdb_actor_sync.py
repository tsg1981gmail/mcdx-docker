import asyncio
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from mdcx.config.resources import (
    COL_BIO,
    COL_BIRTH_DATE,
    COL_JP,
    COL_KEYWORD,
    COL_TMDBID,
    COL_ZH_CN,
    DB_HEADERS,
)
from mdcx.core import tmdb_actor
from mdcx.tools.actor_db_tool import sync_from_avdb
from mdcx.utils.xml_avdb import (
    clean_actor_value,
    extract_birth_date,
    parse_avdb_actor_mapping,
    strip_age_and_birth,
)


@pytest.fixture
def _tmp_actor_db(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setattr(tmdb_actor.manager, "data_folder", tmp_path)
    monkeypatch.setattr(tmdb_actor.resources, "actor_db", {})
    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    return userdata / "actor_database.xlsx"


@pytest.fixture(autouse=True)
def _reset_actor_db_row_index():
    with tmdb_actor._ACTOR_DB_ROW_INDEX_LOCK:
        tmdb_actor._ACTOR_DB_ROW_INDEX.clear()


def test_db_headers_has_nine_columns():
    assert len(DB_HEADERS) == 9
    assert DB_HEADERS[7] == "出生日期"
    assert DB_HEADERS[8] == "简介"
    assert COL_BIRTH_DATE == 7
    assert COL_BIO == 8


def test_update_actor_db_row_writes_birth_date_and_bio(_tmp_actor_db: Path):
    status = asyncio.run(
        tmdb_actor.update_actor_db_row(
            jp="新演员",
            zh_cn="新演员",
            tmdbid=99999,
            birth_date="1990-01-01",
            bio="身高160cm",
        )
    )
    assert status == "inserted_new_row"

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=COL_BIRTH_DATE + 1).value == "1990-01-01"
    assert ws.cell(row=2, column=COL_BIO + 1).value == "身高160cm"
    wb.close()


def test_update_actor_db_row_keeps_existing_birth_date(_tmp_actor_db: Path):
    asyncio.run(tmdb_actor.update_actor_db_row(jp="演员A", birth_date="1990-01-01", bio="旧简介"))
    asyncio.run(tmdb_actor.update_actor_db_row(jp="演员A", birth_date="2000-02-02", bio="新简介"))

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=COL_BIRTH_DATE + 1).value == "1990-01-01"
    assert ws.cell(row=2, column=COL_BIO + 1).value == "旧简介"
    wb.close()


_SAMPLE_XML = """<?xml version="1.0" encoding="UTF-8"?>
<actor-mapping>
  <actor>
    <a zh_cn="阿部純子" zh_tw="阿部純子" jp="阿部純子" keyword="Abe Junko,阿部純子" tmdb_id="1417328" verified="1" bio_graphy="安部純子（あべじゅんこ / Abe Junko），1993年06月05日出生，33岁，身高158cm，三围B86/W78/H82，籍贯东京都。" />
    <a zh_cn="阿部涼音" zh_tw="阿部涼音" jp="阿部涼音" keyword="阿部涼音" bio_graphy="阿部涼音（あべすずね），身高157cm，三围B88/W58/H87。" />
    <a jp="无名字段" />
  </actor>
  <actor-blacklist>
    黑名单演员
  </actor-blacklist>
</actor-mapping>
"""


def test_parse_count_and_blacklist_ignored():
    actors = parse_avdb_actor_mapping(_SAMPLE_XML)
    assert len(actors) == 3
    assert actors[0].tmdb_id == "1417328"
    assert actors[1].bio_graphy


def test_parse_missing_fields_are_empty():
    actors = parse_avdb_actor_mapping(_SAMPLE_XML)
    missing = actors[2]
    assert missing.zh_cn == ""
    assert missing.keyword == ""
    assert missing.tmdb_id == ""


def test_parse_invalid_xml_raises_value_error():
    try:
        parse_avdb_actor_mapping("<actor-mapping><actor>")
    except ValueError:
        return
    raise AssertionError("expected ValueError for invalid xml")


def test_extract_birth_date_full_formats():
    assert extract_birth_date("1993年06月05日出生") == "1993-06-05"
    assert extract_birth_date("1993年6月5日出生") == "1993-06-05"
    assert extract_birth_date("1993.10.18 出生") == "1993-10-18"
    assert extract_birth_date("1993/1/5出生") == "1993-01-05"
    assert extract_birth_date("1993-06-05 出生") == "1993-06-05"


def test_extract_birth_date_partial_and_missing():
    assert extract_birth_date("1993年6月出生") == "1993-06"
    assert extract_birth_date("出生于1993年") == "1993"
    assert extract_birth_date("身高158cm，三围B86/W78/H82") == ""
    assert extract_birth_date("") == ""


def test_extract_birth_date_rejects_debut_or_work_dates():
    """出道年份/作品发行日期不是出生日期，宁缺毋滥不提取。"""
    assert extract_birth_date("愛内ハル，2011年出道") == ""
    assert extract_birth_date("2013年11月30日出道作品：人妻の色香") == ""
    assert extract_birth_date("出道作品：初撮りおばさん（2013年03月25日）") == ""
    assert extract_birth_date("出身于三重县，身高160厘米，F罩杯，20") == ""
    assert extract_birth_date("1997年12月03日出生，28岁，身高160cm") == "1997-12-03"
    assert extract_birth_date("（1997-12-03）出道") == ""


def test_strip_age_and_birth_removes_dynamic_parts():
    bio = "安部純子，1993年06月05日出生，33岁，身高158cm，三围B86/W78/H82，籍贯东京都。"
    cleaned = strip_age_and_birth(bio, "1993-06-05")
    assert "1993" not in cleaned
    assert "33岁" not in cleaned
    assert "身高158cm" in cleaned
    assert "三围B86/W78/H82" in cleaned


def test_strip_age_and_birth_keeps_other_age_digits():
    bio = "身高158cm，三围B86/W78/H82。"
    cleaned = strip_age_and_birth(bio)
    assert "158cm" in cleaned
    assert "86/W78" in cleaned


def test_clean_actor_value_decodes_double_entities():
    assert clean_actor_value("&amp;quot;美咲&amp;quot;") == '"美咲"'


def test_clean_actor_value_removes_control_and_backslash_escapes():
    assert clean_actor_value("  美咲\n\t\x00\x1f  ") == "美咲"
    assert clean_actor_value("\\u4f50\\x41\\n山田") == "山田"


def test_clean_actor_value_trim():
    assert clean_actor_value("  三上悠亚  ") == "三上悠亚"
    assert clean_actor_value("") == ""


def _write_db(path: Path, rows, headers=DB_HEADERS):
    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    wb.save(path)
    wb.close()


def _read_rows(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


@pytest.fixture
def _avdb_xml(tmp_path: Path):
    path = tmp_path / "mapping.xml"
    path.write_text(_SAMPLE_XML, encoding="utf-8")
    return path


def test_sync_from_file_creates_rows(_tmp_actor_db: Path, _avdb_xml: Path):
    result = asyncio.run(sync_from_avdb("file", str(_avdb_xml)))
    assert result.downloaded is False
    assert result.parsed == 3
    assert result.created == 3
    assert result.filled == 0
    assert result.merged == 0
    assert result.failed == []

    rows = _read_rows(_tmp_actor_db)
    assert len(rows) == 3
    first = rows[0]
    assert first[COL_JP] == "阿部純子"
    assert first[COL_BIRTH_DATE] == "1993-06-05"
    assert "身高158cm" in str(first[COL_BIO])
    assert first[COL_TMDBID] == 1417328


def test_sync_keeps_existing_local_values(_tmp_actor_db: Path, _avdb_xml: Path):
    _write_db(_tmp_actor_db, [["阿部純子", "本地名", "", "", "", "", "", "", ""]])
    result = asyncio.run(sync_from_avdb("file", str(_avdb_xml)))
    assert result.created == 2
    assert result.filled == 1

    rows = _read_rows(_tmp_actor_db)
    assert len(rows) == 3
    first = rows[0]
    assert first[COL_ZH_CN] == "本地名"
    assert first[COL_BIRTH_DATE] == "1993-06-05"


def test_sync_merges_keywords_and_sorts(_tmp_actor_db: Path, _avdb_xml: Path):
    _write_db(_tmp_actor_db, [["阿部涼音", "", "", "旧别名", "", "", "", "", ""]])
    result = asyncio.run(sync_from_avdb("file", str(_avdb_xml)))
    assert result.filled == 1

    rows = _read_rows(_tmp_actor_db)
    keywords = [k for k in str(rows[0][COL_KEYWORD]).split(",") if k]
    assert keywords == sorted(["旧别名", "阿部涼音"])


def test_sync_keeps_keyword_casefold_dedup(_tmp_actor_db: Path, _avdb_xml: Path):
    _write_db(_tmp_actor_db, [["阿部純子", "", "", "ABE JUNKO", "", "", "", "", ""]])
    asyncio.run(sync_from_avdb("file", str(_avdb_xml)))
    rows = _read_rows(_tmp_actor_db)
    row = next(r for r in rows if r[COL_JP] == "阿部純子")
    kws = [k for k in str(row[COL_KEYWORD]).split(",") if k]
    assert kws == sorted(["ABE JUNKO", "阿部純子"])


def test_sync_tmdbid_conflict_merges_into_existing_row(_tmp_actor_db: Path, _avdb_xml: Path):
    _write_db(_tmp_actor_db, [["演员A", "演员A", "", "", "", 1417328, "", "", ""]])
    result = asyncio.run(sync_from_avdb("file", str(_avdb_xml)))
    assert result.merged == 1
    assert result.created == 2

    rows = _read_rows(_tmp_actor_db)
    assert len(rows) == 3
    assert rows[0][COL_JP] == "演员A"
    keywords = [k for k in str(rows[0][COL_KEYWORD]).split(",") if k]
    assert "阿部純子" in keywords


def test_sync_supports_legacy_seven_column_db(_tmp_actor_db: Path, _avdb_xml: Path):
    _write_db(_tmp_actor_db, [["阿部涼音", "", "", "旧别名", "", "", ""]], headers=DB_HEADERS[:7])
    result = asyncio.run(sync_from_avdb("file", str(_avdb_xml)))
    assert result.failed == []
    rows = _read_rows(_tmp_actor_db)
    assert len(rows) == 3
    assert rows[0][COL_BIRTH_DATE] is None or rows[0][COL_BIRTH_DATE] == ""


def test_sync_upgrades_legacy_headers_to_nine_cols(_tmp_actor_db: Path, _avdb_xml: Path):
    _write_db(_tmp_actor_db, [["阿部涼音", "", "", "", "", "", ""]], headers=DB_HEADERS[:7])
    asyncio.run(sync_from_avdb("file", str(_avdb_xml)))
    wb = load_workbook(_tmp_actor_db)
    headers = [c.value for c in wb.active[1]]
    wb.close()
    assert headers == list(DB_HEADERS)


def test_sync_local_file_missing(_tmp_actor_db: Path, tmp_path: Path):
    result = asyncio.run(sync_from_avdb("file", str(tmp_path / "none.xml")))
    assert result.failed
    assert result.parsed == 0


def test_sync_invalid_xml(_tmp_actor_db: Path, tmp_path: Path):
    bad = tmp_path / "bad.xml"
    bad.write_text("<actor-mapping><actor>", encoding="utf-8")
    result = asyncio.run(sync_from_avdb("file", str(bad)))
    assert result.failed
    assert result.parsed == 0


def test_sync_unknown_source_uses_github_url(_tmp_actor_db: Path, tmp_path: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["阿部涼音", "", "", "", "", "", "", "", ""]])
    captured = {}

    async def _fake_download(url, file_path, folder):
        captured["url"] = url
        file_path.write_text(_SAMPLE_XML, encoding="utf-8")
        return True

    monkeypatch.setattr("mdcx.base.web.download_file_with_filepath", _fake_download)
    result = asyncio.run(sync_from_avdb("github"))
    assert result.downloaded is True
    assert result.failed == []
    assert captured["url"].startswith("https://raw.githubusercontent.com")


def test_sync_jsdelivr_source_uses_mirror_url(_tmp_actor_db: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["阿部涼音", "", "", "", "", "", "", "", ""]])
    captured = {}

    async def _fake_download(url, file_path, folder):
        captured["url"] = url
        file_path.write_text(_SAMPLE_XML, encoding="utf-8")
        return True

    monkeypatch.setattr("mdcx.base.web.download_file_with_filepath", _fake_download)
    result = asyncio.run(sync_from_avdb("jsdelivr"))
    assert result.downloaded is True
    assert result.failed == []
    assert "cdn.jsdelivr.net" in captured["url"]
