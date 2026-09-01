"""校验 scripts/merge_actor_db_duplicates.py 的重复筛查与合并逻辑。"""

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

from mdcx.config.resources import DB_HEADERS

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts import merge_actor_db_duplicates as mod  # noqa: E402


def _make_db(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(DB_HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _read_rows(path: Path):
    wb = load_workbook(path)
    ws = wb["演员数据库"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def _scan_rows(rows):
    return mod.scan(rows)


def test_scan_detects_alias_of_owner(tmp_path):
    """无 id 行别名含对方主名时判定为确证重复。"""
    rows = [
        ["有id演员", "有id", "", "别名甲", "", "1001", "", "", ""],
        ["无id演员", "无id", "", "有id演员,别字乙", "", "", "", "", ""],
    ]
    pairs = _scan_rows(rows)
    assert ("无id演员", "有id演员", "1001") in pairs


def test_scan_ignores_trailing_space_variant(tmp_path):
    """主名带空格（如「白城 リサ」）经规范化后仍能命中。"""
    rows = [
        ["白城 リサ", "白城リサ", "", "Risa", "", "3182698", "", "", ""],
        ["立花愛恵", "立花愛恵", "", "白城リサ,あすか", "", "", "", "", ""],
    ]
    pairs = _scan_rows(rows)
    assert ("立花愛恵", "白城 リサ", "3182698") in pairs


def test_scan_ignores_self_and_common_short_names(tmp_path):
    """无 id 行自身不匹配；短名/常见名不产生假阳性。"""
    rows = [
        ["みちる", "みちる", "", "", "", "5557511", "", "", ""],
        ["高梨静香", "高梨静香", "", "みちる,あや,らむ", "", "", "", "", ""],
    ]
    pairs = _scan_rows(rows)
    # 别名「みちる」= 有id主名 -> 确证
    assert ("高梨静香", "みちる", "5557511") in pairs
    # 有 id 行自身不会出现在无 id 侧
    assert not any(jp == "みちる" for jp, _, _ in pairs)


def test_scan_no_match_when_alias_not_owner(tmp_path):
    """别名只是普通短名、不命中任何有 id 主名时不算重复。"""
    rows = [
        ["有id演员", "有id", "", "はるか", "", "1001", "", "", ""],
        ["无id演员", "无id", "", "はるか", "", "", "", "", ""],
    ]
    # 两行都有「はるか」但无 id 行的别名「はるか」不是有 id 行主名 -> 不重复
    assert _scan_rows(rows) == []


def test_merge_merges_aliases_and_fills_fields(tmp_path):
    """合并：别名并入、空缺字段补全、无 id 行删除。"""
    db = tmp_path / "actor_database.xlsx"
    _make_db(
        db,
        [
            ["有id演员", "", "", "既有别名", "", "1001", "http://url", "", ""],
            ["无id演员", "中文名", "", "有id演员,新别名", "", "", "", "1990-01-01", "简介"],
        ],
    )
    wb = load_workbook(db)
    ws = wb["演员数据库"]
    pairs = mod.scan(mod._rows_of(ws))
    merged, removed = mod.merge_duplicates(ws, pairs)
    wb.save(db)
    wb.close()

    assert merged == 1
    assert removed == 1

    rows = _read_rows(db)
    assert len(rows) == 1
    row = rows[0]
    assert row[0] == "有id演员"
    kw = {k.strip() for k in str(row[3]).split(",") if k.strip()}
    assert {"有id演员", "既有别名", "新别名"} <= kw
    assert row[1] == "中文名"  # 空缺字段补全
    assert row[7] == "1990-01-01"
    assert row[8] == "简介"


def test_merge_two_sources_into_one_owner(tmp_path):
    """两个无 id 行并入同一个有 id 行（多艺名同人）。"""
    db = tmp_path / "actor_database.xlsx"
    _make_db(
        db,
        [
            ["みちる", "みちる", "", "", "", "5557511", "", "", ""],
            ["草薙千沙", "", "", "みちる,川美優香", "", "", "", "", ""],
            ["高梨静香", "", "", "みちる,あや", "", "", "", "", ""],
        ],
    )
    wb = load_workbook(db)
    ws = wb["演员数据库"]
    pairs = mod.scan(mod._rows_of(ws))
    merged, removed = mod.merge_duplicates(ws, pairs)
    wb.save(db)
    wb.close()

    assert merged == 1
    assert removed == 2
    rows = _read_rows(db)
    assert len(rows) == 1
    assert rows[0][0] == "みちる"
    kw = {k.strip() for k in str(rows[0][3]).split(",") if k.strip()}
    assert {"草薙千沙", "高梨静香", "川美優香", "あや"} <= kw
