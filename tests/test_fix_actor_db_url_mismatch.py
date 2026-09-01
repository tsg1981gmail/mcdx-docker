"""校验 scripts/fix_actor_db_url_mismatch.py 的 url 错配修复逻辑。"""

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook  # noqa: E402

from mdcx.config.resources import DB_HEADERS  # noqa: E402
from scripts import fix_actor_db_url_mismatch as mod  # noqa: E402


def _make_db(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(DB_HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _read_rows(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["演员数据库"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def test_clears_url_for_noid_rows(tmp_path):
    """无 id 但有 url 的行：url 被清空，行保留。"""
    db = tmp_path / "actor_database.xlsx"
    _make_db(
        db,
        [
            ["阿部真琴", "阿部真琴", "", "", "", "", "https://www.themoviedb.org/person/2132746", "", ""],  # 无id有url
            [
                "阿部乃みく",
                "阿部乃みく",
                "",
                "",
                "",
                "2132746",
                "https://www.themoviedb.org/person/2132746",
                "",
                "",
            ],  # 正常
        ],
    )
    rc = mod.main(["--db", str(db), "--apply"])
    assert rc == 0
    rows = _read_rows(db)
    assert len(rows) == 2
    noid = next(r for r in rows if r[0] == "阿部真琴")
    assert noid[6] in (None, "")
    normal = next(r for r in rows if r[0] == "阿部乃みく")
    assert normal[6] == "https://www.themoviedb.org/person/2132746"


def test_deletes_jp_empty_rows(tmp_path):
    """无 jp 的垃圾行（只有 url 或全空）被整行删除。"""
    db = tmp_path / "actor_database.xlsx"
    _make_db(
        db,
        [
            ["正常演员", "正常", "", "", "", "1001", "https://www.themoviedb.org/person/1001", "", ""],
            ["", "", "", "", "", "", "https://www.themoviedb.org/person/999", "", ""],  # 无jp有url
            ["", "", "", "", "", "", "", "", ""],  # 全空
        ],
    )
    rc = mod.main(["--db", str(db), "--apply"])
    assert rc == 0
    rows = _read_rows(db)
    assert len(rows) == 1
    assert rows[0][0] == "正常演员"
