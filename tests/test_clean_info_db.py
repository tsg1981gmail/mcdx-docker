"""校验 scripts/clean_info_db.py 的合并与清洗逻辑。"""

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook  # noqa: E402

from scripts import clean_info_db as mod  # noqa: E402


def _make_db(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(["jp", "zh_cn", "zh_tw", "keyword"])
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def test_scan_jp_duplicate():
    rows = [
        ["たとえば熟女がコスッたら!?", "假如熟女玩COS呢!?", "假如熟女玩COS呢!?", ",たとえば熟女がコスッたら!?,"],
        [
            "たとえば熟女がコスッたら！？",
            "假如熟女玩COS呢！？",
            "假如熟女玩COS呢！？",
            ",たとえば熟女がコスッたら！？,",
        ],
    ]
    pairs = mod.scan(rows)
    assert len(pairs) == 1
    main, dup = pairs[0]
    assert main == 0 and dup == 1


def test_scan_skip_delete_rows():
    rows = [
        ["删除", "删除", "删除", ",词A,"],
        ["删除", "删除", "删除", ",词B,"],
    ]
    assert mod.scan(rows) == []


def test_dedup_keyword_parts():
    assert mod._dedup_keyword_parts(",3P,３P,") == ",3P,"
    assert mod._dedup_keyword_parts(",OL,ＯＬ,") == ",OL,"
    assert mod._dedup_keyword_parts(",marx,MARX,") == ",marx,"
    assert mod._dedup_keyword_parts(",3P,") == ",3P,"


def test_build_fix_plan_dedup():
    rows = [["3P", "3P", "3P", ",3P,３P,"]]
    plan = mod.build_fix_plan(rows)
    assert plan[0] == ",3P,"


def test_build_fix_plan_empty_kw():
    rows = [["M女", "M女", "M女", ",,"]]
    plan = mod.build_fix_plan(rows)
    assert plan[0] == ",M女,"


def test_build_fix_plan_jp_merge():
    rows = [
        ["たとえば熟女がコスッたら!?", "假如熟女玩COS呢!?", "假如熟女玩COS呢!?", ",たとえば熟女がコスッたら!?,"],
        [
            "たとえば熟女がコスッたら！？",
            "假如熟女玩COS呢！？",
            "假如熟女玩COS呢！？",
            ",たとえば熟女がコスッたら！？,",
        ],
    ]
    plan = mod.build_fix_plan(rows)
    assert 0 in plan
    assert 1 not in plan
    merged = plan[0]
    assert "たとえば熟女がコスッたら！？" in merged
    assert merged.endswith(",") and merged.startswith(",")


def test_apply_fixes(tmp_path):
    p = tmp_path / "info_database.xlsx"
    rows = [
        ["たとえば熟女がコスッたら!?", "假如熟女玩COS呢!?", "假如熟女玩COS呢!?", ",たとえば熟女がコスッたら!?,"],
        [
            "たとえば熟女がコスッたら！？",
            "假如熟女玩COS呢！？",
            "假如熟女玩COS呢！？",
            ",たとえば熟女がコスッたら！？,",
        ],
        ["3P", "3P", "3P", ",3P,３P,"],
    ]
    _make_db(p, rows)
    wb = __import__("openpyxl").load_workbook(p)
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    plan = mod.build_fix_plan(data_rows)
    fixed, removed = mod.apply_fixes(ws, data_rows, plan, {d for _, d in mod.scan(data_rows)})
    wb.save(p)
    wb.close()
    assert fixed == 2 and removed == 1
    from scripts.check_info_db import check_xlsx

    assert check_xlsx(p) == 0
