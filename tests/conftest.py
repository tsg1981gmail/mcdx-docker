import sys
import tempfile
import types
from pathlib import Path
from types import SimpleNamespace

import pytest

from mdcx.config.enums import Language
from mdcx.config.models import Config

# 生产 Config 的默认实例，用作 _DummyConfig 缺失字段的透明回退来源，
# 避免桩缺字段（枚举列表 / 复杂默认值等）导致测试在访问 config.xxx 时报 AttributeError。
_REAL_CONFIG = Config()


@pytest.fixture(autouse=True)
def _reset_dmm_upgrade_cache():
    # upgrade_dmm_cover 的进程内缓存是模块级全局状态，测试间必须复位，
    # 否则 test_javbus / test_r18dev 的成败会依赖测试执行顺序。
    from mdcx.crawlers.dmm_direct import _clear_dmm_upgrade_cache

    _clear_dmm_upgrade_cache()
    yield
    _clear_dmm_upgrade_cache()


class _DummySignal:
    def emit(self, *args, **kwargs):
        return None

    def connect(self, *args, **kwargs):
        # 启动冒烟测试会执行 Init_Singal 的全部 .connect 调用
        return None

    def disconnect(self, *args, **kwargs):
        return None


class _DummySignals:
    def __init__(self):
        self.stop = False
        self.log_text = _DummySignal()
        self.scrape_info = _DummySignal()
        self.net_info = _DummySignal()
        self.exec_set_main_info = _DummySignal()
        self.change_buttons_status = _DummySignal()
        self.reset_buttons_status = _DummySignal()
        self.set_label_file_path = _DummySignal()
        self.label_result = _DummySignal()
        self.logs_failed_settext = _DummySignal()
        self.view_success_file_settext = _DummySignal()
        self.exec_set_processbar = _DummySignal()
        self.exec_exit_app = _DummySignal()
        self.view_failed_list_settext = _DummySignal()
        self.exec_show_list_name = _DummySignal()
        self.logs_failed_show = _DummySignal()

    def add_log(self, *args, **kwargs):
        return None

    def get_log(self, *args, **kwargs):
        # show_detail_log 是主窗口 timer.timeout 回调，缺此方法会抛
        # AttributeError，PyQt6 下槽内未捕获异常会触发 qFatal abort。
        return ""

    def show_traceback_log(self, *args, **kwargs):
        return None

    def show_log_text(self, *args, **kwargs):
        return None

    def show_scrape_info(self, *args, **kwargs):
        return None

    def show_net_info(self, *args, **kwargs):
        return None

    def set_main_info(self, *args, **kwargs):
        return None

    def show_list_name(self, *args, **kwargs):
        return None


class _DummyConfig:
    def __init__(self):
        self.show_data_log = False
        self.download_files = []
        self.keep_files = []
        self.outline_format = []
        self.main_mode = 1
        self.naming_media = "number title"
        self.update_titletemplate = "number title"
        self.nfo_include_new = []
        self.actor_no_name = "未知演员"
        self.read_mode = []
        self.nfo_tag_series = "series"
        self.nfo_tagline = "release"
        self.add_genre_to_tag = False
        self.extra_genres = False
        self.add_tag_to_genre = False
        self.original_nfo_title = False
        self.use_simple_tag = False
        self.website = ""
        self.tmdb_api_key = ""
        self.tmdb_api_base = ""
        self.server_type = "emby"
        self.folder_name_max = 60
        self.media_path = "./media"
        self.api_key = ""
        self.emby_url = ""
        self.user_id = ""
        self.extrafanart_folder = "extrafanart_copy"
        self.failed_output_folder = "failed"
        self.folders = ["JAV_output", "examples"]
        self.scrape_softlink_path = False
        self.softlink_path = "softlink"
        self.success_output_folder = "JAV_output"

    def __getattr__(self, name):
        # 桩上未显式设置的字段，透明回退到生产 Config 默认值。
        # 这样无论哪个测试访问 config.<任意字段>，都不会因桩缺字段而报 AttributeError。
        try:
            return getattr(_REAL_CONFIG, name)
        except AttributeError:
            raise AttributeError(f"_DummyConfig object has no attribute {name!r}")

    def get_field_config(self, _field):
        return SimpleNamespace(language=Language.JP, translate=False)


class _DummyAsyncClient:
    async def get_json(self, *args, **kwargs):
        raise NotImplementedError("Tests should monkeypatch async_client.get_json")

    async def post_content(self, *args, **kwargs):
        raise NotImplementedError("Tests should monkeypatch async_client.post_content")

    async def request(self, *args, **kwargs):
        raise NotImplementedError("Tests should monkeypatch async_client.request")

    async def _close_response(self, response=None):
        # 真实 AsyncWebClient 的辅助方法（关闭响应），测试无需 mock，提供 no-op。
        return None

    def __getattr__(self, name):
        # 桩上未显式实现的方法（如 get_text / _close_response 等）：
        # 返回一个协程桩，使得 monkeypatch.setattr(async_client, name, ...) 能定位到属性，
        # 而测试若忘记打桩就直接调用时，以清晰的 NotImplementedError 失败（而非 AttributeError）。
        if name.startswith("__") and name.endswith("__"):
            raise AttributeError(name)

        async def _stub(*args, **kwargs):
            raise NotImplementedError(f"Tests should monkeypatch async_client.{name}")

        return _stub


class _DummyLLMClient:
    async def ask(self, *args, **kwargs):
        raise NotImplementedError("Tests should monkeypatch llm_client.ask")

    def __getattr__(self, name):
        if name.startswith("__") and name.endswith("__"):
            raise AttributeError(name)

        async def _stub(*args, **kwargs):
            raise NotImplementedError(f"Tests should monkeypatch llm_client.{name}")

        return _stub


class _DummyComputed:
    def __init__(self):
        self.async_client = _DummyAsyncClient()
        self.llm_client = _DummyLLMClient()
        # get_file_info_v2 等会从 manager.computed.escape_string_list 读取，
        # 桩需提供该属性，否则解析番号时会抛 AttributeError 被吞成空番号。
        self.escape_string_list: list[str] = []
        self.official_websites: dict[str, str] = {}

    def retain(self):
        return None

    def release(self):
        return None


class _DummyComputedLease:
    def __init__(self, computed):
        self._computed = computed

    async def __aenter__(self):
        return self._computed

    async def __aexit__(self, *exc):
        return None


class _DummyManager:
    def __init__(self):
        self.config = _DummyConfig()
        self.data_folder = Path(tempfile.gettempdir()) / "mdcx-tests"
        self.file = "config.json"
        self.path = self.data_folder / self.file
        self.computed = _DummyComputed()

    def load(self):
        # 主窗口启动冒烟测试 load_config 会调用；返回空错误列表。
        return []

    def reset(self):
        # 主窗口启动冒烟测试：写默认配置文件，保证 load_config 重入后走
        # 「配置存在」分支，避免 init_config → reset → load_config 无限递归。
        self.data_folder.mkdir(parents=True, exist_ok=True)
        self.path.write_text("{}", encoding="utf-8")

    def list_configs(self):
        # 列出配置目录下的 *.json；目录不存在时回退默认项。
        try:
            files = sorted(p.name for p in self.data_folder.glob("*.json"))
        except OSError:
            files = []
        return files if files else ["config.json"]

    def acquire_computed(self):
        return _DummyComputedLease(self.computed)


class _DummyResources:
    # 主窗口启动冒烟测试用到的图标资源（空路径即可，QIcon 静默忽略）。
    # 注意：不能用 __getattr__ 兜底返回 "" —— 生产代码（如 is_male_actor）
    # 依赖缺失资源方法时抛 AttributeError 做优雅降级，兜底会把方法调用变成
    # "'str' object is not callable"。
    _ICON_ATTRS = (
        "clear_tree_icon",
        "del_file_icon",
        "del_folder_icon",
        "help_icon",
        "hide_boss_icon",
        "hide_logs_icon",
        "home_icon",
        "icon_ico",
        "input_number_icon",
        "input_website_icon",
        "log_icon",
        "net_icon",
        "open_folder_icon",
        "open_nfo_icon",
        "play_icon",
        "right_menu",
        "save_failed_list_icon",
        "setting_icon",
        "show_logs_icon",
        "start_icon",
        "stop_icon",
        "tool_icon",
    )

    def __init__(self):
        for name in self._ICON_ATTRS:
            setattr(self, name, "")
        self.actor_db = {}
        self.actor_db_reverse_index = None
        self.info_db = []
        self.info_mapping_data = None

    def u(self, relative_path):
        return manager_module.manager.data_folder / "userdata" / relative_path

    def get_fonts(self):
        # 主窗口启动冒烟测试会调用（真实实现读资源字体文件）
        return None

    def start_data_loading(self):
        # 主窗口启动冒烟测试会调用（真实实现后台线程加载数据库）
        return None

    def qtr(self, relative_path):
        return str(self.u(relative_path))

    def get_actor_data(self, actor):
        from mdcx.core.tmdb_actor import search_actor_db_reverse

        actor_data = {
            "zh_cn": actor,
            "zh_tw": actor,
            "jp": actor,
            "keyword": [actor],
            "href": "",
            "has_name": False,
        }

        if self.actor_db is not None:
            row = search_actor_db_reverse(actor)
            if row:
                actor_data["zh_cn"] = row.get("zh_cn") or actor
                actor_data["zh_tw"] = row.get("zh_tw") or actor
                actor_data["jp"] = row.get("jp") or actor
                kw = row.get("keyword") or ""
                actor_data["keyword"] = [k.strip() for k in kw.split(",") if k.strip()] if kw else [actor_data["jp"]]
                actor_data["href"] = row.get("href") or ""
                actor_data["has_name"] = True
        return actor_data

    def get_info_data(self, info):
        info_data = {
            "zh_cn": info,
            "zh_tw": info,
            "jp": info,
            "keyword": [info],
            "has_name": False,
        }

        info_key = info.upper()
        for row in self.info_db or []:
            matched = False
            if row.get("keyword"):
                for kw in row["keyword"].split(","):
                    kw = kw.strip()
                    if kw and kw.upper() == info_key:
                        matched = True
                        break
            if not matched:
                for attr in ("zh_cn", "zh_tw", "jp"):
                    if (row.get(attr) or "").upper() == info_key:
                        matched = True
                        break
            if matched:
                info_data["zh_cn"] = row.get("zh_cn") or info
                info_data["zh_tw"] = row.get("zh_tw") or info
                info_data["jp"] = row.get("jp") or info
                kw = row.get("keyword") or ""
                info_data["keyword"] = [k.strip() for k in kw.split(",") if k.strip()] if kw else [info]
                info_data["has_name"] = True
                return info_data
        return info_data

    def reload_actor_db(self):
        import openpyxl

        db_path = self.u("actor_database.xlsx")
        if not db_path.exists():
            self.actor_db = None
            self.actor_db_reverse_index = None
            return

        old_actor_db = self.actor_db
        try:
            wb = openpyxl.load_workbook(db_path, read_only=True, data_only=True)
            ws = wb.active
            db = {}
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    continue
                jp = str(row[0] or "").strip()
                if not jp:
                    continue
                tmdbid_val = None
                if len(row) > 5:
                    tmdbid_raw = str(row[5] or "").strip()
                    if tmdbid_raw.isdigit():
                        tmdbid_val = int(tmdbid_raw)
                db[jp] = {
                    "zh_cn": str(row[1] or "").strip() if len(row) > 1 else "",
                    "zh_tw": str(row[2] or "").strip() if len(row) > 2 else "",
                    "keyword": str(row[3] or "").strip() if len(row) > 3 else "",
                    "href": str(row[4] or "").strip() if len(row) > 4 else "",
                    "tmdbid": tmdbid_val,
                    "tmdb_url": str(row[6] or "").strip() if len(row) > 6 else "",
                }
            wb.close()
            self.actor_db = db
            self.actor_db_reverse_index = None
        except Exception:
            self.actor_db = old_actor_db


manager_module = types.ModuleType("mdcx.config.manager")
manager_module.manager = _DummyManager()
manager_module.get_new_str = lambda a, wanted=False: a
sys.modules.setdefault("mdcx.config.manager", manager_module)

resources_module = types.ModuleType("mdcx.config.resources")
resources_module.COL_JP = 0
resources_module.COL_ZH_CN = 1
resources_module.COL_ZH_TW = 2
resources_module.COL_KEYWORD = 3
resources_module.COL_HREF = 4
resources_module.COL_TMDBID = 5
resources_module.COL_TMDB_URL = 6
resources_module.COL_BIRTH_DATE = 7
resources_module.COL_BIO = 8
resources_module.DB_HEADERS = ["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"]
resources_module.ACTOR_DB_SHEET = "演员数据库"
resources_module._tmdb_person_url = lambda tid: f"https://www.themoviedb.org/person/{tid}"
resources_module.read_actor_db_xlsx = lambda db_path: {}
resources_module.get_actor_db_sheet = lambda wb: (
    wb["演员数据库"] if "演员数据库" in getattr(wb, "sheetnames", []) else wb.active
)
resources_module.resources = _DummyResources()
sys.modules.setdefault("mdcx.config.resources", resources_module)

signals_module = types.ModuleType("mdcx.signals")
signals_module.signal = _DummySignals()
signals_module.signal_qt = signals_module.signal
signals_module.set_signal = lambda signal_instance: None
sys.modules.setdefault("mdcx.signals", signals_module)
