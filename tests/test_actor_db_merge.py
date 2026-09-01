from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

import mdcx.config.resources as res_module

_HEADERS = list(res_module.DB_HEADERS)


def _load_merge_func():
    """从源码提取真实 merge 相关函数（绕过 conftest 模块替换与 Resources 初始化）。"""
    import ast

    # tests/ 与 mdcx/ 同属项目根，用 __file__ 相对定位源码（CI 与本地一致）
    src = Path(__file__).resolve().parent.parent / "mdcx" / "config" / "resources.py"
    src_text = src.read_text(encoding="utf-8")
    tree = ast.parse(src_text)
    g = dict(vars(res_module))
    g["Path"] = Path
    g["openpyxl"] = __import__("openpyxl")
    g["LogBuffer"] = mock.MagicMock()
    # 先执行依赖的辅助函数，再执行主函数
    for node in tree.body:
        if isinstance(node, (ast.FunctionDef,)) and node.name in (
            "_alias_sort_key",
            "_merge_keyword_union",
            "merge_actor_db_from_backup",
        ):
            exec(ast.get_source_segment(src_text, node), g)
    return g["merge_actor_db_from_backup"]


_MERGE_FUNC = None


def _merge(backup_path, local_path):
    global _MERGE_FUNC
    if _MERGE_FUNC is None:
        _MERGE_FUNC = _load_merge_func()
    return _MERGE_FUNC(backup_path, local_path)


def _make_db(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(_HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _read_rows(path: Path):
    wb = load_workbook(path)
    ws = wb["演员数据库"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def _mock_paths(tmp_path: Path, local_rows, backup_rows):
    """构造出厂库(backup)与用户库(local)。"""
    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    local_path = userdata / "actor_database.xlsx"
    backup_path = tmp_path / "backup" / "actor_database.xlsx"
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    _make_db(local_path, local_rows)
    _make_db(backup_path, backup_rows)
    return backup_path, local_path


def test_merge_adds_new_entries_only(tmp_path):
    backup_path, local_path = _mock_paths(
        tmp_path,
        local_rows=[["已有演员", "已有", "", "", "", "1001", "", "", ""]],
        backup_rows=[
            ["已有演员", "已有", "", "", "", "1001", "", "", ""],  # 用户库已有
            ["新增演员", "新演员", "", "", "", "2002", "", "", ""],  # 出厂库新增
        ],
    )
    _merge(backup_path, local_path)
    rows = _read_rows(local_path)
    jps = {r[0] for r in rows}
    assert "新增演员" in jps
    assert "已有演员" in jps
    assert len(rows) == 2


def test_merge_does_not_delete_user_entries(tmp_path):
    backup_path, local_path = _mock_paths(
        tmp_path,
        local_rows=[["用户独有", "用户", "", "", "", "", "", "", ""]],  # 用户库有，出厂库无
        backup_rows=[],
    )
    _merge(backup_path, local_path)
    rows = _read_rows(local_path)
    assert {r[0] for r in rows} == {"用户独有"}  # 出厂库为空，用户库条目不被删


def test_merge_fills_empty_fields_only(tmp_path):
    backup_path, local_path = _mock_paths(
        tmp_path,
        local_rows=[["三上悠亜", "三上", "", "", "", "", "", "", ""]],  # 用户库空 tmdbid/生日
        backup_rows=[["三上悠亜", "三上", "三上", "Yua", "http://x", "9988", "", "1993-08-16", "简介"]],
    )
    _merge(backup_path, local_path)
    rows = _read_rows(local_path)
    row = [r for r in rows if r[0] == "三上悠亜"][0]
    assert str(row[5]) == "9988"  # tmdbid 补全
    assert row[7] == "1993-08-16"  # 生日补全
    assert row[1] == "三上"  # 已有值保留


def test_merge_does_not_overwrite_existing_values(tmp_path):
    backup_path, local_path = _mock_paths(
        tmp_path,
        local_rows=[["橋本ありな", "桥本有菜", "", "", "", "5555", "", "", "用户自己的简介"]],
        backup_rows=[["橋本ありな", "其他中文名", "", "", "", "6666", "", "1990-01-01", "出厂简介"]],
    )
    _merge(backup_path, local_path)
    rows = _read_rows(local_path)
    row = [r for r in rows if r[0] == "橋本ありな"][0]
    assert row[1] == "桥本有菜"  # 中文名不覆盖
    assert str(row[5]) == "5555"  # tmdbid 不覆盖（用户已有值）
    assert row[7] == "1990-01-01"  # 空缺生日被补全
    assert row[8] == "用户自己的简介"  # 简介不覆盖


def test_merge_skipped_when_backup_unchanged(tmp_path):
    backup_path, local_path = _mock_paths(
        tmp_path,
        local_rows=[["已有演员", "已有", "", "", "", "1001", "", "", ""]],
        backup_rows=[["已有演员", "已有", "", "", "", "1001", "", "", ""]],
    )
    _merge(backup_path, local_path)
    # 出厂库更新（发版），md5 变化应重新合并
    _make_db(
        backup_path,
        [["已有演员", "已有", "", "", "", "1001", "", "", ""], ["新二号", "新", "", "", "", "3003", "", "", ""]],
    )
    _merge(backup_path, local_path)
    rows = _read_rows(local_path)
    assert {r[0] for r in rows} == {"已有演员", "新二号"}


def test_merge_marker_prevents_second_scan(tmp_path):
    backup_path, local_path = _mock_paths(
        tmp_path,
        local_rows=[["已有演员", "已有", "", "", "", "1001", "", "", ""]],
        backup_rows=[["已有演员", "已有", "", "", "", "1001", "", "", ""]],
    )
    _merge(backup_path, local_path)
    marker = tmp_path / "userdata" / ".actor_db_merge_marker"
    assert marker.exists()  # 合并后写入 marker
    # 再次合并（出厂库未变）应跳过——不重复写文件
    import os

    mtime_before = os.path.getmtime(local_path)
    _merge(backup_path, local_path)
    assert os.path.getmtime(local_path) == mtime_before  # 文件未重写


def test_merge_unions_aliases(tmp_path):
    """别名列做并集合并：用户库已有别名 + 出厂库新增别名都保留。"""
    backup_path, local_path = _mock_paths(
        tmp_path,
        local_rows=[["橋本ありな", "", "", "新有菜,橋本ありな", "", "1001", "", "", ""]],
        backup_rows=[["橋本ありな", "", "", "Arata Arina,桥本有菜", "", "1001", "", "", ""]],
    )
    _merge(backup_path, local_path)
    rows = _read_rows(local_path)
    row = [r for r in rows if r[0] == "橋本ありな"][0]
    kw = str(row[3] or "")
    # 并集包含用户库与出厂库全部别名
    for alias in ["新有菜", "橋本ありな", "Arata Arina", "桥本有菜"]:
        assert alias in kw.split(","), f"别名 {alias} 应被合并"


def test_merge_alias_sorted_chinese_first(tmp_path):
    """别名并集后按中文(汉字)优先 -> 日文(假名) -> 罗马音排序。"""
    backup_path, local_path = _mock_paths(
        tmp_path,
        local_rows=[["三上悠亜", "", "", "Mikami Yua", "", "1001", "", "", ""]],
        backup_rows=[["三上悠亜", "", "", "三上悠亜,鬼頭桃菜", "", "1001", "", "", ""]],
    )
    _merge(backup_path, local_path)
    rows = _read_rows(local_path)
    row = [r for r in rows if r[0] == "三上悠亜"][0]
    aliases = str(row[3] or "").split(",")
    # 中文/汉字在前，罗马音在后
    assert aliases[0] == "三上悠亜"  # 汉字优先
    assert "Mikami Yua" in aliases
    assert aliases.index("Mikami Yua") > aliases.index("三上悠亜")  # 罗马音最后


def test_merge_alias_dedup(tmp_path):
    """别名并集去重：重复别名只保留一份。"""
    backup_path, local_path = _mock_paths(
        tmp_path,
        local_rows=[["橋本ありな", "", "", "橋本ありな,新有菜", "", "1001", "", "", ""]],
        backup_rows=[["橋本ありな", "", "", "新有菜,橋本ありな", "", "1001", "", "", ""]],
    )
    _merge(backup_path, local_path)
    rows = _read_rows(local_path)
    row = [r for r in rows if r[0] == "橋本ありな"][0]
    aliases = str(row[3] or "").split(",")
    assert len(aliases) == len({a.casefold() for a in aliases})  # 无重复
