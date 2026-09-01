import pytest
from openpyxl import load_workbook

from mdcx.core import amazon_database
from mdcx.models.log_buffer import LogBuffer


@pytest.fixture
def _tmp_asin_db(monkeypatch: pytest.MonkeyPatch, tmp_path):
    from mdcx.config import manager

    monkeypatch.setattr(manager.manager, "data_folder", tmp_path)
    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    return userdata / "amazon_asin_database.xlsx"


@pytest.mark.asyncio
async def test_save_single_asin_record_validates_asin_format(_tmp_asin_db):
    success = await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="invalid",
    )
    assert success is False


@pytest.mark.asyncio
async def test_save_single_asin_record_skips_empty_asin(_tmp_asin_db):
    success = await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="",
    )
    assert success is False


@pytest.mark.asyncio
async def test_save_single_asin_record_inserts_new_row(_tmp_asin_db):
    success = await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="B000000001",
        title="Test Title",
        product_url="https://www.amazon.co.jp/dp/B000000001",
        poster_url="https://m.media-amazon.com/images/I/test.jpg",
    )

    assert success is True

    wb = load_workbook(_tmp_asin_db)
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "ABC-123"
    assert ws.cell(row=2, column=2).value == "B000000001"
    wb.close()


@pytest.mark.asyncio
async def test_save_asin_to_excel_formats_worksheet(_tmp_asin_db):
    records = [
        {
            "number": "ABC-123",
            "asin": "B000000001",
            "product_url": "https://www.amazon.co.jp/dp/B000000001",
            "poster_url": "https://m.media-amazon.com/images/I/test.jpg",
        }
    ]

    await amazon_database.save_asin_to_excel(records, _tmp_asin_db)

    wb = load_workbook(_tmp_asin_db)
    ws = wb.active

    assert ws.title == "ASIN 数据库"
    assert ws.freeze_panes == "B2"
    assert ws.auto_filter.ref == "A1:F2"

    headers = [ws.cell(row=1, column=i).value for i in range(1, 7)]
    assert headers == ["影片番号", "ASIN 编号", "影片链接", "商品标题", "封面 URL", "搜索关键词"]

    assert ws.cell(row=2, column=3).hyperlink.target == "https://www.amazon.co.jp/dp/B000000001"
    assert ws.cell(row=2, column=3).style == "Hyperlink"
    assert ws.cell(row=2, column=5).hyperlink.target == "https://m.media-amazon.com/images/I/test.jpg"
    assert ws.cell(row=2, column=5).style == "Hyperlink"

    wb.close()


@pytest.mark.asyncio
async def test_save_asin_to_excel_dedup_same_number(_tmp_asin_db):
    await amazon_database.save_asin_to_excel(
        [
            {
                "number": "ABC-123",
                "asin": "B000000001",
                "title": "First",
            }
        ],
        _tmp_asin_db,
    )
    # 同番号再次写入（即使 ASIN 不同）也跳过
    await amazon_database.save_asin_to_excel(
        [
            {
                "number": "ABC-123",
                "asin": "B000000002",
                "title": "Second",
            }
        ],
        _tmp_asin_db,
    )

    wb = load_workbook(_tmp_asin_db)
    ws = wb.active
    numbers = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert numbers == ["ABC-123"]
    assert ws.cell(row=2, column=2).value == "B000000001"
    wb.close()


@pytest.mark.asyncio
async def test_save_asin_to_excel_dedup_skips_only_duplicate(_tmp_asin_db):
    await amazon_database.save_asin_to_excel(
        [
            {"number": "ABC-123", "asin": "B000000001"},
            {"number": "DEF-456", "asin": "B000000002"},
            {"number": "ABC-123", "asin": "B000000003"},  # 重复番号，应跳过
        ],
        _tmp_asin_db,
    )

    wb = load_workbook(_tmp_asin_db)
    ws = wb.active
    numbers = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert numbers == ["ABC-123", "DEF-456"]
    wb.close()


@pytest.mark.asyncio
async def test_merge_asin_db_from_backup(_tmp_asin_db, tmp_path):
    from mdcx.core.amazon_database import merge_asin_db_from_backup

    # 出厂库：旧番号(用户已有，且用户已填值) + 空缺番号(用户已有但字段空) + 新番号
    backup_path = tmp_path / "backup.xlsx"
    wb = load_workbook(backup_path) if backup_path.exists() else None
    import openpyxl

    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["影片番号", "ASIN 编号", "影片链接", "商品标题", "封面 URL", "搜索关键词"])
    ws.append(["ABC-123", "B000000001", "https://www.amazon.co.jp/dp/B000000001", "Factory Title", "", ""])
    ws.append(["DEF-456", "B000000002", "", "Factory Title 2", "", ""])
    ws.append(["NEW-001", "B000000003", "https://www.amazon.co.jp/dp/B000000003", "Factory New", "", ""])
    wb.save(backup_path)
    wb.close()

    # 用户库：已有 ABC-123（含值）和 DEF-456（ASIN 空缺）
    await amazon_database.save_asin_to_excel(
        [
            {"number": "ABC-123", "asin": "B999999999", "title": "User Kept Title", "product_url": "https://user.url"},
            {"number": "DEF-456", "asin": ""},
        ],
        _tmp_asin_db,
    )

    merge_asin_db_from_backup(backup_path, _tmp_asin_db)

    wb = load_workbook(_tmp_asin_db)
    ws = wb.active
    rows = {}
    order = []
    for r in range(2, ws.max_row + 1):
        num = ws.cell(row=r, column=1).value
        if num:
            rows[num] = [ws.cell(row=r, column=c).value for c in range(2, 7)]
            order.append(num)
    # 已有番号不覆盖用户值
    assert rows["ABC-123"][0] == "B999999999"  # ASIN 不被出厂覆盖
    assert rows["ABC-123"][1] == "https://user.url"  # 链接不被覆盖
    # 空缺字段被补全
    assert rows["DEF-456"][0] == "B000000002"
    # 新番号追加
    assert rows["NEW-001"][0] == "B000000003"
    assert rows["NEW-001"][1] == "https://www.amazon.co.jp/dp/B000000003"
    # 合并后按番号排序
    assert order == ["ABC-123", "DEF-456", "NEW-001"]
    wb.close()


@pytest.mark.asyncio
async def test_merge_asin_db_from_backup_resorts(_tmp_asin_db, tmp_path):
    """合并新增行后，用户库整体按番号（前缀字母 + 数字）重排。"""
    import openpyxl

    from mdcx.core.amazon_database import merge_asin_db_from_backup

    # 用户库：乱序番号（SSNI 在前、ABC 在后）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["影片番号", "ASIN 编号", "影片链接", "商品标题", "封面 URL", "搜索关键词"])
    ws.append(["SSNI-804", "B000000010", "", "", "", ""])
    ws.append(["ABC-123", "B000000020", "", "", "", ""])
    wb.save(_tmp_asin_db)
    wb.close()

    # 出厂库：新增 MMD-001（应排到 ABC 与 SSNI 之间）
    backup_path = tmp_path / "backup.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["影片番号", "ASIN 编号", "影片链接", "商品标题", "封面 URL", "搜索关键词"])
    ws.append(["MMD-001", "B000000030", "", "", "", ""])
    wb.save(backup_path)
    wb.close()

    merge_asin_db_from_backup(backup_path, _tmp_asin_db)

    wb = load_workbook(_tmp_asin_db)
    ws = wb.active
    numbers = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert numbers == ["ABC-123", "MMD-001", "SSNI-804"]
    wb.close()


@pytest.mark.asyncio
async def test_merge_asin_db_from_backup_marker_skips(_tmp_asin_db, tmp_path):
    import openpyxl

    from mdcx.core.amazon_database import merge_asin_db_from_backup

    backup_path = tmp_path / "backup.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["影片番号", "ASIN 编号", "影片链接", "商品标题", "封面 URL", "搜索关键词"])
    ws.append(["NEW-001", "B000000003", "", "", "", ""])
    wb.save(backup_path)
    wb.close()

    await amazon_database.save_asin_to_excel([{"number": "ABC-123", "asin": "B000000001"}], _tmp_asin_db)
    marker = _tmp_asin_db.parent / ".asin_db_merge_marker"

    merge_asin_db_from_backup(backup_path, _tmp_asin_db)
    assert marker.exists()

    # 出厂库未变，第二次合并应跳过（不再新增重复）
    merge_asin_db_from_backup(backup_path, _tmp_asin_db)

    wb = load_workbook(_tmp_asin_db)
    ws = wb.active
    numbers = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert numbers.count("NEW-001") == 1
    wb.close()


@pytest.mark.asyncio
async def test_query_asin_database_by_number(_tmp_asin_db):
    await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="B000000001",
        title="Test Title",
    )

    results = await amazon_database.query_asin_database(number="ABC-123", excel_path=_tmp_asin_db)

    assert len(results) == 1
    assert results[0]["asin"] == "B000000001"


@pytest.mark.asyncio
async def test_query_asin_database_by_asin(_tmp_asin_db):
    await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="B000000001",
    )

    results = await amazon_database.query_asin_database(asin="B000000001", excel_path=_tmp_asin_db)

    assert len(results) == 1
    assert results[0]["number"] == "ABC-123"


@pytest.mark.asyncio
async def test_query_asin_database_returns_empty_when_file_not_exists():
    from pathlib import Path

    results = await amazon_database.query_asin_database(
        number="NONEXISTENT",
        excel_path=Path("/nonexistent/path/file.xlsx"),
    )

    assert results == []


@pytest.mark.asyncio
async def test_query_asin_database_logs_read_failure(monkeypatch, _tmp_asin_db):
    await amazon_database.save_single_asin_record(
        number="ABC-123",
        asin="B000000001",
    )

    import openpyxl

    def _raise_runtime_error(*args, **kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr(openpyxl, "load_workbook", _raise_runtime_error)
    LogBuffer.log().clear()

    results = await amazon_database.query_asin_database(number="ABC-123", excel_path=_tmp_asin_db)

    assert results == []
    assert "[ASIN 数据库] 读取失败：boom" in LogBuffer.log().get()


def test_format_asin_worksheet_logs_failure():
    LogBuffer.log().clear()

    amazon_database._format_asin_worksheet(object())

    assert "[ASIN 数据库] 工作表格式化失败：" in LogBuffer.log().get()


@pytest.mark.asyncio
async def test_save_asin_to_excel_logs_missing_openpyxl(monkeypatch, _tmp_asin_db):
    import sys

    original_openpyxl = sys.modules.get("openpyxl")
    sys.modules["openpyxl"] = None

    LogBuffer.log().clear()

    with pytest.raises(ImportError):
        await amazon_database.save_asin_to_excel([], _tmp_asin_db)

    assert "[ASIN 数据库] 缺少 openpyxl，无法保存 amazon_asin_database.xlsx" in LogBuffer.log().get()

    if original_openpyxl:
        sys.modules["openpyxl"] = original_openpyxl
    else:
        del sys.modules["openpyxl"]
