"""校验 actor_database.xlsx 的 sheet 结构，防止辅助 sheet（如男优备份）被误读。"""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from mdcx.config.resources import ACTOR_DB_SHEET, get_actor_db_sheet

ROOT = Path(__file__).resolve().parent.parent
FACTORY_XLSX = ROOT / "resources" / "userdata" / "actor_database.xlsx"


def test_factory_db_has_actor_sheet_first():
    """出厂模板首个 sheet 必须是「演员数据库」，辅助 sheet 不得前置。"""
    if not FACTORY_XLSX.exists():
        pytest.skip("出厂数据库不存在")
    wb = load_workbook(FACTORY_XLSX, read_only=True)
    assert ACTOR_DB_SHEET in wb.sheetnames
    assert wb.sheetnames[0] == ACTOR_DB_SHEET
    wb.close()


def test_factory_db_main_sheet_headers():
    """主 sheet 表头应为 9 列固定结构。"""
    if not FACTORY_XLSX.exists():
        pytest.skip("出厂数据库不存在")
    wb = load_workbook(FACTORY_XLSX, read_only=True, data_only=True)
    ws = wb[ACTOR_DB_SHEET]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()
    assert headers == ["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"]


def test_get_actor_db_sheet_picks_named_sheet_not_active():
    """辅助 sheet 排在前面时，get_actor_db_sheet 仍取「演员数据库」。"""
    wb = Workbook()
    main = wb.active
    main.title = ACTOR_DB_SHEET
    backup = wb.create_sheet("男优备份")
    backup.title = "男优备份"
    wb.move_sheet(backup, offset=-1)  # 把辅助 sheet 移到最前

    ws = get_actor_db_sheet(wb)
    assert ws.title == ACTOR_DB_SHEET
    wb.close()


def test_get_actor_db_sheet_falls_back_to_active():
    """工作簿无「演员数据库」名时回退 active（兼容新建/异常文件）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "其他表"
    assert get_actor_db_sheet(wb).title == "其他表"
    wb.close()


def test_clean_male_actors_backup_sheet_kept_after_sync_write(tmp_path: Path):
    """sync 写入后辅助 sheet 保留——模拟含男优备份的库被正常写入不被破坏。"""
    wb = Workbook()
    main = wb.active
    main.title = ACTOR_DB_SHEET
    for col, header in enumerate(
        ["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"], 1
    ):
        main.cell(row=1, column=col, value=header)
    backup = wb.create_sheet("男优备份")
    backup.cell(row=1, column=1, value="加藤鷹")
    backup.cell(row=2, column=1, value="しみけん")

    path = tmp_path / "actor_database.xlsx"
    wb.save(path)
    wb.close()

    # 模拟 sync_from_avdb 的写路径：load_workbook + get_actor_db_sheet 写入
    wb2 = load_workbook(path)
    ws2 = get_actor_db_sheet(wb2)
    ws2.cell(row=ws2.max_row + 1, column=1, value="新演员ZZZ")
    wb2.save(path)
    wb2.close()

    wb3 = load_workbook(path)
    assert wb3.sheetnames == [ACTOR_DB_SHEET, "男优备份"]
    backup3 = wb3["男优备份"]
    backup_rows = [r[0] for r in backup3.iter_rows(min_row=1, values_only=True) if r[0]]
    assert backup_rows == ["加藤鷹", "しみけん"]
    main3 = wb3[ACTOR_DB_SHEET]
    assert "新演员ZZZ" in [r[0] for r in main3.iter_rows(values_only=True)]
    wb3.close()


def test_check_actor_db_rejects_backup_sheet_first(tmp_path: Path):
    """check_actor_db 应拦截「男优备份」前置的文件（防止辅助 sheet 被误读）。"""
    import sys

    sys.path.insert(0, str(ROOT / "scripts"))
    from check_actor_db import check_xlsx

    wb = Workbook()
    main = wb.active
    main.title = ACTOR_DB_SHEET
    main.cell(row=1, column=1, value="日文原名")
    backup = wb.create_sheet("男优备份")
    backup.title = "男优备份"
    wb.move_sheet(backup, offset=-1)

    bad = tmp_path / "bad.xlsx"
    wb.save(bad)
    wb.close()

    assert check_xlsx(bad) == 1


def test_write_goes_to_main_sheet_even_when_backup_first(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    """男优备份排第一位时，update_actor_db_row 仍写入主 sheet 而非备份 sheet。"""
    import asyncio

    from mdcx.core import tmdb_actor

    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    db_path = userdata / "actor_database.xlsx"

    # 构造男优备份在前的库
    wb = Workbook()
    main = wb.active
    main.title = ACTOR_DB_SHEET
    for col, header in enumerate(
        ["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"], 1
    ):
        main.cell(row=1, column=col, value=header)
    backup = wb.create_sheet("男优备份")
    backup.cell(row=1, column=1, value="加藤鷹")
    wb.move_sheet("男优备份", offset=-1)
    wb.save(db_path)
    wb.close()
    assert load_workbook(db_path).sheetnames[0] == "男优备份"

    monkeypatch.setattr(tmdb_actor.manager, "data_folder", tmp_path)
    monkeypatch.setattr(tmdb_actor.resources, "actor_db", {})
    tmdb_actor._ACTOR_DB_ROW_INDEX = {}

    async def _write():
        return await tmdb_actor.update_actor_db_row(jp="测试新演员", zh_cn="", zh_tw="", tmdbid=999)

    status = asyncio.run(_write())
    assert status == "inserted_new_row"

    wb2 = load_workbook(db_path)
    main2 = wb2[ACTOR_DB_SHEET]
    main_names = [r[0] for r in main2.iter_rows(min_row=2, values_only=True) if r[0]]
    backup2 = wb2["男优备份"]
    backup_names = [r[0] for r in backup2.iter_rows(min_row=1, values_only=True) if r[0]]
    wb2.close()

    assert "测试新演员" in main_names  # 写入主 sheet
    assert backup_names == ["加藤鷹"]  # 备份 sheet 未被污染
