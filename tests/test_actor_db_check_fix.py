"""「检查用户库」的检查与自动修复逻辑单元测试。

覆盖 _check_actor_db_issues（分类问题检测）与 auto_fix_actor_db（安全项自动修复）。
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from mdcx.tools.actor_db_tool import _check_actor_db_issues, auto_fix_actor_db

HEADER = ["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"]


def _make_db(tmp_path: Path, rows: list[list]) -> Path:
    p = tmp_path / "actor_database.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(HEADER)
    for row in rows:
        ws.append(row)
    wb.save(p)
    return p


def _categories(issues: dict) -> list[str]:
    return [cat for _, _, cat in issues["errors"]] + [cat for _, _, cat in issues["warnings"]]


def _read_rows(p: Path) -> list[list]:
    from openpyxl import load_workbook

    wb = load_workbook(p)
    ws = wb.active
    return [[c.value if c.value is not None else "" for c in row] for row in ws.iter_rows(min_row=2)]


class TestCheckActorDbIssues:
    def test_normal_db_no_issues(self, tmp_path: Path):
        p = _make_db(
            tmp_path,
            [
                [
                    "测试一",
                    "中文一",
                    "繁体一",
                    "测试一,测试二",
                    "",
                    "123",
                    "https://www.themoviedb.org/person/123",
                    "1995-01-01",
                    "生涯: 2020~ | 出身: 东京都",
                ]
            ],
        )
        issues = _check_actor_db_issues(p)
        assert issues["errors"] == []
        assert issues["warnings"] == []

    def test_missing_file_returns_empty(self, tmp_path: Path):
        issues = _check_actor_db_issues(tmp_path / "nope.xlsx")
        assert issues["errors"] == []
        assert issues["warnings"] == []

    def test_jp_empty_is_error(self, tmp_path: Path):
        p = _make_db(tmp_path, [["", "中文一", "", "kw", "", "", "", "", ""]])
        cats = _categories(_check_actor_db_issues(p))
        assert "jp_empty" in cats

    def test_jp_duplicate_is_error(self, tmp_path: Path):
        p = _make_db(
            tmp_path,
            [["测试一", "中文一", "", "a", "", "", "", "", ""], ["测试一", "中文二", "", "b", "", "", "", "", ""]],
        )
        cats = _categories(_check_actor_db_issues(p))
        assert "jp_dup" in cats
        errors = [e for e in _check_actor_db_issues(p)["errors"] if e[2] == "jp_dup"]
        assert errors[0][0] == 3  # 第二数据行

    def test_keyword_leading_trailing_comma(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", ",测试一,", "", "", "", "", ""]])
        cats = _categories(_check_actor_db_issues(p))
        assert "kw_format" in cats

    def test_keyword_double_comma(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "a,,b", "", "", "", "", ""]])
        cats = _categories(_check_actor_db_issues(p))
        assert "kw_format" in cats

    def test_keyword_duplicate_word(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "测试一,测试一", "", "", "", "", ""]])
        cats = _categories(_check_actor_db_issues(p))
        assert "kw_dup" in cats

    def test_birth_format_invalid(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "1995/01/01", ""]])
        cats = _categories(_check_actor_db_issues(p))
        assert "birth_format" in cats

    def test_birth_range_out_of_bounds(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "1888-01-01", ""]])
        cats = _categories(_check_actor_db_issues(p))
        assert "birth_range" in cats

    def test_birth_range_future(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "2099-01-01", ""]])
        cats = _categories(_check_actor_db_issues(p))
        assert "birth_range" in cats

    def test_birth_valid_old_actor_not_flagged(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "1945-06-15", ""]])
        cats = _categories(_check_actor_db_issues(p))
        assert "birth_range" not in cats
        assert "birth_format" not in cats

    def test_career_no_year(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "", "生涯: 引退済 | 出身: 东京都"]])
        cats = _categories(_check_actor_db_issues(p))
        assert "career_no_year" in cats

    def test_career_with_year_ok(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "", "生涯: 2020~ | 出身: 东京都"]])
        cats = _categories(_check_actor_db_issues(p))
        assert "career_no_year" not in cats

    def test_career_fullwidth_year_ok(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "", "生涯: ２０２１～ | 出身: 东京都"]])
        cats = _categories(_check_actor_db_issues(p))
        assert "career_no_year" not in cats

    def test_tmdb_no_id_but_url(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "https://www.themoviedb.org/person/123", "", ""]])
        cats = _categories(_check_actor_db_issues(p))
        assert "tmdb_no_id" in cats

    def test_tmdb_mismatch_url(self, tmp_path: Path):
        p = _make_db(
            tmp_path, [["测试一", "中文一", "", "", "", "123", "https://wrong.example.com/person/999", "", ""]]
        )
        cats = _categories(_check_actor_db_issues(p))
        assert "tmdb_mismatch" in cats

    def test_tmdb_duplicate(self, tmp_path: Path):
        p = _make_db(
            tmp_path,
            [
                ["测试一", "中文一", "", "", "", "123", "", "", ""],
                ["测试二", "中文二", "", "", "", "123", "", "", ""],
            ],
        )
        cats = _categories(_check_actor_db_issues(p))
        assert "tmdb_dup" in cats

    def test_tmdb_url_duplicate(self, tmp_path: Path):
        p = _make_db(
            tmp_path,
            [
                ["测试一", "中文一", "", "", "", "123", "https://www.themoviedb.org/person/123", "", ""],
                ["测试二", "中文二", "", "", "", "456", "https://www.themoviedb.org/person/123", "", ""],
            ],
        )
        cats = _categories(_check_actor_db_issues(p))
        assert "tmdb_dup_url" in cats

    def test_name_empty_is_warning(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "", "", "kw", "", "", "", "", ""]])
        cats = _categories(_check_actor_db_issues(p))
        assert "name_empty" in cats

    def test_bio_jp_residual_is_warning(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "", "出身: 東京都（とうきょうと）"]])
        cats = _categories(_check_actor_db_issues(p))
        assert "bio_jp" in cats

    def test_bio_jp_ignores_debut_field(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "", "出道: にっぽん発売作品"]])
        cats = _categories(_check_actor_db_issues(p))
        assert "bio_jp" not in cats

    def test_bio_unstructured_is_warning(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "", "自由文本简介没有任何标准字段"]])
        cats = _categories(_check_actor_db_issues(p))
        assert "bio_unstruct" in cats

    def test_bio_with_standard_field_ok(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "", "身高: 160cm"]])
        cats = _categories(_check_actor_db_issues(p))
        assert "bio_unstruct" not in cats


class TestAutoFixActorDb:
    def test_missing_file_returns_empty(self, tmp_path: Path):
        result = auto_fix_actor_db(tmp_path / "nope.xlsx")
        assert result["fixed"] == {}
        assert result["needs_manual"] == []

    def test_jp_empty_row_deleted(self, tmp_path: Path):
        p = _make_db(
            tmp_path, [["测试一", "中文一", "", "a", "", "", "", "", ""], ["", "", "", "", "", "", "", "", ""]]
        )
        result = auto_fix_actor_db(p)
        assert result["fixed"].get("jp_empty") == 1
        rows = _read_rows(p)
        assert len(rows) == 1
        assert rows[0][0] == "测试一"

    def test_jp_duplicate_merged_and_deleted(self, tmp_path: Path):
        p = _make_db(
            tmp_path,
            [
                ["测试一", "中文一", "", "别名A", "", "123", "", "", ""],
                ["测试一", "中文二", "", "别名B", "", "456", "", "", ""],
            ],
        )
        result = auto_fix_actor_db(p)
        assert result["fixed"].get("jp_dup") == 1
        rows = _read_rows(p)
        assert len(rows) == 1
        kw = rows[0][3]
        assert "别名A" in kw and "别名B" in kw

    def test_keyword_normalized_and_deduped(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", ",测试一,,测试二,测试一,", "", "", "", "", ""]])
        result = auto_fix_actor_db(p)
        assert result["fixed"].get("kw_format") == 1
        rows = _read_rows(p)
        assert rows[0][3] == "测试一,测试二"

    def test_birth_range_cleared(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "1888-01-01", ""]])
        result = auto_fix_actor_db(p)
        assert result["fixed"].get("birth_range") == 1
        assert _read_rows(p)[0][7] == ""

    def test_career_no_year_segment_removed(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", "", "生涯: 引退済 | 出身: 东京都"]])
        result = auto_fix_actor_db(p)
        assert result["fixed"].get("career_no_year") == 1
        bio = _read_rows(p)[0][8]
        assert "生涯" not in bio
        assert "出身: 东京都" in bio

    def test_tmdb_mismatch_url_rewritten(self, tmp_path: Path):
        p = _make_db(
            tmp_path, [["测试一", "中文一", "", "", "", "123", "https://wrong.example.com/person/999", "", ""]]
        )
        result = auto_fix_actor_db(p)
        assert result["fixed"].get("tmdb_mismatch") == 1
        assert _read_rows(p)[0][6] == "https://www.themoviedb.org/person/123"

    def test_tmdb_no_id_reported_manual(self, tmp_path: Path):
        p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "https://www.themoviedb.org/person/123", "", ""]])
        result = auto_fix_actor_db(p)
        cats = [cat for _, _, cat in result["needs_manual"]]
        assert "tmdb_no_id" in cats
        assert result["fixed"] == {}

    def test_tmdb_duplicate_reported_manual(self, tmp_path: Path):
        p = _make_db(
            tmp_path,
            [
                ["测试一", "中文一", "", "", "", "123", "", "", ""],
                ["测试二", "中文二", "", "", "", "123", "", "", ""],
            ],
        )
        result = auto_fix_actor_db(p)
        cats = [cat for _, _, cat in result["needs_manual"]]
        assert "tmdb_dup" in cats

    def test_after_fix_errors_cleared(self, tmp_path: Path):
        p = _make_db(
            tmp_path,
            [
                [
                    "测试一",
                    "中文一",
                    "",
                    ",a,,a,",
                    "",
                    "123",
                    "https://wrong.example.com/person/9",
                    "1888-01-01",
                    "生涯: 引退済",
                ],
                ["", "", "", "", "", "", "", "", ""],
            ],
        )
        issues_before = _check_actor_db_issues(p)
        assert issues_before["errors"]
        auto_fix_actor_db(p)
        issues_after = _check_actor_db_issues(p)
        # 仅剩 name_empty 类 warning（中文/繁体名缺失）属建议项
        assert all(cat == "name_empty" for _, _, cat in issues_after["errors"] + issues_after["warnings"])


@pytest.mark.parametrize(
    ("birth", "expected"),
    [
        ("1995-01-01", []),
        ("1995-1", []),
        ("1995", []),
        ("1995.01", ["birth_format"]),
        ("abc", ["birth_format"]),
        ("1945-06-15", []),
    ],
)
def test_birth_format_parametrized(tmp_path: Path, birth: str, expected: list[str]):
    p = _make_db(tmp_path, [["测试一", "中文一", "", "", "", "", "", birth, ""]])
    cats = _categories(_check_actor_db_issues(p))
    for e in expected:
        assert e in cats
