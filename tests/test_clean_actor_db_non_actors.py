"""校验 scripts/clean_actor_db_non_actors.py 的非演员清洗逻辑。"""

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook  # noqa: E402

from mdcx.config.resources import DB_HEADERS  # noqa: E402
from scripts import clean_actor_db_non_actors as mod  # noqa: E402


def _make_db(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(DB_HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _read_jps(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["演员数据库"]
    rows = [str(r[0]) for r in ws.iter_rows(min_row=2, values_only=True) if str(r[0] or "").strip()]
    wb.close()
    return rows


def test_non_av_names_removed(tmp_path, monkeypatch):
    """确证的非 AV 主流人物（有 id）被删除。"""
    db = tmp_path / "actor_database.xlsx"
    _make_db(
        db,
        [
            ["埃里克·坎通纳", "埃里克·坎通纳", "", "", "", "4213880", "", "", ""],
            ["正常女优", "正常女优", "", "", "", "1001", "", "", ""],
        ],
    )
    monkeypatch.setattr(mod, "NON_ACTING_FILE", tmp_path / "empty.txt")  # 无第1类清单
    monkeypatch.setattr(mod, "_NON_AV_NAMES", {"埃里克·坎通纳"})

    rc = mod.main(["--db", str(db), "--apply"])
    assert rc == 0
    jps = _read_jps(db)
    assert "埃里克·坎通纳" not in jps
    assert "正常女优" in jps


def test_desc_rows_removed(tmp_path, monkeypatch):
    """描述词/占位符行被删除，真实女优保留。"""
    db = tmp_path / "actor_database.xlsx"
    _make_db(
        db,
        [
            ["本妻", "本妻", "", "本妻", "", "", "", "", ""],
            ["八乙女かのん", "八乙女", "", "八乙女かのん", "", "", "", "", ""],
        ],
    )
    monkeypatch.setattr(mod, "NON_ACTING_FILE", tmp_path / "empty.txt")
    monkeypatch.setattr(mod, "_DESC_ROWS", {"本妻"})

    rc = mod.main(["--db", str(db), "--apply"])
    assert rc == 0
    jps = _read_jps(db)
    assert "本妻" not in jps
    assert "八乙女かのん" in jps


def test_apply_removes_trailing_empty_rows(tmp_path, monkeypatch):
    """删除后清理 openpyxl 末尾残留的空行。"""
    db = tmp_path / "actor_database.xlsx"
    _make_db(
        db,
        [
            ["演员甲", "甲", "", "", "", "1001", "", "", ""],
            ["演员乙", "乙", "", "", "", "1002", "", "", ""],
        ],
    )
    # 构造一个 jp 为空的末尾行（模拟 delete_rows 残留）
    wb = load_workbook(db)
    ws = wb["演员数据库"]
    ws.append(["", "", "", "", "", "2000", "", "", ""])
    wb.save(db)
    wb.close()

    monkeypatch.setattr(mod, "NON_ACTING_FILE", tmp_path / "empty.txt")
    monkeypatch.setattr(mod, "_NON_AV_NAMES", {"演员甲"})

    rc = mod.main(["--db", str(db), "--apply"])
    assert rc == 0
    wb = load_workbook(db, read_only=True, data_only=True)
    ws = wb["演员数据库"]
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if str(r[0] or "").strip()]
    wb.close()
    assert len(data_rows) == 1
    assert data_rows[0][0] == "演员乙"


def test_placeholder_rows_removed(tmp_path, monkeypatch):
    """前 4 列同值 + 仅链接有值（6/7/9 空）的行视为占位删除；有 id 的行保留。"""
    db = tmp_path / "actor_database.xlsx"
    _make_db(
        db,
        [
            ["安田みう", "安田みう", "安田みう", "安田みう", "", "", "", "", ""],  # 全空占位
            [
                "奥村佳代子",
                "奥村佳代子",
                "奥村佳代子",
                "奥村佳代子",
                "https://www.libredmm.com/actresses/1",
                "",
                "",
                "",
                "",
            ],  # 仅链接：视为占位删除
            [
                "真实女优",
                "真实女优",
                "真实女优",
                "真实女优",
                "",
                "1001",
                "",
                "",
                "",
            ],  # 有 id 保留
        ],
    )
    monkeypatch.setattr(mod, "NON_ACTING_FILE", tmp_path / "empty.txt")

    rc = mod.main(["--db", str(db), "--apply"])
    assert rc == 0
    jps = _read_jps(db)
    assert "安田みう" not in jps
    assert "奥村佳代子" not in jps
    assert "真实女优" in jps


def test_birthdate_only_rows_removed(tmp_path, monkeypatch):
    """前 4 列同值 + 仅生日有值的行被视为占位删除；有 id 的行保留。"""
    db = tmp_path / "actor_database.xlsx"
    _make_db(
        db,
        [
            ["阿香里えな", "阿香里えな", "阿香里えな", "阿香里えな", "", "", "", "1994-02-25", ""],  # 仅生日
            ["真实女优", "真实女优", "真实女优", "真实女优", "", "1001", "", "", ""],  # 有id保留
        ],
    )
    monkeypatch.setattr(mod, "NON_ACTING_FILE", tmp_path / "empty.txt")

    rc = mod.main(["--db", str(db), "--apply"])
    assert rc == 0
    jps = _read_jps(db)
    assert "阿香里えな" not in jps
    assert "真实女优" in jps


def test_hyperlink_rebuilt_after_delete(tmp_path, monkeypatch):
    """删除行后 hyperlink 按 cell 实际坐标重建，不产生孤儿 hyperlink。"""
    import re
    import zipfile

    db = tmp_path / "actor_database.xlsx"
    _make_db(
        db,
        [
            ["演员一", "演员一", "演员一", "演员一", "", "", "", "", ""],  # 占位（行2）
            [
                "演员二",
                "演员二",
                "演员二",
                "演员二",
                "https://libredmm.com/x/1",
                "1001",
                "https://www.themoviedb.org/person/1001",
                "",
                "",
            ],
            [
                "演员三",
                "演员三",
                "演员三",
                "演员三",
                "https://libredmm.com/x/2",
                "1002",
                "https://www.themoviedb.org/person/1002",
                "",
                "",
            ],
        ],
    )
    # 给演员二/三的链接列绑定超链接，模拟真实数据
    wb = load_workbook(db)
    ws = wb["演员数据库"]
    for row in (3, 4):
        ws.cell(row=row, column=5).hyperlink = f"https://libredmm.com/x/{row - 2}"
    wb.save(db)
    wb.close()
    monkeypatch.setattr(mod, "NON_ACTING_FILE", tmp_path / "empty.txt")

    rc = mod.main(["--db", str(db), "--apply"])
    assert rc == 0
    with zipfile.ZipFile(db) as zf:
        sheet = "xl/worksheets/sheet1.xml"
        s = zf.read(sheet).decode("utf-8")
    defined = set(re.findall(r'<c r="([A-Z]+\d+)"', s))
    refs = re.findall(r'<hyperlink [^>]*ref="([A-Z]+\d+)"', s)
    orphans = [r for r in refs if r not in defined]
    assert orphans == []
    assert len(refs) == 2  # 两张有链接的演员超链接保留并重建
