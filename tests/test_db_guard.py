"""校验 scripts/db_guard.py 的安全写入与保存后验证。"""

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook  # noqa: E402

from mdcx.config.resources import DB_HEADERS  # noqa: E402
from scripts import db_guard  # noqa: E402


def _make_ws(tmp_path: Path):
    path = tmp_path / "actor_database.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(DB_HEADERS)
    ws.append(["演员甲", "演员甲", "", "", "", "", "", "", ""])
    return path, wb, ws


def test_safe_write_requires_id(tmp_path):
    """无 id 时拒绝写 url。"""
    _, wb, ws = _make_ws(tmp_path)
    ok = db_guard.safe_write_tmdb(ws, 2, None, "https://www.themoviedb.org/person/1001")
    assert ok is False
    assert not ws.cell(row=2, column=6).value  # 空
    wb.close()


def test_safe_write_rejects_mismatch_url(tmp_path):
    """url 与该 id 不匹配时拒绝。"""
    _, wb, ws = _make_ws(tmp_path)
    ok = db_guard.safe_write_tmdb(ws, 2, "1001", "https://www.themoviedb.org/person/9999")
    assert ok is False
    assert not ws.cell(row=2, column=6).value  # 空
    wb.close()


def test_safe_write_sets_pair(tmp_path):
    """id+url 成对写入，url 缺省时由 id 推导。"""
    _, wb, ws = _make_ws(tmp_path)
    ok = db_guard.safe_write_tmdb(ws, 2, "1001")
    assert ok is True
    assert ws.cell(row=2, column=6).value == 1001
    assert ws.cell(row=2, column=7).value == "https://www.themoviedb.org/person/1001"
    wb.close()


def test_clear_tmdb_clears_both(tmp_path):
    """clear_tmdb 成对清空 id+url。"""
    _, wb, ws = _make_ws(tmp_path)
    db_guard.safe_write_tmdb(ws, 2, "1001")
    db_guard.clear_tmdb(ws, 2)
    assert ws.cell(row=2, column=6).value is None
    assert ws.cell(row=2, column=7).value is None
    wb.close()


def test_validate_after_save_ok(tmp_path):
    """正常库保存后验证通过。"""
    path, wb, ws = _make_ws(tmp_path)
    ws.cell(row=2, column=6).value = 1001
    ws.cell(row=2, column=7).value = "https://www.themoviedb.org/person/1001"
    wb.save(path)
    wb.close()
    assert db_guard.validate_after_save(path) is True
