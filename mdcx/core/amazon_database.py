"""
Amazon ASIN 数据库保存功能
用于保存影片番号与 ASIN 对应关系，方便后续统计和复用
"""

import asyncio
from datetime import datetime
from pathlib import Path
from typing import TypedDict

from ..utils.file import write_file_atomic


class AsinRecord(TypedDict, total=False):
    """ASIN 记录结构"""

    number: str  # 影片番号
    asin: str  # 亚马逊 ASIN
    product_url: str  # 亚马逊商品详情页链接
    title: str  # 商品标题
    poster_url: str  # 封面图片 URL
    search_keyword: str  # 搜索关键词


def _get_default_excel_path() -> Path:
    """获取默认的 Excel 文件路径，位于 userdata 目录下（与 mapping、watermark 等同目录）"""
    from ..config.manager import manager

    userdata_dir = manager.data_folder / "userdata"
    userdata_dir.mkdir(parents=True, exist_ok=True)
    return userdata_dir / "amazon_asin_database.xlsx"


def _asin_sort_key(row: tuple) -> tuple:
    """ASIN 库按番号排序键：前缀字母升序 + 数字升序；无法解析的放最后。"""
    import re

    n = str(row[0]) if row and row[0] is not None else ""
    m = re.match(r"^([A-Za-z]+)[-_]?(\d+)", n)
    if not m:
        return (chr(0x10FFFF), n)
    return (m.group(1).upper(), int(m.group(2)))


def _resort_asin_worksheet(local_path: Path) -> None:
    """读取 ASIN 库全部数据行，按番号排序后重建工作簿并重新格式化。

    用重建（而非原地 sort_rows）避免 delete_rows 的 max_row 虚高/空行残留，
    复用 _format_asin_worksheet 保持表头样式/边框/超链接/auto_filter 一致。
    """
    import openpyxl
    from openpyxl import Workbook

    wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = [
        tuple(r[:6])
        for r in ws.iter_rows(min_row=2, max_col=6, values_only=True)
        if r and r[0] is not None and str(r[0]).strip()
    ]
    wb.close()

    rows.sort(key=_asin_sort_key)

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = ws.title
    new_ws.append(header)
    for r in rows:
        new_ws.append(list(r))
    _format_asin_worksheet(new_ws)
    new_wb.save(local_path)
    new_wb.close()


def merge_asin_db_from_backup(backup_path: Path, local_path: Path) -> None:
    """把出厂 ASIN 库的增量同步进已存在的用户库（按番号去重，只增不删、不覆盖用户已有值）。

    出厂库随软件版本更新（新增/修正番号→ASIN 映射），老用户的用户库不会自动获得
    这些改进。此函数在启动时把出厂库中「用户库没有的番号」完整追加，给「用户库
    已有但字段空缺」的条目补全，绝不覆盖用户已填的值、绝不删除用户库任何行。
    合并产生新增行时，合并后按番号（前缀字母 + 数字）整体重排并重新格式化。

    用出厂库文件 md5 作为合并标记写入 local_path 同目录的 .asin_db_merge_marker，
    出厂库内容未变时跳过，避免每次启动重复扫描。
    """
    from ..models.log_buffer import LogBuffer

    try:
        import openpyxl
    except ImportError:
        LogBuffer.log().write("  ⚠️ [ASIN 数据库] 缺少 openpyxl，无法合并 amazon_asin_database.xlsx")
        return

    if not backup_path.exists() or not local_path.exists():
        return

    import hashlib

    marker_path = local_path.parent / ".asin_db_merge_marker"
    try:
        backup_hash = hashlib.md5(backup_path.read_bytes()).hexdigest()
        if marker_path.exists() and marker_path.read_text(encoding="utf-8").strip() == backup_hash:
            return  # 出厂库未变化，无需合并

        wb = openpyxl.load_workbook(local_path)
        ws = wb.active
        number_row_map: dict[str, int] = {}
        next_row = ws.max_row + 1
        for row_no, row in enumerate(ws.iter_rows(min_row=2, max_col=6, values_only=True), start=2):
            if row and row[0]:
                number_row_map.setdefault(str(row[0]).strip().upper(), row_no)

        added = 0
        filled = 0
        backup_wb = openpyxl.load_workbook(backup_path, read_only=True, data_only=True)
        backup_ws = backup_wb.active
        for row in backup_ws.iter_rows(min_row=2, max_col=6, values_only=True):
            if not row or not row[0]:
                continue
            number = str(row[0]).strip().upper()
            if number in number_row_map:
                # 字段补全：仅填空缺，不覆盖已有值
                existing_row = number_row_map[number]
                for col_idx in range(1, 6):  # 番号列除外
                    cur = ws.cell(row=existing_row, column=col_idx + 1).value
                    new = row[col_idx] if col_idx < len(row) else None
                    if (cur is None or str(cur).strip() == "") and new not in (None, ""):
                        ws.cell(row=existing_row, column=col_idx + 1, value=new)
                        filled += 1
                continue
            ws.append(list(row[:6]))
            number_row_map[number] = next_row
            next_row += 1
            added += 1
        backup_wb.close()

        if added or filled:
            wb.save(local_path)
        wb.close()
        write_file_atomic(marker_path, backup_hash, "utf-8")
        if added:
            # 有新增行才整体重排（纯字段补全不改变行数与顺序，无需重排）
            _resort_asin_worksheet(local_path)
        if added or filled:
            LogBuffer.log().write(f"  ℹ️ [ASIN 数据库] 出厂库增量合并: 新增 {added} 条, 补全 {filled} 个字段")
    except Exception as e:
        LogBuffer.log().write(f"  ⚠️ [ASIN 数据库] 出厂库合并失败: {e}")


async def save_asin_to_excel(
    records: list[AsinRecord],
    excel_path: Path | None = None,
    *,
    sheet_name: str = "ASIN 数据库",
) -> Path:
    """
    保存 ASIN 记录到 Excel 文件

    Args:
        records: ASIN 记录列表
        excel_path: Excel 文件路径，默认保存到运行目录下的 amazon_asin_database.xlsx
        sheet_name: 工作表名称

    Returns:
        Excel 文件路径

    注意：
        需要安装 openpyxl 库：pip install openpyxl
    """
    from ..models.log_buffer import LogBuffer

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        LogBuffer.log().write("  ⚠️ [ASIN 数据库] 缺少 openpyxl，无法保存 amazon_asin_database.xlsx")
        raise ImportError("请安装 openpyxl 库：pip install openpyxl") from None

    if excel_path is None:
        excel_path = _get_default_excel_path()
    elif isinstance(excel_path, str):
        excel_path = Path(excel_path)

    excel_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = sheet_name

    # 去重：以番号为键，用户库已有该番号时跳过不写（避免重复行）
    existing_numbers: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            existing_numbers.add(str(row[0]).strip().upper())

    if not ws["A1"].value:
        headers = [
            "影片番号",
            "ASIN 编号",
            "影片链接",
            "商品标题",
            "封面 URL",
            "搜索关键词",
        ]
        # 使用 cell() 直接设置表头，避免 append() 的空行问题
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="C0C0C0")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for record in records:
        number = str(record.get("number", "") or "").strip().upper()
        if not number or number in existing_numbers:
            continue
        existing_numbers.add(number)
        row_data = [
            record.get("number", ""),
            record.get("asin", ""),
            record.get("product_url", ""),
            record.get("title", ""),
            record.get("poster_url", ""),
            record.get("search_keyword", ""),
        ]
        ws.append(row_data)

    _format_asin_worksheet(ws)

    try:
        wb.save(excel_path)
        wb.close()
    except PermissionError as e:
        raise PermissionError(f"无法保存 Excel 文件，可能文件正被其他程序打开：{excel_path}") from e

    return excel_path


def _format_asin_worksheet(ws) -> None:
    """格式化 ASIN 数据库工作表：固定表头、自动筛选、列宽、边框、超链接、表头样式。"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        ws.freeze_panes = "B2"

        last_col = get_column_letter(6)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

        header_fill = openpyxl.styles.PatternFill("solid", fgColor="F2F2F2")
        header_font = openpyxl.styles.Font(bold=True, size=11)
        header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        thin = openpyxl.styles.Side(style="thin", color="D0D0D0")
        border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = border

        for row in ws.iter_rows(min_row=2, values_only=False):
            for col_idx in [2, 4]:
                cell = row[col_idx]
                val = str(cell.value or "").strip()
                if val and val.startswith("http"):
                    existing_target = cell.hyperlink.target if cell.hyperlink else None
                    if existing_target != val:
                        cell.style = "Hyperlink"
                        cell.hyperlink = val

        # 超链接处理会覆盖边框，重新设置
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = border

        # 数据行字体统一为 11pt
        data_font = openpyxl.styles.Font(size=11)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.font = data_font

        caps = {1: 20, 2: 15, 3: 50, 4: 80, 5: 50, 6: 40}
        col_max = [0] * 7
        for row in ws.iter_rows(min_row=2, values_only=True):
            for ci, cell in enumerate(row, 1):
                if cell is None or ci > 6:
                    continue
                s = str(cell)
                width = sum(2 if "\u3040" <= c <= "\u30ff" or "\u4e00" <= c <= "\u9fff" else 1 for c in s)
                col_max[ci] = max(col_max[ci], width)
        for ci in range(1, 7):
            letter = get_column_letter(ci)
            ws.column_dimensions[letter].width = min(col_max[ci] + 2, caps.get(ci, 80))
    except Exception as e:
        from ..models.log_buffer import LogBuffer

        LogBuffer.log().write(f"  ⚠️ [ASIN 数据库] 工作表格式化失败：{e}")


async def save_single_asin_record(
    number: str,
    asin: str,
    title: str = "",
    product_url: str = "",
    poster_url: str = "",
    search_keyword: str = "",
    excel_path: Path | None = None,
) -> bool:
    """
    保存单条 ASIN 记录

    Args:
        number: 影片番号
        asin: ASIN 编号（必须为 10 位字母数字）
        title: 商品标题
        product_url: 亚马逊商品详情页链接
        poster_url: 封面图片 URL
        search_keyword: 搜索关键词
        excel_path: Excel 文件路径

    Returns:
        bool: 保存成功返回 True，失败或跳过返回 False

    示例:
        success = await save_single_asin_record(
            number="ABC-123",
            asin="B0000001",
            title="作品标题",
            product_url="https://www.amazon.co.jp/dp/B0000001",
            poster_url="https://m.media-amazon.com/images/I/xxx.jpg",
        )
        if success:
            print("保存成功")
        else:
            print("保存失败或跳过")
    """
    import re

    if not asin or not asin.strip():
        return False

    asin = asin.strip().upper()

    if not re.match(r"^[A-Z0-9]{10}$", asin):
        return False

    record: AsinRecord = {
        "number": number,
        "asin": asin,
        "product_url": product_url,
        "title": title,
        "poster_url": poster_url,
        "search_keyword": search_keyword,
    }

    try:
        await save_asin_to_excel([record], excel_path)
        return True
    except Exception:
        return False


async def update_asin_record(
    number: str,
    poster_url: str,
    excel_path: Path | None = None,
) -> bool:
    """
    更新已有 ASIN 记录的 poster_url（原地更新，不新增行）

    Args:
        number: 影片番号
        poster_url: 新的封面 URL
        excel_path: Excel 文件路径

    Returns:
        bool: 更新成功返回 True，未找到记录返回 False
    """
    try:
        import openpyxl
    except ImportError:
        return False

    if excel_path is None:
        excel_path = _get_default_excel_path()

    if not excel_path.exists():
        return False

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    updated = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_number = str(row[0].value or "").upper()
        if row_number == number.upper():
            row[4].value = poster_url
            updated = True
            break

    if updated:
        wb.save(excel_path)
    wb.close()
    return updated


async def query_asin_database(
    number: str | None = None,
    asin: str | None = None,
    excel_path: Path | None = None,
) -> list[AsinRecord]:
    """
    查询 ASIN 数据库

    Args:
        number: 按番号查询
        asin: 按 ASIN 查询
        excel_path: Excel 文件路径

    Returns:
        匹配的记录列表

    示例:
        results = await query_asin_database(number="ABC-123")
        results = await query_asin_database(asin="B0000001")
    """
    try:
        import openpyxl
    except ImportError:
        from ..models.log_buffer import LogBuffer

        LogBuffer.log().write("  ⚠️ [ASIN 数据库] 缺少 openpyxl，无法读取 amazon_asin_database.xlsx")
        return []

    if excel_path is None:
        excel_path = _get_default_excel_path()

    if not excel_path.exists():
        return []

    results: list[AsinRecord] = []

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue

            if len(row) < 6:
                continue

            record = AsinRecord(
                number=str(row[0] or ""),
                asin=str(row[1] or ""),
                product_url=str(row[2] or ""),
                title=str(row[3] or ""),
                poster_url=str(row[4] or ""),
                search_keyword=str(row[5] or ""),
            )

            if number and str(record.get("number", "")).upper() == number.upper():
                results.append(record)
            elif asin and str(record.get("asin", "")).upper() == asin.upper():
                results.append(record)

        wb.close()
    except Exception as e:
        from ..models.log_buffer import LogBuffer

        LogBuffer.log().write(f"  ⚠️ [ASIN 数据库] 读取失败：{e}")

    return results


async def export_asin_statistics(
    excel_path: Path | None = None,
    output_path: Path | None = None,
) -> dict:
    """
    导出 ASIN 数据库统计信息

    Returns:
        统计信息字典
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("请安装 openpyxl 库：pip install openpyxl") from None

    if excel_path is None:
        excel_path = _get_default_excel_path()

    if not excel_path.exists():
        return {}

    if output_path is None:
        output_path = excel_path.parent / "amazon_statistics.txt"

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    total_records = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue

        if len(row) < 2:
            continue

        total_records += 1

    wb.close()

    stats = {
        "total_records": total_records,
    }

    report = (
        "=" * 60
        + "\n"
        + "Amazon ASIN 数据库统计报告\n"
        + f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        + "=" * 60
        + "\n\n"
        + f"总记录数：{total_records}\n"
        + "\n"
        + "=" * 60
        + "\n"
    )
    await asyncio.to_thread(write_file_atomic, output_path, report, "utf-8")

    return stats
