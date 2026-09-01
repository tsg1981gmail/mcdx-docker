import asyncio
import time
import traceback
from pathlib import Path
from typing import TYPE_CHECKING

import aiofiles.os
from PyQt6.QtWidgets import QMessageBox

from ..base.file import (
    _clean_empty_folders,
    check_file,
    copy_trailer_to_theme_videos,
    get_movie_list,
    move_bif,
    move_file_to_failed_folder,
    move_other_file,
    move_torrent,
    newtdisk_creat_symlink,
    pic_some_deal,
    save_success_list,
)
from ..base.image import extrafanart_copy2, extrafanart_extras_copy
from ..config.enums import (
    DownloadableFile,
    EmbyAction,
    FixedScrapingType,
    KeepableFile,
    NfoInclude,
    ReadMode,
    Switch,
    TagInclude,
)
from ..config.extend import get_movie_path_setting, parse_media_paths
from ..config.manager import manager
from ..config.resources import resources
from ..core.scrape_cache import ScrapeStateCache
from ..core.tmdb_actor import _normalize_translation
from ..crawler import CrawlerProvider
from ..models.enums import FileMode
from ..models.flags import JSON_DATA_CACHE_MAX_ENTRIES, FileDoneDict, Flags
from ..models.log_buffer import LogBuffer
from ..models.model_types import CrawlersResult, FileInfo, OtherInfo, ScrapeResult, ShowData
from ..signals import signal
from ..tools.emby_actor_image import update_emby_actor_photo
from ..tools.emby_actor_info import creat_kodi_actors
from ..utils import executor, get_current_time, get_real_time, get_used_time, split_path
from ..utils.dataclass import update
from ..utils.file import copy_file_async, move_file_async
from ..utils.image import compress_images_in_folder_async
from ..utils.path import is_any_descendant
from .file import (
    _generate_file_name,
    _get_folder_path,
    creat_folder,
    deal_old_files,
    get_file_info_v2,
    get_output_name,
    move_movie,
)
from .file_crawler import FileScraper, classify_existing_scrape_result, classify_scrape_task
from .image import add_mark
from .media_resource import MediaResourceContext
from .nfo import get_nfo_data, write_nfo
from .translate import translate_actor, translate_info, translate_title_outline
from .utils import (
    add_definition_tag,
    deal_some_field,
    get_video_size,
    replace_special_word,
    replace_word,
    show_movie_info,
    show_result,
)
from .web import (
    extrafanart_download,
    fanart_download,
    poster_download,
    thumb_download,
    trailer_download,
)

if TYPE_CHECKING:
    from ..crawler import CrawlerProviderProtocol


class StopScrape(Exception): ...


class UnexpectedScrapeCancellation(Exception): ...


class Scraper:
    def __init__(self, crawler_provider: "CrawlerProviderProtocol"):
        self.crawler_provider = crawler_provider
        self._rest_lock = asyncio.Lock()
        self._state_cache: ScrapeStateCache | None = None

    async def _run_tasks_with_limit(self, movie_list: list[Path], task_count: int, thread_number: int) -> None:
        task_iter = iter(enumerate(movie_list, 1))
        running_tasks: set[asyncio.Task[None]] = set()

        def _submit_next_task() -> bool:
            try:
                index, each_file = next(task_iter)
            except StopIteration:
                return False
            task_name = f"scrape-{index}/{task_count}:{each_file.name}"
            running_tasks.add(
                asyncio.create_task(self.process_one_file((each_file, index, task_count)), name=task_name)
            )
            return True

        for _ in range(min(thread_number, task_count)):
            _submit_next_task()

        try:
            while running_tasks:
                done, pending = await asyncio.wait(running_tasks, return_when=asyncio.FIRST_COMPLETED)
                running_tasks = set(pending)

                stop_requested = False
                fatal_error: Exception | None = None
                done_count = 0
                for done_task in done:
                    done_count += 1
                    try:
                        done_task.result()
                    except StopScrape:
                        if signal.stop or Flags.stop_requested:
                            stop_requested = True
                        elif fatal_error is None:
                            fatal_error = UnexpectedScrapeCancellation(
                                f"刮削任务异常停止：{done_task.get_name()}，但未检测到手动停止标识"
                            )
                    except asyncio.CancelledError:
                        if signal.stop or Flags.stop_requested:
                            stop_requested = True
                        elif fatal_error is None:
                            fatal_error = UnexpectedScrapeCancellation(
                                f"刮削任务被异常取消：{done_task.get_name()}，但未检测到手动停止标识"
                            )
                    except Exception as e:
                        if fatal_error is None:
                            fatal_error = e

                if stop_requested or fatal_error is not None:
                    for pending_task in running_tasks:
                        pending_task.cancel()
                    if running_tasks:
                        await asyncio.gather(*running_tasks, return_exceptions=True)
                    if fatal_error is not None:
                        raise fatal_error
                    return

                for _ in range(done_count):
                    _submit_next_task()
        except asyncio.CancelledError:
            for pending_task in running_tasks:
                pending_task.cancel()
            if running_tasks:
                await asyncio.gather(*running_tasks, return_exceptions=True)
            raise

    async def run(self, file_mode: FileMode, movie_list: list[Path] | None) -> None:
        try:
            await self._run(file_mode, movie_list)
        finally:
            await self.crawler_provider.close()
            if self._state_cache is not None:
                self._state_cache.close()
                self._state_cache = None

    async def _run(self, file_mode: FileMode, movie_list: list[Path] | None) -> None:
        Flags.reset()
        if movie_list is None:
            movie_list = []
        Flags.scrape_start_time = time.time()  # 开始刮削时间
        Flags.file_mode = file_mode  # 刮削模式（工具单文件或主界面/日志点开始正常刮削）

        # 初始化刮削状态缓存（断点续刮）。失败回退内存模式；递归调用（Again 模式）复用已打开的缓存。
        if self._state_cache is None:
            cache = ScrapeStateCache(resources.u("scrape_state.db"))
            if cache.open():
                self._state_cache = cache
        else:
            cache = self._state_cache
        signal.show_scrape_info("🔎 正在刮削中...")

        signal.set_main_info()  # 清空主界面显示信息
        thread_number = manager.config.thread_number  # 线程数量
        thread_time = manager.config.thread_time  # 线程延时
        signal.label_result.emit(f" 刮削中：{0} 成功：{Flags.succ_count} 失败：{Flags.fail_count}")
        signal.logs_failed_settext.emit("\n\n\n")

        # 日志页面显示开始时间
        Flags.start_time = time.time()
        if file_mode == FileMode.Single:
            signal.show_log_text("🍯 🍯 🍯 NOTE: 当前是单文件刮削模式！")
        elif file_mode == FileMode.Again:
            signal.show_log_text(f"🍯 🍯 🍯 NOTE: 开始重新刮削！！！ 刮削文件数量（{len(movie_list)})")
            n = 0
            for each_f, each_i in Flags.new_again_dic.items():
                n += 1
                if each_i[0]:
                    signal.show_log_text(f"{n} 🖥 File path: {each_f}\n 🚘 File number: {each_i[0]}")
                else:
                    signal.show_log_text(f"{n} 🖥 File path: {each_f}\n 🌐 File url: {each_i[1]}")

        # 获取设置的媒体目录、失败目录、成功目录
        path_settings = get_movie_path_setting()
        movie_path = path_settings.movie_path
        movie_paths = path_settings.movie_paths

        # 获取待刮削文件列表的相关信息
        if not movie_list:
            signal.show_log_text("\n ⏰ Start time: " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
            movie_list = []
            scan_movie_paths = movie_paths if file_mode == FileMode.Default else [movie_path]
            for media_path in scan_movie_paths:
                current_paths = get_movie_path_setting(movie_path_override=media_path)
                scan_path = current_paths.movie_path
                scan_ignore_dirs = current_paths.ignore_dirs
                if manager.config.scrape_softlink_path:
                    await newtdisk_creat_symlink(
                        Switch.COPY_NETDISK_NFO in manager.config.switch_on,
                        current_paths.movie_path,
                        current_paths.softlink_path,
                    )
                    scan_path = current_paths.softlink_path
                movie_list.extend(await get_movie_list(file_mode, scan_path, scan_ignore_dirs))
        else:
            signal.show_log_text("\n ⏰ Start time: " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))

        # 断点续刮：过滤掉已完成且 mtime 未变的文件；恢复上次失败的未超限文件（跨会话重试）
        # 读取模式（main_mode==4）本质是读已有结果，不参与断点续刮的跳过/恢复逻辑，
        # 否则会导致读取模式只能看到「mtime 变化」的文件，且读完即被标记 done 导致下次读不到。
        if cache.is_usable() and manager.config.main_mode != 4:
            try:
                existing = set(movie_list)
                cache.cleanup_missing(existing)
                force = file_mode != FileMode.Default  # Again/单文件等模式视为强制重新刮削
                if force:
                    skipped = 0
                else:
                    before = len(movie_list)
                    filtered = []
                    for p in movie_list:
                        mtime = await _safe_mtime(p)
                        if not cache.should_skip(p, mtime, force=False):
                            filtered.append(p)
                    skipped = before - len(filtered)
                    if skipped:
                        movie_list = filtered
                        signal.show_log_text(f" ⏭ 断点续刮：跳过 {skipped} 个已刮削且未变化的文件")
                pending = cache.list_pending(existing)
                if pending:
                    movie_list.extend(pending)
                    signal.show_log_text(f" 🔄 恢复 {len(pending)} 个上次失败的文件重新刮削")
            except Exception as e:
                signal.show_log_text(f" ⚠ 刮削状态缓存读取失败，按全量处理: {e}")

        Flags.remain_list = movie_list.copy()
        Flags.can_save_remain = True

        task_count = len(movie_list)
        Flags.total_count = task_count

        if task_count:
            Flags.count_claw = await Flags.increment("count_claw")
            if manager.config.main_mode == 4:
                signal.show_log_text(f" 🕷 当前为读取模式，并发数（{thread_number}），线程延时（0）秒...")
            else:
                if task_count < thread_number:
                    thread_number = task_count
                signal.show_log_text(f" 🕷 开启异步并发，并发数（{thread_number}），线程延时（{thread_time}）秒...")
            if Switch.REST_SCRAPE in manager.config.switch_on and manager.config.main_mode != 4:
                signal.show_log_text(
                    f'<font color="brown"> 🍯 间歇刮削 已启用，连续刮削 {manager.config.rest_count} 个文件后，将自动休息 {Flags.rest_time_convert} 秒...</font>'
                )

            if task_count > thread_number * 1000:
                signal.show_log_text(f" ⚠ 待刮削任务较多（{task_count}），已启用渐进式任务调度以降低内存峰值。")

            Flags.next_start_time = time.time()

            # 异步并发（按并发数渐进投喂任务，避免大列表一次性创建海量协程）
            await self._run_tasks_with_limit(movie_list, task_count, thread_number)
            if Flags.scrape_done < task_count and not (signal.stop or Flags.stop_requested):
                message = f"刮削异常提前结束：已完成 {Flags.scrape_done}/{task_count}，剩余 {task_count - Flags.scrape_done} 个任务未执行"
                signal.show_traceback_log(message)
                signal.show_log_text(f" 🔴 {message}")
                raise UnexpectedScrapeCancellation(message)
            signal.label_result.emit(f" 刮削中：0 成功：{Flags.succ_count} 失败：{Flags.fail_count}")
            await save_success_list()  # 保存成功列表
            if signal.stop or Flags.stop_requested:
                return

        signal.show_log_text("================================================================================")
        for media_path in movie_paths:
            current_paths = (
                path_settings if media_path == movie_path else get_movie_path_setting(movie_path_override=media_path)
            )
            clean_path = current_paths.softlink_path if manager.config.scrape_softlink_path else media_path
            await _clean_empty_folders(clean_path, file_mode)
        end_time = time.time()
        used_time = str(round((end_time - Flags.start_time), 2))
        average_time = str(round((end_time - Flags.start_time) / task_count, 2)) if task_count else used_time
        signal.exec_set_processbar.emit(0)
        signal.set_label_file_path.emit(f"🎉 恭喜！全部刮削完成！共 {task_count} 个文件！用时 {used_time} 秒")
        signal.show_traceback_log(
            f"🎉 All finished!!! Total {task_count} , Success {Flags.succ_count} , Failed {Flags.fail_count} "
        )
        signal.show_log_text(
            f" 🎉🎉🎉 All finished!!! Total {task_count} , Success {Flags.succ_count} , Failed {Flags.fail_count} "
        )
        signal.show_log_text("================================================================================")
        if Flags.failed_list:
            signal.show_log_text("    *** Failed results ****")
            for i in range(len(Flags.failed_list)):
                fail_path, fail_reson = Flags.failed_list[i]
                signal.show_log_text(f" 🔴 {i + 1} {fail_path}\n    {fail_reson}")
                signal.show_log_text("================================================================================")
        signal.show_log_text(
            " ⏰ Start time".ljust(15) + ": " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(Flags.start_time))
        )
        signal.show_log_text(
            " 🏁 End time".ljust(15) + ": " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(end_time))
        )
        signal.show_log_text(" ⏱ Used time".ljust(15) + f": {used_time}S")
        signal.show_log_text(" 📺 Movies num".ljust(15) + f": {task_count}")
        signal.show_log_text(" 🍕 Per time".ljust(15) + f": {average_time}S")
        signal.show_log_text("================================================================================")
        signal.show_scrape_info(f"🎉 刮削完成 {task_count}/{task_count}")

        # auto run after scrape
        if EmbyAction.ACTOR_PHOTO_AUTO in manager.config.emby_on:
            await update_emby_actor_photo(manage_button_state=False)
        if manager.config.actor_photo_kodi_auto:
            await creat_kodi_actors(True, manage_button_state=False)

        signal.reset_buttons_status.emit()
        if len(Flags.again_dic):
            Flags.new_again_dic = Flags.again_dic.copy()
            new_movie_list = list(Flags.new_again_dic.keys())
            Flags.again_dic.clear()
            signal.change_buttons_status.emit()
            await self._run(FileMode.Again, new_movie_list)
            return
        if Switch.AUTO_EXIT in manager.config.switch_on:
            signal.show_log_text("\n\n 🍔 已启用「刮削后自动退出软件」！")
            count = 5
            for i in range(count):
                signal.show_log_text(f" {count - i} 秒后将自动退出！")
                await asyncio.sleep(1)
            await self.crawler_provider.close()
            signal.exec_exit_app.emit()

    async def process_one_file(self, task: tuple[Path, int, int]) -> None:
        # 并发刮削的兄弟任务从这里各自开启新的日志任务组：
        # 本任务派生的子协程（fanart/poster/TMDB 等）与 to_thread 线程写入
        # 都归入本组，get() 只聚合本组——避免别的影片的失败原因混入
        # 本片的 failed_list/断点缓存（跨任务日志污染），finally 整树回收。
        LogBuffer.new_root()
        try:
            await self._process_one_file_impl(task)
        finally:
            LogBuffer.clear_task()

    async def _release_shared_status(self, *numbers: object) -> None:
        """释放共享番号状态：标记为失败并唤醒等待方。

        需释放注册时的全部键（movie_number 与 file_info.number），否则读模式下
        NFO 番号与原始番号不同时，等待方会空转 300 秒超时（原实现只释放 origin_number）。
        """
        for status_number in {n for n in numbers if n}:
            if status_number in Flags.json_get_status and Flags.json_get_status.get(status_number) is None:
                async with Flags._json_get_lock:
                    if Flags.json_get_status.get(status_number) is None:
                        Flags.json_get_status[status_number] = False
                        event = Flags.json_get_events.get(status_number)
                        if event is not None:
                            event.set()

    async def _process_one_file_impl(self, task: tuple[Path, int, int]) -> None:
        # 获取顺序
        file_path, count, count_all = task
        async with Flags._counter_lock:
            Flags.counting_order += 1
            count = Flags.counting_order

        # 名字缩写
        show_name = file_path.name
        if len(show_name) > 40:
            show_name = show_name[:40] + "..."

        # 处理间歇任务
        while (
            manager.config.main_mode != 4
            and Switch.REST_SCRAPE in manager.config.switch_on
            and count - Flags.rest_now_begin_count > manager.config.rest_count
        ):
            self._check_stop(show_name)
            await asyncio.sleep(1)

        # 非第一个加延时
        Flags.scrape_starting = await Flags.increment("scrape_starting")
        count = Flags.scrape_starting
        thread_time = manager.config.thread_time
        if count == 1 or thread_time == 0 or manager.config.main_mode == 4:
            Flags.next_start_time = time.time()
            signal.show_log_text(f" 🕷 {get_current_time()} 开始刮削：{Flags.scrape_starting}/{count_all} {show_name}")
            thread_time = 0
        else:
            async with Flags._counter_lock:
                Flags.next_start_time += thread_time

        # 计算本线程开始剩余时间, 休眠并定时检查是否手动停止
        remain_time = int(Flags.next_start_time - time.time())
        if remain_time > 0:
            signal.show_log_text(
                f" ⏱ {get_current_time()}（{remain_time}）秒后开始刮削：{count}/{count_all} {show_name}"
            )
            for _ in range(remain_time):
                self._check_stop(show_name)
                await asyncio.sleep(1)

        Flags.scrape_started = await Flags.increment("scrape_started")
        if count > 1 and thread_time != 0:
            signal.show_log_text(f" 🕷 {get_current_time()} 开始刮削：{Flags.scrape_started}/{count_all} {show_name}")

        start_time = time.time()
        file_mode = Flags.file_mode

        # 获取文件基础信息
        file_info = await get_file_info_v2(file_path)
        number = file_info.number
        origin_number = number
        folder_old_path = file_info.folder_path
        file_show_name = file_info.file_show_name
        file_show_path = file_info.file_show_path

        # 显示刮削信息
        progress_value = Flags.scrape_started / count_all * 100
        progress_percentage = f"{progress_value:.2f}%"
        signal.exec_set_processbar.emit(int(progress_value))
        signal.set_label_file_path.emit(
            f"正在刮削： {Flags.scrape_started}/{count_all} {progress_percentage} \n {file_show_path}"
        )
        signal.label_result.emit(
            f" 刮削中：{Flags.scrape_started - Flags.succ_count - Flags.fail_count} 成功：{Flags.succ_count} 失败：{Flags.fail_count}"
        )
        LogBuffer.log().write("\n" + "=" * 40)
        LogBuffer.log().write("\n 🙈 [file] " + str(file_info.file_path))
        LogBuffer.log().write("\n 🚘 [number] " + number)

        # 如果指定了单一网站，进行提示
        website_single = manager.config.website_single
        if manager.config.scrape_like == "single" and file_mode != FileMode.Single and manager.config.main_mode != 4:
            LogBuffer.log().write(
                f"\n 😸 [Note] You specified 「 {website_single} 」, some videos may not have results! "
            )

        # 获取刮削数据
        json_data = None
        other = None
        scrape_error: str | None = None
        try:
            json_data, other = await self._process_one_file(file_info, file_mode)
            if json_data and other:
                if manager.config.main_mode == 4:
                    number = json_data.number
                async with Flags._json_get_lock:
                    Flags.json_data_dic[number] = ScrapeResult(file_info, json_data, other)
                    Flags.json_data_dic.move_to_end(number)
                    while len(Flags.json_data_dic) > JSON_DATA_CACHE_MAX_ENTRIES:
                        Flags.json_data_dic.popitem(last=False)
                    for status_number in (origin_number, number):
                        if status_number in Flags.json_get_status and Flags.json_get_status[status_number] is None:
                            Flags.json_get_status[status_number] = True
                            event = Flags.json_get_events.get(status_number)
                            if event is not None:
                                event.set()
            else:
                await self._release_shared_status(origin_number, number, getattr(file_info, "shared_number", None))
        except Exception as e:
            scrape_error = str(e)
            await self._release_shared_status(origin_number, number, getattr(file_info, "shared_number", None))
            self._check_stop(show_name)
            signal.show_traceback_log(traceback.format_exc())
            signal.show_log_text(traceback.format_exc())
            LogBuffer.error().write("scrape file error: " + scrape_error)
            LogBuffer.log().write("\n" + traceback.format_exc())

        # 显示刮削数据
        try:
            show_data = ShowData.empty()
            show_data.file_info = file_info
            if json_data and other:
                show_data.data = json_data
                show_data.other = other
                Flags.succ_count = await Flags.increment("succ_count")
                show_data.show_name = (
                    str(Flags.count_claw)
                    + "-"
                    + str(Flags.succ_count)
                    + "."
                    + file_show_name.replace(number, file_info.number)
                    + ("-" if file_info.definition else "")
                    + file_info.definition
                )
                signal.show_list_name("succ", show_data, number)
                # 读取模式不写缓存：读取模式只读已有 NFO/结果，不应标记文件为 done，
                # 否则下次读取模式或普通模式会被断点续刮跳过。
                if self._state_cache and self._state_cache.is_usable() and manager.config.main_mode != 4:
                    try:
                        summary = {
                            "number": number,
                            "title": json_data.title,
                            "tags": list(json_data.tags),
                            "series": json_data.series,
                            "studio": json_data.studio,
                            "actors": list(json_data.actors),
                            "release": json_data.release,
                            "runtime": json_data.runtime,
                            "mosaic": json_data.mosaic,
                            "publisher": json_data.publisher,
                            "directors": list(json_data.directors),
                            "score": json_data.score,
                        }
                        self._state_cache.set_done(
                            file_info.file_path,  # 移动后路径：move_movie 已更新 file_info.file_path
                            await _safe_mtime(file_info.file_path),
                            number=number,
                            summary=summary,
                            commit=False,
                        )
                    except Exception:
                        pass
            else:
                Flags.fail_count = await Flags.increment("fail_count")
                show_data.show_name = (
                    str(Flags.count_claw)
                    + "-"
                    + str(Flags.fail_count)
                    + "."
                    + file_show_name.replace(number, file_info.number)
                    + ("-" if file_info.definition else "")
                    + file_info.definition
                )
                signal.show_list_name("fail", show_data, number)
                error_msg = LogBuffer.error().get() or scrape_error or "未知错误"
                LogBuffer.log().write(f"\n 🔴 [Failed] Reason: {error_msg}")
                if "WinError 5" in error_msg:
                    LogBuffer.log().write(
                        "\n 🔴 该问题为权限问题：请尝试以管理员身份运行，同时关闭其他正在运行的Python脚本！"
                    )
                failed_folder = get_movie_path_setting(file_path).failed_folder
                fail_file_path = await move_file_to_failed_folder(failed_folder, file_path, folder_old_path)
                Flags.failed_list.append((fail_file_path, error_msg))
                await self._failed_file_info_show(str(Flags.fail_count), fail_file_path, error_msg)
                signal.view_failed_list_settext.emit(f"失败 {Flags.fail_count}")
                if self._state_cache and self._state_cache.is_usable() and manager.config.main_mode != 4:
                    try:
                        self._state_cache.set_failed(
                            fail_file_path,  # 移动后路径：文件已被移到 failed_folder
                            await _safe_mtime(fail_file_path),
                            error=error_msg,
                            commit=False,
                        )
                    except Exception:
                        pass
        except Exception as e:
            self._check_stop(show_name)
            signal.show_traceback_log(traceback.format_exc())
            signal.show_log_text(traceback.format_exc())
            signal.show_log_text(str(e))

        # 显示刮削结果
        try:
            Flags.scrape_done = await Flags.increment("scrape_done")
            count = Flags.scrape_done
            progress_value = count / count_all * 100
            progress_percentage = f"{progress_value:.2f}%"
            used_time = get_used_time(start_time)
            scrape_info_begin = f"{count:d}/{count_all:d} ({progress_percentage}) round({Flags.count_claw}) {split_path(file_path)[1]}    新的刮削线程"
            scrape_info_begin = "\n\n\n" + "=" * 40 + "\n" + scrape_info_begin
            scrape_info_after = f"\n 🕷 {get_current_time()} {count}/{count_all} {split_path(file_path)[1]} 刮削完成！用时 {used_time} 秒！"
            if manager.config.show_web_log:
                signal.show_log_text(scrape_info_begin + LogBuffer.log().get() + scrape_info_after)
            else:
                fail_reason = LogBuffer.error().get()
                if fail_reason:
                    signal.show_log_text(
                        scrape_info_begin + f"\n 🔴 [Failed] Reason: {fail_reason}" + scrape_info_after
                    )
                else:
                    signal.show_log_text(scrape_info_begin + scrape_info_after)
            remain_count = max(0, Flags.scrape_started - count)  # 避免瞬时负数显示"刮削中：-1"
            if Flags.scrape_started == count_all:
                signal.show_log_text(f" 🕷 剩余正在刮削的线程：{remain_count}")
            signal.label_result.emit(f" 刮削中：{remain_count} 成功：{Flags.succ_count} 失败：{Flags.fail_count}")
            signal.show_scrape_info(f"🔎 已刮削 {count}/{count_all}")
        except Exception as e:
            self._check_stop(show_name)
            signal.show_traceback_log(traceback.format_exc())
            signal.show_log_text(traceback.format_exc())
            signal.show_log_text(str(e))

        # 更新剩余任务
        try:
            try:
                Flags.remain_list.remove(file_path)
                Flags.can_save_remain = True
            except Exception as e1:
                signal.show_log_text(f"remove:  {file_path}\n {e1!s}\n {traceback.format_exc()}")
        except Exception as exc:
            self._check_stop(show_name)
            signal.show_traceback_log(traceback.format_exc())
            signal.show_log_text(traceback.format_exc())
            signal.show_log_text(str(exc))

        # 处理间歇刮削
        try:
            if manager.config.main_mode != 4 and Switch.REST_SCRAPE in manager.config.switch_on:
                async with self._rest_lock:
                    time_note = f" 🏖 已累计刮削 {count}/{count_all}，已连续刮削 {count - Flags.rest_now_begin_count}/{manager.config.rest_count}..."
                    signal.show_log_text(time_note)
                    if count - Flags.rest_now_begin_count >= manager.config.rest_count:
                        if Flags.sleep_end.is_set():
                            # 达到阈值且未在休息 → 启动休息
                            Flags.sleep_end.clear()
                            Flags.rest_next_begin_time = time.time()  # 下一轮倒计时开始时间
                            time_note = f'\n ⏸ 休息 {Flags.rest_time_convert} 秒，将在 <font color="red">{get_real_time(Flags.rest_next_begin_time + Flags.rest_time_convert)}</font> 继续刮削剩余的 {count_all - count} 个任务...\n'
                            signal.show_log_text(time_note)
                            while (
                                Switch.REST_SCRAPE in manager.config.switch_on
                                and time.time() - Flags.rest_next_begin_time < Flags.rest_time_convert
                            ):
                                if Flags.scrape_starting > count:  # 如果突然调大了文件数量，这时跳出休眠
                                    break
                                await asyncio.sleep(1)
                            Flags.rest_now_begin_count = count  # 休息周期结束，重置计数
                            Flags.sleep_end.set()  # 休眠结束，下一轮开始
                            Flags.next_start_time = time.time() - manager.config.thread_time
                        else:
                            await Flags.sleep_end.wait()  # 正在休息 → 等待休眠结束
                    # 未达阈值：继续刮削，无需处理
        except Exception as e:
            self._check_stop(show_name)
            signal.show_traceback_log(traceback.format_exc())
            signal.show_log_text(traceback.format_exc())
            signal.show_log_text(str(e))

    async def _download_images(
        self,
        res: CrawlersResult,
        other: OtherInfo,
        file_info: FileInfo,
        folder_new_path: Path,
        thumb_final_path: Path,
        fanart_final_path: Path,
        poster_final_path: Path,
        media_context: MediaResourceContext | None = None,
        single_folder_catched: bool = False,
    ) -> bool:
        if not await thumb_download(res, other, file_info.cd_part, folder_new_path, thumb_final_path, media_context):
            return False

        fanart_task = asyncio.create_task(fanart_download(res.number, other, file_info.cd_part, fanart_final_path))
        poster_task = asyncio.create_task(
            poster_download(res, other, file_info.cd_part, folder_new_path, poster_final_path, media_context)
        )
        extrafanart_task = (
            asyncio.create_task(extrafanart_download(res.extrafanart, res.extrafanart_from, folder_new_path))
            if single_folder_catched
            else None
        )

        try:
            await asyncio.gather(fanart_task, poster_task)
        except Exception:
            # fanart/poster 下载异常时取消 extrafanart 后台任务，避免任务泄漏到批次结束
            if extrafanart_task is not None and not extrafanart_task.done():
                extrafanart_task.cancel()
            raise

        if not poster_task.result() or not fanart_task.result():
            if extrafanart_task is not None and not extrafanart_task.done():
                extrafanart_task.cancel()
            return False

        await pic_some_deal(res.number, thumb_final_path, fanart_final_path)
        await add_mark(other, file_info, res.mosaic)

        if extrafanart_task is not None:
            await extrafanart_task
            await extrafanart_copy2(folder_new_path)
            await extrafanart_extras_copy(folder_new_path)

        return True

    async def _process_one_file(
        self, file_info: FileInfo, file_mode: FileMode
    ) -> tuple[CrawlersResult | None, OtherInfo | None]:
        media_context = MediaResourceContext()
        try:
            return await self._process_one_file_with_context(file_info, file_mode, media_context)
        finally:
            media_context.close()

    async def _process_one_file_with_context(
        self,
        file_info: FileInfo,
        file_mode: FileMode,
        media_context: MediaResourceContext,
    ) -> tuple[CrawlersResult | None, OtherInfo | None]:
        # 处理单个文件刮削
        # 初始化所需变量
        start_time = time.time()
        read_mode = manager.config.read_mode
        file_escape_size = float(manager.config.file_size)
        file_path = file_info.file_path

        # 获取文件信息
        movie_number = file_info.number
        folder_old_path = file_info.folder_path
        file_name = file_info.file_name
        file_ex = file_info.file_ex
        sub_list = file_info.sub_list

        # 获取设置的媒体目录、失败目录、成功目录
        paths = get_movie_path_setting(file_path)
        success_folder = paths.success_folder
        movie_path = paths.movie_path

        # 检查文件大小
        result = await check_file(file_path, file_escape_size)
        if not result:
            return None, None

        is_nfo_existed = False
        res = CrawlersResult.empty()  # todo 保证所有路径上均有 res 值
        file_classification = None
        # 读取模式
        file_can_download = True
        if manager.config.main_mode == 4:
            nfo_data, info = await get_nfo_data(file_path, movie_number)
            if nfo_data:  # 有nfo
                is_nfo_existed = True
                res = nfo_data
                movie_number = nfo_data.number
                file_info.shared_number = movie_number  # 供释放方释放共享番号双键
                file_classification = classify_existing_scrape_result(file_info.crawl_task(), res, manager.config)

                has_nfo_update = ReadMode.HAS_NFO_UPDATE in read_mode
                should_update_nfo = ReadMode.READ_UPDATE_NFO in read_mode
                redownload = ReadMode.READ_DOWNLOAD_AGAIN in read_mode
                if not has_nfo_update and not should_update_nfo and not redownload:  # 都不勾才跳过
                    show_result(res, start_time)
                    show_movie_info(file_info, nfo_data)
                    LogBuffer.log().write(f"\n 🙉 [Movie] {file_path}")
                    await save_success_list(file_path, file_path)
                    return nfo_data, info

                if not redownload:
                    file_can_download = False
            else:
                if "no_nfo_scrape" not in read_mode:  # 无 nfo 且未勾选「本地没有nfo的文件重新刮削」
                    return None, None

        # 判断是否write_nfo
        update_nfo = True
        # 不写nfo的情况：
        if manager.config.main_mode == 2 and Switch.SORT_DEL in manager.config.switch_on:
            # 2模式勾选“删除本地已下载的nfo文件”（暂无效，会直接return）
            update_nfo = False
        elif manager.config.main_mode in [1, 2, 3] or (
            manager.config.main_mode == 4 and not is_nfo_existed and ReadMode.NO_NFO_SCRAPE in read_mode
        ):
            # 1、2、3模式，或4模式启用了“本地没有nfo的文件重新刮削”（变量命名有点问题，存在"no_nfo_scrape"意思其实是要刮削）
            # 且
            if DownloadableFile.NFO not in manager.config.download_files:
                # [下载]处不勾选下载nfo时
                update_nfo = False
            if KeepableFile.NFO in manager.config.keep_files and is_nfo_existed:
                # [下载]处勾选保留nfo且nfo存在时
                update_nfo = False
        elif manager.config.main_mode == 4:
            # 读模式下，由"允许更新 nfo 文件"独立控制
            update_nfo = ReadMode.READ_UPDATE_NFO in read_mode

        # 读取模式下，补充缺失的演员 tmdbid（NFO已有→跳过，xlsx缓存→命中，API查询→补充）
        if update_nfo and is_nfo_existed and NfoInclude.ACTOR_TMDBID in manager.config.nfo_include_new:
            tmdb_api_base = manager.config.tmdb_api_base
            tmdb_api_key = manager.config.tmdb_api_key
            if tmdb_api_base and tmdb_api_key:
                existing_tmdb_ids = res.actor_tmdb_ids or {}
                all_actors = [a.strip() for a in (res.actor or "").split(",") if a.strip()]
                missing_actors = [a for a in all_actors if a not in existing_tmdb_ids]
                if missing_actors:
                    from .tmdb_actor import query_single_actor_cached, search_actor_db_reverse, update_actor_db_row

                    # 用 xlsx 反向搜索：从 NFO 中的演员名（可能是中文名/日文名/繁体名）查找 tmdbid 和日文原名
                    for actor_name in missing_actors:
                        row = search_actor_db_reverse(actor_name)
                        if row and row.get("tmdbid"):
                            existing_tmdb_ids[actor_name] = row["tmdbid"]
                            LogBuffer.log().write(f"  ℹ️ [TMDB] {actor_name} -> tmdbid={row['tmdbid']} (xlsx缓存)")
                    # 缓存命中部分必须立即回写 res，否则当 still_missing 为空（全命中）时
                    # res.actor_tmdb_ids 不会被更新，NFO 仍缺这些 tmdbid
                    res.actor_tmdb_ids = existing_tmdb_ids

                    # 仍未命中的演员，尝试用 xlsx 找到日文原名后查 TMDB API
                    still_missing = [a for a in missing_actors if a not in existing_tmdb_ids]
                    if still_missing:
                        try:
                            import aiohttp
                        except ImportError:
                            LogBuffer.log().write(
                                f"  ⚠️ [TMDB] 缺少 aiohttp 库，读取模式下无法通过 API 补充 {len(still_missing)} 个演员的 tmdbid"
                            )
                            res.actor_tmdb_ids = existing_tmdb_ids
                        else:
                            # 预加载 workbook，所有查询结果的写操作先攒在内存，最后一次性落盘
                            _wb, _db_path = _load_actor_db_wb()
                            try:
                                async with aiohttp.ClientSession() as client:
                                    protocol = "https://"
                                    base = tmdb_api_base.strip()
                                    if base.startswith("http://"):
                                        protocol = "http://"
                                        base = base[7:]
                                    elif base.startswith("https://"):
                                        protocol = "https://"
                                        base = base[8:]
                                    base_url = f"{protocol}{base}" if base else "https://api.tmdb.org"

                                    _read_mode_semaphore = asyncio.Semaphore(3)

                                    async def _query_one_actor(actor_name: str) -> None:
                                        row = search_actor_db_reverse(actor_name)
                                        jp_name = str(row.get("jp")) if row and row.get("jp") else actor_name
                                        try:
                                            query_result = await query_single_actor_cached(
                                                jp_name, base_url, tmdb_api_key, client
                                            )
                                            if query_result:
                                                tmdbid = query_result["pid"]
                                                existing_tmdb_ids[actor_name] = tmdbid
                                                translations = query_result.get("translations", {})
                                                aka = query_result.get("also_known_as", [])
                                                jp_original = query_result.get("original_name", "") or jp_name
                                                write_status = await update_actor_db_row(
                                                    jp=jp_original,
                                                    zh_cn=_normalize_translation(translations.get("zh_cn", "")),
                                                    zh_tw=_normalize_translation(translations.get("zh_tw", "")),
                                                    keyword=",".join(aka) if aka else "",
                                                    tmdbid=tmdbid,
                                                    append_keyword=True,
                                                    _wb=_wb,
                                                )
                                                LogBuffer.log().write(
                                                    f"  ✅ [TMDB] {actor_name} -> tmdbid={tmdbid} (读取模式补充)"
                                                )
                                                if write_status == "inserted_tmdbid":
                                                    LogBuffer.log().write(
                                                        f"  ✅ [演员数据库] 已写入 {jp_name} -> tmdbid={tmdbid}"
                                                    )
                                                elif write_status == "inserted_new_row":
                                                    LogBuffer.log().write(
                                                        f"  ✅ [演员数据库] 已新增 {jp_name}，并写入 tmdbid={tmdbid}"
                                                    )
                                                elif write_status == "kept_existing_tmdbid":
                                                    LogBuffer.log().write(
                                                        f"  ℹ️ [演员数据库] {jp_name} 已存在 tmdbid，保留原值"
                                                    )
                                                elif write_status == "missing_openpyxl":
                                                    LogBuffer.log().write(
                                                        f"  ⚠️ [演员数据库] 缺少 openpyxl，未写入 {jp_name} 的 tmdbid"
                                                    )
                                                elif write_status == "file_locked":
                                                    LogBuffer.log().write(
                                                        f"  ⚠️ [演员数据库] 文件被占用，未写入 {jp_name} 的 tmdbid"
                                                    )
                                                elif write_status.startswith("write_failed:"):
                                                    LogBuffer.log().write(
                                                        f"  ⚠️ [演员数据库] 写入失败，未保存 {jp_name} 的 tmdbid: {write_status.split(':', 1)[1]}"
                                                    )
                                            else:
                                                LogBuffer.log().write(f"  ⚠️ [TMDB] {actor_name} 未找到匹配的 TMDB 演员")
                                        except Exception as e:
                                            LogBuffer.log().write(f"  ❌ [TMDB] {actor_name} 查询失败: {e}")

                                    async def _limited_query(actor_name: str) -> None:
                                        async with _read_mode_semaphore:
                                            await _query_one_actor(actor_name)

                                    tasks = [asyncio.create_task(_limited_query(a)) for a in still_missing]
                                    await asyncio.gather(*tasks)

                                    _flush_actor_db_wb(_wb, _db_path)
                                    res.actor_tmdb_ids = existing_tmdb_ids
                            finally:
                                _wb.close()

        # 刮削json_data
        # 获取已刮削的json_data
        if file_classification is None:
            file_classification = classify_scrape_task(file_info.crawl_task(), manager.config)
        enable_shared_json = "." not in movie_number and file_classification.scraping_type != FixedScrapingType.GUOCHAN
        if enable_shared_json:
            # 首次发现该番号时原子性标记为“正在刮削”，避免两个协程同时走“首次”分支导致重复刮削
            async with Flags._json_get_lock:
                if movie_number not in Flags.json_get_status:
                    Flags.json_get_set.add(movie_number)
                    Flags.json_get_status[movie_number] = None
                    Flags.json_get_events[movie_number] = asyncio.Event()
                    # 读模式下 movie_number 可能被 nfo_data.number 覆盖，与原 file_info.number 不同。
                    # 释放方 _process_one_file_impl 的异常路径用 origin_number 释放，若二者不同需同时注册。
                    if movie_number != file_info.number:
                        Flags.json_get_set.add(file_info.number)
                        Flags.json_get_status[file_info.number] = None
                        Flags.json_get_events[file_info.number] = asyncio.Event()
                    LogBuffer.log().write(f"\n 🟡 [Same Number] 首次刮削，开始共享番号数据：{movie_number}")
                    is_first = True
                else:
                    is_first = False
            if not is_first:
                # 同番号任务等待首个任务完成；若首个任务失败，直接结束等待，避免线程卡死
                wait_timeout = 300
                waited = 0
                event = Flags.json_get_events.get(movie_number)
                while Flags.json_get_status.get(movie_number) is None:
                    if Flags.stop_requested or signal.stop:
                        LogBuffer.log().write(f"\n 🟡 [Same Number] 检测到停止请求，取消等待：{movie_number}")
                        return None, None
                    if waited >= wait_timeout:
                        LogBuffer.error().write(f"同番号等待超时（{wait_timeout}秒），取消等待：{movie_number}")
                        async with Flags._json_get_lock:
                            Flags.json_get_status[movie_number] = False
                            event = Flags.json_get_events.get(movie_number)
                            if event is not None:
                                event.set()
                        return None, None
                    if event is None:
                        await asyncio.sleep(1)
                    else:
                        try:
                            await asyncio.wait_for(event.wait(), timeout=1)
                        except TimeoutError:
                            pass
                    waited += 1
                if Flags.json_get_status.get(movie_number) is False:
                    LogBuffer.error().write(f"同番号任务失败，取消等待：{movie_number}")
                    return None, None

        pre_data = Flags.json_data_dic.get(movie_number)
        # 已存在该番号数据时直接使用该数据
        if pre_data and enable_shared_json:
            pre_res = pre_data.data
            res = update(pre_res, file_info)

            tags = pre_res.tag.split(",")
            tags = [
                tag
                for tag in tags
                if tag
                not in (  # 移除与具体文件相关的 tag; 分辨率相关 tag 在 add_definition_tag 中会移除; codec tag 无法穷举, 移除常见类型
                    # todo 所有文件相关的 tag 推迟到 write_nfo 时从 file_info 生成, json_data_dic 只存储通用的 tag
                    "国产",
                    "國產",
                    "里番",
                    "裏番",
                    "动漫",
                    "動漫",
                    "H264",
                    "HEVC",
                    "MPEG4",
                    "VP8",
                    "VP9",
                )
            ]
            tags.append(file_info.mosaic)
            if file_info.has_sub and TagInclude.CNWORD in manager.config.nfo_tag_include:
                tags.append("中文字幕")
            res.tag = ",".join(tags)

        elif not is_nfo_existed:
            # ========================= call crawlers =========================
            # res = await crawl(file_info.crawl_task(), file_mode)

            crawl_task = file_info.crawl_task()
            crawl_task.media_context = media_context  # type: ignore[attr-defined]
            scraper = FileScraper(manager.config, self.crawler_provider)
            crawl_result = await scraper.run(crawl_task, file_mode)
            if crawl_result is None:
                return None, None
            res = crawl_result
            # 处理 FileInfo 和 CrawlersResult 的共同字段, 即 number/mosaic/letters
            # todo 理想情况, crawl 后应该以 res 为准, 后续不应再访问 file_info 的相关字段
            # todo 注意, 实际上目前各 crawler 返回的 mosaic 和 number 字段并未被使用
            # 1. number 在 crawl 中被更新, 当前只可能取 file_info.number/short_number/appoint_number
            # 2. letters 在 crawl 过程不会变化, 直接取 file_info 的值
            res.letters = file_info.letters
            # 3. res.mosaic 在 crawl 中被更新, 实际上完全是由 file_info 的某些字段决定的, 和初始化 file_info.mosaic 的逻辑存在重复
            file_info.mosaic = res.mosaic

        # 显示json_data结果或日志
        show_result(res, start_time)

        # 映射或翻译
        # 无已刮削数据时执行；演员/标签等映射服务于命名变量（{actor}/{tag} 等），与是否写 NFO 解耦
        if not pre_data:
            deal_some_field(res)  # 处理字段
            replace_special_word(res)  # 替换特殊字符
            if update_nfo:
                await translate_title_outline(res, file_info.cd_part, movie_number)  # 翻译json_data（标题/介绍）
                deal_some_field(res)  # 再处理一遍字段，翻译后可能出现要去除的内容
                # 查询演员 TMDB ID（在演员名映射前，使用原始名）
                if NfoInclude.ACTOR_TMDBID in manager.config.nfo_include_new:
                    from .tmdb_actor import fetch_actor_tmdb_ids

                    res.actor_tmdb_ids = await fetch_actor_tmdb_ids(res.actors, self.crawler_provider.client)
            # 保存原始演员名（映射前），供读取模式反向查找使用
            res.original_actors = res.actors.copy() if res.actors else []
            await translate_actor(res)  # 映射输出演员名/信息
            translate_info(res, file_info.has_sub)  # 映射输出标签等信息
            replace_word(res)

        # 更新视频分辨率
        definition, codec = await get_video_size(file_path, file_info.number)
        file_info.definition, file_info.codec = definition, codec
        add_definition_tag(res, definition, codec)

        # 显示json_data内容
        show_movie_info(file_info, res)

        # 读模式不勾"重新整理分类"时跳过路径计算
        skip_reorganize = manager.config.main_mode == 4 and is_nfo_existed and ReadMode.HAS_NFO_UPDATE not in read_mode

        if skip_reorganize:
            naming_rule = _generate_file_name(file_info.cd_part, file_info, res)
            file_new_name = naming_rule + file_ex.lower()
            file_new_path = success_folder / file_new_name
            folder_new_path = success_folder
            nfo_new_path = success_folder / (naming_rule + ".nfo")
            poster_new_path_with_filename = success_folder / (naming_rule + "-poster.jpg")
            thumb_new_path_with_filename = success_folder / (naming_rule + "-thumb.jpg")
            fanart_new_path_with_filename = success_folder / (naming_rule + "-fanart.jpg")
            _, folder_name = _get_folder_path(success_folder, file_info, res)
            if manager.config.pic_simple_name and folder_name:
                poster_final_path = success_folder / "poster.jpg"
                thumb_final_path = success_folder / "thumb.jpg"
                fanart_final_path = success_folder / "fanart.jpg"
            else:
                poster_final_path = poster_new_path_with_filename
                thumb_final_path = thumb_new_path_with_filename
                fanart_final_path = fanart_new_path_with_filename
        else:
            # 生成输出文件夹和输出文件的路径
            (
                folder_new_path,
                file_new_path,
                nfo_new_path,
                poster_new_path_with_filename,
                thumb_new_path_with_filename,
                fanart_new_path_with_filename,
                naming_rule,
                poster_final_path,
                thumb_final_path,
                fanart_final_path,
            ) = get_output_name(file_info, res, success_folder, file_ex)

        # 判断输出文件的路径是否重复
        if not skip_reorganize:
            if manager.config.soft_link == 0:
                async with Flags._file_path_lock:
                    done_file_new_path_list = Flags.file_new_path_dic.get(file_new_path)
                    if not done_file_new_path_list:
                        Flags.file_new_path_dic[file_new_path] = [file_path]
                    else:
                        done_file_new_path_list.append(file_path)
                        done_file_new_path_list.sort(reverse=True)
                        LogBuffer.error().write(
                            "存在重复文件（指刮削后的文件路径相同！），请检查:\n    🍁 "
                            + "\n    🍁 ".join(str(path) for path in done_file_new_path_list)
                        )
                        res.outline = split_path(str(file_path))[1]
                        res.tag = str(file_path)
                        return None, None

        # 不移动文件时，NFO、图片等写入原目录
        if not manager.config.success_file_move:
            nfo_new_path = file_path.with_suffix(".nfo")
            folder_new_path = folder_old_path
            poster_new_path_with_filename = folder_old_path / (file_name + "-poster.jpg")
            thumb_new_path_with_filename = folder_old_path / (file_name + "-thumb.jpg")
            fanart_new_path_with_filename = folder_old_path / (file_name + "-fanart.jpg")
            _, folder_name = _get_folder_path(folder_old_path, file_info, res)
            if manager.config.pic_simple_name and folder_name:
                poster_final_path = folder_old_path / "poster.jpg"
                thumb_final_path = folder_old_path / "thumb.jpg"
                fanart_final_path = folder_old_path / "fanart.jpg"
            else:
                poster_final_path = poster_new_path_with_filename
                thumb_final_path = thumb_new_path_with_filename
                fanart_final_path = fanart_new_path_with_filename

        # 判断输出文件夹和文件是否已存在，如无则创建输出文件夹
        other = OtherInfo.empty()
        if not skip_reorganize:
            if not await creat_folder(
                other,
                res,
                folder_new_path,
                file_path,
                file_new_path,
                thumb_new_path_with_filename,
                poster_new_path_with_filename,
            ):
                return None, None

        # 初始化图片已下载地址的字典
        async with Flags._file_done_lock:
            if not Flags.file_done_dic.get(res.number):
                Flags.file_done_dic[res.number] = FileDoneDict(
                    poster=None,
                    thumb=None,
                    fanart=None,
                    trailer=None,
                    local_poster=None,
                    local_thumb=None,
                    local_fanart=None,
                    local_trailer=None,
                )

        # 视频模式（原来叫整理模式）
        # 视频模式（仅根据刮削数据把电影命名为番号并分类到对应目录名称的文件夹下）
        if manager.config.main_mode == 2:
            # 移动文件
            if await move_movie(other, file_info, file_path, file_new_path):
                if Switch.SORT_DEL in manager.config.switch_on:
                    await deal_old_files(
                        res.number,
                        other,
                        folder_old_path,
                        folder_new_path,
                        file_path,
                        thumb_new_path_with_filename,
                        poster_new_path_with_filename,
                        fanart_new_path_with_filename,
                        nfo_new_path,
                        poster_final_path,
                        thumb_final_path,
                        fanart_final_path,
                        naming_rule,
                    )  # 清理旧的thumb、poster、fanart、nfo
                await save_success_list(file_path, file_new_path)  # 保存成功列表
                return res, other
            # 返回MDCx1_1main, 继续处理下一个文件
            return None, None

        # 清理旧的thumb、poster、fanart、extrafanart、nfo
        pic_final_catched = False
        single_folder_catched = False
        pic_final_catched, single_folder_catched = await deal_old_files(
            res.number,
            other,
            folder_old_path,
            folder_new_path,
            file_path,
            thumb_new_path_with_filename,
            poster_new_path_with_filename,
            fanart_new_path_with_filename,
            nfo_new_path,
            poster_final_path,
            thumb_final_path,
            fanart_final_path,
            naming_rule,
        )

        # 如果 final_pic_path 没处理过，这时才需要下载和加水印
        if pic_final_catched and file_can_download:
            if not await self._download_images(
                res,
                other,
                file_info,
                folder_new_path,
                thumb_final_path,
                fanart_final_path,
                poster_final_path,
                media_context,
                single_folder_catched,
            ):
                return None, None

        if file_can_download:
            # trailer 有带文件名、不带文件名两种命名方式，不能依赖图片处理权。
            await trailer_download(res, folder_new_path, folder_old_path, naming_rule)
            if single_folder_catched:
                await copy_trailer_to_theme_videos(folder_new_path, naming_rule)

        # 生成nfo文件
        await write_nfo(file_info, res, nfo_new_path, folder_new_path, update_nfo)

        # 移动字幕、种子、bif、trailer、其他文件（配置允许时才执行）
        if manager.config.success_file_move:
            if file_info.has_sub:
                await move_sub(folder_old_path, folder_new_path, file_name, sub_list, naming_rule)
            await move_torrent(folder_old_path, folder_new_path, file_name, movie_number, naming_rule)
            await move_bif(folder_old_path, folder_new_path, file_name, naming_rule)
            await move_other_file(res.number, folder_old_path, folder_new_path, file_name, naming_rule)

            # 移动文件
            if not await move_movie(other, file_info, file_path, file_new_path):
                return None, None
        await save_success_list(file_path, file_new_path)

        # 创建软链接及复制文件（由 auto_link 独立控制）
        if manager.config.auto_link:
            if manager.config.success_file_move:
                target_dir = Path(manager.config.localdisk_path) / folder_new_path.relative_to(
                    success_folder, walk_up=True
                )
            else:
                target_dir = Path(manager.config.localdisk_path) / folder_old_path.relative_to(movie_path, walk_up=True)
            copy = Switch.COPY_NETDISK_NFO in manager.config.switch_on
            await newtdisk_creat_symlink(copy, folder_new_path, target_dir)

        # json添加封面缩略图路径（仅在路径重算后有值）
        if poster_final_path is not None:
            other.poster_path = poster_final_path
            other.thumb_path = thumb_final_path
            other.fanart_path = fanart_final_path
            if not await aiofiles.os.path.exists(thumb_final_path) and await aiofiles.os.path.exists(fanart_final_path):
                other.thumb_path = fanart_final_path

        # 所有图片及相关文件处理完成后，最后统一压缩最终输出目录中的图片。
        await compress_images_in_folder_async(folder_new_path, manager.config.compress_downloaded_images)

        return res, other

    def _check_stop(self, show_name: str) -> None:
        if signal.stop or Flags.stop_requested:
            Flags.now_kill += 1
            signal.show_log_text(
                f" 🕷 {get_current_time()} 已停止刮削：{Flags.now_kill}/{Flags.total_kills} {show_name}"
            )
            signal.set_label_file_path.emit(
                f"⛔️ 正在停止刮削...\n   正在停止已在运行的任务线程（{Flags.now_kill}/{Flags.total_kills}）..."
            )
            raise StopScrape("手动停止刮削")

    async def _failed_file_info_show(self, count: str, p: Path, error_info: str) -> None:
        info_str = f"{'🔴 ' + count + '.':<3} {p} \n    所在目录: {p.parent} \n    失败原因: {error_info} \n"
        if await aiofiles.os.path.islink(p):
            info_str = f"{'🔴 ' + count + '.':<3} {p} \n    指向文件: {p.resolve()} \n    失败原因: {error_info} \n"
        signal.logs_failed_show.emit(info_str)


async def _safe_mtime(file_path: Path) -> float:
    """获取文件 mtime；失败返回 0（视为文件已变化，避免误跳过）。"""
    try:
        stat = await aiofiles.os.stat(file_path)
        return float(getattr(stat, "st_mtime", 0.0) or 0.0)
    except Exception:
        return 0.0


def _load_actor_db_wb():
    """加载演员数据库 workbook，返回 (wb, db_path) 或 (None, None)。"""
    import openpyxl

    from ..config.resources import DB_HEADERS
    from .tmdb_actor import _get_db_path

    db_path = _get_db_path()
    if db_path.exists():
        wb = openpyxl.load_workbook(db_path)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "演员数据库"
        for col, header in enumerate(DB_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
    return wb, db_path


def _flush_actor_db_wb(wb, db_path):
    """将 workbook 落盘并重载缓存。"""
    from ..config.resources import resources
    from .tmdb_actor import _ACTOR_DB_ROW_INDEX, _ACTOR_DB_ROW_INDEX_LOCK, _format_db_worksheet

    try:
        ws = wb.active
        _format_db_worksheet(ws)
        wb.save(db_path)
        wb.close()
        with _ACTOR_DB_ROW_INDEX_LOCK:
            _ACTOR_DB_ROW_INDEX.clear()
        resources.reload_actor_db()
    except Exception as e:
        LogBuffer.log().write(f" ❌ [演员数据库] 落盘失败: {e}")


def start_new_scrape(file_mode: FileMode, movie_list: list[Path] | None = None) -> None:
    Flags.stop_requested = False
    signal.stop = False
    signal.change_buttons_status.emit()
    signal.exec_set_processbar.emit(0)
    try:
        Flags.start_time = time.time()
        with manager.acquire_computed() as computed:
            crawler_provider = CrawlerProvider(
                manager.config, computed.async_client, config_getter=lambda: manager.config
            )
        scraper = Scraper(crawler_provider)
        executor.submit(scraper.run(file_mode, movie_list))
    except Exception as e:
        signal.show_traceback_log(traceback.format_exc())
        signal.show_log_text(traceback.format_exc())
        LogBuffer.error().write(f"start_new_scrape error: {e}")


def get_remain_list() -> bool:
    """This function is intended to be sync."""
    remain_list_path = resources.u("remain.txt")
    if not remain_list_path.is_file():
        return False
    with open(remain_list_path, encoding="utf-8", errors="ignore") as f:
        remains = [p for path in f if (line := path.strip()) and (p := Path(line)).suffix]
    Flags.remain_list = remains
    if not len(Flags.remain_list) or Switch.REMAIN_TASK not in manager.config.switch_on:
        return False
    box = QMessageBox(QMessageBox.Icon.Information, "继续刮削", "上次刮削未完成，是否继续刮削剩余任务？")
    box.setStandardButtons(
        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
    )
    yes_button = box.button(QMessageBox.StandardButton.Yes)
    assert yes_button is not None
    yes_button.setText("继续刮削剩余任务")
    no_button = box.button(QMessageBox.StandardButton.No)
    assert no_button is not None
    no_button.setText("从头刮削")
    cancel_button = box.button(QMessageBox.StandardButton.Cancel)
    assert cancel_button is not None
    cancel_button.setText("取消")
    box.setDefaultButton(QMessageBox.StandardButton.No)
    reply = box.exec()
    if reply == QMessageBox.StandardButton.Yes:
        signal.show_log_text("🍯 🍯 🍯 NOTE: 用户选择继续刮削剩余任务。")
    elif reply == QMessageBox.StandardButton.No:
        signal.show_log_text("🍯 🍯 🍯 NOTE: 用户选择从头刮削。")
        return False  # 从头刮削
    else:
        signal.show_log_text("🍯 🍯 🍯 NOTE: 已取消本次刮削启动。")
        return True  # 不刮削（包括点取消、ESC、右上角关闭）

    movie_paths = parse_media_paths()

    p = Flags.remain_list[0]
    if not is_any_descendant(p, *movie_paths):
        box = QMessageBox(
            QMessageBox.Icon.Warning,
            "提醒",
            f"很重要！！请注意：\n当前待刮削目录：{';'.join(str(path) for path in movie_paths)}\n剩余任务文件路径：{p.resolve()}\n"
            "文件不在当前待刮削目录中, 可能是使用其他配置扫描的！\n"
            "请确认成功输出目录和失败目录是否正确！如果配置不正确，继续刮削可能会导致文件被移动到新配置的输出位置！\n是否继续刮削？",
        )
        box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        yes_button = box.button(QMessageBox.StandardButton.Yes)
        assert yes_button is not None
        yes_button.setText("继续")
        no_button = box.button(QMessageBox.StandardButton.No)
        assert no_button is not None
        no_button.setText("取消")
        box.setDefaultButton(QMessageBox.StandardButton.No)
        reply = box.exec()
        if reply == QMessageBox.StandardButton.No:
            return True
    signal.show_log_text(f"🍯 🍯 🍯 NOTE: 继续刮削未完成任务！！！ 剩余未刮削文件数量（{len(Flags.remain_list)})")
    start_new_scrape(FileMode.Default, Flags.remain_list)
    return True


def again_search() -> None:
    Flags.new_again_dic = Flags.again_dic.copy()
    new_movie_list = list(Flags.new_again_dic.keys())
    Flags.again_dic.clear()
    start_new_scrape(FileMode.Again, new_movie_list)


async def move_sub(
    folder_old_path: Path,
    folder_new_path: Path,
    file_name: str,
    sub_list: list[str],
    naming_rule: str,
) -> None:
    copy_flag = False

    # 更新模式 或 读取模式
    if manager.config.main_mode > 2:
        if manager.config.update_mode == "c" and not manager.config.success_file_rename:
            return

    # 软硬链接开时，复制字幕（EMBY 显示字幕）
    elif manager.config.soft_link > 0:
        copy_flag = True

    # 成功移动关、成功重命名关时，返回
    elif not manager.config.success_file_move and not manager.config.success_file_rename:
        return

    for sub in sub_list:
        sub_old_path = str(folder_old_path / (file_name + sub))
        sub_new_path = str(folder_new_path / (naming_rule + sub))
        sub_new_path_chs = str(folder_new_path / (naming_rule + ".chs" + sub))
        if manager.config.subtitle_add_chs and ".chs" not in sub:
            sub_new_path = sub_new_path_chs
        if await aiofiles.os.path.exists(sub_old_path) and not await aiofiles.os.path.exists(sub_new_path):
            if copy_flag:
                if not await copy_file_async(sub_old_path, sub_new_path):
                    LogBuffer.log().write(f"\n 🔴 Sub copy failed: {sub_old_path}")
                    continue
                LogBuffer.log().write("\n 🍀 Sub done!")
            elif not await move_file_async(sub_old_path, sub_new_path):
                LogBuffer.log().write(f"\n 🔴 Sub move failed: {sub_old_path}")
                continue
            else:
                LogBuffer.log().write("\n 🍀 Sub done!")
        else:
            if await aiofiles.os.path.exists(sub_old_path):
                LogBuffer.log().write(f"\n 🍀 Sub already exists: {sub_new_path}")
