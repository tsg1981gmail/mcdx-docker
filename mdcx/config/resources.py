import asyncio
import os
import sys
import threading
import traceback
from pathlib import Path
from typing import TYPE_CHECKING

import defusedxml  # noqa: F401  # 探活：ImportError 时给出友好错误（见 L340 提示）
import zhconv
from PyQt6.QtGui import QFontDatabase

from ..consts import IS_PYINSTALLER, MAIN_PATH
from ..manual import ManualConfig
from ..models.log_buffer import LogBuffer
from ..signals import signal
from ..utils import singleton
from ..utils.file import copy_file_sync
from .manager import manager

try:
    import openpyxl
except ImportError:
    openpyxl = None

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# 演员数据库 xlsx 列索引（与 tmdb_actor 共享）
COL_JP = 0
COL_ZH_CN = 1
COL_ZH_TW = 2
COL_KEYWORD = 3
COL_HREF = 4
COL_TMDBID = 5
COL_TMDB_URL = 6
COL_BIRTH_DATE = 7
COL_BIO = 8

DB_HEADERS = ["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"]


def _tmdb_person_url(tmdbid: int | str) -> str:
    return f"https://www.themoviedb.org/person/{tmdbid}"


ACTOR_DB_SHEET = "演员数据库"


def _normalize_info_key(value: str) -> str:
    key = value.upper()
    for source, target in ManualConfig.FULL_HALF_CHAR:
        key = key.replace(source, target)
    return key


def _build_info_db_index(rows: list[dict]) -> dict[str, dict]:
    index: dict[str, dict] = {}
    for item in rows:
        for keyword in (item.get("keyword") or "").split(","):
            keyword = keyword.strip()
            if keyword:
                index.setdefault(_normalize_info_key(keyword), item)
        for attr in ("zh_cn", "zh_tw", "jp"):
            value = item.get(attr) or ""
            if value:
                index.setdefault(_normalize_info_key(value), item)
    return index


def get_actor_db_sheet(wb) -> "Worksheet":
    """显式取「演员数据库」sheet，不依赖 sheet 顺序（防止男优备份等辅助 sheet 被误读）。

    若工作簿中无该名（新建/异常文件），回退到 active sheet 保持兼容。
    """
    if ACTOR_DB_SHEET in wb.sheetnames:
        return wb[ACTOR_DB_SHEET]
    return wb.active


def read_actor_db_xlsx(db_path: Path) -> dict[str, dict]:
    db: dict[str, dict] = {}
    import openpyxl

    wb = openpyxl.load_workbook(db_path, read_only=True, data_only=True)
    ws = get_actor_db_sheet(wb)
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue
        if len(row) < 1:
            continue
        jp = str(row[COL_JP] or "").strip()
        if not jp:
            continue
        tmdbid_val = None
        tmdbid_raw = str(row[COL_TMDBID] or "").strip() if len(row) > COL_TMDBID else ""
        if tmdbid_raw and tmdbid_raw.isdigit():
            tmdbid_val = int(tmdbid_raw)
        tmdb_url = str(row[COL_TMDB_URL] or "").strip() if len(row) > COL_TMDB_URL else ""
        if tmdbid_val is not None:
            tmdb_url = _tmdb_person_url(tmdbid_val)
        db[jp] = {
            "zh_cn": str(row[COL_ZH_CN] or "").strip() if len(row) > COL_ZH_CN else "",
            "zh_tw": str(row[COL_ZH_TW] or "").strip() if len(row) > COL_ZH_TW else "",
            "keyword": str(row[COL_KEYWORD] or "").strip() if len(row) > COL_KEYWORD else "",
            "href": str(row[COL_HREF] or "").strip() if len(row) > COL_HREF else "",
            "tmdbid": tmdbid_val,
            "tmdb_url": tmdb_url,
            "birth_date": str(row[COL_BIRTH_DATE] or "").strip() if len(row) > COL_BIRTH_DATE else "",
            "bio": str(row[COL_BIO] or "").strip() if len(row) > COL_BIO else "",
        }
    wb.close()
    return db


@singleton
class Resources:
    def __init__(self):
        # 获取内置资源路径和用户数据路径
        self._resources_base = MAIN_PATH / "resources"
        if IS_PYINSTALLER:
            # 获取 pyinstaller 打包程序运行时解压资源的临时目录
            try:
                self._resources_base = Path(sys._MEIPASS) / "resources"  # type: ignore
            except Exception:
                signal.show_traceback_log(self._resources_base)
                signal.show_traceback_log(traceback.format_exc())
        self._userdata_base = manager.data_folder / "userdata"
        self._userdata_base.mkdir(parents=True, exist_ok=True)  # 确保用户数据目录存在

        # 获取资源路径
        self.actor_db_backup_path = self.r("userdata/actor_database.xlsx")  # 内置演员数据库 xlsx
        self.info_db_backup_path = self.r("userdata/info_database.xlsx")  # 内置信息映射数据库 xlsx
        self.tmdb_query_cache_path = self.u("actor_tmdb_query_cache.json")  # TMDB 查询缓存（与 actor_database 同目录）

        self.icon_ico = self.qtr("Img/MDCx.ico")  # 任务栏图标
        self.right_menu = self.qtr("Img/menu.svg")  # 主界面菜单按钮
        self.play_icon = self.qtr("Img/play.svg")  # 主界面播放按钮
        self.open_folder_icon = self.qtr("Img/folder.svg")  # 主界面打开文件夹按钮
        self.open_nfo_icon = self.qtr("Img/nfo.svg")  # 主界面打开nfo按钮
        self.input_number_icon = self.qtr("Img/number.svg")  # 主界面输入番号按钮
        self.input_website_icon = self.qtr("Img/website.svg")  # 主界面输入网址按钮
        self.del_file_icon = self.qtr("Img/delfile.svg")  # 主界面删除文件按钮
        self.del_folder_icon = self.qtr("Img/delfolder.svg")  # 主界面删除文件夹按钮
        self.start_icon = self.qtr("Img/start.svg")  # 主界面开始按钮
        self.stop_icon = self.qtr("Img/stop.svg")  # 主界面开始按钮
        self.show_logs_icon = self.qtr("Img/show.svg")  # 日志界面显示日志按钮
        self.hide_logs_icon = self.qtr("Img/hide.svg")  # 日志界面隐藏日志按钮
        self.hide_boss_icon = self.qtr("Img/hide_boss.svg")  # 隐藏界面按钮
        self.save_failed_list_icon = self.qtr("Img/save.svg")  # 保存失败列表按钮
        self.clear_tree_icon = self.qtr("Img/clear.svg")  # 主界面清空结果列表按钮
        self.home_icon = self.qtr("Img/home.svg")
        self.log_icon = self.qtr("Img/log.svg")
        self.tool_icon = self.qtr("Img/tool.svg")
        self.setting_icon = self.qtr("Img/setting.svg")
        self.net_icon = self.qtr("Img/net.svg")
        self.help_icon = self.qtr("Img/help.svg")

        self.mark_4k = self.r("Img/4k.png")
        self.mark_8k = self.r("Img/8k.png")
        self.mark_sub = self.r("Img/sub.png")
        self.mark_youma = self.r("Img/youma.png")
        self.mark_umr = self.r("Img/umr.png")
        self.mark_leak = self.r("Img/leak.png")
        self.mark_wuma = self.r("Img/wuma.png")
        self.icon_4k_path = self.u("watermark/4k.png")
        self.icon_8k_path = self.u("watermark/8k.png")
        self.icon_sub_path = self.u("watermark/sub.png")
        self.icon_youma_path = self.u("watermark/youma.png")
        self.icon_umr_path = self.u("watermark/umr.png")
        self.icon_leak_path = self.u("watermark/leak.png")
        self.icon_wuma_path = self.u("watermark/wuma.png")

        self.actor_db: dict[str, dict] | None = None  # 演员数据库（xlsx 格式）
        self.actor_db_reverse_index: dict[str, str] | None = None  # 规范名/别名 -> jp 名索引
        self.info_db: list[dict] | None = None  # 信息映射数据库（xlsx 格式，有序列表以保持行顺序）
        self.info_db_index: dict[str, dict] | None = None
        self._data_ready = threading.Event()
        self._data_load_lock = threading.Lock()
        self._data_loading = False
        self._data_load_error: Exception | None = None

        self._get_mark_icon()
        zhconv.loaddict(str(self.r("zhconv/zhcdict.json")))  # 加载繁简转换字典

    def start_data_loading(self) -> None:
        """后台准备本地 XLSX 数据，避免构造 Resources 阻塞首屏。"""
        with self._data_load_lock:
            if self._data_loading or self._data_ready.is_set():
                return
            self._data_loading = True
        threading.Thread(target=self._load_local_data, name="ResourcesDataLoader", daemon=True).start()

    def _load_local_data(self) -> None:
        try:
            self._get_or_generate_local_data()
        except Exception as e:
            self._data_load_error = e
            signal.show_traceback_log(f"本地数据库后台加载失败: {e}\n{traceback.format_exc()}")
        finally:
            self._data_ready.set()

    def ensure_data_ready(self) -> None:
        """等待本地 XLSX 数据就绪，供后台业务首次访问时调用。"""
        if not self._data_ready.is_set():
            self.start_data_loading()
            self._data_ready.wait()
        if self._data_load_error is not None:
            raise RuntimeError("本地数据库加载失败") from self._data_load_error

    def r(self, relative_path: str | Path):
        return self._resources_base / relative_path

    def qtr(self, relative_path: str | Path):
        # Qt 内部所有路径都使用正斜杠
        return self.r(relative_path).as_posix()

    def u(self, relative_path: str | Path):
        return self._userdata_base / relative_path

    def get_actor_data(self, actor):
        self.ensure_data_ready()
        # 初始化数据
        actor_data = {
            "zh_cn": actor,
            "zh_tw": actor,
            "jp": actor,
            "keyword": [actor],
            "href": "",
            "birth_date": "",
            "bio": "",
            "has_name": False,
        }

        actor_db = self.actor_db
        if actor_db is not None:
            from ..core.tmdb_actor import search_actor_db_reverse

            row = search_actor_db_reverse(actor)
            if row:
                actor_data["zh_cn"] = row.get("zh_cn") or actor
                actor_data["zh_tw"] = row.get("zh_tw") or actor
                actor_data["jp"] = row.get("jp") or actor
                kw = row.get("keyword") or ""
                actor_data["keyword"] = [k.strip() for k in kw.split(",") if k.strip()] if kw else [actor_data["jp"]]
                actor_data["href"] = row.get("href") or ""
                actor_data["birth_date"] = row.get("birth_date") or ""
                actor_data["bio"] = row.get("bio") or ""
                actor_data["has_name"] = True
                return actor_data
        return actor_data

    def get_info_data(self, info):
        self.ensure_data_ready()
        # 初始化数据
        info_data = {
            "zh_cn": info,
            "zh_tw": info,
            "jp": info,
            "keyword": [info],
            "has_name": False,
        }

        # 查询信息映射数据库 xlsx，索引在加载时建立，查询保持 O(1)。
        # 加锁防并发 reload_info_db 置 None 导致 _build_info_db_index(None) 抛 TypeError
        info_db_index = self.info_db_index
        if info_db_index is None:
            with self._data_load_lock:
                info_db_index = self.info_db_index
                if info_db_index is None and self.info_db is not None:
                    info_db_index = _build_info_db_index(self.info_db)
                    self.info_db_index = info_db_index
        if info_db_index is not None:
            row = info_db_index.get(_normalize_info_key(info))
            if row is not None:
                info_data["zh_cn"] = (row.get("zh_cn") or info).replace("删除", "")
                info_data["zh_tw"] = (row.get("zh_tw") or info).replace("删除", "")
                info_data["jp"] = (row.get("jp") or info).replace("删除", "")
                kw = row.get("keyword") or ""
                info_data["keyword"] = (
                    [k.strip() for k in kw.split(",") if k.strip()] if kw else [row.get("jp") or info]
                )
                info_data["has_name"] = True
                return info_data
        return info_data

    def get_fonts(self):
        font_folder_path = self.qtr("fonts")
        for f in os.listdir(font_folder_path):
            QFontDatabase.addApplicationFont(os.path.join(font_folder_path, f))  # 字体路径

    def _get_or_generate_local_data(self):
        """如果用户数据目录下已有数据则直接读取, 否则根据内置数据生成"""
        # 演员数据库 xlsx：尝试迁移 XML → xlsx，或直接加载
        db_local_path = self.u("actor_database.xlsx")
        if not os.path.exists(db_local_path):
            # 尝试从 XML 迁移
            try:
                from ..core.tmdb_actor import migrate_xml_to_xlsx

                # 同步执行迁移（在事件循环外）
                loop = asyncio.new_event_loop()
                loop.run_until_complete(migrate_xml_to_xlsx())
                loop.close()
            except Exception:
                pass

            # 如果迁移未生成文件，尝试复制内置备份
            if not os.path.exists(db_local_path) and os.path.exists(self.actor_db_backup_path):
                copy_file_sync(self.actor_db_backup_path, db_local_path)

        # 信息映射数据库 xlsx：尝试迁移 XML → xlsx，或直接加载
        info_db_local_path = self.u("info_database.xlsx")
        if not os.path.exists(info_db_local_path):
            # 尝试从 XML 迁移
            try:
                from ..core.tmdb_actor import migrate_info_xml_to_xlsx

                loop = asyncio.new_event_loop()
                loop.run_until_complete(migrate_info_xml_to_xlsx())
                loop.close()
            except Exception:
                pass

            # 如果迁移未生成文件，尝试复制内置备份
            if not os.path.exists(info_db_local_path) and os.path.exists(self.info_db_backup_path):
                copy_file_sync(self.info_db_backup_path, info_db_local_path)

        # 载入 amazon_asin_database.xlsx
        asin_db_local_path = self.u("amazon_asin_database.xlsx")
        asin_db_backup_path = self.r("userdata/amazon_asin_database.xlsx")
        if not os.path.exists(asin_db_local_path):
            copy_file_sync(asin_db_backup_path, asin_db_local_path)

        # 加载数据库 xlsx
        merge_actor_db_from_backup(self.actor_db_backup_path, self.u("actor_database.xlsx"))
        merge_info_db_from_backup(self.info_db_backup_path, self.u("info_database.xlsx"))
        from ..core.amazon_database import merge_asin_db_from_backup

        merge_asin_db_from_backup(asin_db_backup_path, asin_db_local_path)
        self.reload_actor_db()
        self.reload_info_db()

    def reload_actor_db(self):
        """重新加载演员数据库 xlsx（在刮削更新后调用）"""
        if openpyxl is None:
            old = self.actor_db
            self.actor_db = None
            self.actor_db_reverse_index = None
            LogBuffer.log().write("  ❌ [演员数据库] 初始化失败: openpyxl 模块未加载")
            signal.show_traceback_log("[演员数据库] 初始化失败: openpyxl 模块未加载，请确认打包时已包含 openpyxl")
            return
        db_path = self.u("actor_database.xlsx")
        if not db_path.exists():
            # 文件不存在时不重置 actor_db，保留旧值（如空备份尚未创建）
            if self.actor_db is None:
                LogBuffer.log().write(f"  ❌ [演员数据库] 文件不存在: {db_path}")
                signal.show_traceback_log(f"[演员数据库] 文件不存在，等待首次 TMDB 查询后创建: {db_path}")
            return
        old = self.actor_db
        try:
            self.actor_db = read_actor_db_xlsx(db_path)
            self.actor_db_reverse_index = None
            LogBuffer.log().write(f"  ✅ [演员数据库] 已加载 {len(self.actor_db)} 条记录")
            signal.show_traceback_log(f"[演员数据库] 初始化成功: 已加载 {len(self.actor_db)} 条记录 (路径: {db_path})")
        except ImportError as e:
            self.actor_db = old
            self.actor_db_reverse_index = None
            LogBuffer.log().write(f"  ❌ [演员数据库] 模块缺失: {e}")
            signal.show_traceback_log(
                f"[演员数据库] 初始化失败 (模块缺失): {e}\n"
                f"  请确认打包时已包含: {e.name if hasattr(e, 'name') else 'unknown'}\n"
                f"  常见缺失模块: openpyxl, defusedxml, lxml"
            )
        except Exception as e:
            self.actor_db = old
            tb = traceback.format_exc()
            LogBuffer.log().write(f"  ❌ [演员数据库] 重载失败: {e}")
            signal.show_traceback_log(f"[演员数据库] 重载失败，保留当前缓存: {e}\n{tb}")

    def reload_info_db(self):
        """加载信息映射数据库 xlsx"""
        if openpyxl is None:
            self.info_db = None
            self.info_db_index = None
            return
        db_path = self.u("info_database.xlsx")
        if not db_path.exists():
            self.info_db = None
            self.info_db_index = None
            return
        try:
            wb = openpyxl.load_workbook(db_path, read_only=True, data_only=True)
            ws = wb.active
            db: list[dict] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    continue
                if len(row) < 1:
                    continue
                jp = str(row[0] or "").strip()
                if not jp:
                    continue
                item = {
                    "jp": jp,
                    "zh_cn": str(row[1] or "").strip() if len(row) > 1 else "",
                    "zh_tw": str(row[2] or "").strip() if len(row) > 2 else "",
                    "keyword": str(row[3] or "").strip() if len(row) > 3 else "",
                }
                db.append(item)
            wb.close()
            self.info_db = db
            self.info_db_index = _build_info_db_index(db)
        except Exception:
            self.info_db = None
            self.info_db_index = None

    def _get_mark_icon(self):
        mark_folder = self.u("watermark")
        os.makedirs(mark_folder, exist_ok=True)
        mark_names = ["4k", "8k", "sub", "youma", "umr", "leak", "wuma"]
        for name in mark_names:
            attr_src = f"mark_{name}"
            attr_dst = f"icon_{name}_path"
            dst_path = getattr(self, attr_dst)
            if not os.path.isfile(dst_path):
                copy_file_sync(getattr(self, attr_src), dst_path)


def _alias_sort_key(alias: str) -> tuple[int, int, str]:
    """别名排序键：纯中文(全汉字)优先 -> 中日混合 -> 日文(假名) -> 罗马音/英文。

    按中国人习惯，中文译名最易读优先，日文原名次之，罗马音最后。
    """
    han_count = sum(1 for ch in alias if "\u4e00" <= ch <= "\u9fff")
    kana_count = sum(1 for ch in alias if "\u3040" <= ch <= "\u30ff")
    if han_count > 0 and kana_count == 0:
        return (0, 0, alias)
    if han_count > 0 and kana_count > 0:
        return (1, 0, alias)
    if kana_count > 0:
        return (2, 0, alias)
    return (3, 0, alias)


def _merge_keyword_union(local_kw: str, backup_kw: str) -> str:
    """别名并集合并（去重），按中文优先排序。

    用户库别名与出厂库别名取并集，保留全部线索；按「中文→日文→罗马音」排序。
    """
    merged: list[str] = []
    seen: set[str] = set()
    for group in (local_kw, backup_kw):
        if not group:
            continue
        for item in (k.strip() for k in str(group).split(",") if k.strip()):
            key = item.casefold()
            if key not in seen:
                seen.add(key)
                merged.append(item)
    merged.sort(key=_alias_sort_key)
    return ",".join(merged)


def merge_actor_db_from_backup(backup_path: Path, local_path: Path) -> None:
    """把出厂库的增量同步进已存在的用户库（只增不删、不覆盖用户已有值）。

    出厂库随软件版本更新（清洗修正、新增演员），老用户的用户库不会自动获得这些
    改进。此函数在启动时把出厂库中「用户库没有的新条目」完整追加，并给「用户库
    已有但字段空缺」的条目补全（tmdbid/生日等），别名列做并集合并（中文优先），
    绝不覆盖用户已填的值、绝不删除用户库任何行（用户可能有意保留）。

    用出厂库文件 md5 作为合并标记写入 local_path 同目录的 .actor_db_merge_marker，
    出厂库内容未变时跳过，避免每次启动重复扫描。
    """
    if openpyxl is None:
        return
    if not backup_path.exists() or not local_path.exists():
        return

    import hashlib

    marker_path = local_path.parent / ".actor_db_merge_marker"
    try:
        backup_hash = hashlib.md5(backup_path.read_bytes()).hexdigest()
        if marker_path.exists() and marker_path.read_text(encoding="utf-8").strip() == backup_hash:
            return  # 出厂库未变化，无需合并

        wb = openpyxl.load_workbook(local_path)
        ws = get_actor_db_sheet(wb)
        jp_row_map: dict[str, int] = {}
        next_row = ws.max_row + 1
        for row_no, row in enumerate(ws.iter_rows(min_row=2, max_col=len(DB_HEADERS), values_only=True), start=2):
            if row and row[COL_JP]:
                jp_val = str(row[COL_JP]).strip()
                jp_row_map.setdefault(jp_val, row_no)

        added = 0
        filled = 0
        backup_wb = openpyxl.load_workbook(backup_path, read_only=True, data_only=True)
        backup_ws = get_actor_db_sheet(backup_wb)
        for row in backup_ws.iter_rows(min_row=2, max_col=len(DB_HEADERS), values_only=True):
            if not row or not row[COL_JP]:
                continue
            jp = str(row[COL_JP]).strip()
            if jp in jp_row_map:
                # 字段补全：仅填空缺，不覆盖已有值；别名列做并集合并（中文优先）
                existing_row = jp_row_map[jp]
                for col_idx in range(len(DB_HEADERS)):
                    if col_idx == COL_JP or col_idx == COL_TMDB_URL:
                        continue
                    cur = ws.cell(row=existing_row, column=col_idx + 1).value
                    new = row[col_idx] if col_idx < len(row) else None
                    if col_idx == COL_KEYWORD:
                        cur_str = str(cur or "").strip()
                        new_str = str(new or "").strip()
                        if cur_str or new_str:
                            merged = _merge_keyword_union(cur_str, new_str)
                            if merged != cur_str:
                                ws.cell(row=existing_row, column=COL_KEYWORD + 1, value=merged)
                                filled += 1
                        continue
                    if (cur is None or str(cur).strip() == "") and new not in (None, ""):
                        ws.cell(row=existing_row, column=col_idx + 1, value=new)
                        filled += 1
                continue
            ws.append(list(row[: len(DB_HEADERS)]))
            jp_row_map[jp] = next_row
            next_row += 1
            added += 1
        backup_wb.close()

        if added or filled:
            wb.save(local_path)
        wb.close()
        marker_path.write_text(backup_hash, encoding="utf-8")
        if added or filled:
            LogBuffer.log().write(f"  ℹ️ [演员数据库] 出厂库增量合并: 新增 {added} 条, 补全 {filled} 个字段")
    except Exception as e:
        LogBuffer.log().write(f"  ⚠️ [演员数据库] 出厂库合并失败: {e}")


def merge_info_db_from_backup(backup_path: Path, local_path: Path) -> None:
    """把出厂库的 info_database 同步进已存在的用户库（出厂库权威 + 保留用户新增行）。

    info_database 是「标签/关键词 → 日/中/繁」的映射字典，出厂库是唯一权威：
    - 内容行（cn ≠ "删除"）以 cn 为合并键：
      * 用户库无此 cn → 追加新行
      * 用户库已有此 cn → 覆盖 jp/zh_cn/zh_tw/keyword（出厂库权威，避免 jp 变更导致重复）
    - 删除行（jp=删除，cn 恒为"删除"）无语义主键，整体以出厂库为准覆盖用户库删除行
    - 用户库独有的 cn（出厂库没有）→ 保留（用户自定义标签）
    - 用出厂库 md5 作合并标记写入 .info_db_merge_marker，出厂库未变时跳过。

    出厂库权威而非仿 actor 库「不覆盖」，是因为 info_database 是给刮削翻译用的映射字典，
    正确性完全取决于出厂库维护，用户不应手动改（改了也会被下次合并覆盖）。
    """
    if openpyxl is None:
        return
    if not backup_path.exists() or not local_path.exists():
        return

    import hashlib

    marker_path = local_path.parent / ".info_db_merge_marker"
    try:
        backup_hash = hashlib.md5(backup_path.read_bytes()).hexdigest()
        if marker_path.exists() and marker_path.read_text(encoding="utf-8").strip() == backup_hash:
            return  # 出厂库未变化，无需合并

        wb = openpyxl.load_workbook(local_path)
        ws = wb.active
        # 收集用户库现有 cn -> 行号（内容行）
        cn_row_map: dict[str, int] = {}
        del_rows: list[int] = []  # 用户库删除行行号
        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            jp_val = str(row[0]).strip()
            if jp_val == "删除":
                del_rows.append(row_no)
                continue
            cn_val = str(row[1] or "").strip() if len(row) > 1 else ""
            if cn_val:
                cn_row_map.setdefault(cn_val, row_no)

        added = 0
        updated = 0
        backup_wb = openpyxl.load_workbook(backup_path, read_only=True, data_only=True)
        backup_ws = backup_wb.active
        next_row = ws.max_row + 1
        backup_del_used = 0
        for row in backup_ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            jp = str(row[0]).strip()
            zh_cn = str(row[1] or "").strip() if len(row) > 1 else ""
            zh_tw = str(row[2] or "").strip() if len(row) > 2 else ""
            keyword = str(row[3] or "").strip() if len(row) > 3 else ""
            if jp == "删除":
                # 删除行：覆盖用户库第一个删除行，多余追加；出厂库权威
                if backup_del_used < len(del_rows):
                    row_no = del_rows[backup_del_used]
                    ws.cell(row=row_no, column=1, value=jp)
                    ws.cell(row=row_no, column=2, value=zh_cn)
                    ws.cell(row=row_no, column=3, value=zh_tw)
                    ws.cell(row=row_no, column=4, value=keyword)
                    backup_del_used += 1
                    updated += 1
                else:
                    ws.cell(row=next_row, column=1, value=jp)
                    ws.cell(row=next_row, column=2, value=zh_cn)
                    ws.cell(row=next_row, column=3, value=zh_tw)
                    ws.cell(row=next_row, column=4, value=keyword)
                    next_row += 1
                    added += 1
                continue
            # 内容行：以 cn 为键
            if zh_cn in cn_row_map:
                row_no = cn_row_map[zh_cn]
                ws.cell(row=row_no, column=1, value=jp)
                ws.cell(row=row_no, column=2, value=zh_cn)
                ws.cell(row=row_no, column=3, value=zh_tw)
                ws.cell(row=row_no, column=4, value=keyword)
                updated += 1
            else:
                ws.cell(row=next_row, column=1, value=jp)
                ws.cell(row=next_row, column=2, value=zh_cn)
                ws.cell(row=next_row, column=3, value=zh_tw)
                ws.cell(row=next_row, column=4, value=keyword)
                cn_row_map[zh_cn] = next_row
                next_row += 1
                added += 1
        backup_wb.close()

        # 用户库多余的删除行（出厂库删除行更少时）保留：删除行无语义键，多余的不删以免误删用户新增黑名单词

        if added or updated:
            wb.save(local_path)
        wb.close()
        marker_path.write_text(backup_hash, encoding="utf-8")
        if added or updated:
            LogBuffer.log().write(f"  ℹ️ [信息映射库] 出厂库权威合并: 新增 {added} 条, 覆盖 {updated} 条")
    except Exception as e:
        LogBuffer.log().write(f"  ⚠️ [信息映射库] 出厂库合并失败: {e}")


resources = Resources()
