"""
みんなのAV 演员信息爬虫 + 本地缓存模块。

数据源: https://www.minnano-av.com/
缓存文件: userdata/minnano_cache.xlsx（运行时用户数据目录）

工作流程:
1. 先用演员名搜索缓存（按日文名匹配）
2. 命中缓存 → 直接返回
3. 没命中 → 用名字直接搜索みんなのAV → 抓取详情页面
4. 解析个人资料 → 写入缓存 → 返回

数据字段:
- 日文名、中文名、别名
- 生日、星座
- 三围 (T/B/W/H/S)、罩杯
- 血型
- 出身地、事务所
- Twitter ID
- 生涯期间、出道作品
- 标签（熟女、巨乳等）
- Wikipedia 个人经历段落（补充）
"""
# mypy: ignore-errors

import asyncio
import re
import threading
import traceback
import urllib.parse
from pathlib import Path
from typing import Any

from bs4 import BeautifulSoup

from ..config.manager import manager
from ..config.resources import resources
from ..models.emby import EMbyActressInfo
from ..models.log_buffer import LogBuffer

# ============= 常量定义 =============

# 缓存文件（运行时用户数据目录，与 resources.u() 一致）
CACHE_FILE = "userdata/minnano_cache.xlsx"

# 表头
CACHE_HEADERS = [
    "日文名",
    "别名",
    "生日",
    "身高",
    "胸围",
    "腰围",
    "臀围",
    "罩杯",
    "出身地",
    "所属事务所",
    "Twitter",
    "活动年份",
    "出道作品",
    "维基百科简介",
    "minnano url",
]

COL_JP = 0
COL_ALIAS = 1
COL_BIRTHDAY = 2
COL_HEIGHT = 3
COL_BUST = 4
COL_WAIST = 5
COL_HIP = 6
COL_CUP = 7
COL_PLACE = 8
COL_AGENCY = 9
COL_TWITTER = 10
COL_CAREER = 11
COL_DEBUT = 12
COL_WIKI = 13
COL_MINNANO_URL = 14

# 内存缓存 dict 的字段 key（与 load_cache / _build_cache_row / _fill_emby_info 统一使用字符串 key）
CACHE_FIELD_KEYS = (
    "alias",
    "birthday",
    "height",
    "bust",
    "waist",
    "hip",
    "cup",
    "place",
    "agency",
    "twitter",
    "career",
    "debut",
    "wiki",
    "minnano_url",
)


# ============= 缓存读写 =============

_cache_lock = threading.Lock()
_cache_data: dict[str, dict] = {}  # {日文名: {各字段}}


def _get_cache_path() -> Path:
    return manager.data_folder / CACHE_FILE


def load_cache() -> dict[str, dict]:
    """从 xlsx 加载缓存数据。"""
    global _cache_data
    cache_path = _get_cache_path()
    with _cache_lock:  # 与 save_cache_row 并发写 _cache_data 时避免清空/覆盖错乱
        _cache_data.clear()
        if not cache_path.exists():
            return {}

        try:
            import openpyxl

            wb = openpyxl.load_workbook(cache_path, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < COL_JP + 1:
                    continue
                jp = str(row[COL_JP] or "").strip()
                if not jp:
                    continue
                _cache_data[jp] = {
                    "alias": str(row[COL_ALIAS] or "").strip() if len(row) > COL_ALIAS else "",
                    "birthday": str(row[COL_BIRTHDAY] or "").strip() if len(row) > COL_BIRTHDAY else "",
                    "height": str(row[COL_HEIGHT] or "").strip() if len(row) > COL_HEIGHT else "",
                    "bust": str(row[COL_BUST] or "").strip() if len(row) > COL_BUST else "",
                    "waist": str(row[COL_WAIST] or "").strip() if len(row) > COL_WAIST else "",
                    "hip": str(row[COL_HIP] or "").strip() if len(row) > COL_HIP else "",
                    "cup": str(row[COL_CUP] or "").strip() if len(row) > COL_CUP else "",
                    "place": str(row[COL_PLACE] or "").strip() if len(row) > COL_PLACE else "",
                    "agency": str(row[COL_AGENCY] or "").strip() if len(row) > COL_AGENCY else "",
                    "twitter": str(row[COL_TWITTER] or "").strip() if len(row) > COL_TWITTER else "",
                    "career": str(row[COL_CAREER] or "").strip() if len(row) > COL_CAREER else "",
                    "debut": str(row[COL_DEBUT] or "").strip() if len(row) > COL_DEBUT else "",
                    "wiki": str(row[COL_WIKI] or "").strip() if len(row) > COL_WIKI else "",
                    "minnano_url": str(row[COL_MINNANO_URL] or "").strip() if len(row) > COL_MINNANO_URL else "",
                }
            wb.close()
            LogBuffer.log().write(f"  ℹ️ [演员缓存] 已加载 {len(_cache_data)} 条缓存记录")
        except ImportError:
            LogBuffer.log().write("  ⚠️ [演员缓存] 缺少 openpyxl，无法读取缓存")
        except Exception as e:
            LogBuffer.log().write(f"  ⚠️ [演员缓存] 读取失败: {e}")

    return _cache_data


def save_cache_row(row: dict) -> bool:
    """追加一行到缓存 xlsx（带超链接和格式化）。"""
    cache_path = _get_cache_path()
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        data_fill = PatternFill("solid", fgColor="F2F2F2")
        link_font = Font(color="0563C1", u="single")
        data_align = Alignment(vertical="center")
        full_border = Border(
            bottom=Side(style="thin"),
            top=Side(style="thin"),
            left=Side(style="thin"),
            right=Side(style="thin"),
        )
        # 列宽（与手动格式化一致）
        col_widths = {}

        with _cache_lock:
            if cache_path.exists():
                wb = openpyxl.load_workbook(cache_path)
                ws = wb.active
                new_row = ws.max_row + 1
                # 按 jp 去重：已存在则更新该行，避免重复追加
                jp_val = str(row.get("jp", "") or "").strip()
                for r in range(2, ws.max_row + 1):
                    if str(ws.cell(row=r, column=COL_JP + 1).value or "").strip() == jp_val:
                        new_row = r
                        break
                # 准备值
                values = [row.get("jp", "")] + [row.get(key, "") for key in CACHE_FIELD_KEYS]
                # 写入并格式化
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=new_row, column=col_idx, value=val)
                    cell.fill = data_fill
                    cell.alignment = data_align
                    cell.border = full_border
                    # 第15列(minnano url)超链接：col_idx 为 1-based，COL_MINNANO_URL 是 0-based
                    if col_idx == COL_MINNANO_URL + 1 and str(val or "").startswith("http"):
                        cell.hyperlink = val
                        cell.font = link_font
                wb.save(cache_path)
                wb.close()
                jp = row.get("jp", "")
                _cache_data[jp] = row
                LogBuffer.log().write(f"  ✅ [演员缓存] 已追加: {jp}")
                return True
            # 新建文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "演员缓存"

            header_fill = PatternFill("solid", fgColor="D9D9D9")
            header_font = Font(bold=True)
            header_align = Alignment(horizontal="center", vertical="center")

            for col, header in enumerate(CACHE_HEADERS, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align

            # 冻结首行
            ws.freeze_panes = "A2"

            values = [row.get("jp", "")] + [row.get(key, "") for key in CACHE_FIELD_KEYS]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=2, column=col_idx, value=val)
                cell.fill = data_fill
                cell.alignment = data_align
                cell.border = full_border
                if col_idx == COL_MINNANO_URL + 1 and str(val or "").startswith("http"):
                    cell.hyperlink = val
                    cell.font = link_font

            # 设置列宽
            for col, width in col_widths.items():
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

            wb.save(cache_path)
            wb.close()
            jp = row.get("jp", "")
            _cache_data[jp] = row
            LogBuffer.log().write(f"  ✅ [演员缓存] 已创建并追加: {jp}")
            return True
    except Exception as e:
        LogBuffer.log().write(f"  ⚠️ [演员缓存] 写入失败: {e}")
        return False


def get_cached_actor(name: str) -> dict | None:
    """按日文名查找缓存。"""
    with _cache_lock:  # 与 save_cache_row/load_cache 并发时保证 _cache_data 一致
        if name in _cache_data:
            return _cache_data[name]
        # 模糊匹配：别名中包含
        for _jp, data in _cache_data.items():
            alias = data.get("alias", "")
            if name in alias:
                return data
    return None


# ============= 解析函数 =============


def _parse_size(size_str: str) -> dict[str, str]:
    """解析尺寸字符串: T166 / B108(Lカップ) / W60 / H90"""
    result = {
        "height": "",
        "bust": "",
        "waist": "",
        "hip": "",
        "cup": "",
    }
    if not size_str:
        return result

    parts = re.split(r"[ /／]", size_str)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # 身高 T166
        m = re.match(r"T(\d+)", part)
        if m:
            result["height"] = m.group(1)
            continue
        # 胸围 B108(Lカップ)
        m = re.match(r"B(\d+)\((.+?)カップ?\)", part)
        if m:
            result["bust"] = m.group(1)
            result["cup"] = m.group(2)
            continue
        # 腰围 W60
        m = re.match(r"W(\d+)", part)
        if m:
            result["waist"] = m.group(1)
            continue
        # 臀围 H90
        m = re.match(r"H(\d+)", part)
        if m:
            result["hip"] = m.group(1)
            continue

    return result


def _parse_birthday(raw: str) -> tuple[str, str]:
    """解析生日: 1996年11月30日（現在29歳）いて座"""
    birthday = ""
    zodiac = ""
    m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", raw)
    if m:
        year, month, day = m.group(1), m.group(2), m.group(3)
        birthday = f"{year}-{int(month):02d}-{int(day):02d}"
    # 星座：用分组而非字符类（原 [牡羊|金牛|...] 的 | 是字面量，星座名被拆成错误单字）
    zodiac_match = re.search(r"(牡羊|金牛|雙子|巨蟹|獅子|處女|天秤|天蠍|射手|摩羯|水瓶|雙魚)座", raw)
    if zodiac_match:
        zodiac = zodiac_match.group(1)
    elif "いて座" in raw:
        zodiac = "射手座"
    elif "おひつじ座" in raw:
        zodiac = "白羊座"
    elif "おうし座" in raw:
        zodiac = "金牛座"
    elif "ふたご座" in raw:
        zodiac = "双子座"
    elif "かに座" in raw:
        zodiac = "巨蟹座"
    elif "しし座" in raw:
        zodiac = "狮子座"
    elif "おとめ座" in raw:
        zodiac = "处女座"
    elif "てんびん座" in raw:
        zodiac = "天秤座"
    elif "さそり座" in raw:
        zodiac = "天蝎座"
    elif "やぎ座" in raw:
        zodiac = "摩羯座"
    elif "みずがめ座" in raw:
        zodiac = "水瓶座"
    elif "うお座" in raw:
        zodiac = "双鱼座"

    return birthday, zodiac


def _parse_profile_table(table) -> dict[str, str | list[str] | bool]:
    """解析个人信息表格，返回字段字典。"""
    result = {
        "name": "",
        "aliases": [],
        "birthday_raw": "",
        "size_raw": "",
        "blood": "",
        "place": "",
        "agency": "",
        "hobby": "",
        "career": "",
        "debut": "",
        "blog": "",
        "tags": [],
    }

    # 第一行 h2 是名字
    h2 = table.find("h2")
    if h2:
        text = h2.get_text(strip=True)
        name_match = re.match(r"^(.+?)\s+（(.+?)\s*/\s*(.+?)）$", text)
        if name_match:
            result["name"] = name_match.group(1)

    rows = table.find_all("tr")
    for row in rows:
        spans = row.find_all("span")
        ps = row.find_all("p")
        if not spans or not ps:
            continue
        label = spans[0].get_text(strip=True)
        value = ps[0].get_text(strip=True)
        # 去掉 <a> 标签
        for a in ps[0].find_all("a"):
            a.replace_with(a.get_text())
        value = ps[0].get_text(strip=True)

        if label == "別名":
            # 去掉括号里的内容
            alias = re.sub(r"\s*（.+）$", "", value).strip()
            if alias:
                result["aliases"].append(alias)
        elif label == "生年月日":
            result["birthday_raw"] = value
        elif label == "サイズ":
            result["size_raw"] = value
        elif label == "血液型":
            m = re.search(r"([A-Z]+)", value)
            if m:
                result["blood"] = m.group(1) + "型"
        elif label == "出身地":
            result["place"] = value
        elif label == "所属事務所":
            result["agency"] = value
        elif label == "趣味・特技":
            hobby = re.sub(r"趣味\s*", "", value).strip()
            if hobby:
                result["hobby"] = hobby
        elif label == "AV出演期間":
            result["career"] = value
        elif label == "デビュー作品":
            result["debut"] = value
        elif label == "ブログ":
            result["blog"] = value
        elif label == "タグ":
            # 标签单独处理
            result["_tags_found"] = True

    return result


def _parse_tags_from_table(table_or_tagarea) -> list[str]:
    """从标签区域提取标签。"""
    tags = []
    # 如果传入的是 table，找 tagarea
    if table_or_tagarea.name == "table":
        tagarea = table_or_tagarea.find("div", class_="tagarea")
    else:
        tagarea = table_or_tagarea
    if tagarea:
        for a in tagarea.find_all("a"):
            tags.append(a.get_text(strip=True))
    return tags


def parse_minnano_page(html: str, minnano_id: str) -> dict[str, Any] | None:
    """
    解析みんなのAV 演员页面，返回结构化数据。

    Args:
        html: 页面 HTML
        minnano_id: 演员在みんなのAV 的 ID

    Returns:
        字典或 None
    """
    soup = BeautifulSoup(html, "lxml")
    result = {
        "minnano_id": minnano_id,
        "name": "",
        "aliases": [],
        "birthday": "",
        "zodiac": "",
        "height": "",
        "bust": "",
        "waist": "",
        "hip": "",
        "cup": "",
        "blood": "",
        "place": "",
        "agency": "",
        "hobby": "",
        "career": "",
        "debut": "",
        "twitter": "",
        "tags": [],
        "wiki": "",
    }

    # 解析个人信息表
    tables = soup.find_all("table")
    profile_table = None
    for table in tables:
        rows = table.find_all("tr")
        if len(rows) < 2:
            continue
        first_cell = rows[0].find_all(["td", "th"])
        # 部分演员页个人信息表只有 4 行（无生日行），不再强制 rows >= 5
        if first_cell and "（" in first_cell[0].get_text():
            profile_table = table
            break

    if profile_table:
        profile = _parse_profile_table(profile_table)
        result["name"] = profile["name"]
        result["aliases"] = profile["aliases"]
        result["birthday_raw"] = profile.get("birthday_raw", "")
        result["size_raw"] = profile.get("size_raw", "")
        result["blood"] = profile.get("blood", "")
        result["place"] = profile.get("place", "")
        result["agency"] = profile.get("agency", "")
        result["hobby"] = profile.get("hobby", "")
        result["career"] = profile.get("career", "")
        result["debut"] = re.sub(r"（\d{4}年\d{1,2}月\s*\d{1,2}日）$", "", profile.get("debut", "")).strip()
        blog = profile.get("blog", "")
        if blog:
            twitter_match = re.search(r"twitter\.com/([^/\s]+)", blog)
            if twitter_match:
                result["twitter"] = twitter_match.group(1)

        # 解析生日
        if isinstance(result["birthday_raw"], str) and result["birthday_raw"]:
            birthday, zodiac = _parse_birthday(result["birthday_raw"])
            result["birthday"] = birthday
            result["zodiac"] = zodiac

        # 解析尺寸
        if isinstance(result["size_raw"], str) and result["size_raw"]:
            size = _parse_size(result["size_raw"])
            result["height"] = size["height"]
            result["bust"] = size["bust"]
            result["waist"] = size["waist"]
            result["hip"] = size["hip"]
            result["cup"] = size["cup"]

    # 解析标签（直接查找 tagarea）
    tagarea = soup.find("div", class_="tagarea")
    if tagarea:
        result["tags"] = _parse_tags_from_table(tagarea)

    # 如果名字还没解析出来，从页面标题或 h2 中找
    if not result["name"]:
        title = soup.find("title")
        if title:
            title_text = title.get_text(strip=True)
            if "みんなのAV" not in title_text:
                result["name"] = title_text

    return result


def _clean_alias(alias: str) -> str | None:
    """清洗别名，过滤无效内容。

    过滤规则：
    1. 含感叹词/情绪性内容（※、！！、！？、ヤツ等）→ 过滤
    2. 含括号职业标签（DAS!、着エロ等）→ 去掉括号部分
    3. 纯错别字/异常（长度<2、与主名重复）→ 过滤
    4. 【旧名】标注保留
    """
    if not alias:
        return None

    alias = alias.strip()
    if len(alias) < 2:
        return None

    # 超长项基本可判定为作品标题/绰号混入別名列（实测 38 字描述句），演员名业界常见上限远低于 20
    if len(alias) > 20:
        return None

    # 过滤：含情绪性/无意义内容
    noise_patterns = [r"※", r"！.*！.*！", r"ヤツ", r"戻しとけ", r"消した", r"誰だ"]
    for pat in noise_patterns:
        if re.search(pat, alias):
            return None

    # 过滤：与主名重复
    # （由调用方判断）

    # 保留【旧名】标注
    if "【旧名】" in alias or "（旧名）" in alias:
        return alias

    # 去掉括号后缀如 (DAS!)、(着エロ)，但保留（旧名）类
    cleaned = re.sub(r"\s*（[^）]*旧名[^）]*）", r"【旧名】", alias)  # 标准化旧名标注
    cleaned = re.sub(r"\s*\(.*?\)", "", cleaned).strip()  # 去掉半角括号内容
    cleaned = re.sub(r"\s*（[^）]*）", "", cleaned).strip()  # 去掉非旧名的全角括号内容

    if not cleaned or len(cleaned) < 2:
        return None

    return cleaned


def _build_cache_row(parsed: dict) -> dict:
    """将解析结果转换为缓存行。"""
    # 合并别名
    aliases = []
    seen = set()
    for a in parsed.get("aliases", []):
        clean = _clean_alias(a)
        if clean and clean not in seen:
            aliases.append(clean)
            seen.add(clean)
    alias_str = ",".join(aliases)

    minnano_url = (
        f"https://www.minnano-av.com/actress{parsed.get('minnano_id', '')}.html" if parsed.get("minnano_id") else ""
    )

    return {
        "jp": parsed.get("name", ""),
        "alias": alias_str,
        "birthday": parsed.get("birthday", ""),
        "height": parsed.get("height", ""),
        "bust": parsed.get("bust", ""),
        "waist": parsed.get("waist", ""),
        "hip": parsed.get("hip", ""),
        "cup": parsed.get("cup", ""),
        "place": parsed.get("place", ""),
        "agency": parsed.get("agency", ""),
        "twitter": parsed.get("twitter", ""),
        "career": parsed.get("career", ""),
        "debut": parsed.get("debut", ""),
        "wiki": parsed.get("wiki", ""),
        "minnano_url": minnano_url,
    }


def _fill_emby_info(actor_info: EMbyActressInfo, cached: dict, wiki_intro: str = "") -> None:
    """将缓存数据填充到 EMbyActressInfo。"""
    # 生日
    if cached.get("birthday"):
        actor_info.birthday = cached["birthday"]
        actor_info.year = cached["birthday"][:4] if len(cached["birthday"]) >= 4 else ""

    # 标签
    tags = []
    if cached.get("cup"):
        tags.append(f"罩杯: {cached['cup']}")
    if cached.get("height"):
        tags.append(f"身高: {cached['height']}cm")
    if cached.get("bust") or cached.get("waist") or cached.get("hip"):
        parts = []
        if cached.get("bust"):
            parts.append(cached["bust"])
        if cached.get("waist"):
            parts.append(cached["waist"])
        if cached.get("hip"):
            parts.append(cached["hip"])
        tags.append(f"三围: {'/'.join(parts)}")
    if cached.get("career"):
        career_clean = cached["career"].replace("年", "").replace(" ", "").replace("-", "~")
        tags.append(f"生涯: {career_clean}")

    if tags:
        actor_info.tags.extend(tags)

    # ProviderIds
    if cached.get("twitter"):
        actor_info.provider_ids["Twitter"] = cached["twitter"]
    if cached.get("minnano_url"):
        actor_info.provider_ids["minnano-av"] = cached["minnano_url"]
    if cached.get("place"):
        location = "日本·" + cached["place"].replace("県", "县")
        if actor_info.locations == ["日本"] or not actor_info.locations:
            actor_info.locations = [location]

    # 简介（wiki > cache wiki > minnano debut）
    if wiki_intro:
        actor_info.overview = wiki_intro
    elif cached.get("wiki"):
        actor_info.overview = cached["wiki"]
    elif cached.get("debut"):
        actor_info.overview = f"出道作品：{cached['debut']}"

    if not actor_info.taglines:
        actor_info.taglines = ["日本AV女优"]


# ============= 搜索映射 =============


def _name_matches(actor_name: str, minnano_name: str) -> bool:
    """判断 actor_name 和 minnano_name 是否匹配同一人。

    匹配规则（任一满足即匹配）：
    1. actor_name 中的所有字符都出现在 minnano_name 中
    2. minnano_name 中的所有字符都出现在 actor_name 中
    3. 两名字有至少2个重叠字符且至少1个是汉字，且重叠字符数 <= 2
    """
    if not actor_name or not minnano_name:
        return False

    # 规则1&2: 完全包含
    if all(c in minnano_name for c in actor_name):
        return True
    if all(c in actor_name for c in minnano_name):
        return True

    # 规则3: 至少2个共同字符且至少1个是汉字，重叠数<=2，且不超过短名长度的60%
    common = set(actor_name) & set(minnano_name)
    if len(common) >= 2 and len(common) <= 2:
        common_kanji = [c for c in common if re.match(r"[\u4e00-\u9fff]", c)]
        if common_kanji:
            min_len = min(len(set(actor_name)), len(set(minnano_name)))
            if len(common) <= min_len * 0.6:
                return True

    return False


def _lookup_japanese_name(actor_name: str) -> str | None:
    """在演员数据库中查找演员的日文名（Emby 中的可能是中文名/别名）。

    复用 resources.get_actor_data 的内存缓存与反向索引（读运行时用户库
    actor_database.xlsx，支持中文/日文/别名/归一化变体匹配），比逐行扫描
    resources 出厂库更准确且高效。未找到时返回 None（调用方会用原名搜索）。
    """
    try:
        actor_data = resources.get_actor_data(actor_name)
        jp = (actor_data.get("jp") or "").strip()
        if actor_data.get("has_name") and jp:
            return jp
    except Exception:
        from ..signals import signal

        signal.show_log_text(f"⚠️ 演员数据库读取失败: {traceback.format_exc()}")
    return None


async def _search_minnano_by_name(actor_name: str) -> tuple[str | None, str | None]:
    """
    在みんなのAV 上搜索演员，返回 (minnano_id, html) 或 (None, None)。

    搜索策略:
    1. 用 search_result.php API 精确匹配名字
    2. 模糊匹配名字（支持部分重叠，解决异体字/写法差异）
    """
    search_url = f"https://www.minnano-av.com/search_result.php?search_scope=actress&search_word={urllib.parse.quote(actor_name)}&search=+Go+"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        "Referer": "https://www.minnano-av.com/",
    }

    async with manager.acquire_computed() as computed:
        html, error = await computed.async_client.get_text(search_url, headers=headers)

    if not html:
        return None, None

    soup = BeautifulSoup(html, "lxml")

    # 1. 精确匹配名字（在链接文本中）
    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True)
        if (
            text == actor_name
            and "actress" in href
            and "ranking" not in href
            and "works" not in href
            and "list" not in href
        ):
            minnano_id = re.search(r"actress(\d+)", str(href))
            if minnano_id:
                mid = minnano_id.group(1)
                detail_url = f"https://www.minnano-av.com/actress{mid}.html"
                detail_html, detail_error = await computed.async_client.get_text(detail_url)
                if detail_html:
                    # Title 验证：详情页标题应包含演员名
                    detail_soup = BeautifulSoup(detail_html, "lxml")
                    detail_title = (
                        detail_soup.title.string.strip() if detail_soup.title and detail_soup.title.string else ""
                    )
                    if actor_name and actor_name not in detail_title:
                        LogBuffer.log().write(f"  ⚠️ [minnano] 详情页标题不含演员名: '{detail_title[:60]}'")
                        continue  # 标题校验失败, 拒绝该候选, 继续尝试下一个
                    return mid, detail_html

    # 2. 模糊匹配名字（解决异体字/写法差异，如 绫瀬舞菜 vs あやせ舞菜）
    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True)
        if (
            "actress" in href
            and "ranking" not in href
            and "works" not in href
            and "list" not in href
            and text != actor_name
            and text
        ):
            if _name_matches(actor_name, text):
                minnano_id = re.search(r"actress(\d+)", str(href))
                if minnano_id:
                    mid = minnano_id.group(1)
                    detail_url = f"https://www.minnano-av.com/actress{mid}.html"
                    detail_html, detail_error = await computed.async_client.get_text(detail_url)
                    if detail_html:
                        return mid, detail_html

    return None, None


async def fetch_minnano_aliases(actor_name: str) -> list[str]:
    """从みんなのAV 查询演员别名，未找到或匹配失败返回空列表。

    流程: _search_minnano_by_name 精确/模糊搜索 → parse_minnano_page 提取「別名」行
    → _clean_alias 过滤噪声（※/情绪词/括号注释/长度<2），大小写不敏感去重。

    搜不到、解析失败、无别名均静默返回 []，由调用方决定如何降级。
    """
    if not actor_name or not actor_name.strip():
        return []

    try:
        minnano_id, html = await _search_minnano_by_name(actor_name.strip())
        if not html:
            return []
        parsed = parse_minnano_page(html, minnano_id)
        if not parsed:
            return []

        aliases: list[str] = []
        seen_lower: set[str] = set()
        for alias in parsed.get("aliases", []):
            cleaned = _clean_alias(alias)
            if not cleaned:
                continue
            key = cleaned.lower()
            if key in seen_lower:
                continue
            seen_lower.add(key)
            aliases.append(cleaned)
        return aliases
    except Exception:
        LogBuffer.log().write(f"[minnano] 别名查询失败 {actor_name}: {traceback.format_exc()}")
        return []


# ============= 主入口 =============


async def get_minnano_info(actor_info: EMbyActressInfo, wiki_intro: str = "") -> tuple[bool, str]:
    """
    获取みんなのAV 演员信息并填充到 actor_info。

    Args:
        actor_info: EMbyActressInfo 实例
        wiki_intro: 可选的 Wikipedia 简介，会覆盖缓存中的简介

    Returns:
        (success, message)
    """
    actor_name = actor_info.name

    # 1. 查 actor_database.xlsx 获取日文名（Emby 中的可能是中文名）
    search_name = _lookup_japanese_name(actor_name)
    if search_name:
        LogBuffer.log().write(f"  ℹ️ [数据库映射] {actor_name} -> {search_name}")
    else:
        search_name = actor_name
        LogBuffer.log().write(f"  ℹ️ [数据库映射] {actor_name} 未在数据库中映射，使用原名搜索")

    # 2. 查缓存（按日文名匹配）
    cached = get_cached_actor(search_name)
    if cached:
        LogBuffer.log().write(f"  ℹ️ [演员缓存] 命中: {actor_name} ({search_name})")
        _fill_emby_info(actor_info, cached, wiki_intro)
        return True, f"✅ {actor_name} 使用缓存数据"

    # 3. 搜索みんなのAV
    minnano_id, html = await _search_minnano_by_name(search_name)
    if not html:
        return False, f"🔴 {actor_name}: みんなのAV 未找到"

    # 4. 解析
    assert html is not None
    parsed = parse_minnano_page(html, minnano_id)
    if not parsed or not parsed.get("name"):
        return False, f"🔴 {actor_name}: みんなのAV 页面解析失败"

    # 5. 写入缓存（openpyxl load/save 为同步阻塞 IO，移后台线程避免阻塞事件循环）
    cache_row = _build_cache_row(parsed)
    await asyncio.to_thread(save_cache_row, cache_row)

    # 6. 填充 actor_info
    _fill_emby_info(actor_info, cache_row, wiki_intro)
    return True, f"✅ {actor_name} 从みんなのAV 获取信息"
