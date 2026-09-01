"""
演员库维护工具。

从刮削流程剥离出的翻译补全与 LibreDMM 链接补全能力，供「工具」页独立批量触发。
"""

import asyncio
import json
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import aiofiles.os
import aiohttp
import zhconv
from lxml import etree

from ..config.manager import manager
from ..config.resources import (
    COL_BIO,
    COL_BIRTH_DATE,
    COL_JP,
    COL_KEYWORD,
    COL_TMDB_URL,
    COL_TMDBID,
    COL_ZH_CN,
    COL_ZH_TW,
    DB_HEADERS,
    get_actor_db_sheet,
    resources,
)
from ..core.tmdb_actor import (
    _actor_db_write_lock,
    _expand_name_variants,
    _fetch_person_translations,
    _format_db_worksheet,
    _get_db_path,
    _merge_keyword_values,
    _norm_name_set,
    _normalize_translation,
    _resolve_tmdb_config,
    _tmdb_person_url,
    fetch_libredmm_link,
    fetch_person_gender,
    fetch_person_identity,
    search_actor_db_reverse,
    update_actor_db_row,
)
from ..models.log_buffer import LogBuffer
from ..utils import get_used_time
from ..utils.file import write_file_atomic, write_file_atomic_async

_JP_SHINJITAI_TO_SIMPLIFIED = str.maketrans(
    {
        "亜": "亚",
        "亞": "亚",
        "囲": "围",
        "圍": "围",
        "壱": "壹",
        "円": "圆",
        "圓": "圆",
        "厳": "严",
        "嚴": "严",
        "拡": "扩",
        "擴": "扩",
        "楽": "乐",
        "樂": "乐",
        "顕": "显",
        "顯": "显",
        "気": "气",
        "氣": "气",
        "済": "济",
        "濟": "济",
        "桜": "樱",
        "櫻": "樱",
        "雑": "杂",
        "雜": "杂",
        "粛": "肃",
        "肅": "肃",
        "渋": "涩",
        "澀": "涩",
        "終": "终",
        "処": "处",
        "處": "处",
        "緒": "绪",
        "塩": "盐",
        "鹽": "盐",
        "弾": "弹",
        "彈": "弹",
        "滝": "泷",
        "瀧": "泷",
        "単": "单",
        "單": "单",
        "諸": "诸",
        "禅": "禅",
        "禪": "禅",
        "姫": "姬",
        "標": "标",
        "氷": "冰",
        "浜": "滨",
        "濱": "滨",
        "濵": "滨",
        "冨": "富",
        "仏": "佛",
        "橋": "桥",
        "樓": "楼",
        "錦": "锦",
        "髙": "高",
        "﨑": "崎",
        "免": "免",
        "瀬": "濑",
        "瀨": "濑",
        "沢": "泽",
        "澤": "泽",
        "恵": "惠",
        "穂": "穗",
        "絵": "绘",
        "繪": "绘",
        "嶋": "岛",
        "島": "岛",
        "嶌": "岛",
        "鳥": "鸟",
        "黒": "黑",
        "歩": "步",
        "辺": "边",
        "邊": "边",
        "変": "变",
        "變": "变",
        "応": "应",
        "應": "应",
        "栄": "荣",
        "榮": "荣",
        "営": "营",
        "營": "营",
        "県": "县",
        "縣": "县",
        "録": "录",
        "錄": "录",
        "難": "难",
        "渕": "渊",
        "淵": "渊",
        "広": "广",
        "廣": "广",
        "帰": "归",
        "歸": "归",
        "圧": "压",
        "壓": "压",
        "華": "华",
        "紗": "纱",
        "織": "织",
        "聖": "圣",
        "荘": "庄",
        "莊": "庄",
        "対": "对",
        "對": "对",
        "証": "证",
        "證": "证",
        "強": "强",
        "両": "两",
        "兩": "两",
        "専": "专",
        "專": "专",
        "様": "样",
        "樣": "样",
        "駅": "驿",
        "驛": "驿",
        "騒": "骚",
        "騷": "骚",
        "髪": "发",
        "髮": "发",
        "鶴": "鹤",
        "鵬": "鹏",
        "麗": "丽",
        "齢": "龄",
        "齡": "龄",
        "縁": "缘",
        "緣": "缘",
        "蝋": "蜡",
        "蠟": "蜡",
        "欧": "欧",
        "歐": "欧",
        "鴎": "鸥",
        "鷗": "鸥",
        "揺": "摇",
        "搖": "摇",
        "誉": "誉",
        "譽": "誉",
        "謡": "谣",
        "謠": "谣",
        "陥": "陷",
    }
)


def _to_simplified(text: str) -> str:
    """繁体→简体转换，额外覆盖 zhconv 缺失的日文新字体/异体字。"""
    return zhconv.convert(text.translate(_JP_SHINJITAI_TO_SIMPLIFIED), "zh-cn")


@dataclass
class ActorDbToolResult:
    total: int = 0
    translated: int = 0
    linked: int = 0
    skipped: int = 0
    failed: list[tuple[str, str]] = field(default_factory=list)


AVDB_MAPPING_URL = "https://raw.githubusercontent.com/li-peifeng/Jav-Actors-Mapping/main/actor-mapping.xml"
AVDB_MAPPING_URL_MIRROR = "https://cdn.jsdelivr.net/gh/li-peifeng/Jav-Actors-Mapping@main/actor-mapping.xml"


@dataclass
class ActorDbSyncResult:
    downloaded: bool = False
    parsed: int = 0
    created: int = 0
    filled: int = 0
    merged: int = 0
    skipped_male: int = 0
    skipped_tmdbid: int = 0
    failed: list[tuple[str, str]] = field(default_factory=list)


@dataclass
class CleanActorResult:
    checked: int = 0
    removed_male: int = 0
    kept: int = 0
    failed: list[tuple[str, str]] = field(default_factory=list)


@dataclass
class VerifyTmdbIdResult:
    """tmdbid 有效性校验结果。"""

    checked: int = 0
    invalid: int = 0  # 404 失效被清除的 id 数
    recovered: int = 0  # 清除后按名字重搜补回的新 id 数
    valid: int = 0
    kept: int = 0  # 请求失败/限流保守保留
    failed: list[tuple[str, str]] = field(default_factory=list)


@dataclass
class UpdateNfoTmdbIdResult:
    """nfo tmdbid 更新结果。"""

    checked: int = 0
    updated_files: int = 0
    updated_actors: int = 0
    no_change: int = 0
    failed: list[tuple[str, str]] = field(default_factory=list)


def _log_line(message: str) -> None:
    LogBuffer.log().write(message)
    from mdcx.signals import signal_qt

    signal_qt.show_log_text(message)


_male_actor_set: set[str] | None = None
_male_actor_set_path: Path | None = None


def _load_male_actor_set() -> set[str]:
    """加载内置男优名单（resources/userdata/male_actors.txt），懒加载并缓存。

    返回名单的 casefold 归一化集合；文件缺失/空时返回空集。
    """
    global _male_actor_set, _male_actor_set_path
    try:
        path = resources.r("userdata/male_actors.txt")
    except AttributeError:
        path = None
    if _male_actor_set is not None and _male_actor_set_path == path:
        return _male_actor_set
    names: set[str] = set()
    if path is not None:
        try:
            if path.exists():
                names = {
                    line.strip().casefold() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()
                }
        except OSError:
            pass
    _male_actor_set = names
    _male_actor_set_path = path
    return names


def is_male_actor(name: str) -> bool:
    """按内置男优名单判断演员是否为男优（不依赖 TMDB，可命中无 tmdbid / gender=0 的男优）。"""
    if not name or not name.strip():
        return False
    return name.strip().casefold() in _load_male_actor_set()


async def collect_actors_from_nfo_dir(dir_path: Path) -> list[str]:
    """递归扫描 nfo 目录，解析 //actor/name 收集演员名并去重。"""
    if not await aiofiles.os.path.isdir(dir_path):
        return []

    actors: list[str] = []
    seen: set[str] = set()

    async def _walk(current: Path) -> None:
        try:
            entries = await aiofiles.os.scandir(current)
        except OSError:
            return
        for entry in entries:
            try:
                if entry.is_dir(follow_symlinks=True):
                    await _walk(Path(entry.path))
                elif entry.name.lower().endswith(".nfo"):
                    await _collect_from_nfo(Path(entry.path), actors, seen)
            except OSError:
                continue

    await _walk(dir_path)
    return actors


async def _collect_from_nfo(nfo_path: Path, actors: list[str], seen: set[str]) -> None:
    try:
        async with aiofiles.open(nfo_path, encoding="utf-8") as f:
            content = await f.read()
    except (OSError, UnicodeDecodeError):
        return
    parser = etree.XMLParser(encoding="utf-8", recover=True)
    try:
        xml_nfo = etree.fromstring(content.encode("utf-8"), parser)
    except etree.XMLSyntaxError:
        return
    if xml_nfo is None:
        return
    for ae in xml_nfo.xpath("//actor"):
        name = "".join(ae.xpath("name/text()")).strip()
        if name and name not in seen:
            seen.add(name)
            actors.append(name)


async def run(
    actor_names: list[str],
    translate: bool = True,
    link: bool = True,
) -> ActorDbToolResult:
    """批量维护演员库：补全翻译（中文/繁体）与 LibreDMM 链接。"""
    result = ActorDbToolResult(total=len(actor_names))
    names = [a.strip() for a in actor_names if a and a.strip()]
    names = list(dict.fromkeys(names))  # 去重保序
    if not names:
        return result

    base_url, tmdb_api_key = _resolve_tmdb_config()
    if not tmdb_api_key:
        _log_line(" ⚠️ [演员库维护] 未配置 TMDB API Key，仅能执行链接补全（如需翻译补全请先配置 TMDB API）")

    db_path = _get_db_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    _wb = None
    try:
        import openpyxl as _openpyxl

        if db_path.exists():
            _wb = _openpyxl.load_workbook(db_path)
        else:
            _wb = _openpyxl.Workbook()
            _ws = _wb.active
            _ws.title = "演员数据库"
            for _col, _header in enumerate(DB_HEADERS, 1):
                _cell = _ws.cell(row=1, column=_col, value=_header)
                _cell.font = _openpyxl.styles.Font(bold=True)
                _cell.fill = _openpyxl.styles.PatternFill("solid", fgColor="C0C0C0")
                _cell.alignment = _openpyxl.styles.Alignment(horizontal="center")
    except ImportError:
        _wb = None

    start_time = time.time()
    _log_line(f" 🎬 [演员库维护] 开始处理 {len(names)} 个演员 (翻译={translate}, 链接={link})")

    # 确保行索引与 _wb 快照同源：若索引是更早磁盘状态遗留，与快照错位会并发写错行
    if _wb is not None:
        try:
            from mdcx.core.tmdb_actor import (
                _ACTOR_DB_ROW_INDEX,
                _ACTOR_DB_ROW_INDEX_LOCK,
                _rebuild_actor_db_row_index,
            )

            with _ACTOR_DB_ROW_INDEX_LOCK:
                _ACTOR_DB_ROW_INDEX.clear()
                _rebuild_actor_db_row_index(get_actor_db_sheet(_wb))
        except Exception as e:
            _log_line(f" ⚠️ [演员库维护] 行索引重建失败: {e}")

    async with aiohttp.ClientSession() as client:
        semaphore = asyncio.Semaphore(3)

        async def _process_one(actor_name: str) -> None:
            async with semaphore:
                try:
                    row = search_actor_db_reverse(actor_name)
                    tmdbid = (row or {}).get("tmdbid")
                    jp_name = (row or {}).get("jp") or actor_name

                    need_translate = (
                        translate
                        and tmdbid
                        and base_url
                        and tmdb_api_key
                        and (not (row or {}).get("zh_cn") or not (row or {}).get("zh_tw"))
                    )
                    need_link = link and tmdbid and not (row or {}).get("href")

                    if not need_translate and not need_link:
                        reason = "无 tmdbid" if not tmdbid else "数据已完整"
                        result.skipped += 1
                        _log_line(f"  ℹ️ [演员库维护] {actor_name} 跳过 ({reason})")
                        return

                    # 并行拉取翻译和链接
                    async def _fetch_translations():
                        if need_translate:
                            translations = await _fetch_person_translations(tmdbid, base_url, tmdb_api_key, client)
                            zh_cn = _normalize_translation(translations.get("zh_cn", ""))
                            zh_tw = _normalize_translation(translations.get("zh_tw", ""))
                            if not zh_cn and zh_tw:
                                zh_cn = _to_simplified(zh_tw)
                            if not zh_tw and zh_cn:
                                zh_tw = zhconv.convert(zh_cn, "zh-hant")
                            return zh_cn, zh_tw
                        return "", ""

                    async def _fetch_href():
                        if need_link:
                            return await fetch_libredmm_link(jp_name)
                        return ""

                    zh_cn, zh_tw = "", ""
                    href = ""
                    if need_translate and need_link:
                        (zh_cn, zh_tw), href = await asyncio.gather(_fetch_translations(), _fetch_href())
                    elif need_translate:
                        zh_cn, zh_tw = await _fetch_translations()
                    elif need_link:
                        href = await _fetch_href()

                    await update_actor_db_row(
                        jp=jp_name,
                        zh_cn=zh_cn,
                        zh_tw=zh_tw,
                        href=href,
                        tmdbid=tmdbid,
                        overwrite_names=True,
                        _wb=_wb,
                    )

                    if zh_cn or zh_tw:
                        result.translated += 1
                        _log_line(f"  🔄 [演员库维护] {actor_name} 翻译补全: zh_cn={zh_cn or '-'} zh_tw={zh_tw or '-'}")
                    if href:
                        result.linked += 1
                        _log_line(f"  ✅ [演员库维护] {actor_name} -> {href}")
                    elif need_link and not href:
                        _log_line(f"  ⚠️ [演员库维护] {actor_name} 未在 LibreDMM 找到链接")
                except Exception as e:
                    result.failed.append((actor_name, str(e)))
                    _log_line(f"  ❌ [演员库维护] {actor_name} 处理失败: {e}")

        tasks = [asyncio.create_task(_process_one(name)) for name in names]
        await asyncio.gather(*tasks)

    # 落盘 + 重载内存缓存
    if _wb is not None:
        try:
            _ws = get_actor_db_sheet(_wb)
            _format_db_worksheet(_ws)
            _wb.save(db_path)
            _wb.close()
            resources.reload_actor_db()
            from mdcx.core.tmdb_actor import _ACTOR_DB_ROW_INDEX, _ACTOR_DB_ROW_INDEX_LOCK

            with _ACTOR_DB_ROW_INDEX_LOCK:
                _ACTOR_DB_ROW_INDEX.clear()
            _log_line(" ✅ [演员库维护] 已保存 actor_database.xlsx 并重载内存缓存")
        except Exception as e:
            _log_line(f" ❌ [演员库维护] 落盘失败，写入可能未保存: {e}")

    _log_line(
        f" 🎬 [演员库维护] 完成: 共 {result.total} 个, 翻译补全 {result.translated}, "
        f"链接补全 {result.linked}, 跳过 {result.skipped}, 失败 {len(result.failed)} ({get_used_time(start_time)}s)"
    )
    return result


_STRUCTURED_BIO_RE = re.compile(
    r"^(?:身高: [0-9.]+cm|罩杯: [^\s|]+|三围: [0-9]+/[0-9]+(?:/[0-9]+)?|生涯: [\d~\-]+|"
    r"出身: [^\s|]+|血型: [A-O]+型|事务所: [^|]+|爱好: [^|]+|出道: [^|]+|标签: [^|]+)(?: \| |$)"
)


def _is_structured_bio(bio: str) -> bool:
    """判断简介是否已是 minnano 结构化一行格式（用于 overwrite 断点续传）。

    要求以 `身高: 155cm` 等 key 开头并带 ` | ` 分隔；旧自由文本（如
    `身高155cm，三围B83/...`）不会被误判。
    """
    return bool(bio and _STRUCTURED_BIO_RE.search(bio))


def _extract_bio_fields(text: str) -> dict:
    """从出厂库自由中文简介中抽取结构化字段（reformat_minnano 用，纯本地不发请求）。

    返回 dict 与 minnano 解析结果同构（height/cup/bust/waist/hip/career/place/blood/
    agency/hobby/debut/tags），供 _build_bio_line 复用。
    """
    out: dict = {
        "height": "",
        "cup": "",
        "bust": "",
        "waist": "",
        "hip": "",
        "career": "",
        "place": "",
        "blood": "",
        "agency": "",
        "hobby": "",
        "debut": "",
        "tags": [],
    }
    if not text:
        return out

    # 身高：身高172cm / 身高: 165 / 148cm（无前缀）
    m = re.search(r"(?:身高|身長)\s*[:：]?\s*(\d{2,3})\s*(?:cm|CM)?", text)
    if m:
        out["height"] = m.group(1)

    # 罩杯：罩杯I / 罩杯：E / (F罩杯)
    m = re.search(r"罩杯\s*[:：]?\s*([A-Za-z])\s*(?:罩杯)?", text)
    if m:
        out["cup"] = m.group(1).upper()
    m = re.search(r"\(([A-Za-z])\s*罩杯\)", text)
    if m and not out["cup"]:
        out["cup"] = m.group(1).upper()

    # 三围：三围B111/W66/H92 / 三围：B79/W57/H77 / 三围: 65/93 / 三圍 B87/W56/H86（B/W/H 前缀或纯数字段）
    m = re.search(r"三[围圍]\s*[:：]?\s*((?:[BWHSbwhs]\d{2,3}[/／]?)+)", text)
    if m:
        segs = re.findall(r"[BWHSbwhs](\d{2,3})", m.group(1))
        if len(segs) >= 2:
            out["bust"], out["waist"], out["hip"] = segs[0], segs[1], (segs[2] if len(segs) > 2 else "")
    else:
        m = re.search(r"三[围圍]\s*[:：]?\s*(\d{2,3})\s*[/／]\s*(\d{2,3})(?:\s*[/／]\s*(\d{2,3}))?", text)
        if m:
            out["bust"], out["waist"] = m.group(1), m.group(2)
            out["hip"] = m.group(3) or ""

    # 生涯：出演期间2010~2011 / 活动期间 2010-2015 / 生涯: 2013~
    m = re.search(
        r"(?:出演期间|AV出演期間|活动期间|活動期間|生涯|出演期間)\s*[:：]?\s*(\d{4})年?\s*[-~～至]\s*(\d{4})年?", text
    )
    if m:
        out["career"] = f"{m.group(1)}~{m.group(2)}"
    else:
        m = re.search(r"(?:出演期间|活动期间|生涯|出演期間)\s*[:：]?\s*(\d{4})年?\s*[-~～]?", text)
        if m:
            out["career"] = f"{m.group(1)}~"

    # 出身：先匹配"出身于/出生于"，避免出身正则吞掉"于"字
    m = re.search(r"出身于\s*([\u4e00-\u9fff]{2,5})", text)
    if m:
        out["place"] = m.group(1)
    m = re.search(r"出生于\s*([\u4e00-\u9fff]{2,5})", text)
    if m and not out["place"]:
        out["place"] = m.group(1)
    m = re.search(r"(?:出身地|出身|籍贯|籍貫)\s*[:：]?\s*([\u4e00-\u9fff]{2,5})", text)
    if m and not out["place"]:
        out["place"] = m.group(1)

    # 血型：血型B型 / A型血 / 血型：O
    m = re.search(r"血型\s*[:：]?\s*([A-O])\s*型?", text)
    if m:
        out["blood"] = m.group(1).upper() + "型"
    m = re.search(r"([A-O])\s*型\s*血", text)
    if m and not out["blood"]:
        out["blood"] = m.group(1).upper() + "型"

    # 事务所：事务所名東
    m = re.search(r"(?:事务所|所属)\s*[:：]?\s*([\u4e00-\u9fffA-Za-z0-9&()（）\s]{2,20}?)(?:[，,。;；|]|$)", text)
    if m:
        out["agency"] = m.group(1).strip()

    # 爱好：爱好：滑雪 / 兴趣：料理
    m = re.search(r"(?:爱好|興趣|兴趣|趣味)\s*[:：]?\s*([^\s,，。;；|]{2,25})", text)
    if m:
        out["hobby"] = m.group(1).strip()

    # 出道：xxx（仅收短值；"出道作品：长标题" 视为噪音丢弃）
    m = re.search(r"出道\s*[:：]\s*([^\s,，。;；|]{2,15})", text)
    if m and "作品" not in m.group(1):
        out["debut"] = m.group(1).strip()

    # 标签：标签巨乳,美少女
    m = re.search(r"标签\s*[:：]?\s*([^\s|]+)", text)
    if m:
        out["tags"] = [t for t in m.group(1).split(",") if t.strip()]

    return out


def _extract_name_alias(text: str) -> list[str]:
    """从简介提取假名/罗马音/别名。

    兼容形态：
      - `日文名（假名 / 罗马音）`（逗号/未闭合括号变体）
      - `别名：xxx` / `姓名：xxx，别名：yyy`（别名字段）
    返回可并入 keyword 的别名列表，日文名主体不返回。
    """
    aliases: list[str] = []
    if not text:
        return aliases

    def _add(seg: str) -> None:
        seg = seg.strip()
        if len(seg) >= 2 and not seg.isdigit() and "・" not in seg and not re.fullmatch(r"[A-Z]{2,}", seg):
            aliases.append(seg)

    # 提取"别名："后的整段（可含多个，用逗号/顿号/空格分隔）
    for m in re.finditer(r"别名\s*[:：]\s*([^。；;|]+)", text):
        for seg in re.split(r"[，,、\s]+", m.group(1)):
            _add(seg)
    # 去掉"姓名："前缀后的括号段（假名/罗马音）
    name_part = re.sub(r"姓名\s*[:：]\s*", "", text)
    for m in re.finditer(r"[（(]([^（）()]*)[)）]?", name_part):
        inner = m.group(1).strip()
        for seg in re.split(r"[/／，,]", inner):
            seg = seg.strip()
            _add(seg)
    return aliases


def _clean_alias_parens(alias: str) -> str | None:
    """清洗别名中的括号后缀。

    返回清洗后的名字，或 None 表示整条删除。
    规则：
      1. (注)/（注） 开头的注释说明 → 整条删除
      2. 名字(数字/仮/仮名) + 后续内容 → 去括号及之后所有内容，只留括号前名字
      3. 名字(标签) 末尾括号 → 去括号保留名字
      4. 括号在中间（如 "しいなうしお SIR(...)"）→ 整条删除
      5. 括号前为空 → 整条删除
      6. 嵌套括号取最外层
    """
    import re

    s = alias.strip()
    if not s:
        return None
    # 无括号 → 原样返回
    if "(" not in s and "（" not in s:
        return s
    # 统一全角括号
    s = s.replace("（", "(").replace("）", ")")
    # 规则1: (注) 开头的注释说明 → 整条删除
    if re.match(r"^\(注\)", s):
        return None
    # 规则4: 括号前有空格（括号在中间，如 "しいなうしお SIR(...)"）→ 整条删除
    if re.match(r"^\S+\s+\S*\(", s):
        return None
    # 找第一个括号
    m = re.match(r"^(.+?)\s*\(", s)
    if not m:
        return None
    base = m.group(1).strip()
    # 规则5: 括号前为空 → 整条删除
    if not base:
        return None
    # 规则2/3: 去括号及之后所有内容，只留括号前名字
    # 但要确认括号前是"名字"而非"名字 标签"
    return base


def _build_bio_line(parsed: dict) -> str:
    """按 emby 补全风格拼一行简介（身高/罩杯/三围/生涯/出身/血型/事务所/爱好/出道/标签）。

    缺字段则跳过该段；全空时返回空串（调用方不写）。
    """
    parts = []
    if parsed.get("height"):
        parts.append(f"身高: {parsed['height']}cm")
    if parsed.get("cup"):
        parts.append(f"罩杯: {parsed['cup']}")
    if parsed.get("bust") or parsed.get("waist") or parsed.get("hip"):
        segs = [s for s in (parsed.get("bust"), parsed.get("waist"), parsed.get("hip")) if s]
        if len(segs) >= 2:
            parts.append(f"三围: {'/'.join(segs)}")
    if parsed.get("career"):
        m = re.search(r"(20\d{2}|19\d{2})", str(parsed["career"]))
        if m:
            start = m.group(1)
            m2 = re.search(r"(?:~|～|至|[-—])\s*(20\d{2}|19\d{2})", str(parsed["career"]))
            career = f"{start}~{m2.group(1)}" if m2 else f"{start}~"
            parts.append(f"生涯: {career}")
    if parsed.get("place"):
        parts.append(f"出身: {parsed['place']}")
    if parsed.get("blood"):
        parts.append(f"血型: {parsed['blood']}")
    if parsed.get("agency"):
        parts.append(f"事务所: {parsed['agency']}")
    if parsed.get("hobby"):
        parts.append(f"爱好: {parsed['hobby']}")
    if parsed.get("debut"):
        debut = str(parsed["debut"]).strip()
        if len(debut) < 15:
            parts.append(f"出道: {debut}")
    if parsed.get("tags"):
        parts.append(f"标签: {','.join(parsed['tags'])}")
    return " | ".join(parts)


def _clean_struct_segments(bio: str) -> str:
    """规范化结构化简介的字段段（cleanup_bio 与 fill_minnano 兜底共用）。

    规则（只作用于 `键: 值` 段，非结构化自由文本由 _cleanup_bio_residual 处理）：
      1. 出道值 >= 15 字（作品标题）→ 整段删除
      2. 三围单值（"三围: 100" 无 /）→ 整段删除
      3. 事务所/爱好 值循环剥离前缀（"事务所: 事务所为SELECTION" -> "事务所: SELECTION"，
         "爱好: 爱好是按摩" -> "爱好: 按摩"）
      4. 连续标点（"。。。" "，，"）合并为单个
    """
    parts = [p.strip() for p in bio.split("|")] if "|" in bio else [bio.strip()]
    out: list[str] = []
    for p in parts:
        if p.startswith("出道: "):
            v = p[len("出道: ") :].strip()
            if len(v) >= 15:
                continue
            out.append(p)
        elif re.fullmatch(r"三围: \d{2,3}", p):
            continue
        elif p.startswith("事务所: ") or p.startswith("爱好: "):
            prefix, v = p.split(": ", 1)
            while True:
                m = re.match(r"^(?:事务所|所属|爱好|为|是|包括|及)", v)
                if not m:
                    break
                v = v[len(m.group(0)) :].strip()
            v = re.sub(r"^[:：]\s*", "", v)
            if v:
                out.append(f"{prefix}: {v}")
        else:
            out.append(p)
    cleaned = " | ".join(out)
    cleaned = re.sub(r"[，,。;；]{2,}", lambda m: m.group(0)[0], cleaned)
    cleaned = re.sub(r"。。。+", "。", cleaned)
    return cleaned


def _cleanup_bio_residual(bio: str, jp: str) -> str:
    """清理简介中无用残留（cleanup_bio 模式与 fill_minnano 兜底共用）。

    规则：
      1. 抽"xxx年出道" → 生涯字段，删原句；剩余仍是大段自由文本则只留生涯
      2. 删"出道作品/作品/参演作品/出道作为/作品仅出演过/作品为"长标题噪音
      3. 清"サイズ：S"日文尺寸残段
      4. 三围孤值（"三围: 90" 单数字无 /）→ 清空
      5. 空字段残留（"标签:" "出道:" "背景:" 等）
      6. 只剩纯名字（"高宮慶子" / "姓名：宮本冷子"）→ 清空
    返回清理后的简介（可能为空串）。
    """
    new_bio = bio

    m = re.search(r"(\d{4})年(?:\d{1,2}月)?\s*出道", new_bio)
    if m:
        career = f"{m.group(1)}~"
        if career not in new_bio:
            new_bio = f"生涯: {career} | {new_bio}" if new_bio else f"生涯: {career}"
        new_bio = re.sub(r"[，,。;；]?\s*\d{4}年(?:\d{1,2}月)?\s*出道[，,。;；]?", "", new_bio)
        rest = new_bio.replace(f"生涯: {career} |", "", 1).strip()
        if rest and re.search(r"[（()）]|别名|出道作", rest):
            new_bio = f"生涯: {career}"

    new_bio = re.sub(
        r"[，,。;；]?\s*(?:出道作品|参演作品|出道作为|作品仅出演过|作品为|作品)\s*[:：]?[^|]*", "", new_bio
    )
    new_bio = re.sub(r"[，,。;；\s]*サイズ\s*[:：]?\s*[A-Za-z0-9.]*[，,。;；,]?", "", new_bio)
    # 清"鞋码/鞋碼"残留段（历史数据清理项，已从项目移除）
    new_bio = re.sub(
        r"[，,。;；\s]*(?:鞋码|鞋碼|鞋子尺寸|靴サイズ)\s*[:：]?\s*[A-Za-z0-9.]*\s*[，,。;；]?", "", new_bio
    )

    if re.fullmatch(r"(?:三围|三圍|バスト)\s*[:：]?\s*\d{2,3}\s*", new_bio.strip()):
        new_bio = ""

    new_bio = re.sub(r"[，,。;；]?\s*(?:标签|出道|背景|备注)\s*[:：]?\s*(?=\||$)", "", new_bio)

    new_bio = _clean_struct_segments(new_bio)

    cleaned = new_bio.strip(" |,").strip()
    if cleaned and not _is_structured_bio(cleaned):
        m = re.fullmatch(r"(?:姓名|氏名|名字)\s*[:：]?\s*([\u4e00-\u9fffA-Za-z0-9]+)", cleaned)
        if m and m.group(1) == jp:
            return ""
        if re.fullmatch(r"[\u4e00-\u9fffA-Za-z0-9]+", cleaned) and cleaned == jp:
            return ""
    return cleaned


def _resolve_bio_from_parsed(parsed: dict | None, orig_bio: str, jp: str) -> tuple[str, str]:
    """确定简介内容与来源（fill_minnano 用）。

    优先 minnano 结构化数据；minnano 无 bio 字段（或未查到）→ 本地 reformat 原简介；
    仍提不出字段 → 清洗原简介残留。返回 (new_bio, source)，source ∈ minnano/local/clean。
    """
    if parsed:
        bio = _build_bio_line(parsed)
        if bio:
            return bio, "minnano"
    if orig_bio:
        local = _build_bio_line(_extract_bio_fields(orig_bio))
        if local:
            return local, "local"
        cleaned = _cleanup_bio_residual(orig_bio, jp)
        if cleaned != orig_bio:
            return cleaned, "clean"
    return "", ""


# 简介中可直接用 info 库映射翻译的字段（优先 info，未命中再调翻译引擎）
_INFO_MAPPED_FIELDS = ("事务所", "标签")
# 简介中直接调翻译引擎的字段
_ENGINE_FIELDS = ("出身", "爱好")
# 非文本字段（数值/字母，不翻译）
_NON_TEXT_FIELDS = ("身高", "罩杯", "三围", "血型", "生涯", "出道")


async def _translate_bio_text_fields(bio: str) -> str:
    """翻译结构化简介的日文文本字段（出身/爱好/事务所/标签）。

    - 出身/爱好：直接用运行时翻译引擎（info 库无映射）
    - 事务所/标签：先用 info 库映射（get_info_data → zh_cn），未命中再调翻译引擎
    - 单字段翻译失败保留原文；翻译引擎不可用则整体原样返回。
    """
    if not bio or not _is_structured_bio(bio):
        return bio

    from ..config.enums import Language

    translate_by = manager.config.translate_config.translate_by
    if not translate_by:
        return bio

    def _needs_translate(value: str) -> bool:
        """判断字段值是否需要翻译：含假名，或非简体中文（繁体/日文汉字）。"""
        if re.search(r"[ぁ-んァ-ヶ]", value):
            return True
        return zhconv.convert(value, "zh-hans") != value

    async def _translate_via_engine(text: str) -> str | None:
        """遍历用户配置的翻译引擎，第一个成功的生效（与刮削翻译一致）。"""
        import random

        from ..base.translate import get_translator_skip_reason, translate_with_engine

        engines = list(translate_by)
        random.shuffle(engines)
        for each in engines:
            if get_translator_skip_reason(each):
                continue
            try:
                result = await translate_with_engine(
                    each, text, "", title_language=Language.ZH_CN, outline_language=Language.ZH_CN
                )
                if result and not result.error and result.title and result.title.strip() != text:
                    return result.title.strip()
            except Exception:
                continue
        return None

    segs = bio.split("|")
    new_segs: list[str] = []
    changed = False
    for seg in segs:
        seg = seg.strip()
        m = re.match(r"^([^\s:：]+)[:：]\s*(.*)$", seg)
        if not m:
            new_segs.append(seg)
            continue
        field = m.group(1).strip()
        value = m.group(2).strip()
        if not value or field in _NON_TEXT_FIELDS:
            new_segs.append(seg)
            continue
        if not _needs_translate(value):
            new_segs.append(seg)
            continue
        translated = None
        if field in _INFO_MAPPED_FIELDS:
            # 标签字段：逐词用 info 库映射；事务所：整体映射
            if field == "标签":
                new_tags = []
                tag_changed = False
                for t in value.split(","):
                    t = t.strip()
                    if not t:
                        continue
                    info = resources.get_info_data(t)
                    zh = info.get("zh_cn") if info.get("has_name") else None
                    if zh and zh != t:
                        new_tags.append(zh)
                        tag_changed = True
                    else:
                        new_tags.append(t)
                if tag_changed:
                    new_segs.append(f"{field}: {','.join(new_tags)}")
                    changed = True
                    continue
                else:
                    new_segs.append(seg)
                    continue
            else:
                info = resources.get_info_data(value)
                zh = info.get("zh_cn") if info.get("has_name") else None
                if zh and zh != value:
                    new_segs.append(f"{field}: {zh}")
                    changed = True
                    continue
        # 调翻译引擎（info 未命中或出身/爱好）
        translated = await _translate_via_engine(value)
        if translated:
            new_segs.append(f"{field}: {translated}")
            changed = True
        else:
            new_segs.append(seg)
    if changed:
        return " | ".join(s for s in new_segs if s)
    return bio


# ============================================================
# 检查用户库（GUI「检查用户库」按钮用，与出厂库 check_actor_db 共用规则）
# ============================================================


def _check_actor_db_issues(db_path: Path) -> dict:
    """检查 actor_database.xlsx（运行库或出厂库），返回分类问题列表。

    返回结构:
      {"errors": [(行号, 消息, 类别), ...], "warnings": [(行号, 消息, 类别), ...]}
    类别: jp_empty/jp_dup/kw_format/kw_dup/birth_format/birth_range/career_no_year/
          tmdb_no_id/tmdb_mismatch/tmdb_dup/orphan_link/name_empty/bio_jp/bio_unstruct
    """
    import openpyxl as _xl

    issues: dict[str, list[tuple[int, str, str]]] = {"errors": [], "warnings": []}
    if not db_path.exists():
        return issues
    try:
        wb = _xl.load_workbook(db_path, read_only=True, data_only=True)
    except Exception:
        return issues
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    _FULL2HALF = str.maketrans("０１２３４５６７８９", "0123456789")

    def _add(err, warn, row_idx, msg, cat):
        if err:
            issues["errors"].append((row_idx, msg, cat))
        elif warn:
            issues["warnings"].append((row_idx, msg, cat))

    seen_jp: dict[str, int] = {}
    seen_tmdb: dict[str, int] = {}
    seen_url: dict[str, int] = {}
    for idx, row in enumerate(rows, 2):
        jp = str(row[0] or "").strip() if len(row) > 0 else ""
        zh_cn = str(row[1] or "").strip() if len(row) > 1 else ""
        zh_tw = str(row[2] or "").strip() if len(row) > 2 else ""
        kw = str(row[3] or "").strip() if len(row) > 3 else ""
        tmdb = str(row[5] or "").strip() if len(row) > 5 else ""
        url = str(row[6] or "").strip() if len(row) > 6 else ""
        birth = str(row[7] or "").strip() if len(row) > 7 else ""
        bio = str(row[8] or "").strip() if len(row) > 8 else ""

        if not jp:
            _add(True, False, idx, "jp(日文原名) 为空", "jp_empty")
            continue
        # jp 重复
        jp_key = jp.casefold()
        if jp_key in seen_jp:
            _add(True, False, idx, f"jp 与行{seen_jp[jp_key]} 重复: {jp}", "jp_dup")
        else:
            seen_jp[jp_key] = idx
        # keyword 格式
        if kw and (kw.startswith(",") or kw.endswith(",") or ",," in kw):
            _add(True, False, idx, f"keyword 存在首尾/连续逗号: {kw[:30]}", "kw_format")
        # keyword 重复
        if kw:
            parts = [k.strip() for k in kw.split(",") if k.strip()]
            if len(parts) != len({k.casefold() for k in parts}):
                _add(True, False, idx, f"keyword 存在重复词: {kw[:30]}", "kw_dup")
        # zh_cn/zh_tw 空
        if not zh_cn:
            _add(False, True, idx, "zh_cn(中文名) 为空", "name_empty")
        if not zh_tw:
            _add(False, True, idx, "zh_tw(繁体名) 为空", "name_empty")
        # tmdbid 重复
        if tmdb and tmdb.isdigit():
            if tmdb in seen_tmdb:
                _add(True, False, idx, f"tmdbid 与行{seen_tmdb[tmdb]} 重复: {tmdb}", "tmdb_dup")
            else:
                seen_tmdb[tmdb] = idx
        # 出生日期格式
        if birth and not re.fullmatch(r"\d{4}(-\d{1,2}(-\d{1,2})?)?", birth):
            _add(True, False, idx, f"出生日期格式非法: {birth}", "birth_format")
        elif birth:
            m = re.match(r"(\d{4})", birth)
            if m:
                year = int(m.group(1))
                if year < 1900 or year > 2030:
                    _add(True, False, idx, f"出生日期年份异常({year}, 期望 1900-2030): {birth}", "birth_range")
        # tmdbid 空但 url 有值
        if jp and not tmdb and url:
            _add(True, False, idx, f"tmdbid 为空但 tmdb url 有值: {url[:40]}", "tmdb_no_id")
        # tmdbid 与 url 不匹配
        if tmdb and tmdb.isdigit() and url:
            expect = f"https://www.themoviedb.org/person/{int(tmdb)}"
            if url.rstrip("/") != expect:
                _add(True, False, idx, f"tmdbid={tmdb} 与 url 不匹配: {url[:40]}", "tmdb_mismatch")
        # tmdb url 重复
        if jp and url:
            if url in seen_url:
                _add(True, False, idx, f"tmdb url 与行{seen_url[url]} 重复", "tmdb_dup_url")
            else:
                seen_url[url] = idx
        # 生涯无年份
        if bio:
            for seg in bio.split("|"):
                seg = seg.strip()
                m = re.match(r"^生涯\s*[:：]\s*(.*)$", seg)
                if m:
                    value = m.group(1).strip()
                    if value and not re.search(r"(?:19|20)\d{2}", value.translate(_FULL2HALF)):
                        _add(True, False, idx, f"生涯无年份(非年份区间): {value}", "career_no_year")
                # 简介日文残留(排除出道)
                m2 = re.match(r"^([^\s:：]+)\s*[:：]\s*(.*)$", seg)
                if m2:
                    field, value = m2.group(1).strip(), m2.group(2).strip()
                    if field != "出道" and value and re.search(r"[ぁ-んァ-ヶ]", value):
                        _add(False, True, idx, f"简介 {field} 字段含日文: {value[:30]}", "bio_jp")
        # 简介非结构化
        if bio and not re.search(r"身高|罩杯|三围|生涯|出身|血型|事务所|爱好|出道|标签", bio):
            _add(False, True, idx, f"简介非结构化(无标准字段段): {bio[:40]}", "bio_unstruct")

    # 孤儿 hyperlink（读 XML）
    try:
        import zipfile

        with zipfile.ZipFile(db_path) as zf:
            names = set(zf.namelist())
            if "xl/workbook.xml" in names:
                wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
                m = re.search(r'<sheet[^>]*name="演员数据库"[^>]*r:id="(rId\d+)"', wb_xml)
                if m:
                    rid = m.group(1)
                    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
                    rel_m = re.search(rf'<Relationship[^>]*Id="{rid}"[^>]*Target="([^"]+)"', rels)
                    if rel_m:
                        sheet_xml = "xl/" + rel_m.group(1).lstrip("/")
                        if sheet_xml in names:
                            xml_text = zf.read(sheet_xml).decode("utf-8")
                            defined = set(re.findall(r'<c r="([A-Z]+\d+)"', xml_text))
                            for ref in re.findall(r'<hyperlink [^>]*ref="([A-Z]+\d+)"', xml_text):
                                if ref not in defined:
                                    _add(True, False, 0, f"孤儿 hyperlink 引用不存在单元格: {ref}", "orphan_link")
    except Exception:
        pass

    return issues


def auto_fix_actor_db(db_path: Path) -> dict:
    """自动修复运行库的安全项（不修 tmdb 相关，只报告）。

    安全项（可自动修）: jp 空删除 / jp 重复合并 / keyword 格式规范化+去重 /
    出生日期格式+范围 / 生涯无年份 / 孤儿 hyperlink / 简介日文残留翻译(需引擎) /
    简介非结构化重排 / tmdbid 与 url 不匹配(url 按 id 重写)

    返回 {"fixed": {类别: 次数}, "needs_manual": [...]}
    """
    import openpyxl as _xl

    result: dict[str, Any] = {"fixed": {}, "needs_manual": []}
    if not db_path.exists():
        return result
    issues = _check_actor_db_issues(db_path)

    wb = _xl.load_workbook(db_path)
    ws = wb.active
    fixed_counter: dict[str, int] = {}

    def _incr(cat):
        fixed_counter[cat] = fixed_counter.get(cat, 0) + 1

    # 1. jp 空行删除 + jp 重复合并（按行号逆序处理避免索引错乱）
    delete_rows = set()
    for row_idx, _msg, cat in issues["errors"] + issues["warnings"]:
        if cat == "jp_empty":
            delete_rows.add(row_idx)
    if delete_rows:
        for r in sorted(delete_rows, reverse=True):
            ws.delete_rows(r, 1)
        _incr("jp_empty")
    # 2. 逐行修复 keyword 格式/重复、出生日期、生涯、简介
    _FULL2HALF = str.maketrans("０１２３４５６７８９", "0123456789")
    jp_rows: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2):
        jp_val = str(row[0].value or "").strip()
        if jp_val:
            jp_rows.setdefault(jp_val.casefold(), row[0].row)
    # 合并 jp 重复（保留首个，keyword 并集）
    seen_row: dict[str, int] = {}
    merged = 0
    dup_rows_to_delete = []
    for row in ws.iter_rows(min_row=2):
        jp_val = str(row[0].value or "").strip()
        if not jp_val:
            continue
        key = jp_val.casefold()
        if key in seen_row:
            # 并集 keyword
            first_row = seen_row[key]
            cur_kw = str(ws.cell(row=first_row, column=4).value or "").strip()
            this_kw = str(ws.cell(row=row[0].row, column=4).value or "").strip()
            merged_set = set()
            for k in (cur_kw + "," + this_kw).split(","):
                k = k.strip()
                if k:
                    merged_set.add(k)
            ws.cell(row=first_row, column=4, value=",".join(sorted(merged_set)))
            dup_rows_to_delete.append(row[0].row)
            merged += 1
        else:
            seen_row[key] = row[0].row
    if merged:
        # 逆序删除重复行，避免行号偏移
        for r in sorted(dup_rows_to_delete, reverse=True):
            ws.delete_rows(r, 1)
        _incr("jp_dup")
    # keyword 规范化+去重
    kw_fixed = 0
    for row in ws.iter_rows(min_row=2):
        kw = str(ws.cell(row=row[0].row, column=4).value or "").strip()
        if not kw:
            continue
        items = [k.strip() for k in kw.split(",") if k.strip()]
        dedup = []
        seen_kw = set()
        for it in items:
            if it.casefold() not in seen_kw:
                seen_kw.add(it.casefold())
                dedup.append(it)
        new_kw = ",".join(dedup)
        if new_kw != kw:
            ws.cell(row=row[0].row, column=4, value=new_kw)
            kw_fixed += 1
    if kw_fixed:
        _incr("kw_format")
    # 出生日期 + 生涯 + tmdb url 重写
    bio_fixed = 0
    for row in ws.iter_rows(min_row=2):
        row_no = row[0].row
        # 出生日期
        birth = str(ws.cell(row=row_no, column=8).value or "").strip()
        if birth:
            m = re.match(r"(\d{4})", birth)
            if m and (int(m.group(1)) < 1900 or int(m.group(1)) > 2030):
                ws.cell(row=row_no, column=8).value = None
                _incr("birth_range")
        # 生涯无年份
        bio = str(ws.cell(row=row_no, column=9).value or "").strip()
        if bio:
            new_bio = bio
            changed = False
            segs = bio.split("|")
            new_segs = []
            for seg in segs:
                seg = seg.strip()
                m = re.match(r"^生涯\s*[:：]\s*(.*)$", seg)
                if m:
                    value = m.group(1).strip()
                    if value and not re.search(r"(?:19|20)\d{2}", value.translate(_FULL2HALF)):
                        # 提取年份，不行删除该段
                        years = re.findall(r"(?:19|20)\d{2}", value.translate(_FULL2HALF))
                        if years:
                            new_segs.append(
                                f"生涯: {min(years)}~{max(years)}" if len(years) > 1 else f"生涯: {years[0]}~"
                            )
                        changed = True
                        continue
                new_segs.append(seg)
            if changed:
                new_bio = " | ".join(s for s in new_segs if s)
                ws.cell(row=row_no, column=9, value=new_bio)
                bio_fixed += 1
        # tmdbid 与 url 不匹配 -> url 重写
        tmdb = str(ws.cell(row=row_no, column=6).value or "").strip()
        url = str(ws.cell(row=row_no, column=7).value or "").strip()
        if tmdb and tmdb.isdigit() and url:
            expect = f"https://www.themoviedb.org/person/{int(tmdb)}"
            if url.rstrip("/") != expect:
                ws.cell(row=row_no, column=7, value=expect)
                _incr("tmdb_mismatch")
    if bio_fixed:
        _incr("career_no_year")
    # 孤儿 hyperlink 删除（通过重存消除）
    try:
        wb.save(db_path)
    except Exception as e:
        # 保存失败必须上报：避免用户看到"修复了 N 项"但实际未落盘
        result["save_error"] = str(e)
    wb.close()

    result["fixed"] = fixed_counter
    # 删除行后需人工处理的行号会偏移：减去该行之前被删除的行数（jp_empty 删除 + jp 重复合并删除）
    all_deleted = sorted(delete_rows | set(dup_rows_to_delete))

    def _adjust_row(row_idx: int) -> int:
        return row_idx - sum(1 for r in all_deleted if r < row_idx)

    # 需要人工处理
    for row_idx, msg, cat in issues["errors"]:
        if cat in ("tmdb_no_id", "tmdb_dup", "tmdb_dup_url"):
            result["needs_manual"].append((_adjust_row(row_idx), msg, cat))
    return result


async def run_actor_db_xlsx(
    mode: str,
    *,
    limit: int = 5000,
    overwrite: bool = False,
    offset: int = 0,
    alias_source: str = "tmdb",
) -> None:
    """直接扫描 actor_database.xlsx 执行维护，无需演员名单。

    mode:
      'translate'    — 补全缺中文名的条目
      'link'         — 补全缺 LibreDMM 链接的条目
      'sync_aliases' — 从来源同步别名到 keyword 列
      'fill_minnano' — 从 minnano-av 补全生日/简介的条目（别名并入 keyword、生日、简介）
      'reformat_minnano' — 从出厂库原有自由中文简介中抽字段，本地重排成统一结构化格式
                          （不发请求，覆盖 fill_minnano 搜不到的行的历史简介）
      'merge_name_alias' — 把纯名字简介（日文名（假名 / 罗马音））的假名/罗马音并入 keyword
                          别名列，然后清空简介（出厂库清理用，不发请求）
      'cleanup_bio'  — 清理简介中无用残留：单字段残值（三围: 90）、日文尺寸写法（サイズ：S）、
                      空标签段（标签:）等，能抽出生涯/maggie 的先抽出再清
      'cleanup_aliases' — 清洗 keyword 列中带括号后缀的别名：去括号保留名字，
                        去括号后与主名或其他别名重复的整条删除（不发请求）
      'fill_zh_javdb'  — 从 JavDB 移动端 API 补全「中文名/繁体名」：
                        当「中文名为空」或「中文名 == 日文原名」时，查 JavDB 的 name/name_zht 字段，
                        拿到正式中文名后转为简体更新中文名列、繁体更新繁体名列。
                        不需要 tmdbid（按日文原名搜索）。

    overwrite:
      仅 fill_minnano 与 sync_aliases 生效。
      fill_minnano: False=只补空缺（运行时维护）；True=用 minnano 数据覆盖已有生日/简介。
      sync_aliases: False=只处理别名列为空的条目；True=所有条目都并入来源别名（不覆盖本地）。

    alias_source:
      仅 sync_aliases 生效。'tmdb'=从 TMDB 同步（需 API Key）；
      'avwiki'=从みんなのAV (minnano-av.com) 同步（原 AVWikiDB 改经 Cloudflare 拦截后下线）。

    offset:
      跳过数据文件前 offset 行（不含表头）再扫描，配合 limit 实现分片推进，
      避免大批量时反复停留在文件头部无法处理的行上。
    """
    db_path = _get_db_path()
    if not db_path.exists():
        _log_line(" 🔴 actor_database.xlsx 不存在")
        return

    import openpyxl as _xl

    from ..models.flags import Flags
    from ..signals import signal

    def _is_stop_requested() -> bool:
        return signal.stop or Flags.stop_requested

    wb = _xl.load_workbook(db_path)
    ws = get_actor_db_sheet(wb)
    base_url, tmdb_api_key = _resolve_tmdb_config()
    if not tmdb_api_key:
        _log_line(" ⚠️ 未配置 TMDB API Key，部分功能不可用")

    rows_to_process = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=9, values_only=True), start=2):
        if offset and row_idx - 2 < offset:
            continue
        jp = str(row[0] or "").strip()
        tmdbid_val = str(row[5] or "").strip()
        if not jp:
            continue
        # fill_minnano/reformat_minnano/merge_name_alias/cleanup_bio/cleanup_aliases 不要求 tmdbid
        # sync_aliases 仅在 javdb 来源时不要求 tmdbid（javdb 按演员名搜索，无需 id）
        # fill_zh_javdb 同样按日文原名搜索，不要求 tmdbid
        _sync_aliases_no_id = mode == "sync_aliases" and alias_source == "javdb"
        if (
            mode
            not in (
                "fill_minnano",
                "reformat_minnano",
                "merge_name_alias",
                "cleanup_bio",
                "cleanup_aliases",
                "fill_zh_javdb",
            )
            and not _sync_aliases_no_id
            and not tmdbid_val.isdigit()
        ):
            continue
        tmdbid = int(tmdbid_val) if tmdbid_val.isdigit() else 0
        zh_cn = str(row[1] or "").strip()
        zh_tw = str(row[2] or "").strip()
        href = str(row[4] or "").strip()

        if mode == "translate":
            if not zh_cn or not zh_tw:
                rows_to_process.append((jp, tmdbid, row_idx))
        elif mode == "link":
            if not href:
                rows_to_process.append((jp, tmdbid, row_idx))
        elif mode == "sync_aliases":
            if overwrite or not str(row[3] or "").strip():
                rows_to_process.append((jp, tmdbid, row_idx))
        elif mode == "fill_minnano":
            birth = str(row[7] or "").strip() if len(row) > 7 else ""
            bio = str(row[8] or "").strip() if len(row) > 8 else ""
            if overwrite:
                if not _is_structured_bio(bio):
                    rows_to_process.append((jp, tmdbid, row_idx))
            elif not birth or not bio:
                rows_to_process.append((jp, tmdbid, row_idx))
        elif mode == "reformat_minnano":
            bio = str(row[8] or "").strip() if len(row) > 8 else ""
            # 仅重排有文本且未结构化的行；空简介行无源数据可整理
            if bio and not _is_structured_bio(bio):
                rows_to_process.append((jp, tmdbid, row_idx))
        elif mode == "merge_name_alias":
            bio = str(row[8] or "").strip() if len(row) > 8 else ""
            # 处理无身体数据字段的简介（纯名字/别名），有身体数据的行保留不动
            if bio and not re.search(
                r"身高|三围|罩杯|出身|血型|生涯|出道|标签|事务所|爱好|籍贯|作品|サイズ|バスト|鞋|靴", bio
            ):
                rows_to_process.append((jp, tmdbid, row_idx))
        elif mode == "cleanup_bio":
            # 仅处理非结构化且有文本的行（已结构化的行已经过 minnano 校验，不动）
            bio = str(row[8] or "").strip() if len(row) > 8 else ""
            if bio and not _is_structured_bio(bio):
                rows_to_process.append((jp, tmdbid, row_idx))
        elif mode == "cleanup_aliases":
            kw = str(row[3] or "").strip()
            if kw and re.search(r"\([^)]*\)", kw):
                rows_to_process.append((jp, tmdbid, row_idx))
        elif mode == "fill_zh_javdb":
            # 处理两类行：1) 中文名为空；2) 中文名 == 日文原名（未做中文化）
            # 均要求日文原名含汉字（纯假名无中文形式，跳过避免无意义请求）
            if (not zh_cn or zh_cn == jp) and any("\u4e00" <= c <= "\u9fff" for c in jp):
                rows_to_process.append((jp, tmdbid, row_idx))

    if limit and len(rows_to_process) > limit:
        rows_to_process = rows_to_process[:limit]
        _log_line(f" ℹ️ 本次限量处理前 {limit} 条，可再次运行继续处理剩余")

    # 显式打印分片参数，便于用户中断后记下行号续跑
    if offset or (limit and limit < 5000):
        _log_line(f" 📌 分片参数: offset={offset}, limit={limit if limit else '不限'}")

    _log_line(f" 🎬 扫描完成：{len(rows_to_process)} 个演员需要处理 (模式: {mode})")
    if not rows_to_process:
        _log_line(" ✅ 没有需要处理的数据")
        wb.close()
        return

    # reformat_minnano 纯本地同步处理，不走网络/并发通道
    if mode == "reformat_minnano":
        reformatted = 0
        for _jp, _tmdbid, row_idx in rows_to_process:
            if _is_stop_requested():
                break
            raw = str(ws.cell(row=row_idx, column=COL_BIO + 1).value or "").strip()
            parsed = _extract_bio_fields(raw)
            bio = _build_bio_line(parsed)
            if bio and bio != raw:
                ws.cell(row=row_idx, column=COL_BIO + 1, value=bio)
                reformatted += 1
        wb.save(db_path)
        _log_line(f" ✅ 重排完成：{reformatted} 行简介已重排为结构化格式")
        wb.close()
        return

    # merge_name_alias 纯本地同步处理，不走网络/并发通道
    if mode == "merge_name_alias":
        merged = 0
        for jp, _tmdbid, row_idx in rows_to_process:
            if _is_stop_requested():
                break
            raw = str(ws.cell(row=row_idx, column=COL_BIO + 1).value or "").strip()
            aliases = _extract_name_alias(raw)
            aliases = [a for a in aliases if a != jp]
            if aliases:
                existing = str(ws.cell(row=row_idx, column=COL_KEYWORD + 1).value or "").strip()
                merged_list: list[str] = []
                seen_lower: set[str] = set()
                for kw in [x.strip() for x in existing.split(",") if x.strip()] + aliases:
                    key = kw.lower()
                    if key not in seen_lower:
                        seen_lower.add(key)
                        merged_list.append(kw)
                ws.cell(row=row_idx, column=COL_KEYWORD + 1, value=",".join(merged_list))
            ws.cell(row=row_idx, column=COL_BIO + 1, value="")
            merged += 1
        wb.save(db_path)
        _log_line(f" ✅ 名字合入别名完成：{merged} 行已处理，简介已清空")
        wb.close()
        return

    # cleanup_bio 纯本地同步处理
    if mode == "cleanup_bio":
        cleaned = 0
        for jp, _tmdbid, row_idx in rows_to_process:
            if _is_stop_requested():
                break
            bio = str(ws.cell(row=row_idx, column=COL_BIO + 1).value or "").strip()
            new_bio = _cleanup_bio_residual(bio, jp)

            # 规则 7: 仍非结构化 → 尝试提取结构化字段；有则重建，无则清空并并入别名到 keyword
            _cleaned = new_bio.strip(" |,").strip()
            if _cleaned and not _is_structured_bio(_cleaned):
                fields = _extract_bio_fields(_cleaned)
                rebuilt = _build_bio_line(fields)
                if rebuilt:
                    new_bio = rebuilt
                    _log_line(f"  {jp}: 提取字段重建 -> {new_bio}")
                else:
                    aliases = [
                        a
                        for a in _extract_name_alias(_cleaned)
                        if a != jp and re.search(r"[ぁ-んァ-ヶ]|[A-Za-z]", a) and "・" not in a
                    ]
                    if aliases:
                        existing = str(ws.cell(row=row_idx, column=COL_KEYWORD + 1).value or "").strip()
                        _merged: list[str] = []
                        _seen_lower: set[str] = set()
                        for kw in [x.strip() for x in existing.split(",") if x.strip()] + aliases:
                            key = kw.lower()
                            if key not in _seen_lower:
                                _seen_lower.add(key)
                                _merged.append(kw)
                        ws.cell(row=row_idx, column=COL_KEYWORD + 1, value=",".join(_merged))
                    new_bio = ""
                    _log_line(f"  {jp}: 无法提取字段，清空简介（别名已并入）")

            if new_bio != bio:
                ws.cell(row=row_idx, column=COL_BIO + 1, value=new_bio.strip(" |,"))
                cleaned += 1
        wb.save(db_path)
        _log_line(f" ✅ 清理完成：{cleaned} 行简介已清理")
        wb.close()
        return

    # cleanup_aliases 纯本地同步处理，不走网络/并发通道
    if mode == "cleanup_aliases":
        cleaned = 0
        for jp, _tmdbid, row_idx in rows_to_process:
            if _is_stop_requested():
                break
            raw_kw = str(ws.cell(row=row_idx, column=COL_KEYWORD + 1).value or "").strip()
            if not raw_kw:
                continue
            parts = [p.strip() for p in raw_kw.split(",") if p.strip()]
            cleaned_parts: list[str] = []
            seen_norm: set[str] = {jp.lower()}
            changed = False
            for part in parts:
                base = _clean_alias_parens(part)
                if base is None or base.lower() in seen_norm:
                    changed = True
                    continue
                if base != part:
                    changed = True
                seen_norm.add(base.lower())
                cleaned_parts.append(base)
            if changed:
                ws.cell(row=row_idx, column=COL_KEYWORD + 1, value=",".join(cleaned_parts))
                cleaned += 1
        wb.save(db_path)
        _log_line(f" ✅ 别名清洗完成：{cleaned} 行已处理")
        wb.close()
        return

    start_time = time.time()
    translated_count = 0
    linked_count = 0

    async with aiohttp.ClientSession() as client:
        if mode == "link":
            base_concurrency = 2
        elif mode == "fill_zh_javdb":
            base_concurrency = 5
        else:
            base_concurrency = 5
        current_concurrency = base_concurrency
        consecutive_failures = 0
        consecutive_successes = 0
        task_iter = iter(enumerate(rows_to_process, 1))
        running_tasks: set[asyncio.Task[None]] = set()
        task_row_map: dict[asyncio.Task[None], int] = {}  # task → xlsx 行号，用于跟踪续跑位置

        async def _process_one(jp, tmdbid, row_idx):
            nonlocal translated_count, linked_count, current_concurrency, consecutive_failures, consecutive_successes
            try:
                if mode == "translate" and base_url and tmdb_api_key:
                    translations = await _fetch_person_translations(tmdbid, base_url, tmdb_api_key, client)
                    new_zh_cn = _normalize_translation(translations.get("zh_cn", ""))
                    new_zh_tw = _normalize_translation(translations.get("zh_tw", ""))
                    if not new_zh_cn and new_zh_tw:
                        new_zh_cn = _to_simplified(new_zh_tw)
                    if not new_zh_tw and new_zh_cn:
                        new_zh_tw = zhconv.convert(new_zh_cn, "zh-hant")

                    updated = False
                    if new_zh_cn:
                        ws.cell(row=row_idx, column=2, value=new_zh_cn)
                        updated = True
                    if new_zh_tw:
                        ws.cell(row=row_idx, column=3, value=new_zh_tw)
                        updated = True
                    if updated:
                        translated_count += 1
                        _log_line(f"  ✅ {jp}: zh_cn={new_zh_cn}, zh_tw={new_zh_tw}")

                elif mode == "link":
                    href_val = await fetch_libredmm_link(jp)
                    if href_val:
                        ws.cell(row=row_idx, column=5, value=href_val)
                        linked_count += 1
                        _log_line(f"  ✅ {jp} -> {href_val}")
                    else:
                        _log_line(f"  ⚠️ {jp} 未在 LibreDMM 找到链接")

                elif mode == "sync_aliases":
                    if alias_source == "javdb":
                        from ..crawlers.javdb_app import fetch_javdb_aliases

                        new_keywords = ",".join(await fetch_javdb_aliases(jp))
                    elif alias_source == "avwiki":
                        from ..tools.minnano_crawler import fetch_minnano_aliases

                        new_keywords = ",".join(await fetch_minnano_aliases(jp))
                    elif base_url and tmdb_api_key:
                        from mdcx.core.tmdb_actor import query_single_actor_cached

                        query_result = await query_single_actor_cached(jp, base_url, tmdb_api_key, client)
                        if query_result:
                            new_keywords = _merge_keyword_values(
                                query_result.get("name", ""),
                                query_result.get("original_name", ""),
                                query_result.get("also_known_as", []),
                            )
                        else:
                            new_keywords = ""
                            _log_line(f"  ⚠️ {jp} TMDB 未查询到数据")
                    else:
                        new_keywords = ""

                    if new_keywords:
                        existing_kw = str(ws.cell(row=row_idx, column=4).value or "").strip()
                        existing_parts = [k.strip() for k in existing_kw.split(",") if k.strip()]
                        new_parts = [k.strip() for k in new_keywords.split(",") if k.strip()]
                        # 归一化去重：大小写不敏感比较，已有别名优先保留
                        seen_norm: dict[str, str] = {}
                        for k in existing_parts + new_parts:
                            norm = k.casefold()
                            if norm not in seen_norm:
                                seen_norm[norm] = k
                        merged = sorted(seen_norm.values())
                        ws.cell(row=row_idx, column=4, value=",".join(merged))
                        new_count = len(new_parts)
                        _log_line(f"  ✅ [行{row_idx}] {jp}: 别名已同步 ({new_count} 个)")

                elif mode == "fill_zh_javdb":
                    from ..crawlers.javdb_app import fetch_javdb_actor_info

                    info = await fetch_javdb_actor_info(jp)
                    if info is not None:
                        # 候选中文名优先级: name_zht（繁体正式名）> name（若与日文原名不同，说明是中文名）
                        candidate_zht = info.name_zht
                        candidate_name = info.name
                        # JavDB 的 name 字段若与日文原名相同则无价值（可能就是日文原名本身）
                        if candidate_name == jp:
                            candidate_name = ""

                        # 过滤无效候选：带括号（含别名/标注）、比日文原名短（截断或非完整名）
                        def _is_valid_candidate(s: str) -> bool:
                            if not s:
                                return False
                            if "(" in s or "（" in s or "[" in s or "【" in s:
                                return False
                            # 候选名不应比日文原名短（排除截断的别名误当中文名）
                            if len(s) < len(jp):
                                return False
                            return True

                        if not _is_valid_candidate(candidate_zht):
                            candidate_zht = ""
                        if not _is_valid_candidate(candidate_name):
                            candidate_name = ""

                        new_zh_tw = ""
                        new_zh_cn = ""
                        if candidate_zht:
                            new_zh_tw = candidate_zht
                            new_zh_cn = _to_simplified(candidate_zht)
                        elif candidate_name:
                            # name 可能是繁体中文名，转简体；若转换后与原名相同则无效
                            new_zh_tw = candidate_name
                            new_zh_cn = _to_simplified(candidate_name)
                            if new_zh_cn == jp:
                                new_zh_tw = ""
                                new_zh_cn = ""

                        if new_zh_cn:
                            # 更新单元格（仅当与当前值不同时写）
                            current_zh_cn = str(ws.cell(row=row_idx, column=COL_ZH_CN + 1).value or "").strip()
                            if new_zh_cn != current_zh_cn:
                                ws.cell(row=row_idx, column=COL_ZH_CN + 1, value=new_zh_cn)
                            current_zh_tw = str(ws.cell(row=row_idx, column=COL_ZH_TW + 1).value or "").strip()
                            if new_zh_tw and new_zh_tw != current_zh_tw:
                                ws.cell(row=row_idx, column=COL_ZH_TW + 1, value=new_zh_tw)
                            translated_count += 1
                            _log_line(f"  ✅ [行{row_idx}] {jp}: zh_cn={new_zh_cn}, zh_tw={new_zh_tw}")
                        else:
                            _log_line(f"  ⚪ {jp}: JavDB 无中文名 (name={info.name!r}, name_zht={info.name_zht!r})")
                    else:
                        _log_line(f"  ⚠️ {jp}: JavDB 未匹配到演员")

                elif mode == "fill_minnano":
                    from mdcx.tools.minnano_crawler import _clean_alias, _search_minnano_by_name, parse_minnano_page

                    orig_bio = str(ws.cell(row=row_idx, column=COL_BIO + 1).value or "").strip()
                    minnano_id, html = await _search_minnano_by_name(jp)
                    parsed = None
                    if html:
                        parsed = parse_minnano_page(html, minnano_id)
                        if not parsed or not parsed.get("name"):
                            parsed = None

                    # 1) 简介：minnano 查到且有 bio 字段 → 覆盖；无 bio 字段（或未查到）→
                    #    fallback 本地 reformat 原简介；再提不出 → 清洗原简介残留
                    new_bio, source = _resolve_bio_from_parsed(parsed, orig_bio, jp)
                    if source:
                        # 翻译日文文本字段（出身/爱好/事务所/标签）：info 库优先，引擎兜底
                        new_bio = await _translate_bio_text_fields(new_bio)
                        ws.cell(row=row_idx, column=COL_BIO + 1, value=new_bio)

                    # 2) 别名并入 keyword（minnano 查到才并入，只并入不覆盖，大小写不敏感去重）
                    if parsed:
                        aliases = [a for a in parsed.get("aliases", []) if _clean_alias(a)]
                        if aliases:
                            existing = str(ws.cell(row=row_idx, column=COL_KEYWORD + 1).value or "").strip()
                            merged_list: list[str] = []
                            seen_lower: set[str] = set()
                            for kw in [x.strip() for x in existing.split(",") if x.strip()] + aliases:
                                key = kw.lower()
                                if key not in seen_lower:
                                    seen_lower.add(key)
                                    merged_list.append(kw)
                            ws.cell(row=row_idx, column=COL_KEYWORD + 1, value=",".join(merged_list))

                    # 3) 生日：覆盖或填空
                    if (
                        parsed
                        and parsed.get("birthday")
                        and (overwrite or not str(ws.cell(row=row_idx, column=COL_BIRTH_DATE + 1).value or "").strip())
                    ):
                        ws.cell(row=row_idx, column=COL_BIRTH_DATE + 1, value=parsed["birthday"])

                    if source == "minnano":
                        _log_line(f"  ✅ {jp}: minnano 数据已补全")
                    elif source == "local":
                        _log_line(f"  🔄 {jp}: minnano 无数据，本地重排原简介")
                    elif source == "clean":
                        _log_line(f"  🧹 {jp}: minnano 无数据且原简介无字段，清理残留")
                    else:
                        _log_line(f"  ⚠️ {jp}: minnano 未找到且原简介为空")

            except Exception as e:
                _log_line(f"  ❌ {jp} 处理失败: {e}")
                if mode in ("fill_minnano", "fill_zh_javdb"):
                    consecutive_failures += 1
                    consecutive_successes = 0
            else:
                if mode in ("fill_minnano", "fill_zh_javdb"):
                    consecutive_successes += 1
                    consecutive_failures = 0

        def _submit_next() -> bool:
            try:
                _, (jp, tmdbid, row_idx) = next(task_iter)
            except StopIteration:
                return False
            task = asyncio.create_task(_process_one(jp, tmdbid, row_idx))
            task_row_map[task] = row_idx
            running_tasks.add(task)
            return True

        for _ in range(min(base_concurrency, len(rows_to_process))):
            _submit_next()

        total = len(rows_to_process)
        completed = 0
        last_processed_row: int = 0  # 记录最后一个已完成任务的 xlsx 行号，用于续跑提示
        progress_interval = max(1, total // 10)  # 每 10% 输出一次进度

        while running_tasks:
            if _is_stop_requested():
                for t in running_tasks:
                    t.cancel()
            done, pending = await asyncio.wait(running_tasks, return_when=asyncio.FIRST_COMPLETED)
            running_tasks = set(pending)
            # 自适应限流：连续失败降并发，连续成功恢复
            if mode in ("fill_minnano", "fill_zh_javdb"):
                if consecutive_failures >= 3 and current_concurrency > 1:
                    current_concurrency = max(1, current_concurrency // 2)
                    consecutive_failures = 0
                    _log_line(f"  ⚠️ 连续失败较多，并发降为 {current_concurrency}")
                elif consecutive_successes >= 20 and current_concurrency < base_concurrency:
                    current_concurrency = min(base_concurrency, current_concurrency + 1)
                    consecutive_successes = 0
                    _log_line(f"  ℹ️ 并发恢复为 {current_concurrency}")
            while len(running_tasks) < current_concurrency and not _is_stop_requested():
                if not _submit_next():
                    break
            for done_task in done:
                completed += 1
                row_idx = task_row_map.pop(done_task, 0)
                if row_idx > last_processed_row:
                    last_processed_row = row_idx
                try:
                    done_task.result()
                except asyncio.CancelledError:
                    pass
                except Exception as e:
                    _log_line(f"  🔴 子任务异常: {e}")
            if completed % progress_interval == 0 or completed == total:
                _log_line(f"  📊 进度: {completed}/{total}")

    _format_db_worksheet(ws)
    wb.save(db_path)
    wb.close()
    resources.reload_actor_db()
    from mdcx.core.tmdb_actor import _ACTOR_DB_ROW_INDEX, _ACTOR_DB_ROW_INDEX_LOCK

    with _ACTOR_DB_ROW_INDEX_LOCK:
        _ACTOR_DB_ROW_INDEX.clear()

    if _is_stop_requested():
        # 续跑提示：推荐使用「起始行=已处理末行号-1」，从下一行继续；幂等重扫几行无副作用
        if last_processed_row:
            _log_line(
                f" ⛔️ 已手动停止：已保存已处理部分 ({completed}/{total})，最后处理到 xlsx 第 {last_processed_row} 行。"
                f"续跑提示：将「起始行」填入 {last_processed_row - 1} 即可从下一行继续（重复并入安全）"
            )
        else:
            _log_line(f" ⛔️ 已手动停止：已保存已处理部分 ({completed}/{total})，可再次运行继续处理剩余")
    else:
        _log_line(f" ✅ 完成: 翻译补全={translated_count}, 链接补全={linked_count} ({get_used_time(start_time)}s)")


def _entry_name_of_row(ws, row_idx: int) -> str:
    for col in (COL_JP, COL_ZH_CN, COL_KEYWORD):
        val = str(ws.cell(row=row_idx, column=col + 1).value or "").strip()
        if val:
            return val.split(",")[0]
    return f"第{row_idx}行"


_KATAKANA_RE = re.compile(r"[\u30a0-\u30ff\u30fb]")


def _is_katakana_roman_pair(entry_variants: set[str], tmdb_names: list[str]) -> bool:
    """判断条目名与 TMDB 人物名是否为「片假名音译 ↔ 英文原名」对照（同人）。

    例: キャシー・ヘブン (日文音译) 与 Cathy Heaven (英文原名)。
    """
    for v in entry_variants:
        chars = v.replace("・", "").replace(" ", "")
        if not chars:
            continue
        kat_count = sum(1 for c in chars if "\u30a0" <= c <= "\u30ff")
        if kat_count / len(chars) < 0.5:
            continue
        for n in tmdb_names:
            n2 = n.replace(" ", "")
            if n2 and all(("a" <= c.lower() <= "z") or c in ".-'" for c in n2):
                return True
    return False


def _tmdb_id_matches_entry(entry_variants: set[str], identity: dict | None) -> bool:
    """校验 AVdb 提供的 tmdbid 是否与条目名是同一人。

    - identity 为 None（请求失败/404）：保守放行（无法判断时不拦，避免网络问题误伤）
    - 归一化名字（含 こ/子 变体）有交集：放行
    - 片假名音译 ↔ 英文原名对照：放行（同人）
    - 其余：不匹配，丢弃该 tmdbid（宁缺毋滥）
    """
    if identity is None:
        return True
    tmdb_names: list[str] = []
    for n in (identity.get("name"), identity.get("original_name")):
        if n:
            tmdb_names.append(str(n).strip())
    for a in identity.get("also_known_as") or []:
        if a:
            tmdb_names.append(str(a).strip())
    if not entry_variants or not tmdb_names:
        return True
    tmdb_set = _norm_name_set([n for n in tmdb_names if n])
    if entry_variants & tmdb_set:
        return True
    return _is_katakana_roman_pair(entry_variants, tmdb_names)


def _entry_variants(jp: str, zh_cn: str, kw_list: list[str]) -> set[str]:
    """构建条目名的归一化变体集合（含 こ/子 变体）。"""
    variants: set[str] = set()
    for name in [jp, zh_cn, *kw_list]:
        if name:
            variants |= _expand_name_variants(name)
    return variants


async def sync_from_avdb(
    source: str, value: str = "", *, filter_male: bool = True, verify_tmdbid: bool = True
) -> ActorDbSyncResult:
    """从 AVdb (li-peifeng/Jav-Actors-Mapping) 同步演员映射到本地数据库。

    source:
      'jsdelivr'  — 通过 cdn.jsdelivr.net 拉取 (默认)
      'github'    — 通过 GitHub raw 拉取
      'url'       — 从 value 指定的任意下载地址拉取
      'file'      — 从 value 指定的本地 xml 文件导入

    filter_male: True 时对「待新建」条目做男优过滤：优先命中内置男优名单
    (resources/userdata/male_actors.txt) 直接跳过；名单未命中且带 tmdb_id 的
    再校验 TMDB gender，gender=2 (男) 的条目不写入；TMDB 未配置/请求失败时不
    校验直接写入（不误删）。

    匹配顺序: tmdbid 冲突优先并入 -> jp 精确 -> zh_cn 精确 -> keyword 命中。
    本地已有值优先，仅填充空缺字段；tmdbid 冲突视为同一人并入别名。
    """
    from ..base.web import download_file_with_filepath
    from ..utils.actor_clean import clean_actor_keyword, clean_actor_name
    from ..utils.xml_avdb import clean_actor_value, parse_avdb_actor_mapping

    result = ActorDbSyncResult()
    db_path = _get_db_path()

    def _dedup_keywords(*groups: set[str]) -> list[str]:
        seen: set[str] = set()
        out: list[str] = []
        for group in groups:
            for k in sorted(group):
                key = k.casefold()
                if key not in seen:
                    seen.add(key)
                    out.append(k)
        return out

    # ---- 数据源: 读取 XML 文本 ----
    xml_text = ""
    if source == "file":
        file_path = Path(value)
        if not file_path.exists():
            _log_line(f" ❌ [AVdb同步] 本地文件不存在: {file_path}")
            result.failed.append(("<本地文件>", f"文件不存在: {file_path}"))
            return result
        async with aiofiles.open(file_path, encoding="utf-8") as f:
            xml_text = await f.read()
    else:
        url = value or (AVDB_MAPPING_URL_MIRROR if source == "jsdelivr" else AVDB_MAPPING_URL)
        cache_dir = db_path.parent / "cache"
        cache_dir.mkdir(parents=True, exist_ok=True)
        tmp_file = cache_dir / "avdb_actor_mapping.xml"
        if not await download_file_with_filepath(url, tmp_file, cache_dir):
            _log_line(f" ❌ [AVdb同步] 下载失败: {url}")
            result.failed.append((url, "网络下载失败"))
            return result
        result.downloaded = True
        async with aiofiles.open(tmp_file, encoding="utf-8") as f:
            xml_text = await f.read()

    try:
        actors = parse_avdb_actor_mapping(xml_text)
    except ValueError as exc:
        _log_line(f" ❌ [AVdb同步] XML 解析失败: {exc}")
        result.failed.append(("<xml>", str(exc)))
        return result
    result.parsed = len(actors)
    _log_line(f" 🎬 [AVdb同步] 解析 {len(actors)} 条 AVdb 映射, 开始合并")

    # ---- 男优过滤 / tmdbid 身份校验配置 ----
    # 内置名单过滤不依赖 TMDB，始终可用；TMDB gender 校验与 tmdbid 身份校验仅在配置了 API Key 时启用。
    tmdb_base_url = tmdb_api_key = ""
    if filter_male or verify_tmdbid:
        tmdb_base_url, tmdb_api_key = _resolve_tmdb_config()
        if not tmdb_api_key:
            if filter_male:
                _log_line(" ⚠️ [AVdb同步] 未配置 TMDB API Key，仅使用内置名单过滤男优")
            if verify_tmdbid:
                _log_line(" ⚠️ [AVdb同步] 未配置 TMDB API Key，跳过 tmdbid 身份校验")
    tmdb_session: aiohttp.ClientSession | None = None

    async def _tmdb_client() -> aiohttp.ClientSession:
        nonlocal tmdb_session
        if tmdb_session is None:
            tmdb_session = aiohttp.ClientSession()
        return tmdb_session

    # ---- 加载/创建 数据库 ----
    import openpyxl as _xl

    try:
        async with _actor_db_write_lock:
            db_path.parent.mkdir(parents=True, exist_ok=True)
            if db_path.exists():
                wb = _xl.load_workbook(db_path)
            else:
                wb = _xl.Workbook()
                ws = wb.active
                ws.title = "演员数据库"
                for col, header in enumerate(DB_HEADERS, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = _xl.styles.Font(bold=True)
                    cell.fill = _xl.styles.PatternFill("solid", fgColor="C0C0C0")
                    cell.alignment = _xl.styles.Alignment(horizontal="center")
            ws = get_actor_db_sheet(wb)
            if ws.title != "演员数据库":
                ws.title = "演员数据库"

            # ---- 构建本地索引 ----
            jp_index: dict[str, int] = {}
            zh_cn_index: dict[str, int] = {}
            keyword_index: dict[str, int] = {}
            tmdb_index: dict[str, int] = {}
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=len(DB_HEADERS), values_only=True), start=2):
                if row is None:
                    continue

                def _cell(values, col_idx):
                    return str(values[col_idx] or "").strip() if len(values) > col_idx else ""

                jp_val = _cell(row, COL_JP)
                zh_val = _cell(row, COL_ZH_CN)
                kw_val = _cell(row, COL_KEYWORD)
                tmdb_val = _cell(row, COL_TMDBID)
                if jp_val:
                    jp_index.setdefault(jp_val.casefold(), row_idx)
                if zh_val:
                    zh_cn_index.setdefault(zh_val.casefold(), row_idx)
                for k in [k.strip() for k in kw_val.split(",") if k.strip()]:
                    keyword_index.setdefault(k.casefold(), row_idx)
                if tmdb_val:
                    tmdb_index.setdefault(tmdb_val, row_idx)

            total = len(actors)
            progress_interval = max(1, total // 10)  # 每 10% 输出一次进度

            # ---- 并发预热 TMDB 身份缓存 ----
            # 原实现逐行串行 await fetch_person_identity（每个新 id 一个网络往返），
            # 大量条目时极慢。这里提前并发请求所有需校验的 tmdbid，
            # 结果存 identity_cache 供主循环 O(1) 查询；未预热（如本地新建行）才现场请求。
            identity_cache: dict[int, dict | None] = {}
            if verify_tmdbid and tmdb_api_key:
                pending_ids: set[int] = set()
                for actor in actors:
                    raw_id = clean_actor_value(actor.tmdb_id)
                    if raw_id.isdigit() and raw_id not in tmdb_index:
                        pending_ids.add(int(raw_id))
                if pending_ids:
                    _log_line(f" ⚡ [AVdb同步] 并发预热 {len(pending_ids)} 个 tmdbid 身份校验...")
                    _tmdb_http_client = await _tmdb_client()
                    _semaphore = asyncio.Semaphore(8)

                    async def _prefetch_tmdb_identity(_pid: int) -> None:
                        async with _semaphore:
                            identity_cache[_pid] = await fetch_person_identity(
                                _pid, tmdb_base_url, tmdb_api_key, _tmdb_http_client
                            )

                    await asyncio.gather(*(_prefetch_tmdb_identity(_pid) for _pid in pending_ids))

            for n, actor in enumerate(actors, 1):
                try:
                    jp = clean_actor_value(actor.jp)
                    zh_cn = clean_actor_value(actor.zh_cn)
                    zh_tw = clean_actor_value(actor.zh_tw)
                    keyword = clean_actor_value(actor.keyword)
                    tmdb_id = clean_actor_value(actor.tmdb_id)
                    birth_date = clean_actor_value(actor.birth_date)
                    bio = clean_actor_value(actor.bio)

                    # 语义清洗：剥离名字/别名中的系列标签、年份、标注、作品标题、占位符
                    jp = clean_actor_name(jp)
                    zh_cn = clean_actor_name(zh_cn)
                    zh_tw = clean_actor_name(zh_tw)
                    keyword = clean_actor_keyword(keyword)

                    kw_list = [k.strip() for k in keyword.split(",") if k.strip()]
                    kw_set = set(kw_list)
                    entry_name = jp or zh_cn or (kw_list[0] if kw_list else "<无名字段>")
                    if not jp and not zh_cn and not kw_list:
                        continue

                    def _norm_tmdb(raw: str) -> str:
                        return raw if raw.isdigit() else ""

                    tmdb_key = _norm_tmdb(tmdb_id)

                    # 4) tmdbid 身份校验：AVdb 提供的 id 若与条目名不是同一人则丢弃该 id
                    #    （宁缺毋滥——错误 id 比无 id 更糟，刮削遇同名演员会按名字重新搜索）
                    #    仅当该 id 本地尚不存在时校验（已存在的映射为历史数据，不重复反查）
                    if verify_tmdbid and tmdb_key and tmdb_api_key and tmdb_key not in tmdb_index:
                        _pid = int(tmdb_key)
                        _identity = identity_cache.get(_pid)
                        if _identity is None and _pid not in identity_cache:
                            _identity = await fetch_person_identity(
                                _pid, tmdb_base_url, tmdb_api_key, await _tmdb_client()
                            )
                        if not _tmdb_id_matches_entry(_entry_variants(jp, zh_cn, kw_list), _identity):
                            result.skipped_tmdbid += 1
                            _log_line(f"  ⚠️ [AVdb同步] tmdbid={tmdb_key} 与名字 {entry_name} 不匹配，丢弃该 id")
                            tmdb_key = ""

                    # 1) tmdbid 冲突优先并入
                    target_row = None
                    is_tmdb_conflict = False
                    if tmdb_key:
                        conflict_row = tmdb_index.get(tmdb_key)
                        if conflict_row is not None:
                            target_row = conflict_row
                            is_tmdb_conflict = True

                    # 2) jp -> zh_cn -> keyword 文本匹配
                    if target_row is None:
                        candidate = None
                        if jp:
                            candidate = jp_index.get(jp.casefold())
                        if candidate is None and zh_cn:
                            candidate = zh_cn_index.get(zh_cn.casefold())
                        if candidate is None:
                            for k in kw_list:
                                candidate = keyword_index.get(k.casefold())
                                if candidate is not None:
                                    break
                        target_row = candidate

                    # 3) 未匹配 -> 新建
                    if target_row is None:
                        if filter_male:
                            # 优先按内置男优名单过滤（不依赖 TMDB，可命中无 tmdbid 的男优）；
                            # jp / zh_cn 任一命中即判定男优
                            if (jp and is_male_actor(jp)) or (zh_cn and is_male_actor(zh_cn)):
                                result.skipped_male += 1
                                _log_line(f"  🚫 [AVdb同步] 跳过男优: {entry_name} (名单命中)")
                                continue
                            # 名单未命中时，对带 tmdbid 的条目再校验 TMDB gender
                            if tmdb_key:
                                try:
                                    c = await _tmdb_client()
                                    gender = await fetch_person_gender(int(tmdb_key), tmdb_base_url, tmdb_api_key, c)
                                    if gender == 2:
                                        result.skipped_male += 1
                                        _log_line(f"  🚫 [AVdb同步] 跳过男优: {entry_name} (tmdbid={tmdb_key})")
                                        continue
                                except Exception:
                                    pass
                        tmdb_val = int(tmdb_key) if tmdb_key else ""
                        ws.append(
                            [jp, zh_cn, zh_tw, ",".join(_dedup_keywords(kw_set)), "", tmdb_val, "", birth_date, bio]
                        )
                        new_idx = ws.max_row
                        if jp:
                            jp_index.setdefault(jp.casefold(), new_idx)
                        if zh_cn:
                            zh_cn_index.setdefault(zh_cn.casefold(), new_idx)
                        for k in kw_set:
                            keyword_index.setdefault(k.casefold(), new_idx)
                        if tmdb_key:
                            tmdb_index.setdefault(tmdb_key, new_idx)
                        result.created += 1
                        continue

                    # 4) 命中 -> 只填空缺值，不覆盖本地
                    def _fill(row_idx: int, col_idx: int, new_val: str) -> None:
                        if not new_val:
                            return
                        existing = str(ws.cell(row=row_idx, column=col_idx + 1).value or "").strip()
                        if not existing:
                            ws.cell(row=row_idx, column=col_idx + 1, value=new_val)

                    _fill(target_row, COL_JP, jp)
                    _fill(target_row, COL_ZH_CN, zh_cn)
                    _fill(target_row, COL_ZH_TW, zh_tw)
                    _fill(target_row, COL_BIRTH_DATE, birth_date)
                    _fill(target_row, COL_BIO, bio)
                    if tmdb_key:
                        existing_tmdb = str(ws.cell(row=target_row, column=COL_TMDBID + 1).value or "").strip()
                        if not existing_tmdb:
                            ws.cell(row=target_row, column=COL_TMDBID + 1, value=int(tmdb_key))

                    existing_kw = str(ws.cell(row=target_row, column=COL_KEYWORD + 1).value or "").strip()
                    existing_elems = [k.strip() for k in existing_kw.split(",") if k.strip()]
                    merged = _dedup_keywords(set(existing_elems), kw_set)
                    merged_str = ",".join(merged)
                    if merged_str != existing_kw:
                        ws.cell(row=target_row, column=COL_KEYWORD + 1, value=merged_str)

                    if is_tmdb_conflict:
                        result.merged += 1
                        _log_line(
                            f"  🔀 [AVdb同步] tmdbid={tmdb_key} 冲突: {entry_name} 并入 {_entry_name_of_row(ws, target_row)}"
                        )
                    else:
                        result.filled += 1

                    if n % progress_interval == 0 or n == total:
                        _log_line(f"  📊 [AVdb同步] 进度 {n}/{total}")
                except Exception as e:
                    result.failed.append((actor.jp or actor.zh_cn or "<未知>", str(e)))
                    _log_line(f"  ❌ [AVdb同步] {actor.jp or actor.zh_cn or '<未知>'} 处理失败: {e}")

            _format_db_worksheet(ws)
            wb.save(db_path)
            wb.close()
    except Exception as e:
        result.failed.append(("<落盘>", str(e)))
        _log_line(f" ❌ [AVdb同步] 写入数据库失败: {e}")
        return result
    finally:
        if tmdb_session is not None:
            await tmdb_session.close()

    try:
        resources.reload_actor_db()
        from mdcx.core.tmdb_actor import _ACTOR_DB_ROW_INDEX, _ACTOR_DB_ROW_INDEX_LOCK

        with _ACTOR_DB_ROW_INDEX_LOCK:
            _ACTOR_DB_ROW_INDEX.clear()
    except Exception as e:
        _log_line(f" ⚠️ [AVdb同步] 重载内存缓存失败: {e}")

    _log_line(
        f" 🎬 [AVdb同步] 完成: 解析 {result.parsed}, 新建 {result.created}, 补齐 {result.filled}, "
        f"冲突合并 {result.merged}, 跳过男优 {result.skipped_male}, "
        f"丢弃错误id {result.skipped_tmdbid}, 失败 {len(result.failed)}"
    )
    return result


async def clean_male_actors(*, limit: int = 5000, concurrency: int = 5) -> CleanActorResult:
    """存量清洗：按内置男优名单 + TMDB gender 删除男优行。

    - 优先按内置男优名单 (resources/userdata/male_actors.txt) 命中判定，无 tmdbid
      的男优行也能删除。
    - 名单未命中的含 tmdbid 行再校验 TMDB gender；gender 1/0、请求失败、404 保留。
    - 删除前将男优行追加备份到独立「男优备份」sheet。
    - 支持 limit 限量与手动停止（signal.stop / Flags.stop_requested）。
    """
    from ..core.tmdb_actor import _ACTOR_DB_ROW_INDEX, _ACTOR_DB_ROW_INDEX_LOCK, _format_db_worksheet
    from ..models.flags import Flags
    from ..signals import signal

    def _is_stop_requested() -> bool:
        return signal.stop or Flags.stop_requested

    import openpyxl as _xl

    result = CleanActorResult()
    db_path = _get_db_path()
    base_url, tmdb_api_key = _resolve_tmdb_config()
    if not tmdb_api_key:
        _log_line(" ℹ️ [剔除男演员] 未配置 TMDB API Key，仅按内置名单清理男优")
    if not db_path.exists():
        _log_line(" ❌ [剔除男演员] actor_database.xlsx 不存在")
        return result

    try:
        async with _actor_db_write_lock:
            wb = _xl.load_workbook(db_path)
            ws = get_actor_db_sheet(wb)
            backup_name = "男优备份"
            if backup_name not in wb.sheetnames:
                wb.create_sheet(backup_name)
            backup_ws = wb[backup_name]

            # 收集行（仅前置扫描，不修改）。含 tmdbid 的行进入 TMDB 校验；
            # 无 tmdbid 的行若命中内置男优名单同样标记为男优。
            # 断点续传：已校验过 gender 的 tmdbid 记录在 sidecar 文件，重跑自动跳过
            checked_ids: set[int] = set()
            checked_file = db_path.parent / ".gender_checked.json"
            if checked_file.exists():
                try:
                    checked_ids = {int(x) for x in json.loads(checked_file.read_text(encoding="utf-8"))}
                except (ValueError, json.JSONDecodeError, OSError):
                    checked_ids = set()

            candidate_rows: list[tuple[int, int]] = []
            name_male_rows: set[int] = set()
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if _is_stop_requested():
                    break
                name = str(row[COL_JP] or row[COL_ZH_CN] or "").strip()
                if name and is_male_actor(name):
                    name_male_rows.add(row_idx)
                    continue
                # 别名列中的任意别名命中男优名单 → 同样标记删除
                # 仅对 ≥4 字的别名做匹配，短名(≤3字)做别名匹配易误杀女优
                kw_raw = str(row[COL_KEYWORD] or "").strip() if len(row) > COL_KEYWORD else ""
                if kw_raw and any(
                    len(k.strip()) >= 4 and is_male_actor(k.strip()) for k in kw_raw.split(",") if k.strip()
                ):
                    name_male_rows.add(row_idx)
                    continue
                if len(row) > COL_TMDBID:
                    tmdb_val = str(row[COL_TMDBID] or "").strip()
                    if tmdb_val.isdigit():
                        tmdbid_int = int(tmdb_val)
                        if tmdbid_int in checked_ids:
                            continue
                        candidate_rows.append((row_idx, tmdbid_int))
            if not tmdb_api_key:
                # 未配置 TMDB key：不校验 gender，所有含 tmdbid 行保留
                candidate_rows.clear()
            if limit and len(candidate_rows) > limit:
                candidate_rows = candidate_rows[:limit]
                _log_line(f" ℹ️ [剔除男演员] 本次限量处理前 {limit} 条，可再次运行继续")
            if name_male_rows:
                _log_line(f" 🎬 [剔除男演员] 名单命中男优 {len(name_male_rows)} 人，将直接删除")
            if not candidate_rows:
                _log_line(" ✅ [剔除男演员] 没有需要 TMDB 校验的 tmdbid 行")
                result.checked = len(name_male_rows)
                result.removed_male = len(name_male_rows)
                result.kept = 0
                if name_male_rows:
                    for row_idx in sorted(name_male_rows, reverse=True):
                        backup_ws.append([c.value for c in ws[row_idx]])
                        ws.delete_rows(row_idx)
                    _format_db_worksheet(ws)
                    wb.save(db_path)
                wb.close()
                return result

            _log_line(f" 🎬 [剔除男演员] 开始校验 {len(candidate_rows)} 个 tmdbid (并发 {concurrency})")

            # 阶段一：并发校验 gender，收集需删除的行
            male_row_indexes: set[int] = set()
            kept_indexes: set[int] = set()
            async with aiohttp.ClientSession() as client:
                task_iter = iter(enumerate(candidate_rows, 1))
                running_tasks: set[asyncio.Task[None]] = set()

                async def _check_one(seq, row_idx, tmdbid):
                    try:
                        gender = await fetch_person_gender(tmdbid, base_url, tmdb_api_key, client)
                        checked_ids.add(tmdbid)
                        if gender == 2:
                            male_row_indexes.add(row_idx)
                        else:
                            kept_indexes.add(row_idx)
                    except Exception:
                        kept_indexes.add(row_idx)
                        result.failed.append((str(tmdbid), "校验失败"))

                def _submit_next() -> bool:
                    try:
                        _, (row_idx, tmdbid) = next(task_iter)
                    except StopIteration:
                        return False
                    task = asyncio.create_task(_check_one(0, row_idx, tmdbid))
                    running_tasks.add(task)
                    return True

                for _ in range(min(concurrency, len(candidate_rows))):
                    _submit_next()

                total = len(candidate_rows)
                completed = 0
                progress_interval = max(1, total // 10)
                while running_tasks:
                    if _is_stop_requested():
                        for t in running_tasks:
                            t.cancel()
                    done, pending = await asyncio.wait(running_tasks, return_when=asyncio.FIRST_COMPLETED)
                    running_tasks = set(pending)
                    for _ in range(len(done)):
                        _submit_next()
                    for done_task in done:
                        completed += 1
                        try:
                            done_task.result()
                        except asyncio.CancelledError:
                            pass
                        except Exception as e:
                            _log_line(f"  🔴 [剔除男演员] 子任务异常: {e}")
                    if completed % progress_interval == 0 or completed == total:
                        _log_line(f"  📊 [剔除男演员] 进度: {completed}/{total}")

            # 阶段二：串行删除（降序删除避免行号漂移），删除前备份
            all_male_rows = male_row_indexes | name_male_rows

            # 持久化断点（已校验 gender 的 id 集合），便于限量分片续跑
            if checked_ids:
                try:
                    write_file_atomic(checked_file, json.dumps(sorted(checked_ids)), "utf-8")
                except OSError:
                    pass

            if _is_stop_requested():
                _log_line(" ⛔️ [剔除男演员] 已手动停止，未执行删除")
            else:
                for row_idx in sorted(all_male_rows, reverse=True):
                    if _is_stop_requested():
                        _log_line(" ⛔️ [剔除男演员] 删除阶段被停止，已处理部分")
                        break
                    backup_ws.append([c.value for c in ws[row_idx]])
                    ws.delete_rows(row_idx)
                result.removed_male = len(all_male_rows)

            result.checked = len(candidate_rows) + len(name_male_rows)
            result.kept = result.checked - len(all_male_rows)

            _format_db_worksheet(ws)
            wb.save(db_path)
            wb.close()
    except Exception as e:
        result.failed.append(("<落盘>", str(e)))
        _log_line(f" ❌ [剔除男演员] 失败: {e}")
        return result

    try:
        resources.reload_actor_db()
        with _ACTOR_DB_ROW_INDEX_LOCK:
            _ACTOR_DB_ROW_INDEX.clear()
    except Exception as e:
        _log_line(f" ⚠️ [剔除男演员] 重载内存缓存失败: {e}")

    _log_line(
        f" ✅ [剔除男演员] 完成: 校验 {result.checked}, 删除男优 {result.removed_male}, "
        f"保留 {result.kept}, 失败 {len(result.failed)}"
    )
    return result


async def verify_tmdb_ids(*, limit: int = 5000, concurrency: int = 5) -> VerifyTmdbIdResult:
    """存量校验 actor_database.xlsx 中所有 tmdbid 的有效性。

    TMDB 是公开平台，person id 可能被删除/重建/合并（如「三佳詩」旧 id 6231965 被
    TMDB 删除后重建为 5882313）。库中 id 是静态的、不会自愈，失效 id 被刮削直接采用
    会导致拿错误资料或 404。

    逻辑：
    - 扫描所有有 tmdbid 的行，并发调 TMDB person/{id} 校验
    - 404（person 已删除）-> 清除该行 tmdbid + tmdb url（回到无 id 状态，宁缺毋滥，
      刮削会按名字重新搜索）
    - 网络错误/限流/5xx -> 保守保留（不误清）
    - 支持 limit 限量与手动停止（signal.stop / Flags.stop_requested）
    """
    from ..core.tmdb_actor import (
        _ACTOR_DB_ROW_INDEX,
        _ACTOR_DB_ROW_INDEX_LOCK,
        _format_db_worksheet,
        _tmdb_rate_limiter,
    )
    from ..models.flags import Flags
    from ..signals import signal

    def _is_stop_requested() -> bool:
        return signal.stop or Flags.stop_requested

    import openpyxl as _xl

    result = VerifyTmdbIdResult()
    db_path = _get_db_path()
    base_url, tmdb_api_key = _resolve_tmdb_config()
    if not tmdb_api_key:
        _log_line(" ❌ [校验tmdbid] 未配置 TMDB API Key")
        return result
    if not db_path.exists():
        _log_line(" ❌ [校验tmdbid] actor_database.xlsx 不存在")
        return result

    try:
        async with _actor_db_write_lock:
            wb = _xl.load_workbook(db_path)
            ws = get_actor_db_sheet(wb)

            # 断点续传：已校验的 tmdbid 记录在 sidecar 文件中，重跑自动跳过
            verified_set: set[int] = set()
            verified_file = db_path.parent / ".tmdbid_verified.json"
            if verified_file.exists():
                try:
                    verified_set = {int(x) for x in json.loads(verified_file.read_text(encoding="utf-8"))}
                except (ValueError, json.JSONDecodeError, OSError):
                    verified_set = set()

            candidate_rows: list[tuple[int, int, str, str]] = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if _is_stop_requested():
                    break
                if len(row) > COL_TMDBID:
                    tmdb_val = str(row[COL_TMDBID] or "").strip()
                    if tmdb_val.isdigit():
                        tmdbid_int = int(tmdb_val)
                        if tmdbid_int in verified_set:
                            continue
                        jp = str(row[COL_JP] or "").strip() if len(row) > COL_JP else ""
                        zh = str(row[COL_ZH_CN] or "").strip() if len(row) > COL_ZH_CN else ""
                        candidate_rows.append((row_idx, tmdbid_int, jp, zh))
            if limit and len(candidate_rows) > limit:
                candidate_rows = candidate_rows[:limit]
                _log_line(f" ℹ️ [校验tmdbid] 本次限量处理前 {limit} 条，可再次运行继续")
            if not candidate_rows:
                _log_line(" ✅ [校验tmdbid] 没有需要校验的 tmdbid")
                wb.close()
                return result

            _log_line(f" 🎬 [校验tmdbid] 开始校验 {len(candidate_rows)} 个 tmdbid (并发 {concurrency})")

            invalid_rows: dict[int, int] = {}
            result.checked = len(candidate_rows)

            async with aiohttp.ClientSession() as client:
                task_iter = iter(candidate_rows)
                running_tasks: set[asyncio.Task[None]] = set()

                async def _check(row_idx: int, tmdbid: int) -> None:
                    try:
                        url = f"{base_url}/3/person/{tmdbid}?api_key={tmdb_api_key}"
                        await _tmdb_rate_limiter.acquire()
                        status_code = 0
                        try:
                            async with client.get(url, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                                status_code = resp.status
                                if resp.status == 404:
                                    invalid_rows[row_idx] = tmdbid
                                    if len(invalid_rows) <= 5:
                                        _log_line(f"  ⚠️ [校验tmdbid] 行{row_idx} tmdbid={tmdbid} 已失效(404)")
                                # 200/404 均视为校验完成，记录断点；网络错误(异常)不记录以便重试
                                if resp.status in (200, 404):
                                    verified_set.add(tmdbid)
                        finally:
                            await _tmdb_rate_limiter.finish(status_code)
                    except Exception:
                        return  # 网络失败，保守保留

                def _submit_next() -> bool:
                    try:
                        row_idx, tmdbid, _, _ = next(task_iter)
                    except StopIteration:
                        return False
                    task = asyncio.create_task(_check(row_idx, tmdbid))
                    running_tasks.add(task)
                    return True

                for _ in range(min(concurrency, len(candidate_rows))):
                    _submit_next()

                total = len(candidate_rows)
                completed = 0
                progress_interval = max(1, total // 10)
                while running_tasks:
                    if _is_stop_requested():
                        for t in running_tasks:
                            t.cancel()
                    done, pending = await asyncio.wait(running_tasks, return_when=asyncio.FIRST_COMPLETED)
                    running_tasks = set(pending)
                    for _ in range(len(done)):
                        if not _is_stop_requested():
                            _submit_next()
                    for done_task in done:
                        completed += 1
                        try:
                            done_task.result()
                        except asyncio.CancelledError:
                            pass
                        except Exception as e:
                            _log_line(f"  🔴 [校验tmdbid] 子任务异常: {e}")
                    if completed % progress_interval == 0 or completed == total:
                        _log_line(f"  📊 [校验tmdbid] 进度: {completed}/{total}")

            # 持久化断点（校验完成的 id 集合），便于限量分片续跑
            if verified_set:
                try:
                    write_file_atomic(verified_file, json.dumps(sorted(verified_set)), "utf-8")
                except OSError:
                    pass

            if invalid_rows:
                # 清除失效 id 前，记录待补回的行信息（jp/zh）
                invalid_meta = {row_idx: (jp, zh) for row_idx, tid, jp, zh in candidate_rows if row_idx in invalid_rows}
                for row_idx in sorted(invalid_rows):
                    ws.cell(row=row_idx, column=COL_TMDBID + 1).value = None
                    ws.cell(row=row_idx, column=COL_TMDB_URL + 1).value = None
                _format_db_worksheet(ws)
                wb.save(db_path)
                _log_line(f" 🗑️ [校验tmdbid] 清除失效 id {len(invalid_rows)} 个")

                # 按名字重搜补回新 id（TMDB 可能重建了档案）；并发查询提高恢复速度
                from mdcx.core.tmdb_actor import query_single_actor_cached

                recovered = 0
                sem = asyncio.Semaphore(5)

                async def _recover(row_idx: int, jp: str, zh: str) -> int:
                    if not jp and not zh:
                        return 0
                    query_name = jp or zh
                    try:
                        qr = await query_single_actor_cached(query_name, base_url, tmdb_api_key, client)
                        if qr and qr.get("adult") and qr.get("pid"):
                            ws.cell(row=row_idx, column=COL_TMDBID + 1).value = int(qr["pid"])
                            ws.cell(row=row_idx, column=COL_TMDB_URL + 1).value = _tmdb_person_url(int(qr["pid"]))
                            _log_line(f"  🔁 [校验tmdbid] {jp or zh} 补回新 id {qr['pid']}")
                            return 1
                    except Exception:
                        pass
                    return 0

                async def _limited(args):
                    async with sem:
                        return await _recover(*args)

                async with aiohttp.ClientSession() as client:
                    recover_items = [
                        (row_idx, jp, zh) for row_idx, (jp, zh) in invalid_meta.items() if not _is_stop_requested()
                    ]
                    results = await asyncio.gather(*(_limited(a) for a in recover_items), return_exceptions=True)
                    recovered = sum(r for r in results if isinstance(r, int))
                if recovered:
                    _format_db_worksheet(ws)
                    wb.save(db_path)
                    _log_line(f" 🔁 [校验tmdbid] 按名重搜补回 {recovered} 个新 id")
                result.recovered = recovered
            wb.close()
            result.invalid = len(invalid_rows)
            result.valid = len(candidate_rows) - len(invalid_rows)
    except Exception as e:
        result.failed.append(("<落盘>", str(e)))
        _log_line(f" ❌ [校验tmdbid] 失败: {e}")
        return result

    try:
        resources.reload_actor_db()
        with _ACTOR_DB_ROW_INDEX_LOCK:
            _ACTOR_DB_ROW_INDEX.clear()
    except Exception as e:
        _log_line(f" ⚠️ [校验tmdbid] 重载内存缓存失败: {e}")

    if _is_stop_requested():
        _log_line(
            f" ⛔️ [校验tmdbid] 已手动停止：校验 {completed}/{result.checked}，失效清除 {result.invalid}，可再次运行继续"
        )
    else:
        _log_line(
            f" ✅ [校验tmdbid] 完成: 校验 {result.checked}, 失效清除 {result.invalid}, "
            f"有效 {result.valid}, 失败 {len(result.failed)}"
        )
    return result


def _update_nfo_tmdbids_text(text: str, id_map: dict[str, int]) -> tuple[str, int]:
    """文本级替换 nfo 中 actor 的 tmdbid。

    对每个 <actor> 块，按 <name> 匹配 id_map 中的新 id：
    - 块内已有 <tmdbid> 且与新 id 不同 -> 替换为新 id
    - 块内无 <tmdbid> 但库有 id -> 在 </type> 后插入 <tmdbid>（补上缺失的 id）
    仅改动 tmdbid 值，保留 nfo 其他所有内容与格式。

    返回 (更新后的文本, 更新的 actor 数)。
    """
    import re

    def _norm(s: str) -> str:
        return re.sub(r"\s+", "", s or "").strip()

    updated = 0

    def _replace_actor(match):
        nonlocal updated
        block = match.group(0)
        name_m = re.search(r"<name>(.*?)</name>", block, re.S)
        if not name_m:
            return block
        name = name_m.group(1).strip()
        new_id = id_map.get(name) or id_map.get(_norm(name))
        if new_id is None:
            return block
        tmdb_m = re.search(r"<tmdbid>\s*(\d+)\s*</tmdbid>", block)
        if tmdb_m:
            if int(tmdb_m.group(1)) == int(new_id):
                return block
            block = block[: tmdb_m.start(1)] + str(new_id) + block[tmdb_m.end(1) :]
            updated += 1
        else:
            type_m = re.search(r"</type>", block)
            if type_m:
                insert_at = type_m.end()
                block = block[:insert_at] + f"\n    <tmdbid>{new_id}</tmdbid>" + block[insert_at:]
            else:
                insert_at = name_m.end()
                block = block[:insert_at] + f"\n    <tmdbid>{new_id}</tmdbid>" + block[insert_at:]
            updated += 1
        return block

    new_text = re.sub(r"<actor>.*?</actor>", _replace_actor, text, flags=re.S)
    return new_text, updated


async def update_nfo_tmdb_ids(dir_path: Path, *, limit: int = 5000, concurrency: int = 5) -> "UpdateNfoTmdbIdResult":
    """批量更新指定目录下所有 nfo 中 actor 的 tmdbid。

    对每个 nfo：
    1. 读取文本，解析所有 <actor> 的 name
    2. 用本地演员库（search_actor_db_reverse）查每个 actor 的当前 tmdbid
    3. 文本级替换/补入 <tmdbid>（仅改该值，保留 nfo 其他内容）
    4. 有变更则写回

    支持 limit 限量与手动停止（signal.stop / Flags.stop_requested）。
    """
    from ..core.tmdb_actor import search_actor_db_reverse
    from ..models.flags import Flags
    from ..signals import signal

    def _is_stop_requested() -> bool:
        return signal.stop or Flags.stop_requested

    result = UpdateNfoTmdbIdResult()
    if not await aiofiles.os.path.isdir(dir_path):
        _log_line(f" ❌ [更新nfo] 目录不存在: {dir_path}")
        return result

    nfo_files: list[Path] = []

    async def _walk(current: Path) -> None:
        if _is_stop_requested():
            return
        try:
            entries = await aiofiles.os.scandir(current)
        except OSError:
            return
        for entry in entries:
            if _is_stop_requested():
                return
            try:
                if entry.is_dir(follow_symlinks=True):
                    await _walk(Path(entry.path))
                elif entry.name.lower().endswith(".nfo"):
                    nfo_files.append(Path(entry.path))
            except OSError:
                continue

    await _walk(dir_path)
    if limit and len(nfo_files) > limit:
        nfo_files = nfo_files[:limit]
        _log_line(f" ℹ️ [更新nfo] 本次限量处理前 {limit} 个 nfo，可再次运行继续")
    if not nfo_files:
        _log_line(" ✅ [更新nfo] 目录下没有 nfo 文件")
        return result

    _log_line(f" 🎬 [更新nfo] 开始处理 {len(nfo_files)} 个 nfo (并发 {concurrency})")
    result.checked = len(nfo_files)

    sem = asyncio.Semaphore(concurrency)

    async def _process_one(nfo_path: Path) -> None:
        async with sem:
            if _is_stop_requested():
                return
            try:
                async with aiofiles.open(nfo_path, encoding="utf-8") as f:
                    content = await f.read()
            except (OSError, UnicodeDecodeError):
                return
            # 解析 actor names
            import re as _re

            names = _re.findall(r"<name>(.*?)</name>", content, _re.S)
            names = [n.strip() for n in names if n and n.strip()]
            if not names:
                return
            # 查库拿 id（并发/顺序均可，量小）
            id_map: dict[str, int] = {}
            for nm in names:
                if _is_stop_requested():
                    return
                row = search_actor_db_reverse(nm)
                raw_tmdbid = row.get("tmdbid") if row else None
                if raw_tmdbid in (None, ""):
                    continue
                try:
                    id_map[nm] = int(str(raw_tmdbid).strip())
                except (TypeError, ValueError):
                    continue
            if not id_map:
                return
            new_content, cnt = _update_nfo_tmdbids_text(content, id_map)
            if cnt:
                try:
                    await write_file_atomic_async(nfo_path, new_content)
                    result.updated_files += 1
                    result.updated_actors += cnt
                    _log_line(f"  ✅ [更新nfo] {nfo_path.name}: 更新 {cnt} 个 actor tmdbid")
                except OSError:
                    result.failed.append((str(nfo_path), "写入失败"))
            else:
                result.no_change += 1

    tasks = [asyncio.create_task(_process_one(p)) for p in nfo_files]
    await asyncio.gather(*tasks)

    _log_line(
        f" ✅ [更新nfo] 完成: 检查 {result.checked}, 更新 {result.updated_files} 个文件/"
        f"{result.updated_actors} 个 actor, 无变化 {result.no_change}, 失败 {len(result.failed)}"
    )
    return result
