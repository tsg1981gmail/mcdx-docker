#!/usr/bin/env python3
"""tenhow 增量行抽查：软校验法验证图片与番号匹配度。

背景：amazon_asin_database.xlsx 中 31,982 条 tenhow 增量行（标题列含
"tenhow" 标记、无封面 URL），其番号↔ASIN 映射源自 tenhow 页面的
cid 绑定，未经图像验证。本脚本随机抽样，用图像相似度软校验
（阈值 0.82）抽验映射正确率。

裁判链（与生产 web.py::_load_dmm_official_reference 严格同源，勿改回
thejavdb api/直构枚举的旧链）:
1. DMM 高清直链竖版 ps (build_aws_poster_candidates) —— 首选裁判
2. DMM 横版 pl (build_aws_cover_candidates) + 右半裁剪 —— 降级
3. javdb_app 搜索 thumb_url —— 兜底（经 upgrade_dmm_cover 同源的
   DMM 候选判定，thumb 本身即 DMM 图的重压版，仍可与 DMM 直链互证）

基准图: tenhow.net/images/{ASIN}.jpg (<4KB 占位图拒收)。
相似度 >= 0.82 判定匹配（与 _verify_soft_amazon_poster 同阈值）。

用法:
    python scripts/sample_check_tenhow.py --sample 30            # 随机抽 30 条
    python scripts/sample_check_tenhow.py --sample 30 --seed 42  # 固定种子复现
"""

from __future__ import annotations

import argparse
import asyncio
import random
import sys
from io import BytesIO
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import httpx  # noqa: E402
import openpyxl  # noqa: E402
from PIL import Image  # noqa: E402

from mdcx.core.web import _cover_similarity, _cut_thumb_right_image  # noqa: E402
from mdcx.crawlers.dmm_direct import (  # noqa: E402
    build_aws_cover_candidates,
    build_aws_poster_candidates,
)

DB_PATH = ROOT / "userdata" / "amazon_asin_database.xlsx"
TENHOW_URL = "https://tenhow.net/images/{asin}.jpg"
SCORE_THRESHOLD = 0.82
PLACEHOLDER_MAX_BYTES = 4096
JAVDB_APP_SEARCH = "https://javdb573.com/api/v2/search"

_http: httpx.AsyncClient | None = None


async def _fetch_image(url: str) -> tuple[Image.Image | None, int]:
    global _http
    if _http is None:
        _http = httpx.AsyncClient(timeout=20, follow_redirects=True, headers={"User-Agent": "Mozilla/5.0"})
    try:
        r = await _http.get(url)
        if r.status_code != 200 or len(r.content) < PLACEHOLDER_MAX_BYTES:
            return None, len(r.content)
        img = Image.open(BytesIO(r.content))
        img.load()
        return img, len(r.content)
    except Exception:
        return None, 0


async def _dmm_reference(number: str) -> tuple[Image.Image | None, str]:
    """DMM 高清直链裁判（与生产 _load_dmm_official_reference 同源）。"""
    for url in build_aws_poster_candidates(number):
        img, _sz = await _fetch_image(url)
        if img is not None:
            return img, "dmm-ps"
    for url in build_aws_cover_candidates(number):
        img, _sz = await _fetch_image(url)
        if img is None:
            continue
        cropped = _cut_thumb_right_image(img)
        if cropped is not None:
            return cropped, "dmm-pl-crop"
    return None, ""


async def _javdb_app_thumb(number: str) -> tuple[Image.Image | None, str]:
    """javdb_app 搜索兜底（与判定脚本 check_asin_poster_reachability 同口径：
    /api/v2/search 按 number 匹配 thumb_url）。"""
    global _http
    if _http is None:
        _http = httpx.AsyncClient(timeout=20, follow_redirects=True)

    from mdcx.crawlers.javdb_app import (  # noqa: E402
        _API_BASE,
        _API_FALLBACKS,
        _get_api_url,
        make_signature,
    )

    headers = {
        "jdsignature": make_signature(),
        "accept-language": "zh",
        "User-Agent": "Dart/3.5 (dart:io)",
    }

    def norm(s: str) -> tuple:
        import re

        m = re.match(r"^([a-z0-9]+)[-_.]?(\d+)$", s.lower().replace(" ", ""))
        return (m.group(1), int(m.group(2))) if m else ("", -1)

    key = norm(number)
    poster_url = None
    for host in [_API_BASE, *_API_FALLBACKS]:
        try:
            url = _get_api_url(host, "/api/v2/search", {"q": number, "page": "1"})
            r = await _http.get(url, headers=headers, timeout=20)
            if r.status_code != 200:
                continue
            data = r.json()
            if not data.get("success"):
                continue
            movies = (data.get("data") or {}).get("movies") or []
            for mv in movies:
                kn = norm(str(mv.get("number") or ""))
                if key[1] >= 0 and kn[1] >= 0 and kn == key:
                    poster_url = str(mv.get("thumb_url") or "").strip()
                    break
            if poster_url:
                break
        except Exception:
            continue
    if not poster_url:
        return None, ""
    img, _sz = await _fetch_image(poster_url)
    if img is None:
        return None, ""
    return img, "javdb-thumb"


def _normalize_for_score(img: Image.Image) -> Image.Image:
    """横版图裁右半（与软校验同几何）。"""
    if img.size[0] > img.size[1]:
        cut = _cut_thumb_right_image(img)
        if cut is not None:
            return cut
    return img


def _score(base: Image.Image, ref: Image.Image) -> tuple[float, float, float]:
    """与生产软校验同口径：返回 (score, hash_sim, hist_sim) 三元组。

    生产判定（_verify_soft_amazon_poster）为三阈值同时满足：
        score >= 0.82 AND hash >= 0.86 AND hist >= 0.70
    """
    return _cover_similarity(_normalize_for_score(base), _normalize_for_score(ref))


def _verdict_strict(score: float, hash_sim: float, hist_sim: float) -> str:
    """生产同款三阈值严格判定。"""
    if score >= SCORE_THRESHOLD and hash_sim >= 0.86 and hist_sim >= 0.70:
        return "匹配"
    return "疑似错配"


async def _libredmm_reference(number: str) -> tuple[Image.Image | None, str]:
    """libredmm 裁判：详情页 /movies/{番号} 直接给 DMM 大图 URL（真实 cid）。

    下载到的就是 DMM 原图，可作严格裁判——绕过番号→cid 猜测，
    对 DMM 直链猜不中的下架/变体番号特别有效。
    """
    global _http
    if _http is None:
        _http = httpx.AsyncClient(timeout=20, follow_redirects=True)
    try:
        url = f"https://www.libredmm.com/movies/{number.lower()}"
        r = await _http.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=25)
        if r.status_code != 200:
            return None, ""
        html = r.text
        # 复刻 libredmm.py::get_cover / get_poster 的 xpath
        import re

        # cover: //div[@class="col-md-8"]/img/@src
        m = re.search(r'<div class="col-md-8">\s*<img[^>]+src="([^"]+)"', html)
        if not m:
            # poster: //dt[text()="Thumbnail Image"]/following-sibling::dd[1]/img/@src
            m = re.search(r'<dt>Thumbnail Image</dt>\s*<dd[^>]*>\s*<img[^>]+src="([^"]+)"', html)
        if not m:
            return None, ""
        img_url = m.group(1)
        if img_url.startswith("//"):
            img_url = "https:" + img_url
        img, _sz = await _fetch_image(img_url)
        if img is None:
            return None, ""
        return img, "libredmm"
    except Exception:
        return None, ""


async def _r18dev_reference(number: str) -> tuple[Image.Image | None, str]:
    """r18.dev 裁判：API 搜番号拿 jacket_full_url（pics.dmm 低清直链）。

    图片本身来自 DMM 图床（原路径），与 DMM 直链同源只是分辨率低，
    可作严格裁判（软校验几何会缩放到统一尺寸比对）。
    """
    global _http
    if _http is None:
        _http = httpx.AsyncClient(timeout=20, follow_redirects=True)
    try:
        r = await _http.get(
            "https://r18.dev/api/v4.fmc/movies",
            params={"search": number},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=25,
        )
        if r.status_code != 200:
            return None, ""
        data = r.json()
        items = data.get("data") or []
        if not items:
            return None, ""
        # 番号精确匹配（r18 搜索可能返回相近条目）
        import re

        def norm(s):
            m2 = re.match(r"^([a-z0-9]+)[-_.]?(\d+)$", str(s or "").lower().replace(" ", ""))
            return (m2.group(1), int(m2.group(2))) if m2 else ("", -1)

        target = norm(number)
        img_url = None
        for mv in items:
            if norm(mv.get("id")) == target:
                img_url = str(mv.get("jacket_full_url") or "").strip()
                break
        if not img_url:
            return None, ""
        img, _sz = await _fetch_image(img_url)
        if img is None:
            return None, ""
        return img, "r18dev"
    except Exception:
        return None, ""


async def _javbus_reference(number: str) -> tuple[Image.Image | None, str]:
    """javbus 镜像裁判：搜索页直接给 cover（重压横幅图）。

    重压图与 DMM 原图相似度仅 0.5~0.7，只作低置信度参考（与 javdb
    thumb 兜底同规格）。镜像池按 javbus.py::_JAVBUS_DOMAINS 顺序轮换。
    """
    from mdcx.crawlers.javbus import _JAVBUS_DOMAINS

    global _http
    if _http is None:
        _http = httpx.AsyncClient(timeout=20, follow_redirects=True)
    try:
        import re as _re

        key = _re.match(r"^([A-Za-z0-9]+)-?(\d+)$", number)
        if not key:
            return None, ""
        base_num = number
        for domain in _JAVBUS_DOMAINS:
            try:
                search = f"{domain}/search/{base_num}&type=1"
                r = await _http.get(search, headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
                if r.status_code != 200:
                    continue
                # 封面缩略: <img class="img" ... src="...">
                m = _re.search(r'<a[^>]+href="([^"]+/\w+)"[^>]*>\s*<img[^>]+src="([^"]+)"', r.text)
                if not m:
                    continue
                cover = m.group(2)
                if cover.startswith("//"):
                    cover = "https:" + cover
                img, _sz = await _fetch_image(cover)
                if img is not None:
                    return img, f"javbus({domain.split('//')[1]})"
            except Exception:
                continue
    except Exception:
        pass
    return None, ""


async def main(sample: int, seed: int, only: list[tuple[str, str]] | None = None) -> int:
    wb = openpyxl.load_workbook(DB_PATH, read_only=True)
    ws = wb["Sheet"]
    rows = ws.iter_rows(values_only=True)
    next(rows)

    tenhow_rows = []
    for row in rows:
        if not row or not row[1]:
            continue
        number = str(row[0] or "").strip()
        asin = str(row[1] or "").strip()
        title = str(row[3] or "")
        poster = str(row[4] or "")
        if number and asin and not poster and "tenhow" in title.lower():
            tenhow_rows.append((number, asin))
    wb.close()

    total_tenhow = len(tenhow_rows)
    if only:
        picks = only
        print(f"指定补判: {len(picks)} 条 (tenhow 总行 {total_tenhow})\n")
    else:
        rng = random.Random(seed)
        picks = rng.sample(tenhow_rows, min(sample, total_tenhow))
        print(f"tenhow 总行: {total_tenhow}, 随机抽样: {len(picks)} (seed={seed})\n")

    sem = asyncio.Semaphore(3)

    async def check(number: str, asin: str) -> dict:
        async with sem:
            base, base_sz = await _fetch_image(TENHOW_URL.format(asin=asin))
            if base is None:
                return {
                    "number": number,
                    "asin": asin,
                    "verdict": "无基准图",
                    "score": None,
                    "ref": "",
                    "note": f"tenhow 图不可取({base_sz}B)",
                }
            # 主裁判：DMM 高清直链（与生产 _load_dmm_official_reference 同源）
            ref, ref_src = await _dmm_reference(number)
            if ref is None:
                # 严格裁判扩展：libredmm（页面给真实 DMM 大图 URL）
                ref, ref_src = await _libredmm_reference(number)
            if ref is None:
                # 严格裁判扩展：r18.dev（jacket_full_url 为 pics.dmm 直链）
                ref, ref_src = await _r18dev_reference(number)
            if ref is not None:
                score, hash_sim, hist_sim = _score(base, ref)
                verdict = _verdict_strict(score, hash_sim, hist_sim)
                return {
                    "number": number,
                    "asin": asin,
                    "verdict": verdict,
                    "score": round(score, 3),
                    "hash": round(hash_sim, 3),
                    "hist": round(hist_sim, 3),
                    "ref": ref_src,
                    "note": "",
                }
            # 低置信兜底 1：javdb_app thumb（重压图）
            ref, ref_src = await _javdb_app_thumb(number)
            # 低置信兜底 2：javbus 镜像 cover（同为重压图，与 javdb 同规格）
            if ref is None:
                ref, ref_src = await _javbus_reference(number)
            if ref is not None:
                score, hash_sim, hist_sim = _score(base, ref)
                low_conf = score >= 0.60 and hash_sim >= 0.62  # 重压图放宽阈值
                return {
                    "number": number,
                    "asin": asin,
                    "verdict": "低置信参考" if low_conf else "疑似错配",
                    "score": round(score, 3),
                    "hash": round(hash_sim, 3),
                    "hist": round(hist_sim, 3),
                    "ref": ref_src,
                    "note": "重压图兜底(javdb/javbus), 不计严格判定",
                }
            return {
                "number": number,
                "asin": asin,
                "verdict": "无裁判图",
                "score": None,
                "ref": "",
                "note": f"{number} DMM/javdb 裁判图均不可取",
            }

    results = []
    for i, (number, asin) in enumerate(picks, 1):
        r = await check(number, asin)
        results.append(r)
        extra = f" hash={r.get('hash')} hist={r.get('hist')}" if r.get("hash") is not None else ""
        print(
            f"[{i:2d}/{len(picks)}] {r['number']:<14s} {r['asin']:<12s} {r['verdict']:6s} "
            f"score={r['score']}{extra} ref={r['ref']} {r['note']}"
        )

    n_match = sum(1 for r in results if r["verdict"] == "匹配")
    n_mismatch = sum(1 for r in results if r["verdict"] == "疑似错配")
    n_low = sum(1 for r in results if r["verdict"].startswith("低置信"))
    n_no_base = sum(1 for r in results if r["verdict"] == "无基准图")
    n_no_ref = sum(1 for r in results if r["verdict"] == "无裁判图")
    judged = n_match + n_mismatch

    print("\n" + "=" * 60)
    print(f"抽样 {len(results)} 条:")
    print(f"  匹配(DMM直链严格判定): {n_match}")
    print(f"  疑似错配:              {n_mismatch}")
    print(f"  低置信参考(javdb兜底): {n_low} (重压图, 不计严格判定)")
    print(f"  无基准图:              {n_no_base} (tenhow 无此 ASIN 图, 运行时回退日亚)")
    print(f"  无裁判图:              {n_no_ref} (DMM/javdb 均无此番号)")
    if judged:
        print(f"  严格判定样本中匹配率: {n_match}/{judged} = {n_match / judged * 100:.1f}%")
    if n_mismatch:
        print("\n疑似错配清单(需人工复核):")
        for r in results:
            if r["verdict"] == "疑似错配":
                print(f"  {r['number']} <-> {r['asin']} (score={r['score']}, hash={r.get('hash')}, ref={r['ref']})")

    # 结果落盘（长任务规范：批间落盘，tail 截断不再丢数据）
    import json

    out = Path("/tmp/opencode/sample_check_result.json")
    out.write_text(
        json.dumps({"seed": seed, "count": len(results), "results": results}, ensure_ascii=False, indent=1),
        encoding="utf-8",
    )
    print(f"\n结果已落盘: {out}")
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--sample", type=int, default=30, help="抽样数量")
    parser.add_argument("--seed", type=int, default=42, help="随机种子(复现)")
    parser.add_argument("--only", type=str, default="", help="只补判指定番号, 逗号分隔, 如 MKMP-017,YST-193")
    args = parser.parse_args()
    only_pairs = None
    if args.only:
        want = {n.strip().upper() for n in args.only.split(",") if n.strip()}
        # ASIN 从库里按番号找
        wb2 = openpyxl.load_workbook(DB_PATH, read_only=True)
        ws2 = wb2["Sheet"]
        rows2 = ws2.iter_rows(values_only=True)
        next(rows2)
        only_pairs = []
        for row in rows2:
            if not row or not row[1]:
                continue
            if str(row[0] or "").strip().upper() in want:
                only_pairs.append((str(row[0] or "").strip(), str(row[1] or "").strip()))
        wb2.close()
    raise SystemExit(asyncio.run(main(args.sample, args.seed, only_pairs)))
