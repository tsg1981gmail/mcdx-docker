#!/usr/bin/env python3
"""
演员数据库静态校验脚本。

仅检查仓库内出厂 `resources/userdata/actor_database.xlsx`，把关本地数据质量。
检查项：
  1. jp 空字段（主键必须非空）         [error]
  2. 同 jp 名重复（大小写不敏感）      [error]
  3. keyword 首尾逗号 / 连续逗号       [error]
  4. keyword 重复词（大小写不敏感）    [error]
  5. zh_cn / zh_tw 空字段              [warning]
  6. tmdbid 重复                       [error]
  7. 出生日期列格式（空或 YYYY[-MM[-DD]]） [error]
  8. 出生日期年份超合理范围(<1900 或 >2030) [error]
  9. 生涯字段无任何 4 位年份(含全角转半角)  [error]
  10. 简介非出道字段含日文假名(出道作品标题保留) [warning]
  11. 简介非结构化(有文本但无标准字段段)   [warning]
  12. tmdbid 空但 tmdb url 有值（错配）   [error]
  13. tmdbid 与 tmdb url 不匹配（url 非标准格式） [error]
  14. tmdb url 重复（同一 url 多行）     [error]
  15. 孤儿 hyperlink（引用不存在的单元格） [error]

发现任一 error 返回码 1；仅 warning 返回码 0。老 7 列文件缺失新增列时跳过对应检查。

注意：openpyxl 的 read_only 模式只按实际 XML 读取，孤儿 hyperlink（引用不存在单元格的
超链接）会被物化成幽灵数据或撑大文件维度，read_only 校验恰好绕过。因此第 11 项直接
解析 sheet XML 检测，确保普通/只读两种读取路径下数据一致。
"""

import argparse
import re
import sys
import zipfile
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover
    sys.stderr.write("缺少依赖 openpyxl，请先 uv sync\n")
    sys.exit(2)

MAIN_PATH = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = MAIN_PATH / "resources" / "userdata" / "actor_database.xlsx"

ACTOR_DB_SHEET = "演员数据库"

BIRTH_DATE_PATTERN = re.compile(r"^\d{4}(-\d{1,2}(-\d{1,2})?)?$")


def _check_jp_empty(rows):
    errors = []
    for idx, row in enumerate(rows, 2):
        jp = str(row[0] or "").strip()
        if not jp:
            errors.append(f"  行{idx}: jp(日文原名) 为空")
    return errors


def _check_jp_duplicate(rows):
    seen: dict[str, int] = {}
    errors = []
    for idx, row in enumerate(rows, 2):
        jp = str(row[0] or "").strip().casefold()
        if not jp:
            continue
        if jp in seen:
            errors.append(f"  行{idx}: jp 与行{seen[jp]} 重复: {row[0]}")
        else:
            seen[jp] = idx
    return errors


def _check_keyword_format(rows):
    errors = []
    for idx, row in enumerate(rows, 2):
        kw = str(row[3] or "").strip()
        if not kw:
            continue
        if kw.startswith(",") or kw.endswith(",") or ",," in kw:
            errors.append(f"  行{idx}: keyword 存在首尾/连续逗号: {kw}")
    return errors


def _check_keyword_duplicate(rows):
    errors = []
    for idx, row in enumerate(rows, 2):
        kw = str(row[3] or "").strip()
        if not kw:
            continue
        parts = [k.strip() for k in kw.split(",") if k.strip()]
        if len(parts) != len({k.casefold() for k in parts}):
            errors.append(f"  行{idx}: keyword 存在重复词: {kw}")
    return errors


def _check_name_empty(rows):
    warnings = []
    for idx, row in enumerate(rows, 2):
        zh_cn = str(row[1] or "").strip() if len(row) > 1 else ""
        zh_tw = str(row[2] or "").strip() if len(row) > 2 else ""
        jp = str(row[0] or "").strip()
        if jp and not zh_cn:
            warnings.append(f"  行{idx}: zh_cn(中文名) 为空")
        if jp and not zh_tw:
            warnings.append(f"  行{idx}: zh_tw(繁体名) 为空")
    return warnings


def _check_tmdbid_duplicate(rows):
    seen: dict[str, int] = {}
    errors = []
    for idx, row in enumerate(rows, 2):
        if len(row) <= 5:
            continue
        tmdb = str(row[5] or "").strip()
        if not tmdb or not tmdb.isdigit():
            continue
        if tmdb in seen:
            errors.append(f"  行{idx}: tmdbid 与行{seen[tmdb]} 重复: {tmdb}")
        else:
            seen[tmdb] = idx
    return errors


def _check_birth_date(rows):
    errors = []
    for idx, row in enumerate(rows, 2):
        if len(row) <= 7:
            return []  # 老文件无出生日期列，跳过
        birth = str(row[7] or "").strip()
        if birth and not BIRTH_DATE_PATTERN.match(birth):
            errors.append(f"  行{idx}: 出生日期格式非法(期望 YYYY[-MM[-DD]]): {birth}")
    return errors


def _check_birth_date_range(rows):
    """出生日期年份超出合理范围（<1900 或 >2030），疑似填写错误。"""
    errors = []
    for idx, row in enumerate(rows, 2):
        if len(row) <= 7:
            continue
        birth = str(row[7] or "").strip()
        m = re.match(r"(\d{4})", birth)
        if m:
            year = int(m.group(1))
            if year < 1900 or year > 2030:
                errors.append(f"  行{idx}: 出生日期年份异常({year}, 期望 1900-2030): {birth}")
    return errors


def _check_career_no_year(rows):
    """生涯字段无任何 4 位年份（含全角数字转半角后判定），疑似非年份内容。"""
    errors = []
    _FULL2HALF = str.maketrans("０１２３４５６７８９", "0123456789")
    for idx, row in enumerate(rows, 2):
        if len(row) <= 8:
            continue
        bio = str(row[8] or "").strip()
        if not bio:
            continue
        for seg in bio.split("|"):
            seg = seg.strip()
            m = re.match(r"^生涯\s*[:：]\s*(.*)$", seg)
            if not m:
                continue
            value = m.group(1).strip()
            if not value:
                continue
            half = value.translate(_FULL2HALF)
            if not re.search(r"(?:19|20)\d{2}", half):
                errors.append(f"  行{idx}: 生涯无年份(非年份区间): {value}")
    return errors


def _check_bio_jp_residual(rows):
    """简介非出道字段含日文假名（出道作品标题保留日文，不检查）。"""
    warnings = []
    for idx, row in enumerate(rows, 2):
        if len(row) <= 8:
            continue
        bio = str(row[8] or "").strip()
        if not bio:
            continue
        for seg in bio.split("|"):
            seg = seg.strip()
            m = re.match(r"^([^\s:：]+)\s*[:：]\s*(.*)$", seg)
            if not m:
                continue
            field, value = m.group(1).strip(), m.group(2).strip()
            if field == "出道":
                continue
            if value and re.search(r"[ぁ-んァ-ヶ]", value):
                warnings.append(f"  行{idx}: 简介 {field} 字段含日文: {value[:30]}")
    return warnings


def _check_bio_unstructured(rows):
    """简介非结构化：有文本但无任何标准字段段（身高/罩杯/三围等）。"""
    warnings = []
    _FIELD_RE = re.compile(r"身高|罩杯|三围|生涯|出身|血型|事务所|爱好|出道|标签")
    for idx, row in enumerate(rows, 2):
        if len(row) <= 8:
            continue
        bio = str(row[8] or "").strip()
        if bio and not _FIELD_RE.search(bio):
            warnings.append(f"  行{idx}: 简介非结构化(无标准字段段): {bio[:40]}")
    return warnings


def _check_tmdb_url_no_id(rows):
    """tmdbid 为空但 tmdb url 有值：url 错配（url 指向的人物与行无关）。"""
    errors = []
    for idx, row in enumerate(rows, 2):
        if len(row) <= 6:
            return []
        jp = str(row[0] or "").strip()
        tid = str(row[5] or "").strip()
        url = str(row[6] or "").strip()
        if jp and not tid and url:
            errors.append(f"  行{idx}: tmdbid 为空但 tmdb url 有值（url 错配）: {url}")
    return errors


def _check_tmdb_url_mismatch(rows):
    """tmdbid 与 tmdb url 不匹配（url 不是该 id 的标准 person url）。"""
    errors = []
    for idx, row in enumerate(rows, 2):
        if len(row) <= 6:
            return []
        jp = str(row[0] or "").strip()
        tid = str(row[5] or "").strip()
        url = str(row[6] or "").strip()
        if jp and tid.isdigit() and url:
            expect = f"https://www.themoviedb.org/person/{int(tid)}"
            if url.rstrip("/") != expect:
                errors.append(f"  行{idx}: tmdbid={tid} 与 url 不匹配: {url} (期望 {expect})")
    return errors


def _check_tmdb_url_duplicate(rows):
    """同一 tmdb url 被多行使用（复制污染）。"""
    seen: dict[str, int] = {}
    errors = []
    for idx, row in enumerate(rows, 2):
        if len(row) <= 6:
            return []
        jp = str(row[0] or "").strip()
        url = str(row[6] or "").strip()
        if jp and url:
            if url in seen:
                errors.append(f"  行{idx}: tmdb url 与行{seen[url]} 重复: {url}")
            else:
                seen[url] = idx
    return errors


_HYPERLINK_PATTERN = re.compile(r'<hyperlink [^>]*ref="([A-Z]+\d+)"[^>]*>')
_CELL_PATTERN = re.compile(r'<c r="([A-Z]+\d+)"')


def _check_orphan_hyperlinks(xlsx: Path):
    """孤儿 hyperlink：引用 XML 中不存在的单元格的超链接。

    孤儿 hyperlink 在 openpyxl 普通模式下会被物化为幽灵值/幽灵行，而 read_only 模式
    恰好绕过，导致写入路径（import/merge/db_guard）一保存就暴露污染。此检查直接
    解析 sheet XML，把「hyperlink 引用集合」与「实际 cell 定义集合」做差集。
    """
    errors = []
    try:
        with zipfile.ZipFile(xlsx) as zf:
            names = set(zf.namelist())
            # 定位「演员数据库」sheet 对应的 sheetN.xml
            sheet_xml = _locate_sheet_xml(zf, names)
            if not sheet_xml:
                return errors  # 结构异常由 read_only 阶段负责
            xml_text = zf.read(sheet_xml).decode("utf-8")
    except (zipfile.BadZipFile, OSError, KeyError):
        return errors

    defined = set(_CELL_PATTERN.findall(xml_text))
    orphans = []
    for ref in _HYPERLINK_PATTERN.findall(xml_text):
        if ref not in defined:
            orphans.append(ref)
    for ref in sorted(orphans):
        errors.append(f"  单元格 {ref}: 孤儿 hyperlink 引用了不存在的单元格（应清除）")
    return errors


def _locate_sheet_xml(zf: zipfile.ZipFile, names: set[str]) -> str | None:
    """从 workbook.xml + rels 映射 sheet 名 -> sheetN.xml 路径。"""
    try:
        if "xl/workbook.xml" not in names:
            return None
        wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
        # 找 sheet 名与 rId
        m = re.search(r'<sheet[^>]*name="演员数据库"[^>]*r:id="(rId\d+)"', wb_xml)
        if not m:
            return None
        rid = m.group(1)
        rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        m2 = re.search(rf'<Relationship [^>]*Id="{rid}"[^>]*>', rels_xml)
        if not m2:
            return None
        # Target 属性可能在 Id 前或后，从匹配片段内提取
        tm = re.search(r'Target="([^"]+)"', m2.group(0))
        if not tm:
            return None
        target = tm.group(1)
        if target.startswith("/"):
            return target.lstrip("/")
        return f"xl/{target}"
    except (KeyError, OSError):
        return None


def check_xlsx(xlsx: Path) -> int:
    if not xlsx.exists():
        print(f"[check_actor_db] 出厂数据库不存在，跳过: {xlsx}")
        return 0

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        display_path = xlsx.relative_to(MAIN_PATH)
    except ValueError:
        display_path = xlsx
    # 显式取「演员数据库」sheet，不依赖 sheet 顺序（防止男优备份等辅助 sheet 被误读）
    if ACTOR_DB_SHEET not in wb.sheetnames:
        print(f"[check_actor_db] {display_path} 缺少「{ACTOR_DB_SHEET}」sheet，结构异常")
        wb.close()
        return 1
    if wb.sheetnames[0] != ACTOR_DB_SHEET:
        print(
            f"[check_actor_db] {display_path} 首个 sheet 应为「{ACTOR_DB_SHEET}」，"
            f"实际为「{wb.sheetnames[0]}」，防止辅助 sheet 被误读"
        )
        wb.close()
        return 1
    ws = wb[ACTOR_DB_SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    errors: list[str] = []
    warnings: list[str] = []
    for check in (
        _check_jp_empty,
        _check_jp_duplicate,
        _check_keyword_format,
        _check_keyword_duplicate,
        _check_tmdbid_duplicate,
        _check_birth_date,
        _check_birth_date_range,
        _check_career_no_year,
        _check_tmdb_url_no_id,
        _check_tmdb_url_mismatch,
        _check_tmdb_url_duplicate,
    ):
        errors.extend(check(rows))
    errors.extend(_check_orphan_hyperlinks(xlsx))
    for check in (_check_name_empty, _check_bio_jp_residual, _check_bio_unstructured):
        warnings.extend(check(rows))

    print(f"[check_actor_db] {display_path} 共 {len(rows)} 行数据")
    if errors:
        print("[check_actor_db] 发现 error 级问题:")
        for item in errors:
            print(item)
    if warnings:
        print("[check_actor_db] 发现 warning 级问题(不阻断):")
        for item in warnings:
            print(item)
    if not errors and not warnings:
        print("[check_actor_db] 校验通过")
    elif not errors:
        print("[check_actor_db] 无 error，仅 warning")
    else:
        print(f"[check_actor_db] 校验失败: {len(errors)} 个 error")
    return 1 if errors else 0


def main() -> int:
    parser = argparse.ArgumentParser(description="演员数据库静态校验")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="目标 xlsx 路径")
    args = parser.parse_args()
    return check_xlsx(args.xlsx)


if __name__ == "__main__":
    sys.exit(main())
