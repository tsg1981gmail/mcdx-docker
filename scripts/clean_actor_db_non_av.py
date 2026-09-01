"""清除 actor_database.xlsx 中误收录的非 AV 人物。

AV 演员库混入非 AV 人物（好莱坞影星、知名声优、偶像、主流演员）会导致刮削时
把错误人物资料匹配到 AV 演员。本脚本用 TMDB 标记与作品复核识别并删除这类行：

1. 遍历库中所有有 tmdbid 的演员，调 TMDB `person/{id}` 取 adult 标记
2. adult=True -> 真 AV，保留
3. adult=False -> 拉 `combined_credits` 作品列表复核：
   - 作品含强成人特征词（痴漢/SM/乱交 等）-> AV/粉红电影，保留
   - 作品以主流内容为主（动画/演唱会/剧集/好莱坞大片）-> 非 AV，删除
   - 无法判断 -> 保守保留（宁缺毋滥）
4. 确认非 AV 的行整行删除（删除前备份到同目录 .bak）

支持断点续传（进度文件记录已完成 id）与进度展示。
用法:
    python scripts/clean_actor_db_non_av.py            # 只复核，打印统计
    python scripts/clean_actor_db_non_av.py --apply    # 复核 + 删除确认的非AV行
"""

from __future__ import annotations

import json
import shutil
import sys
import time
import urllib.request
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

DB_PATH = ROOT / "resources" / "userdata" / "actor_database.xlsx"
PROGRESS_PATH = Path("/tmp/opencode/clean_nonav_progress.json")
RESULT_PATH = Path("/tmp/opencode/clean_nonav_result.jsonl")

# TMDB API Key 读取：环境变量 MCAI_TMDB_KEY 或默认值
TMDB_KEY = "0b01619322f8002bddf32f680fb55ed2"
TMDB_HEADERS = {"Host": "api.themoviedb.org"}

# 强成人特征词（作品标题含任一即强烈暗示成人向）
STRONG_AV = [
    "痴漢",
    "SM",
    "縄",
    "責め",
    "夜這い",
    "熟女",
    "不倫",
    "情事",
    "欲情",
    "痴女",
    "愛人",
    "未亡人",
    "淫",
    "官能",
    "乱交",
    "セクハラ",
    "バイブ",
    "オナニー",
    "アダルト",
    "無修正",
    "裏ビデオ",
    "ポルノ",
    "生ハメ",
    "中出し",
    "顔射",
    "フェラ",
    "クンニ",
    "電マ",
    "手コキ",
    "緊縛",
    "陵辱",
    "調教",
    "奴隷",
    "スカトロ",
    "秘蔵",
    "本番",
    "porno",
    "hardcore",
    "gangbang",
    "milf",
    "パコ",
    "素人",
    "風俗",
]
# 明确主流作品标志（作品标题含任一即强烈暗示非AV）
STRONG_MAINSTREAM = [
    "アンパンマン",
    "ルパン三世",
    "名探偵コナン",
    "ワンピース",
    "ドラゴンボール",
    "ポケモン",
    "ドラえもん",
    "千と千尋",
    "となりのトトロ",
    "魔女の宅急便",
    "ハウルの動く城",
    "君の膵臓",
    "SPY×FAMILY",
    "リリカルなのは",
    "ブラッククローバー",
    "かいけつゾロリ",
    "ハロー!モーニング",
    "ハロモ",
    "Hello! Project",
    "AKB",
    "乃木坂",
    "IZ*ONE",
    "IVE",
    "K-POP",
    "バービー",
    "Barbie",
    "ディズニー",
    "ピクサー",
    "マーベル",
    "MCU",
    "劇場版",
    "ワルキューレ",
    "LIVE",
    "コンサート",
    "ツアー",
    "Dream Concert",
    "アニメ",
    "OVA",
    "仮面ライダー",
    "プリキュア",
    "ドラマ",
    "大河",
]


def fetch_person(pid: int) -> dict:
    """拉取 person 详情（adult 标记等）。"""
    url = f"https://api.tmdb.org/3/person/{pid}?api_key={TMDB_KEY}"
    req = urllib.request.Request(url, headers=TMDB_HEADERS)
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())


def fetch_credits(pid: int) -> list[str]:
    """拉 combined_credits 返回去重标题列表。"""
    url = f"https://api.tmdb.org/3/person/{pid}/combined_credits?api_key={TMDB_KEY}&language=ja"
    req = urllib.request.Request(url, headers=TMDB_HEADERS)
    with urllib.request.urlopen(req, timeout=15) as r:
        data = json.loads(r.read())
    titles: list[str] = []
    for c in data.get("cast", []) or []:
        t = str(c.get("title") or c.get("name") or "").strip()
        if t and t not in titles:
            titles.append(t)
    return titles


def judge(titles: list[str]) -> str:
    """根据作品标题判定 av / non_av / unknown。"""
    if not titles:
        return "unknown"
    av_count = sum(1 for t in titles if any(k in t for k in STRONG_AV))
    ms_count = sum(1 for t in titles if any(k in t for k in STRONG_MAINSTREAM))
    total = len(titles)
    # 作品很多且成人特征占比极低 -> 主流演员
    if total >= 30 and av_count / total < 0.1:
        return "non_av"
    # 主流作品占比高 -> 非AV
    if ms_count > 0 and ms_count / total >= 0.5:
        return "non_av"
    # 成人作品占比高 -> AV
    if av_count > 0 and av_count / total >= 0.5:
        return "av"
    # 相对比较
    if ms_count > av_count and ms_count > 0:
        return "non_av"
    if av_count > ms_count and av_count > 0:
        return "av"
    return "unknown"


def main() -> int:
    if not DB_PATH.exists():
        print(f"数据库不存在: {DB_PATH}")
        return 1

    wb = openpyxl.load_workbook(DB_PATH, read_only=True)
    ws = wb["演员数据库"]
    actors: list[tuple[int, str, str]] = []  # (行号, jp, tid)
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] and row[5]:
            actors.append((row_no, str(row[0]).strip(), str(row[5]).strip()))
    wb.close()
    print(f"有 tmdbid 的演员: {len(actors)}")

    done: set[int] = set()
    if PROGRESS_PATH.exists():
        done = set(json.loads(PROGRESS_PATH.read_text()))

    results = Counter()
    verdicts: dict[int, str] = {}
    if RESULT_PATH.exists():
        for line in RESULT_PATH.read_text(encoding="utf-8").splitlines():
            r = json.loads(line)
            verdicts[int(r["id"])] = r["verdict"]
            results[r["verdict"]] += 1

    start_all = time.time()
    with RESULT_PATH.open("a", encoding="utf-8") as f:
        for i, (row_no, jp, tid) in enumerate(actors, 1):
            pid = int(tid)
            if pid in done or pid in verdicts:
                continue
            try:
                person = fetch_person(pid)
                adult = str(person.get("adult", "")).lower() == "true"
                if adult:
                    verdict = "av"
                    titles = []
                else:
                    titles = fetch_credits(pid)
                    verdict = judge(titles)
            except Exception as e:
                verdict = "error"
                titles = []
                results["error"] += 1
                f.write(json.dumps({"row": row_no, "jp": jp, "id": pid, "verdict": "error", "error": str(e)}) + "\n")
                f.flush()
                done.add(pid)
                PROGRESS_PATH.write_text(json.dumps(sorted(done)))
                continue

            f.write(json.dumps({"row": row_no, "jp": jp, "id": pid, "verdict": verdict, "titles": titles[:8]}) + "\n")
            f.flush()
            done.add(pid)
            verdicts[pid] = verdict
            results[verdict] += 1

            if i % 20 == 0 or i == len(actors):
                elapsed = time.time() - start_all
                rate = len(done) / elapsed if elapsed > 0 else 0
                remain = (len(actors) - i) / rate if rate > 0 else 0
                print(
                    f"  进度 {i}/{len(actors)} ({i * 100 // len(actors)}%) "
                    f"av={results['av']} non_av={results['non_av']} unknown={results['unknown']} "
                    f"err={results['error']} 耗时{int(elapsed)}s 剩余{int(remain)}s"
                )
            PROGRESS_PATH.write_text(json.dumps(sorted(done)))
            time.sleep(0.2)

    print(f"\n复核完成: {dict(results)}")

    # 删除确认 non_av 的行
    if "--apply" in sys.argv:
        non_av_rows = sorted(
            int(json.loads(line)["row"])
            for line in RESULT_PATH.read_text(encoding="utf-8").splitlines()
            if json.loads(line).get("verdict") == "non_av"
        )
        if non_av_rows:
            backup = DB_PATH.with_suffix(".bak.xlsx")
            shutil.copy(DB_PATH, backup)
            print(f"备份: {backup}")
            wb = openpyxl.load_workbook(DB_PATH)
            ws = wb["演员数据库"]
            for i in reversed(non_av_rows):
                ws.delete_rows(i, 1)
            wb.save(DB_PATH)
            wb.close()
            print(f"删除 {len(non_av_rows)} 行确认非 AV 的演员")
        else:
            print("无确认非 AV 的行可删除")
    else:
        print("（预览模式，加 --apply 删除确认 non_av 的行）")

    return 0


if __name__ == "__main__":
    sys.exit(main())
