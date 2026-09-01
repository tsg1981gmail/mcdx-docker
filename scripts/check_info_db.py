#!/usr/bin/env python3
"""
信息映射数据库静态校验脚本。

仅检查仓库内出厂 `resources/userdata/info_database.xlsx`，把关本地数据质量。
检查项：
  1. jp 空字段（主键必须非空）                     [error]
  2. 同 jp 名重复（全半角归一 + 大小写不敏感）     [error]
  3. keyword 首尾逗号 / 连续逗号                   [error]
  4. keyword 重复词（归一化不敏感）                [error]
  5. zh_cn 重复（merge_info_db_from_backup 以 cn 为合并键，重复会导致用户库丢行）[error]
  6. zh_cn / zh_tw 空字段                          [warning]

「删除」行（jp=删除，cn=删除）是黑名单机制，无语义主键，允许重复，跳过所有检查。

发现任一 error 返回码 1；仅 warning 返回码 0。

用法:
    python scripts/check_info_db.py [--xlsx <path>]
"""

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover
    sys.stderr.write("缺少依赖 openpyxl，请先 uv sync\n")
    sys.exit(2)

MAIN_PATH = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = MAIN_PATH / "resources" / "userdata" / "info_database.xlsx"

DELETE_JP = "删除"


def _norm(text: str) -> str:
    """全半角归一 + 大写，用于匹配。与 mdcx.config.resources.get_info_data 一致。"""
    from mdcx.manual import ManualConfig

    s = str(text).upper()
    for full, half in ManualConfig.FULL_HALF_CHAR:
        s = s.replace(full, half)
    return s


def _is_delete(row) -> bool:
    return str(row[0] or "").strip() == DELETE_JP


def _check_jp_empty(rows):
    errors = []
    for idx, row in enumerate(rows, 2):
        jp = str(row[0] or "").strip()
        if not jp:
            errors.append(f"  行{idx}: jp(日文名) 为空")
    return errors


def _check_jp_duplicate(rows):
    seen: dict[str, int] = {}
    errors = []
    for idx, row in enumerate(rows, 2):
        if _is_delete(row):
            continue
        jp = str(row[0] or "").strip()
        if not jp:
            continue
        key = _norm(jp)
        if key in seen:
            errors.append(f"  行{idx}: jp 与行{seen[key]} 重复(归一化后): {row[0]}")
        else:
            seen[key] = idx
    return errors


def _check_keyword_format(rows):
    """info 库约定：keyword 必须以逗号开头和结尾（get_info_data 用 ,词, 逗号包裹匹配），且无连续逗号。"""
    errors = []
    for idx, row in enumerate(rows, 2):
        kw = str(row[3] or "").strip() if len(row) > 3 else ""
        if not kw:
            continue
        if not (kw.startswith(",") and kw.endswith(",")):
            errors.append(f"  行{idx}: keyword 缺少首尾逗号(应形如 ,词,词,): {kw}")
        elif ",," in kw:
            errors.append(f"  行{idx}: keyword 存在连续逗号: {kw}")
    return errors


def _norm_alnum(text: str) -> str:
    """仅归一大小写 + 全半角字母数字，用于 keyword 词级重复检测。

    刻意不归一 `・`/`·`/`、`/`、` 等标点分隔符：出厂库的标签聚合行会特意收集
    这些标点变体（如 調教・奴隷 / 調教·奴隸）作匹配用，属设计使然，不算重复。
    """
    s = str(text).upper()
    for full, half in (
        ("Ａ", "A"),
        ("Ｂ", "B"),
        ("Ｃ", "C"),
        ("Ｄ", "D"),
        ("Ｅ", "E"),
        ("Ｆ", "F"),
        ("Ｇ", "G"),
        ("Ｈ", "H"),
        ("Ｉ", "I"),
        ("Ｊ", "J"),
        ("Ｋ", "K"),
        ("Ｌ", "L"),
        ("Ｍ", "M"),
        ("Ｎ", "N"),
        ("Ｏ", "O"),
        ("Ｐ", "P"),
        ("Ｑ", "Q"),
        ("Ｒ", "R"),
        ("Ｓ", "S"),
        ("Ｔ", "T"),
        ("Ｕ", "U"),
        ("Ｖ", "V"),
        ("Ｗ", "W"),
        ("Ｘ", "X"),
        ("Ｙ", "Y"),
        ("Ｚ", "Z"),
        ("ａ", "a"),
        ("ｂ", "b"),
        ("ｃ", "c"),
        ("ｄ", "d"),
        ("ｅ", "e"),
        ("ｆ", "f"),
        ("ｇ", "g"),
        ("ｈ", "h"),
        ("ｉ", "i"),
        ("ｊ", "j"),
        ("ｋ", "k"),
        ("ｌ", "l"),
        ("ｍ", "m"),
        ("ｎ", "n"),
        ("ｏ", "o"),
        ("ｐ", "p"),
        ("ｑ", "q"),
        ("ｒ", "r"),
        ("ｓ", "s"),
        ("ｔ", "t"),
        ("ｕ", "u"),
        ("ｖ", "v"),
        ("ｗ", "w"),
        ("ｘ", "x"),
        ("ｙ", "y"),
        ("ｚ", "z"),
        ("０", "0"),
        ("１", "1"),
        ("２", "2"),
        ("３", "3"),
        ("４", "4"),
        ("５", "5"),
        ("６", "6"),
        ("７", "7"),
        ("８", "8"),
        ("９", "9"),
        ("＋", "+"),
        ("－", "-"),
    ):
        s = s.replace(full, half)
    return s


def _check_keyword_duplicate(rows):
    errors = []
    for idx, row in enumerate(rows, 2):
        if _is_delete(row):
            continue  # 黑名单行有意收集变体写法（如 30%OFF/30％OFF），不算重复
        kw = str(row[3] or "").strip() if len(row) > 3 else ""
        if not kw:
            continue
        parts = [k.strip() for k in kw.split(",") if k.strip()]
        if len(parts) != len({_norm_alnum(k) for k in parts}):
            errors.append(f"  行{idx}: keyword 存在重复词: {kw}")
    return errors


def _check_cn_duplicate(rows):
    """zh_cn 重复：merge_info_db_from_backup 以 cn 为合并键，重复会导致同步时互相覆盖、用户库丢行。"""
    seen: dict[str, int] = {}
    errors = []
    for idx, row in enumerate(rows, 2):
        if _is_delete(row):
            continue
        cn = str(row[1] or "").strip() if len(row) > 1 else ""
        if not cn:
            continue
        if cn in seen:
            errors.append(f"  行{idx}: zh_cn 与行{seen[cn]} 重复: {cn}")
        else:
            seen[cn] = idx
    return errors


def _check_name_empty(rows):
    warnings = []
    for idx, row in enumerate(rows, 2):
        jp = str(row[0] or "").strip()
        if not jp:
            continue
        zh_cn = str(row[1] or "").strip() if len(row) > 1 else ""
        zh_tw = str(row[2] or "").strip() if len(row) > 2 else ""
        if not zh_cn:
            warnings.append(f"  行{idx}: zh_cn(中文名) 为空")
        if not zh_tw:
            warnings.append(f"  行{idx}: zh_tw(繁体名) 为空")
    return warnings


def _check_delete_rows_front(rows):
    """删除行（jp=删除）必须全部排在最前面。

    get_info_data 按行顺序匹配、返回第一个命中行：删除行（黑名单）在前才能保证
    黑名单词优先命中并返回空翻译；若删除行散落在内容行之后，同词会被内容行抢先匹配，
    导致黑名单失效。
    """
    errors = []
    seen_content = False
    for idx, row in enumerate(rows, 2):
        jp = str(row[0] or "").strip()
        if not jp:
            continue
        if jp == DELETE_JP:
            if seen_content:
                errors.append(f"  行{idx}: 删除行出现在内容行之后（应全部排在前面）")
        else:
            seen_content = True
    return errors


def check_xlsx(xlsx: Path) -> int:
    if not xlsx.exists():
        print(f"[check_info_db] 出厂数据库不存在，跳过: {xlsx}")
        return 0

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    try:
        display_path = xlsx.relative_to(MAIN_PATH)
    except ValueError:
        display_path = xlsx

    errors: list[str] = []
    warnings: list[str] = []
    for check in (
        _check_jp_empty,
        _check_jp_duplicate,
        _check_keyword_format,
        _check_keyword_duplicate,
        _check_cn_duplicate,
        _check_delete_rows_front,
    ):
        errors.extend(check(rows))
    warnings.extend(_check_name_empty(rows))

    print(f"[check_info_db] {display_path} 共 {len(rows)} 行数据")
    if errors:
        print("[check_info_db] 发现 error 级问题:")
        for item in errors:
            print(item)
    if warnings:
        print("[check_info_db] 发现 warning 级问题(不阻断):")
        for item in warnings:
            print(item)
    if not errors and not warnings:
        print("[check_info_db] 校验通过")
    elif not errors:
        print("[check_info_db] 无 error，仅 warning")
    else:
        print(f"[check_info_db] 校验失败: {len(errors)} 个 error")
    return 1 if errors else 0


def main() -> int:
    parser = argparse.ArgumentParser(description="信息映射数据库静态校验")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="目标 xlsx 路径")
    args = parser.parse_args()
    return check_xlsx(args.xlsx)


if __name__ == "__main__":
    sys.exit(main())
