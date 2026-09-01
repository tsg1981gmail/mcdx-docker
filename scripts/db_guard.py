"""演员数据库安全写入与保存后验证（改库脚本共享防线）。

两类防护：
1. 源头防写：safe_write_tmdb 强制「写 tmdb url 必须成对写 tmdbid」，且
   url 必须与该 id 的标准 person url 一致，违反则拒绝。防止直接操作 ws 的
   脚本产生「无 id 有 url」「id-url 不匹配」的复制污染。
2. 保存后验证：validate_after_save 在 wb.save 后自动运行 check_actor_db，
   发现 error 级问题则打印并返回 False，调用方应中止，避免脏数据进入出厂库。
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

from mdcx.config.resources import COL_TMDB_URL, COL_TMDBID, get_actor_db_sheet  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _tmdb_person_url(tmdbid: int | str) -> str:
    return f"https://www.themoviedb.org/person/{tmdbid}"


def safe_write_tmdb(ws, row_idx: int, tmdbid: int | str | None, url: str | None = None) -> bool:
    """安全写入某行的 tmdbid / tmdb url 列。

    - tmdbid 与 url 必须成对出现；单独写 url 拒绝。
    - url 缺省时由 tmdbid 推导标准 url。
    - url 显式给定时必须与该 id 的标准 url 一致，否则拒绝。
    返回是否写入成功。
    """
    if tmdbid is None or str(tmdbid).strip() == "":
        return False  # 无 id：拒绝单独写 url

    tid = str(tmdbid).strip()
    if not tid.isdigit():
        return False

    if url is None:
        url = _tmdb_person_url(tid)
    elif str(url).strip() != _tmdb_person_url(int(tid)):
        return False  # url 与该 id 不匹配，拒绝

    ws.cell(row=row_idx, column=COL_TMDBID + 1).value = int(tid)
    ws.cell(row=row_idx, column=COL_TMDB_URL + 1).value = url
    ws.cell(row=row_idx, column=COL_TMDB_URL + 1).hyperlink = url
    return True


def clear_tmdb(ws, row_idx: int) -> None:
    """同时清空某行的 tmdbid 与 tmdb url（保持成对）。"""
    ws.cell(row=row_idx, column=COL_TMDBID + 1).value = None
    ws.cell(row=row_idx, column=COL_TMDB_URL + 1).value = None


def validate_after_save(db_path: Path) -> bool:
    """保存后运行 check_actor_db，返回是否通过（无 error 级问题）。

    注意：openpyxl 的 read_only 模式会丢失部分行数据，check_actor_db 内部
    用非只读判定关键列，这里直接用其 check_xlsx。
    """
    sys.path.insert(0, str(ROOT / "scripts"))
    from scripts.check_actor_db import check_xlsx

    return check_xlsx(db_path) == 0


def load_ws(db_path: Path):
    """加载工作簿并取主 sheet（非只读，供写操作）。"""
    wb = openpyxl.load_workbook(db_path)
    return wb, get_actor_db_sheet(wb)
