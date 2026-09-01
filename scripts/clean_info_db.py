#!/usr/bin/env python3
"""
清洗 info_database.xlsx 出厂库的重复与冗余（合并后校验）。

修复两类问题：
1. jp 重复（全半角归一后相同，如标点 ？/? 差异导致的同系列拆行）：
   保留第一个出现的行，把冗余行的 jp + keyword 并入主行 keyword（alias 合并），删除冗余行。
   与 merge_actor_db_duplicates.py 的方法论一致：确证重复 -> 合并 keyword -> 删除冗余行。
2. keyword 内部重复词（全半角/大小写变体冗余，如 3P/３P、OL/ＯＬ、marx/MARX）：
   归一化后相同的词只保留第一个。删除冗余词不影响运行时匹配（get_info_data 会归一化）。
3. 空 keyword（如 M女 的 ,,）补为主名。

「删除」黑名单行（jp=删除）跳过：有意收集变体写法，不算问题。

用法:
    python scripts/clean_info_db.py            # 预览
    python scripts/clean_info_db.py --apply    # 备份 + 修复 + 校验
    python scripts/clean_info_db.py --xlsx <path> --apply  # 测试用副本
"""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover
    sys.stderr.write("缺少依赖 openpyxl，请先 uv sync\n")
    sys.exit(2)

from scripts.check_info_db import _norm, _norm_alnum  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "resources" / "userdata" / "info_database.xlsx"

DELETE_JP = "删除"


def _is_delete(row) -> bool:
    return str(row[0] or "").strip() == DELETE_JP


def _dedup_keyword_parts(kw: str) -> str:
    """归一化去重 keyword 词，保留首个出现顺序。"""
    parts = [k.strip() for k in kw.split(",")]
    seen: set[str] = set()
    kept: list[str] = []
    for p in parts:
        if not p:
            continue
        key = _norm_alnum(p)
        if key in seen:
            continue
        seen.add(key)
        kept.append(p)
    if not kept:
        return ""
    return "," + ",".join(kept) + ","


def scan(rows: list[tuple]) -> list[tuple[int, int]]:
    """返回确证 jp 重复对（归一化后），形如 (主行idx, 冗余行idx)。idx 为枚举序号（从 0 起）。"""
    jp_seen: dict[str, int] = {}
    pairs: list[tuple[int, int]] = []
    for idx, row in enumerate(rows):
        if _is_delete(row):
            continue
        jp = str(row[0] or "").strip()
        if not jp:
            continue
        key = _norm(jp)
        if key in jp_seen:
            pairs.append((jp_seen[key], idx))
        else:
            jp_seen[key] = idx
    return pairs


def build_fix_plan(rows: list[tuple]) -> dict[int, str]:
    """返回需要重写 keyword 的行: 行idx -> 新 keyword。"""
    plan: dict[int, str] = {}

    for idx, row in enumerate(rows):
        if _is_delete(row):
            continue
        kw = str(row[3] or "").strip() if len(row) > 3 else ""
        jp = str(row[0] or "").strip()
        parts = [k.strip() for k in kw.split(",") if k.strip()]
        if not kw or not parts:
            plan[idx] = f",{jp},"
        else:
            dedup = _dedup_keyword_parts(kw)
            if dedup != kw:
                plan[idx] = dedup

    # jp 重复：冗余行的 jp + keyword 并入主行 keyword
    for main_idx, dup_idx in scan(rows):
        main_kw = plan.get(main_idx, str(rows[main_idx][3] or "").strip() if len(rows[main_idx]) > 3 else "")
        dup_jp = str(rows[dup_idx][0] or "").strip()
        dup_kw = str(rows[dup_idx][3] or "").strip() if len(rows[dup_idx]) > 3 else ""
        merged = f"{main_kw},{dup_jp},{dup_kw}"
        plan[main_idx] = _dedup_keyword_parts(merged)
        plan.pop(dup_idx, None)
    return plan


def apply_fixes(ws, rows: list[tuple], plan: dict[int, str], delete_rows: set[int]) -> tuple[int, int]:
    """写入修复：改写 keyword 的行 + 删除冗余行。返回 (修复行数, 删除行数)。"""
    fixed = 0
    for idx, new_kw in plan.items():
        ws.cell(row=idx + 2, column=4, value=new_kw)
        fixed += 1
    for idx in sorted(delete_rows, reverse=True):
        ws.delete_rows(idx + 2, 1)
    return fixed, len(delete_rows)


def main(argv: list[str] | None = None) -> int:
    argv = sys.argv if argv is None else argv
    parser = argparse.ArgumentParser(description="清洗 info_database.xlsx 出厂库重复/冗余")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--apply", action="store_true", help="备份 + 修复 + 校验")
    args = parser.parse_args(argv[1:])

    if not args.xlsx.exists():
        print(f"数据库不存在: {args.xlsx}")
        return 1

    wb = load_workbook(args.xlsx)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    jp_pairs = scan(rows)
    print(f"jp 归一化重复 {len(jp_pairs)} 对:")
    for m, d in jp_pairs:
        print(f"  {rows[m][0]}  <->  {rows[d][0]}")

    plan = build_fix_plan(rows)
    delete_rows = {d for _, d in jp_pairs}
    kw_fixes = {idx: kw for idx, kw in plan.items() if idx not in delete_rows}
    print(f"keyword 修复 {len(kw_fixes)} 行:")
    for idx, kw in kw_fixes.items():
        print(f"  row{idx + 2} {rows[idx][0]}: {kw}")

    if not jp_pairs and not kw_fixes:
        print("无需修复。")
        wb.close()
        return 0

    if "--apply" not in argv:
        print("\n（预览模式，加 --apply 执行）")
        wb.close()
        return 0

    backup = args.xlsx.with_suffix(".bak.xlsx")
    shutil.copy(args.xlsx, backup)
    print(f"备份: {backup}")

    fixed, removed = apply_fixes(ws, rows, plan, delete_rows)
    wb.save(args.xlsx)
    wb.close()

    from scripts.check_info_db import check_xlsx

    ok = check_xlsx(args.xlsx) == 0
    print(f"修复 {fixed} 行，删除 {removed} 行")
    if not ok:
        print("⚠️ 修复后校验发现 error 级问题，请检查！")
        return 1
    print("修复后校验通过")
    return 0


if __name__ == "__main__":
    sys.exit(main())
