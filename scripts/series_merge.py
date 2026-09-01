#!/usr/bin/env python3
"""
将翻译后的系列映射写入 info_database.xlsx 出厂库。

输入: 一个 JSON 文件，形如:
    {
        "日文系列名": {"zh_cn": "中文简体", "zh_tw": "中文繁体", "alias": ["别名1", "别名2"]},
        ...
    }
    alias 可选：同系列的其他写法(如带/不带空格、括号全半角差异)，会并入 keyword 便于匹配。

规则:
- 新系列行追加到出厂库末尾，格式与其他行一致:
    jp=日文系列名, zh_cn=简体, zh_tw=繁体, keyword=,系列名,别名...,简体,繁体,
- 若出厂库已存在同 jp 的系列行(含 keyword 精确匹配)，跳过避免重复。
- keyword 首尾逗号、无连续逗号，与出厂库现有格式一致。
- cn 必须与出厂库既有内容行 cn 唯一（merge_info_db_from_backup 以 cn 为合并键，重复会导致用户库丢行）：
  冲突时打印警告并跳过该条。
- 只在仓库内出厂 resources/userdata/info_database.xlsx 上操作。

用法:
    uv run python scripts/series_merge.py --mapping /path/to/series_translation.json [--dry-run]
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover
    sys.stderr.write("缺少依赖 openpyxl，请先 uv sync\n")
    sys.exit(2)

try:
    from mdcx.manual import ManualConfig
except ModuleNotFoundError:  # pragma: no cover
    sys.stderr.write("缺少依赖 mdcx，请先 uv sync\n")
    sys.exit(2)

MAIN_PATH = Path(__file__).resolve().parent.parent
INFO_DB = MAIN_PATH / "resources" / "userdata" / "info_database.xlsx"


def _norm(text: str) -> str:
    """全半角归一 + 大写，用于匹配。

    与 mdcx.config.resources.get_info_data 的匹配归一完全一致（复用
    ManualConfig.FULL_HALF_CHAR），保证「合并去重」与「运行时匹配」判定一致。
    """
    s = text.upper()
    for full, half in ManualConfig.FULL_HALF_CHAR:
        s = s.replace(full, half)
    return s


def _load_existing() -> tuple[set[str], set[str]]:
    """返回 (existing_norm, existing_cn)。existing_cn 收集出厂库内容行的 zh_cn，用于冲突检测。"""
    wb = load_workbook(INFO_DB, read_only=True, data_only=True)
    ws = wb.active
    existing: set[str] = set()
    existing_cn: set[str] = set()
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        if not row or not row[0]:
            continue
        jp = str(row[0]).strip()
        cn = str(row[1] or "").strip() if len(row) > 1 else ""
        kw = str(row[3] or "").strip() if len(row) > 3 else ""
        existing.add(_norm(jp))
        if cn and jp != "删除":
            existing_cn.add(_norm(cn))
        for part in kw.split(","):
            part = part.strip()
            if part:
                existing.add(_norm(part))
    wb.close()
    return existing, existing_cn


def _build_keyword(jp: str, zh_cn: str, zh_tw: str, aliases: list[str]) -> str:
    parts = [jp, *aliases, zh_cn, zh_tw]
    seen: set[str] = set()
    unique: list[str] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        key = _norm(p)
        if key not in seen:
            seen.add(key)
            unique.append(p)
    return "," + ",".join(unique) + ","


def main() -> int:
    parser = argparse.ArgumentParser(description="系列映射写入 info 出厂库")
    parser.add_argument("--mapping", type=Path, required=True, help="系列翻译映射 json")
    parser.add_argument("--dry-run", action="store_true", help="只打印将写入的行，不修改文件")
    args = parser.parse_args()

    mapping = json.loads(args.mapping.read_text(encoding="utf-8"))
    if not isinstance(mapping, dict) or not mapping:
        sys.stderr.write("映射文件为空或格式错误\n")
        return 2

    existing, existing_cn = _load_existing()
    to_add: list[tuple[str, str, str, str]] = []
    skipped = 0
    cn_conflicts: list[str] = []
    for jp, trans in mapping.items():
        jp = jp.strip()
        if not jp:
            continue
        zh_cn = (trans.get("zh_cn") or "").strip() if isinstance(trans, dict) else ""
        zh_tw = (trans.get("zh_tw") or "").strip() if isinstance(trans, dict) else ""
        aliases = (
            [str(a).strip() for a in (trans.get("alias") or []) if str(a).strip()] if isinstance(trans, dict) else []
        )
        if not zh_cn:
            zh_cn = zh_cn or jp
        if not zh_tw:
            zh_tw = zh_tw or zh_cn
        if _norm(jp) in existing or any(_norm(a) in existing for a in aliases):
            skipped += 1
            continue
        if zh_cn and _norm(zh_cn) in existing_cn:
            cn_conflicts.append(f"{jp} -> {zh_cn}")
            continue
        keyword = _build_keyword(jp, zh_cn, zh_tw, aliases)
        to_add.append((jp, zh_cn, zh_tw, keyword))

    if cn_conflicts:
        print(f"⚠️ cn 冲突跳过 {len(cn_conflicts)} 条（需修正映射使 cn 唯一）:")
        for c in cn_conflicts[:30]:
            print(f"  {c}")

    if not to_add:
        print(f"无新增系列（跳过已存在 {skipped}，cn 冲突 {len(cn_conflicts)}）")
        return 0

    print(f"将新增 {len(to_add)} 个系列行（跳过已存在 {skipped}，cn 冲突 {len(cn_conflicts)}）")
    for jp, zh_cn, zh_tw, kw in to_add[:10]:
        print(f"  {jp} -> {zh_cn} / {zh_tw} | {kw}")

    if args.dry_run:
        print("[dry-run] 未修改文件")
        return 0

    wb = load_workbook(INFO_DB)
    ws = wb.active
    next_row = ws.max_row + 1
    for jp, zh_cn, zh_tw, kw in to_add:
        ws.cell(row=next_row, column=1, value=jp)
        ws.cell(row=next_row, column=2, value=zh_cn)
        ws.cell(row=next_row, column=3, value=zh_tw)
        ws.cell(row=next_row, column=4, value=kw)
        next_row += 1
    wb.save(INFO_DB)
    wb.close()
    print(f"已写入 {len(to_add)} 行 -> {INFO_DB}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
