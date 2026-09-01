#!/usr/bin/env python3
"""从 tenhow 全量索引构建新增 ASIN 库 xlsx 并用 thejavdb api 校验番号。

Phase A (build, 本地秒级):
    读取 tenhow_index_full.json，剔除已在主库/待修正/番号已在库的 ASIN，
    cid→番号清洗（去厂商前缀，数字段 int 后至少补零到 3 位），同番号去重，
    写出 userdata/tenhow_asin_new.xlsx（列与 amazon_asin_database.xlsx 一致）。
    无法解析的 cid 进「待人工」sheet。

Phase B (validate, 长任务):
    对新表逐行调 thejavdb api 校验番号存在性，断点续传（state json），
    分批落盘；进程可用 --max-seconds 限制单次运行时长，由外层 shell
    循环重启直到全部完成，规避云环境超时杀进程。

用法:
    python scripts/build_tenhow_asin_xlsx.py build
    python scripts/build_tenhow_asin_xlsx.py validate [--max-seconds 2850]
    python scripts/build_tenhow_asin_xlsx.py status
"""

import asyncio
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
INDEX_JSON = Path("/tmp/opencode/tenhow_index_full.json")
DB_XLSX = ROOT / "userdata" / "amazon_asin_database.xlsx"
OUT_XLSX = ROOT / "userdata" / "tenhow_asin_new.xlsx"
STATE_JSON = Path("/tmp/opencode/tenhow_validate_state.json")
DONE_FLAG = Path("/tmp/opencode/tenhow_validate_done")
JAVDB_DONE_FLAG = Path("/tmp/opencode/tenhow_validate_javdb_done")

THEJAVDB_API = "https://api.thejavdb.net/v1/movies?q={number}"
CONCURRENCY = 8
BATCH_SAVE = 50

HEADERS = ["影片番号", "ASIN 编号", "影片链接", "商品标题", "封面 URL", "搜索关键词"]

# cid 形态: 可选厂商数字前缀 + 系列字母 + 数字 + 可选变体字母
_CID_RE = re.compile(r"^(\d*)([a-z]+)(\d+)([a-z])?$")


def cid_to_number(cid: str) -> str | None:
    """cid → 番号。去厂商数字前缀，数字段 int 后至少补零到 3 位（DMM 官方约定）。

    24ped00030→PED-030, 13gvg00564→GVG-564, onsg00064→ONSG-064。
    末尾变体字母（1iene577f）与本体同番号，返回 (番号, 是否变体) 由调用方合并。
    返回 None 表示无法解析。
    """
    m = _CID_RE.match(cid.strip().lower())
    if not m:
        return None
    series, digits = m.group(2), m.group(3)
    return f"{series.upper()}-{int(digits):03d}"


def _num_key(number: str) -> tuple[str, int] | None:
    """番号归一化 key：(系列字母, int(数字))，用于与主库去重比对（兼容库存两种补零风格）。"""
    m = re.match(r"^([A-Za-z]+)-(\d+)$", number.strip())
    if not m:
        return None
    return m.group(1).upper(), int(m.group(2))


def phase_build() -> None:
    import openpyxl

    index: dict[str, list[str]] = json.loads(INDEX_JSON.read_text())["asin2cids"]

    wb = openpyxl.load_workbook(DB_XLSX, read_only=True)
    ws = wb.active
    db_asins: set[str] = set()
    db_keys: set[tuple[str, int]] = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        a = str(r[1] or "").strip()
        if a:
            db_asins.add(a)
        k = _num_key(str(r[0] or ""))
        if k:
            db_keys.add(k)
    fix_asins = {str(r[1] or "").strip() for r in wb["待修正"].iter_rows(min_row=2, values_only=True) if r[1]}

    today = datetime.now().strftime("%Y-%m-%d")
    rows: list[list[str]] = []  # 主 sheet
    pending: list[list[str]] = []  # 待人工 sheet（cid 无法解析）
    seen_keys: dict[tuple[str, int], str] = {}  # 番号 key → 已用 ASIN
    dup_rows: list[list[str]] = []  # 同番号被去重的记录
    skipped_db = skipped_fix = skipped_num = 0

    for asin in sorted(index):
        if asin in db_asins:
            skipped_db += 1
            continue
        if asin in fix_asins:
            skipped_fix += 1
            continue
        numbers = set()
        bad_cids = []
        for cid in index[asin]:
            n = cid_to_number(cid)
            (numbers.add if n else bad_cids.append)(n if n else cid)
        if not numbers:
            pending.append(["", asin, today, ",".join(index[asin]), "", ""])
            continue
        number = sorted(numbers)[0]
        key = _num_key(number)
        if key and key in db_keys:
            skipped_num += 1
            continue
        if key and key in seen_keys:
            dup_rows.append([number, asin, today, f"同番号已用 {seen_keys[key]}", "", ""])
            continue
        if key:
            seen_keys[key] = asin
        rows.append([number, asin, today, "tenhow", "", ""])

    out = openpyxl.Workbook()
    sm = out.active
    sm.title = "Sheet"
    sm.append(HEADERS)
    for r in rows:
        sm.append(r)
    sp = out.create_sheet("待人工")
    sp.append(HEADERS)
    for r in pending:
        sp.append(r)
    sd = out.create_sheet("重复剔除")
    sd.append(HEADERS)
    for r in dup_rows:
        sd.append(r)
    out.save(OUT_XLSX)

    print(f"索引 ASIN {len(index)}")
    print(f"跳过: 已在主库 {skipped_db}, 已在待修正 {skipped_fix}, 番号已在库 {skipped_num}")
    print(f"主 sheet {len(rows)} 行; 待人工 {len(pending)}; 同番号去重剔除 {len(dup_rows)}")
    print(f"已写出 {OUT_XLSX}")


def _load_validate_rows() -> list[tuple[str, str]]:
    import openpyxl

    wb = openpyxl.load_workbook(OUT_XLSX, read_only=True)
    ws = wb["Sheet"]
    return [
        (str(r[0] or "").strip(), str(r[1] or "").strip()) for r in ws.iter_rows(min_row=2, values_only=True) if r[1]
    ]


async def phase_validate(max_seconds: int) -> None:
    import aiohttp

    rows = _load_validate_rows()
    state: dict[str, dict] = {}
    if STATE_JSON.exists():
        state = json.loads(STATE_JSON.read_text())
    todo = [(num, asin) for num, asin in rows if asin not in state]
    print(
        f"validate: 总 {len(rows)}, 已完成 {len(state)}, 本次待跑 {len(todo)}",
        flush=True,
    )
    if not todo:
        print("已全部完成，写回 xlsx", flush=True)
        _write_validation_back(state)
        DONE_FLAG.touch()
        return

    deadline = time.time() + max_seconds
    sem = asyncio.Semaphore(CONCURRENCY)
    done_since_save = 0
    counters = {"ok": 0, "notfound": 0, "error": 0}

    async def worker(num: str, asin: str, session) -> tuple[str, str]:
        nonlocal done_since_save
        async with sem:
            for attempt in range(3):
                try:
                    async with session.get(
                        THEJAVDB_API.format(number=num),
                        timeout=aiohttp.ClientTimeout(total=20),
                    ) as r:
                        if r.status == 200:
                            return asin, "ok"
                        if r.status == 404:
                            return asin, "notfound"
                        if r.status == 429:
                            await asyncio.sleep(2 + attempt * 3)
                            continue
                        return asin, f"http{r.status}"
                except Exception as e:
                    if attempt == 2:
                        return asin, f"err:{type(e).__name__}"
                    await asyncio.sleep(1 + attempt * 2)
            return asin, "err:retry_exhausted"

    timeout = aiohttp.ClientTimeout(total=30)
    async with aiohttp.ClientSession(headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout) as session:
        pending = list(todo)
        while pending and time.time() < deadline:
            batch, pending = pending[:BATCH_SAVE], pending[BATCH_SAVE:]
            results = await asyncio.gather(*[worker(num, asin, session) for num, asin in batch])
            for asin, st in results:
                state[asin] = {"status": st, "ts": datetime.now().isoformat()}
                counters[st if st in ("ok", "notfound") else "error"] += 1
            STATE_JSON.write_text(json.dumps(state))
            done_since_save += len(results)
            print(
                f"progress: {len(state)}/{len(rows)} "
                f"(ok {counters['ok']}, notfound {counters['notfound']}, err {counters['error']}) "
                f"剩余批次 {len(pending) // BATCH_SAVE}",
                flush=True,
            )
    if not pending:
        print("校验全部完成，写回 xlsx", flush=True)
        _write_validation_back(state)
        DONE_FLAG.touch()
    else:
        print(f"到达 max_seconds，已落盘断点 {len(state)}/{len(rows)}，等待外层重启", flush=True)


async def _javdb_app_search(number: str) -> str:
    """用 javdb app 移动端 api 补查番号是否存在。返回 ok/notfound/err:xxx。"""
    import aiohttp

    sys.path.insert(0, str(ROOT))
    from mdcx.crawlers.javdb_app import _API_BASE, _API_FALLBACKS, _get_api_url, make_signature

    headers = {
        "jdsignature": make_signature(),
        "accept-language": "zh",
        "User-Agent": "Dart/3.5 (dart:io)",
    }
    paths = [_API_BASE, *_API_FALLBACKS]
    last = "err:no_host"
    async with aiohttp.ClientSession() as session:
        for host in paths:
            try:
                url = _get_api_url(host, "/api/v2/search", {"q": number, "page": "1"})
                async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=15)) as r:
                    if r.status == 200:
                        data = await r.json()
                        if not data.get("success"):
                            last = f"api_fail:{str(data.get('action'))[:30]}"
                            continue
                        movies = (data.get("data") or {}).get("movies") or []
                        for mv in movies:
                            kn, km = _num_key(number), _num_key(str(mv.get("number") or ""))
                            if kn and km and kn == km:
                                return "ok"
                        return "notfound"
                    if r.status in (301, 302, 404):
                        continue
                    return f"http{r.status}"
            except Exception as e:
                last = f"err:{type(e).__name__}"
        return last


async def phase_validate_javdb(max_seconds: int) -> None:
    """对 thejavdb 判定 notfound 的行用 javdb app 补查，断点续传。"""
    state: dict[str, dict] = {}
    if STATE_JSON.exists():
        state = json.loads(STATE_JSON.read_text())
    nf_rows = _notfound_rows_from_state(state)
    st2: dict[str, dict] = state

    pending = [(n, a) for n, a in nf_rows if st2[a].get("javdb") is None]
    print(f"javdb 补查: notfound 共 {len(nf_rows)}, 本次待跑 {len(pending)}", flush=True)
    if not pending:
        print("javdb 补查已全部完成", flush=True)
        JAVDB_DONE_FLAG.touch()
        return

    deadline = time.time() + max_seconds
    counters = {"ok": 0, "notfound": 0, "error": 0}
    sem = asyncio.Semaphore(3)  # javdb app 反爬敏感，低并发

    async def throttled(number: str) -> str:
        async with sem:
            await asyncio.sleep(0.8 + 1.5 * (asyncio.get_running_loop().time() % 1))
            return await _javdb_app_search(number)

    while pending and time.time() < deadline:
        batch, pending = pending[:BATCH_SAVE], pending[BATCH_SAVE:]
        results = await asyncio.gather(*[throttled(num) for num, _ in batch])
        for (_, asin), st in zip(batch, results, strict=True):
            st2[asin]["javdb"] = st
            counters[st if st in ("ok", "notfound") else "error"] += 1
        STATE_JSON.write_text(json.dumps(st2))
        print(
            f"javdb progress: 补查完成 {len(nf_rows) - len(pending)}/{len(nf_rows)} "
            f"(ok {counters['ok']}, notfound {counters['notfound']}, err {counters['error']})",
            flush=True,
        )
    if not pending:
        _write_validation_back(st2)
        JAVDB_DONE_FLAG.touch()
        print("javdb 补查全部完成，已写回 xlsx", flush=True)
    else:
        print("到达 max_seconds，断点已落盘，等待外层重启", flush=True)


def _notfound_rows_from_state(state: dict[str, dict]) -> list[tuple[str, str]]:
    """从 state 反查 (番号, ASIN)：遍历 Sheet 取 status == notfound 且 javdb 未判定的行。"""
    rows = _load_validate_rows()
    return [(n, a) for n, a in rows if state.get(a, {}).get("status") == "notfound"]


def _write_validation_back(state: dict[str, dict]) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(OUT_XLSX)
    ws = wb["Sheet"]
    ws.cell(row=1, column=7, value="thejavdb校验")
    ws.cell(row=1, column=8, value="javdb补查")
    counts = {"ok": 0, "notfound": 0, "other": 0, "missing": 0}
    jb_counts = {"ok": 0, "notfound": 0, "other": 0, "": 0}
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        asin = str(r[1] or "").strip()
        st = state.get(asin, {}).get("status", "")
        if st == "ok":
            counts["ok"] += 1
        elif st == "notfound":
            counts["notfound"] += 1
        elif st:
            counts["other"] += 1
        else:
            counts["missing"] += 1
        ws.cell(row=i, column=7, value=st)
        jb = state.get(asin, {}).get("javdb", "")
        ws.cell(row=i, column=8, value=jb)
        key = jb if jb in ("ok", "notfound") else ("other" if jb else "")
        jb_counts[key] += 1
    wb.save(OUT_XLSX)
    print(f"写回完成: thejavdb {counts}; javdb 补查 {jb_counts}", flush=True)


def phase_status() -> None:
    rows = _load_validate_rows()
    state: dict = json.loads(STATE_JSON.read_text()) if STATE_JSON.exists() else {}
    from collections import Counter

    c = Counter(v.get("status", "?") for v in state.values())
    jb = Counter(v["javdb"] for v in state.values() if "javdb" in v)
    print(f"thejavdb 进度: {len(state)}/{len(rows)} ({100 * len(state) / max(len(rows), 1):.1f}%) 状态: {dict(c)}")
    print(f"javdb 补查: 已完成 {sum(jb.values())} 状态: {dict(jb)}")
    print(
        f"done 标记: thejavdb={'有' if DONE_FLAG.exists() else '无'} javdb={'有' if JAVDB_DONE_FLAG.exists() else '无'}"
    )


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        return
    cmd = sys.argv[1]
    if cmd == "build":
        phase_build()
    elif cmd == "validate":
        max_seconds = 2850
        if "--max-seconds" in sys.argv:
            max_seconds = int(sys.argv[sys.argv.index("--max-seconds") + 1])
        asyncio.run(phase_validate(max_seconds))
    elif cmd == "validate-javdb":
        max_seconds = 2850
        if "--max-seconds" in sys.argv:
            max_seconds = int(sys.argv[sys.argv.index("--max-seconds") + 1])
        asyncio.run(phase_validate_javdb(max_seconds))
    elif cmd == "status":
        phase_status()
    else:
        print(f"未知命令: {cmd}")
        print(__doc__)


if __name__ == "__main__":
    main()
