"""筛查并修复 actor_database.xlsx 中「无 tmdbid 行」与「有 tmdbid 行」的重复。

无 tmdbid 的演员行可能与有 tmdbid 的演员行是同一人（多个艺名被拆成多行）。
确证标准：无 id 行的别名(keyword 列)里精确包含另一个有 id 行的主名(日文原名，
经全半角/大小写/繁简规范化)。这是「同一人多艺名」的强证据。

修复动作（--apply）：
1. 备份原库到同目录 .bak.xlsx
2. 把无 id 行的主名 + 全部别名并入有 id 行的 keyword 列（中文优先排序去重）
3. 有 id 行空缺字段（中文名/繁体名/出生日期/简介）用无 id 行的值补全
4. 删除无 id 行

用法:
    python scripts/merge_actor_db_duplicates.py            # 只读预览
    python scripts/merge_actor_db_duplicates.py --apply    # 备份 + 合并 + 删除
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl

# 列索引与表头硬编码，避免依赖 mdcx.config.resources（测试环境会用桩模块替换）。
COL_JP = 0
COL_ZH_CN = 1
COL_ZH_TW = 2
COL_KEYWORD = 3
COL_TMDBID = 5
COL_BIRTH_DATE = 7
COL_BIO = 8
DB_HEADERS = ["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"]

# 空缺字段补全时覆盖的列（不含 tmdbid/tmdb_url，那些以有 id 行为准）
_FILL_COLS = (COL_ZH_CN, COL_ZH_TW, COL_BIRTH_DATE, COL_BIO)


def _alias_sort_key(alias: str) -> tuple[int, int, str]:
    """别名排序键：纯中文(全汉字)优先 -> 中日混合 -> 日文(假名) -> 罗马音/英文。

    与 mdcx.config.resources._alias_sort_key 保持一致。
    """
    han_count = sum(1 for ch in alias if "\u4e00" <= ch <= "\u9fff")
    kana_count = sum(1 for ch in alias if "\u3040" <= ch <= "\u30ff")
    if han_count > 0 and kana_count == 0:
        return (0, 0, alias)
    if han_count > 0 and kana_count > 0:
        return (1, 0, alias)
    if kana_count > 0:
        return (2, 0, alias)
    return (3, 0, alias)


def _merge_keyword_union(local_kw: str, backup_kw: str) -> str:
    """别名并集合并（去重），按中文优先排序。与 resources 同名函数逻辑一致。"""
    merged: list[str] = []
    seen: set[str] = set()
    for group in (local_kw, backup_kw):
        if not group:
            continue
        for item in (k.strip() for k in str(group).split(",") if k.strip()):
            key = item.casefold()
            if key not in seen:
                seen.add(key)
                merged.append(item)
    merged.sort(key=_alias_sort_key)
    return ",".join(merged)


def _get_actor_db_sheet(wb):
    """显式取「演员数据库」sheet，不依赖 sheet 顺序。"""
    if "演员数据库" in wb.sheetnames:
        return wb["演员数据库"]
    return wb.active


from mdcx.core.tmdb_actor import _full_to_half, norm_name  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DB_PATH = ROOT / "resources" / "userdata" / "actor_database.xlsx"


def _norm(s: str) -> str:
    return _full_to_half(norm_name(s or "")).upper()


def scan(rows: list[tuple]) -> list[tuple[str, str, str]]:
    """返回确证重复对: (无id jp, 有id jp, 有id tmdbid)。"""
    id_jp: dict[str, list[str]] = {}
    id_tid: dict[str, str] = {}
    for row in rows:
        jp = str(row[COL_JP] or "").strip()
        tid = str(row[COL_TMDBID] or "").strip() if len(row) > COL_TMDBID else ""
        if not jp or not tid.isdigit():
            continue
        id_jp.setdefault(_norm(jp), []).append(jp)
        id_tid[jp] = tid

    pairs: list[tuple[str, str, str]] = []
    for row in rows:
        jp = str(row[COL_JP] or "").strip()
        tid = str(row[COL_TMDBID] or "").strip() if len(row) > COL_TMDBID else ""
        if not jp or tid.isdigit():
            continue
        kw = str(row[COL_KEYWORD] or "").strip() if len(row) > COL_KEYWORD else ""
        names = [jp] + [k.strip() for k in kw.split(",") if k.strip()]
        seen: set[str] = set()
        for name in names:
            n = _norm(name)
            if not n or n in seen:
                continue
            seen.add(n)
            for tjp in id_jp.get(n, []):
                pairs.append((jp, tjp, id_tid[tjp]))
    return pairs


def _rows_of(ws) -> list[tuple]:
    return list(ws.iter_rows(min_row=2, max_col=len(DB_HEADERS), values_only=True))


def merge_duplicates(ws, pairs: list[tuple[str, str, str]]) -> tuple[int, int]:
    """执行合并：返回 (并入的有 id 行数, 删除的无 id 行数)。

    无 id 行主名 + 全部别名并入有 id 行 keyword 列（中文优先去重）；
    有 id 行空缺字段用无 id 行补全；随后删除无 id 行。
    """
    rows = _rows_of(ws)
    jp_row: dict[str, int] = {}
    for i, row in enumerate(rows, start=2):
        jp = str(row[COL_JP] or "").strip()
        if jp:
            jp_row.setdefault(jp, i)
    row_data = {i: rows[i - 2] for i in range(2, len(rows) + 2)}

    merge_plan: dict[int, list[tuple[str, str]]] = {}
    noid_rows: set[int] = set()
    for jp, tjp, _tid in pairs:
        target = jp_row[tjp]
        src_row_no = jp_row[jp]
        src_kw = str(row_data[src_row_no][COL_KEYWORD] or "").strip()
        merge_plan.setdefault(target, []).append((jp, src_kw))
        noid_rows.add(src_row_no)

    for target, srcs in merge_plan.items():
        trow = row_data[target]
        cur_kw = str(trow[COL_KEYWORD] or "").strip()
        existing = {k.strip() for k in cur_kw.split(",") if k.strip()}
        extra: list[str] = []
        for src_jp, src_kw in srcs:
            for name in [src_jp] + [k.strip() for k in src_kw.split(",") if k.strip()]:
                if name and name not in existing:
                    existing.add(name)
                    extra.append(name)
        if extra:
            ws.cell(row=target, column=COL_KEYWORD + 1, value=_merge_keyword_union(cur_kw, ",".join(extra)))
        for src_jp, _src_kw in srcs:
            srow = row_data[jp_row[src_jp]]
            for col in _FILL_COLS:
                cur = ws.cell(row=target, column=col + 1).value
                new = srow[col] if len(srow) > col else None
                if (cur is None or str(cur).strip() == "") and new not in (None, ""):
                    ws.cell(row=target, column=col + 1, value=new)

    for i in sorted(noid_rows, reverse=True):
        ws.delete_rows(i, 1)

    return len(merge_plan), len(noid_rows)


def main(argv: list[str] | None = None) -> int:
    argv = sys.argv if argv is None else argv
    if not DB_PATH.exists():
        print(f"数据库不存在: {DB_PATH}")
        return 1

    wb = openpyxl.load_workbook(DB_PATH)
    ws = _get_actor_db_sheet(wb)
    pairs = scan(_rows_of(ws))

    if not pairs:
        print("无确证重复。")
        wb.close()
        return 0

    print(f"确证重复 {len(pairs)} 对:")
    for jp, tjp, tid in pairs:
        print(f"  {jp} -> {tjp} (id={tid})")

    if "--apply" not in argv:
        print("\n（预览模式，加 --apply 执行合并）")
        wb.close()
        return 0

    backup = DB_PATH.with_suffix(".bak.xlsx")
    shutil.copy(DB_PATH, backup)
    print(f"备份: {backup}")

    merged, removed = merge_duplicates(ws, pairs)
    wb.save(DB_PATH)
    wb.close()
    print(f"合并 {merged} 个有 id 行，删除 {removed} 个无 id 行")
    from scripts.db_guard import validate_after_save

    ok = validate_after_save(DB_PATH)
    if not ok:
        print("⚠️ 保存后校验发现 error 级问题，请检查！")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
