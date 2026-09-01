"""清洗 actor_database.xlsx 字段噪声。

针对女优库中混入的多类污染：
1. 日文原名/中文名/繁体名 = 占位符或描述词（素人/人妻/女優情報/复元 等非人名）
2. 中文名/繁体名 = 名字+年龄标注（如「涼子 20歳」）或作品系列名
3. 别名 = 混入作品标题（ラグジュTV / ママ友喰い / VOL.xx 等整段标题）
4. 名字 = 混入系列/站点标签（パコパコママ / FC2 / ロリ主婦 / 1000人斬り / 天然むすめ 等）
5. 名字 = 混入年份（（2015）/【2016年】等）
6. 名字 = 纯占位符（FC2 / 素人 / 抜群なアイドル店員 等非人名）
7. 别名 = 残留人妻/熟女/着エロ 等类型标签
8. 别名 = 悬空斜杠/残括号（ただえりさ /、真東愛 / Mahigashi Ai））
9. 简介 = 1-2 字符碎片 → 置空

仅对发生污染的行做字段级修正，不删除任何行。输出统计报告。
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from mdcx.utils import actor_clean as ac  # noqa: E402

DB_PATH = ROOT / "resources" / "userdata" / "actor_database.xlsx"


def _was_placeholder(name: str) -> bool:
    """判断清洗前是否纯占位符（用于统计置空）。"""
    from mdcx.utils.actor_clean import _is_placeholder_name

    return _is_placeholder_name(name)


def main() -> int:
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["演员数据库"]

    stat = {
        "jp_placeholder": 0,
        "cn_placeholder": 0,
        "tw_placeholder": 0,
        "alias_title": 0,
        "name_tag": 0,
        "name_year": 0,
        "name_placeholder": 0,
        "alias_tag": 0,
        "bio_short": 0,
        "alias_extract": 0,
        "slash_fix": 0,
        "name_annotation": 0,
        "orphan_url_clear": 0,
    }
    detail: list[tuple[str, int, str, str]] = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        jp = row[0].value
        cn = row[1].value
        tw = row[2].value
        alias = row[3].value
        bio = row[8].value
        name_label = jp or cn or tw or "(无名)"

        # 1) 名字字段统一语义清洗（占位符置空 + 系列标签/年份/标注/描述剥离）
        for idx in (1, 2, 3):
            cur = row[idx - 1].value
            if cur is None:
                continue
            cleaned = ac.clean_actor_name(str(cur))
            if not cleaned:
                if _was_placeholder(str(cur)):
                    row[idx - 1].value = None
                    stat[["jp_placeholder", "cn_placeholder", "tw_placeholder"][idx - 1]] += 1
                    detail.append((name_label, idx, str(cur)[:50], "(置空)"))
            elif cleaned != str(cur):
                row[idx - 1].value = cleaned
                stat["name_tag"] += 1
                detail.append((name_label, idx, str(cur)[:50], cleaned[:50]))

        # 3) 名字含已知真别名（括号内容移入别名列）
        for idx, val in ((1, jp), (2, cn), (3, tw)):
            if val:
                m = ac._KNOWN_ALIAS_IN_PAREN.search(str(val))
                if m:
                    alias_part = m.group(1)
                    cleaned = ac.clean_actor_name(str(val))
                    cleaned = cleaned.replace(f"（{alias_part}）", "").replace(f"({alias_part})", "").strip()
                    if cleaned:
                        row[idx - 1].value = cleaned
                        existing_alias = row[3].value or ""
                        new_alias = f"{existing_alias},{alias_part}".strip(",") if existing_alias else alias_part
                        row[3].value = new_alias
                        stat["alias_extract"] += 1
                        detail.append((name_label, idx, str(val)[:40], f"{cleaned[:20]} +别名{alias_part}"))

        # 4) 别名统一语义清洗（作品标题/标签/悬空斜杠/标注/描述剥离）
        if alias:
            # 先记录是否有斜杠异常
            had_slash_issue = False
            for _seg in str(alias).split(","):
                _seg = _seg.strip()
                if "/" not in _seg and "／" not in _seg:
                    continue
                _rhs = _seg.split("/")[-1].split("／")[-1].strip()
                if not _rhs or _rhs in {")", "）", "]", "】"} or re.search(r"[）)】\]\u3000]$", _rhs):
                    had_slash_issue = True
                    break
            cleaned = ac.clean_actor_keyword(str(alias))
            if cleaned != str(alias):
                row[3].value = cleaned
                stat["alias_title"] += 1
                if had_slash_issue:
                    stat["slash_fix"] += 1
                detail.append((name_label, 4, str(alias)[:50], cleaned[:50] or "(置空)"))

        # 6) 简介 1-2 字符碎片 → 置空
        if bio and len(str(bio).strip()) <= 2:
            row[8].value = None
            stat["bio_short"] += 1

        # 7) 无 tmdbid 但 url 有值 → 清除（AVdb 源数据错误映射残留，反推 id 会固化错误）
        tid = row[5].value
        if not tid and row[6].value:
            row[6].value = None
            stat["orphan_url_clear"] += 1

    wb.save(DB_PATH)

    print(f"📊 清洗报告 (共处理 {ws.max_row - 1} 行)")
    print(f"  日文原名占位符置空: {stat['jp_placeholder']}")
    print(f"  中文名占位符置空: {stat['cn_placeholder']}")
    print(f"  繁体名占位符置空: {stat['tw_placeholder']}")
    print(f"  名字混入系列标签/年份剥离: {stat['name_tag']}")
    print(f"  名字标注括号剥离: {stat['name_annotation']}")
    print(f"  名字真别名移入别名列: {stat['alias_extract']}")
    print(f"  别名剔除作品标题/标签: {stat['alias_title']}")
    print(f"  别名悬空斜杠/残括号修复: {stat['slash_fix']}")
    print(f"  简介碎片置空: {stat['bio_short']}")
    print(f"  无tid孤儿url清除: {stat['orphan_url_clear']}")
    print()
    print("详情(前 25 条):")
    for name, col, old, new in detail[:25]:
        col_name = {1: "日文原名", 2: "中文名", 3: "繁体名", 4: "别名"}[col]
        print(f"  [{name}] {col_name}: {old} -> {new}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
