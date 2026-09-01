"""ASIN 数据库冲突清洗：图像相似度裁决错配行，移入「待修正」sheet。

适用场景：amazon_asin_database.xlsx 中同一 ASIN 被多个番号占用（如 BEST 合集
页挂多个番号、或历史导入 bug 产生的错配），通过封面图像相似度判断哪个番号
真正对应该 ASIN，错配行移到「待修正」sheet 保留待补数据，主表不留重复。

裁决方法（重要，勿回退到标题文本匹配）：
- 标题文本比对不可靠（BEST/合集标题归一化误判多），已废弃
- javdb 竖版 cover 是重压图，与 DMM 原图相似度仅 0.5~0.7，不可作裁判
- 可信裁判链：DMM 竖版 ps → DMM 横版 pl 裁右半（0.97+）→ javdb 仅兜底参考

基准图来源优先级：tenhow.net/images/{ASIN}.jpg → 库内 poster_url(media-amazon)。

用法:
    python scripts/clean_asin_db_conflicts.py            # 扫描并预览
    python scripts/clean_asin_db_conflicts.py --apply    # 执行移动
"""

from __future__ import annotations

import argparse
import asyncio
import shutil
import sys
import tempfile
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path

import httpx
import openpyxl
from PIL import Image

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from mdcx.core.web import _cover_similarity, _cut_thumb_right_image  # noqa: E402

DB_PATH = ROOT / "userdata" / "amazon_asin_database.xlsx"
FIX_SHEET_NAME = "待修正"
SCORE_THRESHOLD = 0.82  # 与 _verify_soft_amazon_poster 软校验同阈值

TENHOW_URL = "https://tenhow.net/images/{asin}.jpg"
PICS_DMM_MONO = "https://pics.dmm.co.jp/mono/movie/adult/{cid}/{cid}{suf}.jpg"
AWSIMGSRC = "https://awsimgsrc.dmm.co.jp/pics_dig/digital/video/{cid}/{cid}{suf}.jpg"
THEJAVDB_API = "https://api.thejavdb.net/v1/movies"

_http: httpx.AsyncClient | None = None


async def _fetch_image(url: str) -> tuple[Image.Image | None, int]:
    """下载图片, 返回 (Image, 字节数)。非图/占位图(<4KB)返回 None。"""
    global _http
    if _http is None:
        _http = httpx.AsyncClient(timeout=15, follow_redirects=True)
    try:
        r = await _http.get(url)
        if r.status_code != 200 or len(r.content) < 4096:
            return None, len(r.content)
        img = Image.open(BytesIO(r.content)).convert("RGB")
        img.load()
        return img, len(r.content)
    except Exception:
        return None, 0


async def _dmm_reference_via_thejavdb(number: str) -> tuple[Image.Image | None, str]:
    """优先源: 查 thejavdb api 拿 DMM 图 URL, 优先 frontcover(竖版), 失败降级 fullcover 裁右半.

    thejavdb 覆盖主流番号但极少数缺(如下架/新番). 返回的就是 pics.dmm 真链, 免去试 cid 前缀.
    """
    global _http
    if _http is None:
        _http = httpx.AsyncClient(timeout=15, follow_redirects=True)
    try:
        r = await _http.get(THEJAVDB_API, params={"q": number}, timeout=10)
        if r.status_code != 200:
            return None, ""
        data = r.json()
        for key, label in (("frontcover_url", "thejavdb.front"), ("fullcover_url", "thejavdb.full裁右半")):
            url = data.get(key)
            if not url or "pics.dmm" not in str(url):
                continue
            img, _ = await _fetch_image(url)
            if img is None:
                continue
            if "full" in key and img.size[0] > img.size[1]:
                img2 = _cut_thumb_right_image(img)
                img.close()
                img = img2
            return img, label
    except Exception:
        pass
    return None, ""


async def _dmm_reference(number: str) -> tuple[Image.Image | None, str]:
    """按番号取 DMM 官方图作裁判.

    裁决链(实测可信次序):
    1. thejavdb api (直接返回 DMM 图 URL, 无需猜 cid)
    2. 直构 DMM cid 前缀枚举 (兼容老规则)
    """
    # 1. thejavdb api
    img, src = await _dmm_reference_via_thejavdb(number)
    if img is not None:
        return img, src

    # 2. 直构 cid 逐个试
    norm = number.lower().replace("-", "").replace(" ", "")
    prefixes = ["", "1", "13", "49", "118", "55", "57", "83", "436", "5042", "5642"]
    for suf in ("ps", "pl"):
        for pre in prefixes:
            cid = f"{pre}{norm}"
            for base in (PICS_DMM_MONO, AWSIMGSRC):
                img, _ = await _fetch_image(base.format(cid=cid, suf=suf))
                if img is None:
                    continue
                if suf == "pl":
                    img = _cut_thumb_right_image(img)
                    return img, f"dmm_{suf}({cid}裁右半)"
                return img, f"dmm_{suf}({cid})"
    return None, ""


async def _judge_group(asin: str, numbers: list[str], poster_url: str | None) -> tuple[str, list[dict]]:
    """裁决一组同 ASIN 番号。返回 (verdict, scores)。verdict: unique/all_match/none"""
    # 基准图：tenhow → 库内 poster_url
    ref = None
    img, _ = await _fetch_image(TENHOW_URL.format(asin=asin))
    if img is None and poster_url and "media-amazon" in str(poster_url):
        img, _ = await _fetch_image(str(poster_url))
    if img is not None:
        ref = img
    if ref is None:
        return "none", [{"number": n, "score": None, "src": "", "note": "无基准图"} for n in numbers]

    scores = []
    for num in numbers:
        img, src = await _dmm_reference(num)
        if img is None:
            scores.append({"number": num, "score": None, "src": "", "note": "无DMM图"})
            continue
        s, _, _ = _cover_similarity(ref, img)
        img.close()
        scores.append({"number": num, "score": round(s, 3), "src": src, "note": ""})
    ref.close()

    hits = [s for s in scores if s["score"] is not None and s["score"] >= SCORE_THRESHOLD]
    if len(hits) == 1:
        return "unique", scores
    if len(hits) > 1:
        return "all_match", scores  # 同一商品多番号发行(原版+再版), 两边都对
    return "none", scores


async def run(db_path: Path, apply: bool, limit: int = 0) -> int:
    wb = openpyxl.load_workbook(db_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    col_number, col_asin, col_poster = 0, 1, 4  # 影片番号/ASIN/封面URL

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        num = str(row[col_number] or "").strip()
        asin = str(row[col_asin] or "").strip()
        poster = str(row[col_poster] or "").strip() if len(row) > col_poster else ""
        if num and asin:
            rows.append({"row": i, "number": num, "asin": asin, "poster": poster})

    # 1. 完全重复行(番号+ASIN+标题一致)
    seen, dup_full = {}, []
    for r in rows:
        key = (r["number"], r["asin"])
        if key in seen:
            dup_full.append(r)
        else:
            seen[key] = r

    # 2. 同 ASIN 挂多番号(剔除完全重复后)
    by_asin: dict[str, list[dict]] = defaultdict(list)
    for r in seen.values():
        by_asin[r["asin"]].append(r)
    conflicts = {a: rs for a, rs in by_asin.items() if len(rs) > 1}
    if limit > 0:
        conflicts = dict(list(conflicts.items())[:limit])

    print(f"总行数: {len(rows)}, 完全重复行: {len(dup_full)}, 一ASIN多番号: {len(conflicts)} 组")
    verdicts: dict[str, tuple[str, list[dict], list[dict]]] = {}
    for i, (asin, rs) in enumerate(conflicts.items(), 1):
        poster = next((r["poster"] for r in rs if r["poster"]), None)
        verdict, scores = await _judge_group(asin, [r["number"] for r in rs], poster)
        verdicts[asin] = (verdict, scores, rs)
        print(f"[{i}/{len(conflicts)}] {asin}: {verdict} " + ", ".join(f"{s['number']}={s['score']}" for s in scores))

    if not apply:
        print("\n（预览模式，加 --apply 执行移动）")
        return 0

    # 仅处理 unique 组：错配行移到待修正 sheet
    if FIX_SHEET_NAME in wb.sheetnames:
        fix_ws = wb[FIX_SHEET_NAME]
    else:
        fix_ws = wb.create_sheet(FIX_SHEET_NAME)
        fix_ws.append(header + ["备注", "来源"])

    # 先收集所有待移行(读值入待修正), 再统一倒序删除, 避免行号位移
    to_delete: set[int] = set()
    moved = 0
    for asin, (verdict, scores, rs) in verdicts.items():
        if verdict != "unique":
            continue
        matched = next(s["number"] for s in scores if s["score"] and s["score"] >= SCORE_THRESHOLD)
        rowmap = {r["number"]: r["row"] for r in rs}
        for s in scores:
            if s["number"] == matched:
                continue
            rowno = rowmap[s["number"]]
            if rowno in to_delete:
                continue
            vals = [ws.cell(rowno, c).value for c in range(1, len(header) + 1)]
            if vals[0] is None:
                continue
            sc = f"相似度{s['score']}" if s["score"] is not None else "无对比图"
            note = f"该番号与ASIN {asin} 封面不符({sc})，正确番号疑似为 {matched}"
            fix_ws.append(vals + [note, "clean_asin_db_conflicts"])
            to_delete.add(rowno)
            moved += 1

    to_delete.update(r["row"] for r in dup_full)
    for rowno in sorted(to_delete, reverse=True):
        ws.delete_rows(rowno, 1)

    # 备份后保存
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = db_path.with_suffix(f".bak_{stamp}.xlsx")
    shutil.copy2(db_path, backup)
    wb.save(db_path)
    print(f"完成: 移动错配 {moved} 行→「{FIX_SHEET_NAME}」, 删完全重复 {len(dup_full)} 行, 备份 {backup.name}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="ASIN 数据库冲突清洗(图像相似度裁决)")
    parser.add_argument("--apply", action="store_true", help="执行(默认仅预览)")
    parser.add_argument("--db", type=Path, default=DB_PATH, help="目标库")
    parser.add_argument("--limit", type=int, default=0, help="只处理前N组冲突(冒烟测试用)")
    args = parser.parse_args()
    if not args.db.exists():
        print(f"数据库不存在: {args.db}")
        return 1
    # 在临时副本上跑, 避免锁/半写污染原库
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tmp = Path(tf.name)
    shutil.copy2(args.db, tmp)
    try:
        return asyncio.run(run(args.db if args.apply else tmp, args.apply, args.limit))
    finally:
        tmp.unlink(missing_ok=True)


if __name__ == "__main__":
    sys.exit(main())
