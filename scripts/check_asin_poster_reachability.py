#!/usr/bin/env python3
"""判定 ASIN 库中哪些番号的缓存"永远不会被读".

依据 web.py `_should_skip_amazon_for_existing_poster` 的前置 Poster 大小校验:
当前刮削得到的 Poster 已达标 (DMM 宽>=700 / 字节>=400KB) 时整段跳过日亚,
对应 ASIN 缓存行成为死数据。

判定链 (与真实流程一致):
  1. DMM ps.jpg 竖幅高清直链 (dmm_direct.build_aws_poster_candidates):
     量宽/字节, 宽>=700 或 字节>=400KB -> 达标 (真实流程 poster 即此图)
  2. DMM pl.jpg 横幅 + 右裁剪 (image._right_crop_box 几何):
     裁剪结果 宽>=700 或 字节>=400KB -> 达标
  3. javdb app 搜索 thumb_url 兜底: 字节>=400KB -> 达标
  4. 都不达标/查不到 -> 保留主表 (宁多勿删)

用法:
  python scripts/check_asin_poster_reachability.py --limit 100        # 试跑
  python scripts/check_asin_poster_reachability.py --max-seconds N    # 长任务模式

断点续传: /tmp/opencode/asin_reach_state.json
结果:     /tmp/opencode/asin_reach_result.xlsx (试跑) / 由汇总脚本写回主库
"""

from __future__ import annotations

import argparse
import asyncio
import io
import json
import signal
import sys
import time
from pathlib import Path

import aiohttp
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from PIL import Image  # noqa: E402

from mdcx.crawlers.dmm_direct import build_aws_cover_candidates, build_aws_poster_candidates  # noqa: E402
from mdcx.crawlers.javdb_app import _API_BASE, _API_FALLBACKS, _get_api_url, make_signature  # noqa: E402

DB_PATH = ROOT / "userdata/amazon_asin_database.xlsx"
STATE_PATH = Path("/tmp/opencode/asin_reach_state.json")
RESULT_PATH = Path("/tmp/opencode/asin_reach_result.xlsx")
DONE_FLAG = Path("/tmp/opencode/asin_reach_done")

MIN_BYTES = 400 * 1024
MIN_WIDTH = 700

SEM_DMM = asyncio.Semaphore(8)
SEM_JAVDB = asyncio.Semaphore(3)

_stop = asyncio.Event()


def _on_sig(*_: object) -> None:
    _stop.set()


def _num_key(number: str) -> tuple[str, int]:
    """去前导零的比对 key: (系列字母大写, int 数字). 与清洗规则一致."""
    import re

    m = re.match(r"^([A-Za-z]+)-?(\d+)$", (number or "").strip())
    if not m:
        return (number or "").upper(), -1
    return m.group(1).upper(), int(m.group(2))


def load_rows(limit: int, skip: int = 0) -> list[dict]:
    wb = load_workbook(DB_PATH, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    out = []
    for r in rows:
        number = str(r[0] or "").strip()
        asin = str(r[1] or "").strip()
        if number and asin:
            out.append({"number": number, "asin": asin})
    if skip > 0:
        out = out[skip:]
    if limit > 0:
        out = out[:limit]
    return out


def load_state() -> dict:
    if STATE_PATH.exists():
        return json.loads(STATE_PATH.read_text())
    return {}


def save_state(state: dict) -> None:
    tmp = STATE_PATH.with_suffix(".tmp")
    tmp.write_text(json.dumps(state, ensure_ascii=False))
    tmp.replace(STATE_PATH)


def _right_crop_box(width: int, height: int) -> tuple[int, int, int, int]:
    """与 mdcx/core/image.py::_right_crop_box 保持同一几何."""
    ax, ay, bx, by = width / 1.9, 0, width, height
    if width == 800:
        if height == 439:
            ax, ay, bx, by = 420, 0, width, height
        elif 499 <= height <= 503:
            ax, ay, bx, by = 437, 0, width, height
        else:
            ax, ay, bx, by = 421, 0, width, height
    elif width == 840 and height == 472:
        ax, ay, bx, by = 473, 0, 788, height
    return int(ax), int(ay), int(bx), int(by)


def _judge(width: int, height: int, size: int) -> bool:
    return width >= MIN_WIDTH or size >= MIN_BYTES


async def _fetch_bytes(session: aiohttp.ClientSession, url: str, sem: asyncio.Semaphore) -> bytes | None:
    async with sem:
        try:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=20)) as r:
                if r.status != 200:
                    return None
                return await r.read()
        except Exception:
            return None


async def check_dmm_ps(session: aiohttp.ClientSession, number: str) -> dict | None:
    """遍历全部 ps 候选直到达标; awsimgsrc 同一 URL 可能返回小图或高清, 必须逐个实测."""
    last: dict | None = None
    for url in build_aws_poster_candidates(number):
        if _stop.is_set():
            return last
        data = await _fetch_bytes(session, url, SEM_DMM)
        if not data:
            continue
        try:
            img = Image.open(io.BytesIO(data))
            w, h = img.size
            img.close()
        except Exception:
            continue
        if w < 100:
            continue  # 占位图
        if _judge(w, h, len(data)):
            return {"hit": "dmm_ps", "w": w, "h": h, "bytes": len(data), "url": url}
        last = {"hit": "", "dmm_ps_w": w, "dmm_ps_bytes": len(data), "url": url}
    return last


async def check_dmm_pl_crop(session: aiohttp.ClientSession, number: str) -> dict | None:
    """遍历全部 pl 候选, 裁剪后判定; 与 ps 同样逐候选实测."""
    last: dict | None = None
    for url in build_aws_cover_candidates(number):
        if _stop.is_set():
            return last
        data = await _fetch_bytes(session, url, SEM_DMM)
        if not data:
            continue
        try:
            img = Image.open(io.BytesIO(data)).convert("RGB")
            w, h = img.size
            if w < 100:
                img.close()
                continue  # 占位图
            prop = h / w if w else 0
            if prop >= 1.4:
                cropped = img
            else:
                ax, ay, bx, by = _right_crop_box(w, h)
                cropped = img.crop((ax, ay, bx, by))
            buf = io.BytesIO()
            cropped.save(buf, "JPEG")
            cw, ch = cropped.size
            cbytes = buf.tell()
            cropped.close()
            img.close()
        except Exception:
            continue
        if _judge(cw, ch, cbytes):
            return {"hit": "dmm_pl_crop", "w": cw, "h": ch, "bytes": cbytes, "url": url}
        last = {"hit": "", "dmm_pl_w": cw, "dmm_pl_bytes": cbytes, "url": url}
    return last


async def check_javdb_app(session: aiohttp.ClientSession, number: str) -> dict | None:
    headers = {
        "jdsignature": make_signature(),
        "accept-language": "zh",
        "User-Agent": "Dart/3.5 (dart:io)",
    }
    key = _num_key(number)
    poster_url = None
    for host in [_API_BASE, *_API_FALLBACKS]:
        if _stop.is_set():
            return None
        try:
            url = _get_api_url(host, "/api/v2/search", {"q": number, "page": "1"})
            async with SEM_JAVDB:
                async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=15)) as r:
                    if r.status != 200:
                        continue
                    data = await r.json()
            if not data.get("success"):
                continue
            movies = (data.get("data") or {}).get("movies") or []
            for mv in movies:
                kn = _num_key(str(mv.get("number") or ""))
                if key[1] >= 0 and kn[1] >= 0 and kn == key:
                    poster_url = str(mv.get("thumb_url") or "").strip()
                    break
            if poster_url:
                break
        except Exception:
            continue
    if not poster_url:
        return None
    data = await _fetch_bytes(session, poster_url, SEM_JAVDB)
    if not data:
        return {"hit": "", "javdb_err": "img_fail", "url": poster_url}
    return {"hit": "javdb" if len(data) >= MIN_BYTES else "", "javdb_bytes": len(data), "url": poster_url}


async def check_row(session: aiohttp.ClientSession, number: str) -> dict:
    rec: dict = {"hit": "", "checked_at": time.strftime("%Y-%m-%d %H:%M:%S")}
    r1 = await check_dmm_ps(session, number)
    if r1:
        rec.update(r1)
        if r1.get("hit"):
            return rec
    r2 = await check_dmm_pl_crop(session, number)
    if r2:
        rec.update(r2)
        if r2.get("hit"):
            return rec
    r3 = await check_javdb_app(session, number)
    if r3:
        rec.update(r3)
    if not rec.get("hit"):
        rec["hit"] = "keep"
    return rec


def flush_result_xlsx(state: dict, rows: list[dict]) -> None:
    """把当前 state 写成结果 xlsx: 命中(移走) + 保留 两个 sheet."""
    from openpyxl import Workbook

    wb = Workbook()
    ws_move = wb.active
    ws_move.title = "poster达标不走日亚"
    ws_keep = wb.create_sheet("保留主表")
    header = ["影片番号", "ASIN 编号", "判定", "宽", "高", "字节", "证据URL"]
    ws_move.append(header)
    ws_keep.append(header)
    for row in rows:
        st = state.get(row["asin"])
        if not st:
            continue
        hit = st.get("hit") or ""
        out = ws_move if hit and hit != "keep" else ws_keep
        out.append(
            [
                row["number"],
                row["asin"],
                hit,
                st.get("w", ""),
                st.get("h", ""),
                st.get("bytes", st.get("javdb_bytes", "")),
                st.get("url", ""),
            ]
        )
    wb.save(RESULT_PATH)


async def run(limit: int, max_seconds: int, skip: int = 0) -> None:
    rows = load_rows(limit, skip)
    state = load_state()
    pending = [r for r in rows if r["asin"] not in state]
    print(f"总数 {len(rows)}, 已完成 {len(rows) - len(pending)}, 待跑 {len(pending)}", flush=True)
    if not pending or _stop.is_set():
        return

    start = time.time()
    conn = aiohttp.TCPConnector(limit=16)
    done_count = 0
    async with aiohttp.ClientSession(connector=conn) as session:
        queue: asyncio.Queue = asyncio.Queue()
        for r in pending:
            queue.put_nowait(r)

        async def worker() -> None:
            nonlocal done_count
            while not _stop.is_set():
                try:
                    row = queue.get_nowait()
                except asyncio.QueueEmpty:
                    return
                rec = await check_row(session, row["number"])
                state[row["asin"]] = rec
                done_count += 1
                if done_count % 20 == 0:
                    save_state(state)
                    el = time.time() - start
                    speed = done_count / el * 60 if el > 0 else 0
                    print(f"进度 {done_count}/{len(pending)} 速度 {speed:.0f}行/分", flush=True)

        workers = [asyncio.create_task(worker()) for _ in range(10)]
        try:
            await asyncio.wait_for(asyncio.gather(*workers), timeout=max_seconds if max_seconds > 0 else None)
        except TimeoutError:
            print("到达 max-seconds, 停止", flush=True)
        _stop.set()

    save_state(state)
    hits = {}
    for rec in state.values():
        h = rec.get("hit") or "unknown"
        hits[h] = hits.get(h, 0) + 1
    print(f"完成累计 {len(state)}: {hits}", flush=True)
    flush_result_xlsx(state, rows)
    if all(r["asin"] in state for r in rows):
        DONE_FLAG.write_text(time.strftime("%Y-%m-%d %H:%M:%S"))
        print("全部完成", flush=True)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0, help="只处理前 N 行, 0=全量")
    parser.add_argument("--skip", type=int, default=0, help="跳过前 N 行")
    parser.add_argument("--max-seconds", type=int, default=0, help="本次最长运行秒数, 0=不限")
    args = parser.parse_args()
    signal.signal(signal.SIGINT, _on_sig)
    signal.signal(signal.SIGTERM, _on_sig)
    asyncio.run(run(args.limit, args.max_seconds, args.skip))


if __name__ == "__main__":
    main()
