"""从出厂演员库删除「非演员」混入行。

第 1 类：导演/幕后人员——TMDB known_for_department != Acting 的 AV 从业者。
第 2 类：描述词/占位符——非人名的标签词（本妻/多数/女優不明 等）。
第 3 类：确证的非 AV 主流人物（埃里克·坎通纳/艾琳娜·拉尼娜 等）。

删除前在系统临时目录保留原始库副本（工作区安全网，不写入仓库）。
用法:
    python scripts/clean_actor_db_non_actors.py            # 预览
    python scripts/clean_actor_db_non_actors.py --apply    # 执行删除
    python scripts/clean_actor_db_non_actors.py --db <path> --apply  # 测试用副本
"""

from __future__ import annotations

import argparse
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl

from mdcx.config.resources import get_actor_db_sheet  # noqa: E402
from scripts.db_guard import validate_after_save  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DB_PATH = ROOT / "resources" / "userdata" / "actor_database.xlsx"
NON_ACTING_FILE = Path(tempfile.gettempdir()) / "non_acting.txt"

# 第 2 类：描述词/占位符（人工甄别，非人名）
_DESC_ROWS = {
    "愛嬌抜群のおんな",
    "愛人2",
    "坂道系のツンデレ女子大生",
    "本妻",
    "多数",
    "高学歴娘",
    "宮崎 後藤夫妻 デビ みずき はるか りえ",
    "金持ちを食い荒らす女",
    "可愛すぎる美女",
    "女優不明",
    "女優多数",
    "人物不明",
    "貧乳美女",
    "神待ち娘",
    "円光娘",
    "S級インテリ美女",
}

# 第 3 类：确证的非 AV 主流人物（有 id，人工确认后硬删）
_NON_AV_NAMES = {
    "埃里克·坎通纳",  # 足球明星
    "艾琳娜·拉尼娜",  # 主流演员
    # gender=0（TMDB 漏标性别）的漏网男优，男性名 + adult=True + 作品佐证
    "阿部次郎",
    "白川久雄",
    "本理一郎",
    "蒼武蔵",
    "城山剛",
    "春原一郎",
    "剛",
    "剛力望",
    "光村明男",
    "光一郎",
    "花笠義男",
    "金本慎吾",
    "井草雄二",
    "康介 Kosuke",
    "鈴木洋介",
    "龍八郎太",
    "南大介",
    "青山新太郎",
    "山本吾郎",
    "神谷有樹彦",
    "堂本聖吾",
    "宇野涼助",
    "重枝久雄",
}


def collect_non_acting_ids() -> set[int]:
    """读第 1 类清单，返回待删 tmdbid。"""
    ids: set[int] = set()
    if not NON_ACTING_FILE.exists():
        return ids
    for line in NON_ACTING_FILE.read_text(encoding="utf-8").splitlines():
        parts = line.rstrip().split("\t")
        if len(parts) == 3 and parts[1].isdigit():
            ids.add(int(parts[1]))
    return ids


def _rebuild_hyperlinks(ws) -> None:
    """按 cell 实际坐标重建全部超链接。

    openpyxl 的 delete_rows 移动单元格值时不同步 hyperlink 的 ref：下方行超链接
    ref 仍指向旧行号（错位），被删行的超链接对象残留（孤儿）。重建方式：遍历所有
    绑定 hyperlink 的 cell，清空后按 cell 真实坐标重新绑定，使保存时 XML 中
    hyperlink ref 与实际 cell 位置一致。
    """
    hlinks: list[tuple[str, str]] = []
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.hyperlink and cell.hyperlink.target:
                hlinks.append((cell.coordinate, cell.hyperlink.target))
    ws._hyperlinks = []
    for coord, target in hlinks:
        ws[coord].hyperlink = target


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="删除出厂库中的非演员混入行")
    parser.add_argument("--apply", action="store_true", help="执行删除")
    parser.add_argument("--db", type=Path, default=DB_PATH, help="目标库路径（默认出厂库）")
    args = parser.parse_args(argv)

    db_path = args.db
    if not db_path.exists():
        print(f"数据库不存在: {db_path}")
        return 1

    non_acting_ids = collect_non_acting_ids()
    if not non_acting_ids:
        print(f"⚠️ 第 1 类清单为空: {NON_ACTING_FILE}（仅处理第 2/3 类）")

    wb = openpyxl.load_workbook(db_path)
    ws = get_actor_db_sheet(wb)

    # 收集待删行号
    del_rows: set[int] = set()
    desc_found: list[tuple[int, str]] = []
    non_acting_found: list[tuple[int, str, int]] = []
    placeholder_found: list[tuple[int, str]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        jp = str(row[0] or "").strip()
        tid_val = str(row[5] or "").strip() if len(row) > 5 else ""
        if not jp:
            continue
        if tid_val.isdigit() and int(tid_val) in non_acting_ids:
            del_rows.add(row_idx)
            non_acting_found.append((row_idx, jp, int(tid_val)))
        elif jp in _NON_AV_NAMES:
            del_rows.add(row_idx)
            non_acting_found.append((row_idx, jp, int(tid_val) if tid_val.isdigit() else 0))
        elif not tid_val.isdigit() and jp in _DESC_ROWS:
            del_rows.add(row_idx)
            desc_found.append((row_idx, jp))
        else:
            # 第 4 类：占位/冗余行——前 4 列完全相同（日文/中文/繁体/别名同值），
            # 且无有效信息（第 5 列链接可有值、第 8 列生日可有可无，第 6/7/9 列
            # tmdbid/tmdb url/简介 必须全空）。仅名字+来源链接无任何数据的行无意义。
            zh = str(row[1] or "").strip() if len(row) > 1 else ""
            zt = str(row[2] or "").strip() if len(row) > 2 else ""
            kw = str(row[3] or "").strip() if len(row) > 3 else ""
            tid = str(row[5] or "").strip() if len(row) > 5 else ""
            url = str(row[6] or "").strip() if len(row) > 6 else ""
            bio = str(row[8] or "").strip() if len(row) > 8 else ""
            if jp == zh == zt == kw and not (tid or url or bio):
                del_rows.add(row_idx)
                placeholder_found.append((row_idx, jp))

    print(f"第 1 类（导演/幕后）待删: {len(non_acting_found)}")
    print(f"第 2 类（描述词）待删: {len(desc_found)}")
    print(f"第 4 类（占位/冗余）待删: {len(placeholder_found)}")
    print(f"合计: {len(del_rows)} 行")

    if not args.apply:
        print("\n（预览模式，加 --apply 执行删除）")
        wb.close()
        return 0

    # 系统临时目录保留原始库副本（工作区安全网，不进仓库）
    shutil.copy(db_path, Path(tempfile.gettempdir()) / "actor_db_before_clean_nonactors.xlsx")

    # 待删行合并为连续区间，一次 delete_rows(start, count) 批量删除
    # （逐行 delete_rows 每次 O(n) 移动，几百行会超时且中途 kill 会写坏文件）
    sorted_rows = sorted(del_rows, reverse=True)
    intervals: list[tuple[int, int]] = []  # (start, count)
    for row_idx in sorted_rows:
        if intervals and intervals[-1][0] - 1 == row_idx:
            intervals[-1] = (row_idx, intervals[-1][1] + 1)
        else:
            intervals.append((row_idx, 1))
    for start, count in intervals:
        ws.delete_rows(start, count)

    # openpyxl delete_rows 会在末尾保留带格式/超链接的空行，导致 max_row 延伸出空行，
    # 需二次清理（从后往前删 jp 为空的物理行）
    while ws.max_row >= 2:
        last = ws.cell(row=ws.max_row, column=1).value
        if last is None or str(last).strip() == "":
            ws.delete_rows(ws.max_row, 1)
        else:
            break

    # openpyxl delete_rows 移动单元格值时不会同步 hyperlink 的 ref，导致下方行所有
    # 超链接 ref 错位、被删行超链接残留（孤儿 hyperlink）。这里按 cell 实际坐标
    # 重建全部超链接：ref 跟随绑定 cell 的真实位置，孤儿归零。
    _rebuild_hyperlinks(ws)

    wb.save(db_path)
    wb.close()
    print(f"✅ 已删除 {len(del_rows)} 行")
    ok = validate_after_save(db_path)
    if not ok:
        print("⚠️ 保存后校验发现 error 级问题，请检查！")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
